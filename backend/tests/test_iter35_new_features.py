"""Iteration 35 - Intercloud Portal new features regression.

Covers:
- 2FA TOTP end-to-end (setup, enable, login-2fa, recovery replay, disable, admin reset)
- Admin impersonation
- Lead Form Builder + public form submit (kontak)
- IP Pool allocation + utilization
- Ticket internal notes visibility + timeline + view=active/archive
- Server-side pagination backward compat
- Quotation convert-to-invoice idempotency
- Invoice send
- Backup trigger + download
- UTM links CRUD
- Auto follow-up for new lead
- Excel formulas in finance reports
- Auto-renew toggle
- Order edit provision_log
"""
import io
import os
import time
import pyotp
import pytest
import requests

def _read_env():
    p = "/app/frontend/.env"
    if os.path.exists(p):
        for line in open(p):
            if line.startswith("REACT_APP_BACKEND_URL="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("REACT_APP_BACKEND_URL", "")

BASE = (os.environ.get("REACT_APP_BACKEND_URL") or _read_env()).rstrip("/")
assert BASE, "REACT_APP_BACKEND_URL required"
API = BASE + "/api/portal"
ADMIN = ("admin@intercloud-digital.com", "AdminIntercloud2026!")
CLIENT = ("demo@client.com", "ClientDemo2026!")


def _login(email, password, code=None):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=30)
    assert r.status_code == 200, f"login failed {r.status_code}: {r.text}"
    j = r.json()
    if j.get("require_2fa"):
        assert code, "TOTP code required"
        r2 = requests.post(f"{API}/auth/login/2fa",
                           json={"mfa_token": j["mfa_token"], "code": code}, timeout=30)
        assert r2.status_code == 200, r2.text
        return r2.json()["token"]
    return j["token"]


def _h(tok):
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def admin_token():
    return _login(*ADMIN)


@pytest.fixture(scope="module")
def client_token():
    return _login(*CLIENT)


# ---------------- 2FA ----------------

def test_2fa_full_flow_client(admin_token):
    # Login as client (no 2FA yet)
    tok = _login(*CLIENT)
    # Setup
    r = requests.post(f"{API}/auth/2fa/setup", headers=_h(tok), timeout=15)
    assert r.status_code == 200, r.text
    secret = r.json()["secret"]
    assert r.json().get("qr", "").startswith("data:image")
    totp = pyotp.TOTP(secret)

    # Verify-enable
    r = requests.post(f"{API}/auth/2fa/verify-enable",
                      headers=_h(tok), json={"code": totp.now()}, timeout=15)
    assert r.status_code == 200, r.text
    rc = r.json()["recovery_codes"]
    assert isinstance(rc, list) and len(rc) == 10

    # Login now returns require_2fa
    r = requests.post(f"{API}/auth/login",
                      json={"email": CLIENT[0], "password": CLIENT[1]}, timeout=15)
    assert r.status_code == 200
    body = r.json()
    assert body.get("require_2fa") is True
    mfa_tok = body["mfa_token"]

    # Invalid code 000000
    r = requests.post(f"{API}/auth/login/2fa",
                      json={"mfa_token": mfa_tok, "code": "000000"}, timeout=15)
    assert r.status_code == 401

    # Valid TOTP
    time.sleep(1)
    r = requests.post(f"{API}/auth/login/2fa",
                      json={"mfa_token": mfa_tok, "code": totp.now()}, timeout=15)
    assert r.status_code == 200, r.text
    full_tok = r.json()["token"]
    # Sanity /auth/me
    me = requests.get(f"{API}/auth/me", headers=_h(full_tok), timeout=15)
    assert me.status_code == 200

    # Recovery code single-use
    # need new mfa_token
    r = requests.post(f"{API}/auth/login",
                      json={"email": CLIENT[0], "password": CLIENT[1]}, timeout=15)
    mfa_tok2 = r.json()["mfa_token"]
    r = requests.post(f"{API}/auth/login/2fa",
                      json={"mfa_token": mfa_tok2, "code": rc[0]}, timeout=15)
    assert r.status_code == 200, f"recovery should work first time: {r.text}"

    # Reuse same recovery code → should fail
    r = requests.post(f"{API}/auth/login",
                      json={"email": CLIENT[0], "password": CLIENT[1]}, timeout=15)
    mfa_tok3 = r.json()["mfa_token"]
    r = requests.post(f"{API}/auth/login/2fa",
                      json={"mfa_token": mfa_tok3, "code": rc[0]}, timeout=15)
    assert r.status_code == 401, "recovery code replay must be rejected"

    # Disable via TOTP
    time.sleep(1)
    r = requests.post(f"{API}/auth/2fa/disable",
                      headers=_h(full_tok), json={"code": totp.now()}, timeout=15)
    assert r.status_code == 200, r.text

    # Admin reset (idempotent even when off)
    # find client uid
    r = requests.get(f"{API}/admin/users", headers=_h(admin_token), timeout=15)
    uid = next(u["id"] for u in r.json() if u["email"] == CLIENT[0])
    r = requests.post(f"{API}/admin/users/{uid}/reset-2fa",
                      headers=_h(admin_token), timeout=15)
    assert r.status_code == 200


# ---------------- Impersonation ----------------

def test_admin_impersonate(admin_token):
    r = requests.get(f"{API}/admin/users", headers=_h(admin_token), timeout=15)
    users = r.json()
    client_uid = next(u["id"] for u in users if u["email"] == CLIENT[0])
    # Try to find a staff (non-client) - admin himself is staff
    staff_uid = next((u["id"] for u in users if u.get("role") and u["role"] != "client"), None)

    r = requests.post(f"{API}/admin/users/{client_uid}/impersonate",
                      headers=_h(admin_token), timeout=15)
    assert r.status_code == 200, r.text
    imp_tok = r.json()["token"]
    me = requests.get(f"{API}/auth/me", headers=_h(imp_tok), timeout=15)
    assert me.status_code == 200
    assert me.json()["email"] == CLIENT[0]

    # Staff must be rejected
    if staff_uid:
        r = requests.post(f"{API}/admin/users/{staff_uid}/impersonate",
                          headers=_h(admin_token), timeout=15)
        assert r.status_code == 400, r.text


# ---------------- Form Builder + Public form ----------------

def test_form_builder_and_public_submit(admin_token):
    r = requests.get(f"{API}/admin/form-builder", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    forms = r.json()
    kontak = next((f for f in forms if f["slug"] == "kontak"), None)
    assert kontak, "form 'kontak' should be seeded"

    # Update: add a field
    fields = list(kontak["fields"])
    fields.append({"key": "budget", "label": "Budget", "type": "text",
                   "required": False, "order": 99})
    r = requests.put(f"{API}/admin/form-builder/{kontak['id']}",
                     headers=_h(admin_token), json={"fields": fields}, timeout=15)
    assert r.status_code == 200, r.text
    assert any(f["key"] == "budget" for f in r.json()["fields"])

    # Public config
    r = requests.get(f"{API}/portal-public/forms/kontak", timeout=15)
    assert r.status_code == 200
    assert r.json()["slug"] == "kontak"

    # Invalid submit (bad email + missing required name)
    r = requests.post(f"{API}/portal-public/forms/kontak/submit",
                      json={"email": "invalid-email"}, timeout=15)
    assert r.status_code == 422, r.text
    body = r.json()
    errs = body.get("detail", {}).get("errors", {})
    assert "email" in errs
    assert "name" in errs

    # Valid submit
    uniq = f"TEST_lead_{int(time.time())}@example.com"
    r = requests.post(f"{API}/portal-public/forms/kontak/submit",
                      json={"name": "TEST Lead", "email": uniq,
                            "phone": "081234567890", "message": "hai"}, timeout=15)
    assert r.status_code == 200, r.text
    lead_id = r.json()["lead_id"]

    # Verify lead source form:kontak
    r = requests.get(f"{API}/admin/leads", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    hit = [l for l in r.json() if l.get("id") == lead_id]
    assert hit, "lead not found"
    assert hit[0].get("source") == "form:kontak"

    # Verify CRM prospect
    # not all APIs expose crm search; skip if not present
    # Verify auto follow-up
    r = requests.get(f"{API}/admin/followups", headers=_h(admin_token), timeout=15)
    # public /portal-public/leads (older endpoint) is what creates followups per requirements,
    # form-builder submit path does not. Test that path below.


def test_public_leads_creates_followup(admin_token):
    uniq_name = f"TEST FollowUpLead {int(time.time())}"
    r = requests.post(f"{API}/portal-public/leads",
                      json={"name": uniq_name, "email": f"TEST_{int(time.time())}@ex.com",
                            "phone": "081234567890", "company": "PT Test",
                            "need": "vps", "message": "hai"}, timeout=15)
    assert r.status_code == 200, r.text
    time.sleep(0.5)
    r = requests.get(f"{API}/admin/followups", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    tasks = [f.get("task", "") for f in r.json()]
    assert any(t.startswith("Follow up lead baru") and uniq_name in t for t in tasks), tasks[:5]


# ---------------- IP Pool ----------------

def test_dcim_ip_pool_allocate(admin_token):
    r = requests.get(f"{API}/admin/dcim/prefixes", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    v4 = [p for p in r.json() if p.get("family") == 4 or ":" not in p.get("prefix", "")]
    assert v4, "no IPv4 prefix"
    pid = v4[0]["id"]

    r1 = requests.post(f"{API}/admin/dcim/prefixes/{pid}/allocate",
                       headers=_h(admin_token),
                       json={"hostname": "TEST-host1", "customer": "TEST"}, timeout=15)
    assert r1.status_code == 200, r1.text
    ip1 = r1.json()["address"]

    r2 = requests.post(f"{API}/admin/dcim/prefixes/{pid}/allocate",
                       headers=_h(admin_token),
                       json={"hostname": "TEST-host2", "customer": "TEST"}, timeout=15)
    assert r2.status_code == 200, r2.text
    ip2 = r2.json()["address"]
    assert ip1 != ip2

    r = requests.get(f"{API}/admin/dcim/prefixes/{pid}/utilization",
                     headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    assert "utilization_pct" in r.json()


# ---------------- Ticket internal notes ----------------

def test_ticket_internal_note_visibility(admin_token, client_token):
    # get client ticket list
    r = requests.get(f"{API}/client/tickets", headers=_h(client_token), timeout=15)
    assert r.status_code == 200
    tickets = r.json()
    if not tickets:
        pytest.skip("no client tickets available")
    tid = tickets[0]["id"]

    marker = f"TEST_INTERNAL_{int(time.time())}"
    r = requests.post(f"{API}/admin/tickets/{tid}/replies",
                      headers=_h(admin_token),
                      json={"message": marker, "internal": True}, timeout=15)
    assert r.status_code == 200, r.text

    # Client must NOT see internal reply
    r = requests.get(f"{API}/client/tickets", headers=_h(client_token), timeout=15)
    ct = next(t for t in r.json() if t["id"] == tid)
    assert not any(marker in rr.get("message", "") for rr in ct.get("replies", [])), \
        "internal note leaked to client"

    # Timeline shows internal_note event
    r = requests.get(f"{API}/admin/tickets/{tid}/timeline",
                     headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    kinds = {e["kind"] for e in r.json()["events"]}
    assert "internal_note" in kinds
    assert "created" in kinds

    # view=active vs archive
    r_active = requests.get(f"{API}/admin/tickets?view=active",
                            headers=_h(admin_token), timeout=15).json()
    r_archive = requests.get(f"{API}/admin/tickets?view=archive",
                             headers=_h(admin_token), timeout=15).json()
    # active returns list (no page param)
    assert isinstance(r_active, list)
    assert all(t["status"] != "closed" for t in r_active)
    assert all(t["status"] == "closed" for t in r_archive) if r_archive else True


# ---------------- Pagination ----------------

def test_pagination_backward_compat(admin_token):
    plain = requests.get(f"{API}/admin/invoices", headers=_h(admin_token), timeout=15)
    assert plain.status_code == 200
    assert isinstance(plain.json(), list)

    paged = requests.get(f"{API}/admin/invoices?page=1&limit=5",
                         headers=_h(admin_token), timeout=15)
    assert paged.status_code == 200
    j = paged.json()
    assert isinstance(j, dict)
    for k in ("items", "total", "page", "pages"):
        assert k in j
    assert len(j["items"]) <= 5


# ---------------- Quotation convert-to-invoice ----------------

def test_quotation_convert_idempotent(admin_token):
    r = requests.get(f"{API}/admin/quotations", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    qs = r.json() if isinstance(r.json(), list) else r.json().get("items", [])
    unconv = [q for q in qs if not q.get("converted_invoice_id")]
    if not unconv:
        pytest.skip("no unconverted quotation available")
    qid = unconv[0]["id"]

    r = requests.post(f"{API}/admin/quotations/{qid}/convert-to-invoice",
                      headers=_h(admin_token), json={}, timeout=30)
    assert r.status_code == 200, r.text
    iid = r.json()["id"]

    r2 = requests.post(f"{API}/admin/quotations/{qid}/convert-to-invoice",
                       headers=_h(admin_token), json={}, timeout=15)
    assert r2.status_code == 400, "second convert must be rejected"

    r = requests.get(f"{API}/admin/invoices/{iid}", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200


# ---------------- Invoice send ----------------

def test_invoice_send(admin_token):
    r = requests.get(f"{API}/admin/invoices?page=1&limit=1",
                     headers=_h(admin_token), timeout=15)
    items = r.json()["items"]
    if not items:
        pytest.skip("no invoices")
    iid = items[0]["id"]
    r = requests.post(f"{API}/admin/invoices/{iid}/send",
                      headers=_h(admin_token), timeout=30)
    assert r.status_code == 200, r.text
    j = r.json()
    assert j.get("ok") is True
    assert "email_sent" in j
    assert "wa_link" in j


# ---------------- Backup ----------------

def test_backup_trigger_and_download(admin_token):
    r = requests.post(f"{API}/admin/backup/trigger", headers=_h(admin_token), timeout=120)
    assert r.status_code == 200, r.text
    bid = r.json()["id"]

    r = requests.get(f"{API}/admin/backup/history", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    assert any(b.get("id") == bid for b in r.json())

    r = requests.get(f"{API}/admin/backup/history/{bid}/download",
                     headers=_h(admin_token), timeout=60)
    assert r.status_code == 200
    assert r.headers.get("content-type", "").startswith("application/gzip") or \
           "gzip" in r.headers.get("content-type", ""), r.headers.get("content-type")


# ---------------- UTM links ----------------

def test_utm_links_crud(admin_token):
    # invalid
    r = requests.post(f"{API}/admin/utm-links",
                      headers=_h(admin_token), json={"url": "no-scheme.com"}, timeout=15)
    assert r.status_code == 400

    r = requests.post(f"{API}/admin/utm-links",
                      headers=_h(admin_token),
                      json={"url": "https://intercloud.com/?utm_source=test",
                            "label": "TEST_utm"}, timeout=15)
    assert r.status_code == 200
    lid = r.json()["id"]

    r = requests.get(f"{API}/admin/utm-links", headers=_h(admin_token), timeout=15)
    assert any(x["id"] == lid for x in r.json())

    r = requests.delete(f"{API}/admin/utm-links/{lid}", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200


# ---------------- Excel formulas ----------------

def test_finance_reports_have_formulas(admin_token):
    from openpyxl import load_workbook
    for url in (f"{API}/admin/finance/report/monthly/2026-07",
                f"{API}/admin/finance/report/annual/2026"):
        r = requests.get(url, headers=_h(admin_token), timeout=60)
        assert r.status_code == 200, url
        wb = load_workbook(io.BytesIO(r.content))
        has_formula = False
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("=SUM"):
                        has_formula = True
                        break
                if has_formula:
                    break
            if has_formula:
                break
        assert has_formula, f"no =SUM formula found in {url}"


# ---------------- Auto-renew ----------------

def test_auto_renew_toggle(client_token):
    r = requests.get(f"{API}/client/services", headers=_h(client_token), timeout=15)
    assert r.status_code == 200
    svcs = r.json() if isinstance(r.json(), list) else r.json().get("items", [])
    if not svcs:
        pytest.skip("no services")
    sid = svcs[0]["id"]
    try:
        r = requests.put(f"{API}/client/services/{sid}/auto-renew",
                         headers=_h(client_token), json={"auto_renew": False}, timeout=15)
        assert r.status_code == 200, r.text
        r = requests.get(f"{API}/client/services/{sid}",
                         headers=_h(client_token), timeout=15)
        assert r.status_code == 200
        assert r.json().get("auto_renew") is False
    finally:
        # restore
        requests.put(f"{API}/client/services/{sid}/auto-renew",
                     headers=_h(client_token), json={"auto_renew": True}, timeout=15)


# ---------------- Order edit ----------------

def test_admin_order_edit_provision_log(admin_token):
    r = requests.get(f"{API}/admin/orders", headers=_h(admin_token), timeout=15)
    assert r.status_code == 200
    orders = r.json() if isinstance(r.json(), list) else r.json().get("items", [])
    if not orders:
        pytest.skip("no orders")
    oid = orders[0]["id"]
    r = requests.put(f"{API}/admin/orders/{oid}",
                     headers=_h(admin_token),
                     json={"notes": f"TEST edit {int(time.time())}"}, timeout=15)
    assert r.status_code == 200, r.text
    log = r.json().get("provision_log") or []
    assert any(e.get("step") == "order_updated" for e in log), log[-3:]

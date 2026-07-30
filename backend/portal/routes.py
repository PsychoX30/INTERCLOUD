"""All portal routes (auth + client + admin) under /api/portal."""
import os
import secrets
import re
from urllib.parse import quote


from fastapi import APIRouter, Depends, HTTPException, Request
from bson import ObjectId
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from . import models as m
from .auth import (
    verify_password, hash_password, create_access_token,
    get_current_user, get_current_admin, get_current_staff, get_current_content,
    require_roles, sales_can_access,
    STAFF_ROLES, FINANCE_ROLES, BILLING_ROLES, CATALOG_ROLES,
    OPS_ROLES, USER_MGMT_ROLES, TICKET_ROLES, CONTENT_ROLES,
)
from .audit import log_audit, serialize as _serialize_audit

router = APIRouter(prefix="/api/portal")


# ---------- helpers ----------
def _iso(dt: datetime | str) -> str:
    if isinstance(dt, str):
        return dt
    return dt.isoformat()


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _user_public(u: dict) -> dict:
    return {
        "id": str(u["_id"]) if "_id" in u else u["id"],
        "email": u["email"],
        "name": u.get("name", ""),
        "role": u["role"],
        "company": u.get("company"),
        "phone": u.get("phone"),
        "created_at": _iso(u.get("created_at", _now())),
        "assigned_client_ids": [str(x) for x in (u.get("assigned_client_ids") or [])],
        "twofa_enabled": bool(u.get("totp_enabled")),
        "billing_emails": list(u.get("billing_emails") or []),
        "attention": u.get("attention"),
        "address_line1": u.get("address_line1"),
        "address_line2": u.get("address_line2"),
        "city": u.get("city"),
        "province": u.get("province"),
        "postal_code": u.get("postal_code"),
        "country": u.get("country") or "Indonesia",
        "npwp": u.get("npwp"),
        "menu_keys": u.get("menu_keys"),
        "feature_flags": list(u.get("feature_flags") or []),
        "is_active": u.get("is_active", True),
        "must_change_password": bool(u.get("must_change_password", False)),
    }


async def _get_db():
    from server import db
    return db


async def _load_user(db, user_id: str) -> dict:
    try:
        u = await db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        u = None
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    return u


def _oid(id_str: str) -> ObjectId:
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid id: {id_str}")


def _sales_scope_filter(staff: dict, key: str = "user_id") -> dict:
    """Return an extra Mongo filter clause that restricts sales to their
    assigned clients. Returns {} for non-sales roles (no restriction).

    For sales with an empty assignment list we deliberately match nothing so
    the endpoint returns an empty list rather than leaking global data.
    """
    if staff.get("role") != "sales":
        return {}
    assigned = staff.get("assigned_client_ids") or []
    if not assigned:
        return {"_id": None}  # matches nothing
    return {key: {"$in": [ObjectId(x) for x in assigned]}}


async def _sales_visible_crm_ids(db, staff: dict) -> list | None:
    """Return the list of crm_customers._id that a sales staff can access.

    Returns None if the staff is not a sales role (i.e., no restriction).
    For sales, we consider a CRM row "theirs" if its `user_id` is in their
    assigned_client_ids. This is what powers scoping on the Follow-ups table.
    """
    if staff.get("role") != "sales":
        return None
    assigned = [ObjectId(x) for x in (staff.get("assigned_client_ids") or [])]
    if not assigned:
        return []
    cur = db.crm_customers.find({"user_id": {"$in": assigned}}, {"_id": 1})
    return [d["_id"] async for d in cur]


async def _next_number(db, coll: str, prefix: str) -> str:
    """Race-safe sequential document numbering via an atomic counter doc.
    The old count_documents()+1 approach produced duplicate numbers (and 500s
    on the unique index) under concurrent creation."""
    from pymongo import ReturnDocument
    year = datetime.now(timezone.utc).year
    key = f"number:{coll}"
    doc = await db.counters.find_one_and_update(
        {"_id": key}, {"$inc": {"seq": 1}},
        upsert=True, return_document=ReturnDocument.AFTER,
    )
    seq = int(doc["seq"])
    if seq == 1:
        # First use on an existing dataset: fast-forward past legacy numbers.
        last = await db[coll].find_one({"number": {"$regex": f"^{prefix}-"}},
                                       sort=[("number", -1)])
        if last:
            try:
                legacy = int(str(last.get("number", "")).rsplit("-", 1)[-1])
            except (TypeError, ValueError):
                legacy = await db[coll].count_documents({})
            if legacy >= seq:
                doc = await db.counters.find_one_and_update(
                    {"_id": key}, {"$max": {"seq": legacy + 1}},
                    return_document=ReturnDocument.AFTER,
                )
                seq = int(doc["seq"])
    return f"{prefix}-{year}-{seq:05d}"


async def _insert_numbered(db, coll: str, prefix: str, doc: dict):
    """Insert a document carrying a unique sequential `number`.
    Retries with a fresh number on the (rare) DuplicateKeyError - e.g. when a
    DB restore rolls the counter back underneath concurrent writers."""
    from pymongo.errors import DuplicateKeyError
    for _ in range(5):
        doc["number"] = await _next_number(db, coll, prefix)
        try:
            r = await db[coll].insert_one(doc)
            doc["_id"] = r.inserted_id
            return doc
        except DuplicateKeyError:
            # Fast-forward the counter past the current max and retry
            last = await db[coll].find_one({"number": {"$regex": f"^{prefix}-"}},
                                           sort=[("number", -1)])
            if last:
                try:
                    legacy = int(str(last.get("number", "")).rsplit("-", 1)[-1])
                    await db.counters.update_one({"_id": f"number:{coll}"},
                                                 {"$max": {"seq": legacy}}, upsert=True)
                except (TypeError, ValueError):
                    pass
            doc.pop("_id", None)
    raise HTTPException(status_code=500, detail=f"Could not allocate a unique {prefix} number")


async def _mark_overdue(db):
    """Auto-mark unpaid invoices past due as 'overdue'."""
    today = datetime.now(timezone.utc).date().isoformat()
    await db.invoices.update_many(
        {"status": "unpaid", "due_date": {"$lt": today}},
        {"$set": {"status": "overdue"}},
    )


# ---------- Global settings helpers (settings collection: {key, value}) ----------
async def _get_setting_value(db, key: str, default=None):
    doc = await db.settings.find_one({"key": key})
    return doc.get("value") if doc and "value" in doc else default


async def _set_setting_value(db, key: str, value) -> None:
    await db.settings.update_one(
        {"key": key},
        {"$set": {"key": key, "value": value, "updated_at": _now()}},
        upsert=True,
    )


# Billing defaults - `default_tax_percent` is only ever a SUGGESTED initial
# value pre-filled into invoice/quotation forms and the renewal auto-invoice
# generator. It is stored per-document and stays fully manual/overridable
# (down to 0). Nothing ever recalculates tax_percent after creation.
BILLING_SETTING_DEFAULTS = {
    "default_tax_percent": 11.0,
    "renewal_lead_days": 7,
    "enable_extra_payment_gateways": False,
    "noc_alert_recipients": [],
}

# Extra gateways stay implemented (classes untouched in integrations_v2) but
# hidden from all admin/client surfaces unless the flag above is turned on.
_EXTRA_PAYMENT_MODULES = {"midtrans", "xendit"}


# ============================================================
# AUTH
# ============================================================
@router.get("/auth/config")
async def auth_config():
    """Public config exposed to unauthenticated login/register pages.

    Frontend uses this to decide whether to load the Google reCAPTCHA v3
    script and which site_key to pass to `grecaptcha.execute()`.
    Secrets are never included here.
    """
    db = await _get_db()
    from portal import integrations_v2 as _iv2
    doc = await _iv2.get_recaptcha_settings(db)
    if not doc:
        return {"recaptcha": {"enabled": False, "site_key": None}}
    site_key = ((doc.get("credentials") or {}).get("site_key") or "").strip()
    return {
        "recaptcha": {
            "enabled": bool(site_key),
            "site_key": site_key or None,
        }
    }


async def _log_login_attempt(db, *, email: str, action: str, success: bool, reason: str,
                             ip: str, user_agent: str = "", recaptcha_score: float | None = None,
                             recaptcha_enabled: bool = False):
    """Append a document to `login_attempts` for the Security Analytics dashboard.
    Best-effort - never raise into the caller. On failure, also runs auto-block check."""
    try:
        await db.login_attempts.insert_one({
            "email": (email or "").lower(),
            "action": action,
            "success": bool(success),
            "reason": reason,
            "ip": ip or "unknown",
            "user_agent": user_agent[:400],
            "recaptcha_enabled": bool(recaptcha_enabled),
            "recaptcha_score": recaptcha_score,
            "created_at": _now(),
        })
        if not success and ip and ip != "unknown":
            await _maybe_auto_block(db, ip)
    except Exception:
        import logging
        logging.getLogger("portal.security").warning("[login_attempts] insert failed for %s", email)


DEFAULT_SECURITY_SETTINGS = {
    "auto_block_enabled": True,
    "fail_threshold": 10,        # failures to trigger a block
    "window_minutes": 15,        # sliding window to count failures
    "ban_minutes": 30,           # block duration
    "notify_emails": [],         # recipients for block notifications
    "whitelist_ips": [],         # IPs / CIDRs that are never blocked
    "email_notify_enabled": True,
    "telegram_notify_enabled": True,
}


def _ip_in_whitelist(ip: str, whitelist: list[str]) -> bool:
    """Return True if `ip` matches any entry in the whitelist. Entries may be
    exact IPs, CIDR ranges, or hostnames (exact string match fallback)."""
    if not ip or not whitelist:
        return False
    import ipaddress as _ipaddr
    try:
        target = _ipaddr.ip_address(ip)
    except ValueError:
        target = None
    for raw in whitelist:
        entry = (raw or "").strip()
        if not entry:
            continue
        if entry == ip:
            return True
        if target is None:
            continue
        try:
            if "/" in entry:
                if target in _ipaddr.ip_network(entry, strict=False):
                    return True
            else:
                if target == _ipaddr.ip_address(entry):
                    return True
        except ValueError:
            continue
    return False


async def _get_security_settings(db) -> dict:
    doc = await db.settings.find_one({"_id": "security"})
    if not doc:
        return dict(DEFAULT_SECURITY_SETTINGS)
    merged = dict(DEFAULT_SECURITY_SETTINGS)
    merged.update({k: v for k, v in doc.items() if k != "_id"})
    return merged


async def _maybe_auto_block(db, ip: str):
    """After each failed login, check if this IP has crossed the threshold and,
    if so, upsert a `blocked_ips` doc + emit a `security_notifications` event."""
    s = await _get_security_settings(db)
    if not s.get("auto_block_enabled", True):
        return
    if _ip_in_whitelist(ip, s.get("whitelist_ips") or []):
        return
    window_iso = (datetime.now(timezone.utc) - timedelta(minutes=int(s["window_minutes"]))).isoformat()
    fails = await db.login_attempts.count_documents({
        "ip": ip, "success": False, "created_at": {"$gte": window_iso},
    })
    if fails < int(s["fail_threshold"]):
        return
    now_dt = datetime.now(timezone.utc)
    expires = now_dt + timedelta(minutes=int(s["ban_minutes"]))
    # Only insert a notification if this IP wasn't already actively blocked
    existing = await db.blocked_ips.find_one({"ip": ip})
    existing_exp = existing.get("expires_at") if existing else None
    # Normalize both string ISO and naive-datetime forms to offset-aware.
    if isinstance(existing_exp, str):
        try:
            existing_exp = datetime.fromisoformat(existing_exp.replace("Z", "+00:00"))
        except Exception:
            existing_exp = None
    if isinstance(existing_exp, datetime) and existing_exp.tzinfo is None:
        existing_exp = existing_exp.replace(tzinfo=timezone.utc)
    if existing and existing_exp and existing_exp > now_dt and not existing.get("unblocked_at"):
        # Extend the ban by another window
        await db.blocked_ips.update_one({"ip": ip}, {"$set": {
            "expires_at": expires, "hits": fails, "last_seen_at": now_dt.isoformat(),
        }})
        return
    await db.blocked_ips.update_one(
        {"ip": ip},
        {"$set": {
            "ip": ip,
            "blocked_at": now_dt.isoformat(),
            "expires_at": expires,
            "reason": "auto_block_threshold",
            "hits": fails,
            "unblocked_at": None,
        }},
        upsert=True,
    )
    await db.security_notifications.insert_one({
        "kind": "ip_auto_blocked",
        "ip": ip,
        "hits": fails,
        "window_minutes": int(s["window_minutes"]),
        "ban_minutes": int(s["ban_minutes"]),
        "created_at": now_dt.isoformat(),
        "read": False,
    })
    # Fire-and-forget alerts - never let a mail/Telegram outage break /auth/login
    try:
        import asyncio as _asyncio
        _asyncio.create_task(_dispatch_block_alerts(db, ip, fails, int(s["ban_minutes"]), s))
    except Exception:
        pass


async def _dispatch_block_alerts(db, ip: str, hits: int, ban_minutes: int, settings: dict):
    """Best-effort: send email(s) via SMTP + a Telegram DM.
    Runs in the background - failures are swallowed."""
    import logging
    log = logging.getLogger("portal.security")
    from portal import integrations_v2 as _iv2

    now_iso = datetime.now(timezone.utc).isoformat()
    subject = f"[Security] IP {ip} auto-blocked - {hits} failed logins"
    text = (f"An IP has been auto-blocked by the portal.\n\n"
            f"IP:      {ip}\n"
            f"Hits:    {hits} failed logins within window\n"
            f"Blocked: {ban_minutes} minute(s)\n"
            f"Time:    {now_iso}\n\n"
            f"Unblock via Admin ▸ Security ▸ Blocked IPs, or DELETE "
            f"/api/portal/admin/security/blocked-ips/{ip}")
    html = (f"<h3>IP auto-blocked</h3>"
            f"<p>An IP has been auto-blocked by the portal.</p>"
            f"<ul>"
            f"<li><b>IP:</b> <code>{ip}</code></li>"
            f"<li><b>Failed logins:</b> {hits}</li>"
            f"<li><b>Blocked for:</b> {ban_minutes} minute(s)</li>"
            f"<li><b>Time:</b> {now_iso}</li>"
            f"</ul>"
            f"<p>Unblock via <b>Admin ▸ Security ▸ Blocked IPs</b>.</p>")

    # Email dispatch
    if settings.get("email_notify_enabled", True):
        recipients = [r for r in (settings.get("notify_emails") or []) if r]
        smtp_doc = await _iv2.get_settings(db, "smtp")
        if smtp_doc and smtp_doc.get("enabled") and recipients:
            try:
                mailer = _iv2.SMTPMailer(smtp_doc)
                loop = __import__("asyncio").get_event_loop()
                for to in recipients:
                    try:
                        await loop.run_in_executor(None, lambda t=to: mailer.send(
                            to=t, subject=subject, html=html, text=text))
                    except Exception as e:
                        log.warning("[security] email to %s failed: %s", to, e)
            except Exception as e:
                log.warning("[security] SMTP init failed: %s", e)

    # Telegram dispatch
    if settings.get("telegram_notify_enabled", True):
        tg_doc = await _iv2.get_telegram_settings(db)
        if tg_doc:
            try:
                tg = _iv2.TelegramNotifier(tg_doc)
                await tg.send(
                    f"*⛔ IP auto-blocked*\n"
                    f"`{ip}` after *{hits}* failed logins\n"
                    f"Ban duration: *{ban_minutes} min*"
                )
            except Exception as e:
                log.warning("[security] Telegram send failed: %s", e)


async def _is_ip_blocked(db, ip: str) -> bool:
    if not ip or ip == "unknown":
        return False
    # Whitelist short-circuit - makes the guard idempotent even if a stale
    # block doc still exists for an IP that was just added to the whitelist.
    s = await _get_security_settings(db)
    if _ip_in_whitelist(ip, s.get("whitelist_ips") or []):
        return False
    now_dt = datetime.now(timezone.utc)
    doc = await db.blocked_ips.find_one({"ip": ip, "unblocked_at": None})
    if not doc:
        return False
    expires_at = doc.get("expires_at")
    # Support both ISO strings and datetime objects
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
        except Exception:
            return False
    # MongoDB returns naive UTC datetimes - normalize to offset-aware.
    if isinstance(expires_at, datetime) and expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at and expires_at > now_dt:
        return True
    return False


from portal.security import (
    limiter as _rl_limiter,
    AUTH_LOGIN_LIMIT, AUTH_REGISTER_LIMIT,
    AUTH_FORGOT_LIMIT, AUTH_RESET_LIMIT, PUBLIC_STATUS_LIMIT,
)


@router.post("/auth/login", response_model=m.LoginOut)
@_rl_limiter.limit(AUTH_LOGIN_LIMIT)
async def login(payload: m.LoginIn, request: Request):
    db = await _get_db()
    from portal import integrations_v2 as _iv2
    from portal.security import _client_ip as _rl_client_ip
    ip = _rl_client_ip(request)
    ua = request.headers.get("user-agent", "")
    email = payload.email.lower().strip()

    # Auto-block short-circuit
    if await _is_ip_blocked(db, ip):
        raise HTTPException(status_code=429, detail="IP temporarily blocked due to repeated failures")

    recap_doc = await _iv2.get_recaptcha_settings(db)
    recap_score = None

    if recap_doc:
        try:
            result = await _iv2.RecaptchaV3Verifier(recap_doc).verify(
                payload.recaptcha_token, "login", ip
            )
            recap_score = float(result.get("score", 0.0))
        except HTTPException as e:
            reason = ("recaptcha_missing" if "Missing" in str(e.detail)
                      else "recaptcha_low_score" if getattr(e, "status_code", 0) == 403
                      else "recaptcha_failed")
            await _log_login_attempt(db, email=email, action="login", success=False, reason=reason,
                                     ip=ip, user_agent=ua, recaptcha_enabled=True,
                                     recaptcha_score=recap_score)
            raise

    u = await db.users.find_one({"email": email})
    if not u or not verify_password(payload.password, u["password_hash"]):
        await _log_login_attempt(db, email=email, action="login", success=False,
                                 reason="invalid_credentials", ip=ip, user_agent=ua,
                                 recaptcha_enabled=bool(recap_doc), recaptcha_score=recap_score)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if u.get("totp_enabled"):
        from portal import twofa as _tf
        await _log_login_attempt(db, email=email, action="login", success=True, reason="mfa_challenge",
                                 ip=ip, user_agent=ua, recaptcha_enabled=bool(recap_doc),
                                 recaptcha_score=recap_score)
        return {"require_2fa": True, "mfa_token": _tf.make_mfa_token(str(u["_id"]))}
    token = create_access_token(str(u["_id"]), u["email"], u["role"])
    await _log_login_attempt(db, email=email, action="login", success=True, reason="ok",
                             ip=ip, user_agent=ua, recaptcha_enabled=bool(recap_doc),
                             recaptcha_score=recap_score)
    return {"token": token, "user": _user_public(u)}


@router.post("/auth/login/2fa", response_model=m.LoginOut)
@_rl_limiter.limit(AUTH_LOGIN_LIMIT)
async def login_2fa(payload: dict, request: Request):
    """Langkah kedua login: verifikasi kode TOTP atau recovery code."""
    from portal import twofa as _tf
    from portal.security import _client_ip as _rl_client_ip
    db = await _get_db()
    ip = _rl_client_ip(request)
    ua = request.headers.get("user-agent", "")
    try:
        uid = _tf.decode_mfa_token(payload.get("mfa_token", ""))
    except Exception:
        raise HTTPException(status_code=401, detail="Sesi 2FA kedaluwarsa, silakan login ulang")
    u = await db.users.find_one({"_id": _oid(uid)})
    if not u or not u.get("totp_enabled"):
        raise HTTPException(status_code=400, detail="2FA tidak aktif untuk akun ini")
    code = str(payload.get("code", "")).strip()
    ok = _tf.verify_totp(_tf.decrypt_secret(u["totp_secret"]), code)
    if not ok:
        idx = _tf.check_recovery_code(code, u.get("recovery_codes") or [])
        if idx >= 0:
            rcs = u["recovery_codes"]
            rcs[idx]["used"] = True
            await db.users.update_one({"_id": u["_id"]}, {"$set": {"recovery_codes": rcs}})
            ok = True
    if not ok:
        await db.users.update_one({"_id": u["_id"]}, {"$inc": {"failed_2fa": 1}})
        await _log_login_attempt(db, email=u["email"], action="login_2fa", success=False,
                                 reason="invalid_2fa_code", ip=ip, user_agent=ua)
        raise HTTPException(status_code=401, detail="Kode 2FA tidak valid")
    await db.users.update_one({"_id": u["_id"]}, {"$set": {"failed_2fa": 0}})
    await _log_login_attempt(db, email=u["email"], action="login_2fa", success=True, reason="ok",
                             ip=ip, user_agent=ua)
    token = create_access_token(str(u["_id"]), u["email"], u["role"])
    return {"token": token, "user": _user_public(u)}


@router.get("/auth/2fa/status")
async def twofa_status(user=Depends(get_current_user)):
    db = await _get_db()
    u = await db.users.find_one({"_id": ObjectId(user["id"])})
    codes = u.get("recovery_codes") or []
    return {"enabled": bool(u.get("totp_enabled")),
            "recovery_codes_left": sum(1 for c in codes if not c.get("used"))}


@router.post("/auth/2fa/setup")
async def twofa_setup(user=Depends(get_current_user)):
    """Mulai aktivasi 2FA: buat secret pending + QR (belum aktif sampai diverifikasi)."""
    from portal import twofa as _tf
    db = await _get_db()
    secret = _tf.new_totp_secret()
    await db.users.update_one({"_id": ObjectId(user["id"])},
                              {"$set": {"pending_totp_secret": _tf.encrypt_secret(secret)}})
    uri = _tf.provisioning_uri(secret, user["email"])
    return {"otpauth_uri": uri, "qr": _tf.qr_data_url(uri), "secret": secret}


@router.post("/auth/2fa/verify-enable")
async def twofa_verify_enable(payload: dict, request: Request, user=Depends(get_current_user)):
    from portal import twofa as _tf
    db = await _get_db()
    u = await db.users.find_one({"_id": ObjectId(user["id"])})
    pend = u.get("pending_totp_secret")
    if not pend:
        raise HTTPException(status_code=400, detail="Belum ada setup 2FA yang berjalan")
    if not _tf.verify_totp(_tf.decrypt_secret(pend), payload.get("code", "")):
        raise HTTPException(status_code=400, detail="Kode tidak valid, coba lagi")
    plaintext, docs = _tf.new_recovery_codes(10)
    await db.users.update_one({"_id": u["_id"]}, {
        "$set": {"totp_secret": pend, "totp_enabled": True,
                 "recovery_codes": docs, "failed_2fa": 0},
        "$unset": {"pending_totp_secret": ""}})
    await log_audit(db, actor=user, action="user.2fa_enabled", category="security",
                    target_type="user", target_id=user["id"], target_label=user["email"],
                    request=request)
    return {"ok": True, "recovery_codes": plaintext}


@router.post("/auth/2fa/disable")
async def twofa_disable(payload: dict, request: Request, user=Depends(get_current_user)):
    from portal import twofa as _tf
    db = await _get_db()
    u = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not u.get("totp_enabled"):
        raise HTTPException(status_code=400, detail="2FA belum aktif")
    code = str(payload.get("code", "")).strip()
    ok = _tf.verify_totp(_tf.decrypt_secret(u["totp_secret"]), code)
    if not ok and _tf.check_recovery_code(code, u.get("recovery_codes") or []) < 0:
        raise HTTPException(status_code=401, detail="Kode 2FA tidak valid")
    await db.users.update_one({"_id": u["_id"]}, {
        "$set": {"totp_enabled": False, "recovery_codes": [], "failed_2fa": 0},
        "$unset": {"totp_secret": "", "pending_totp_secret": ""}})
    await log_audit(db, actor=user, action="user.2fa_disabled", category="security",
                    target_type="user", target_id=user["id"], target_label=user["email"],
                    severity="warning", request=request)
    return {"ok": True}


@router.post("/admin/users/{uid}/reset-2fa")
async def admin_reset_2fa(uid: str, request: Request, admin=Depends(get_current_admin)):
    """Admin mereset 2FA staf yang kehilangan authenticator."""
    db = await _get_db()
    target = await db.users.find_one({"_id": _oid(uid)})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    await db.users.update_one({"_id": target["_id"]}, {
        "$set": {"totp_enabled": False, "recovery_codes": [], "failed_2fa": 0},
        "$unset": {"totp_secret": "", "pending_totp_secret": ""}})
    await log_audit(db, actor=admin, action="user.2fa_reset_by_admin", category="security",
                    target_type="user", target_id=str(target["_id"]),
                    target_label=target.get("email", ""), severity="warning", request=request)
    return {"ok": True, "message": f"2FA direset untuk {target.get('email','')}"}


async def _upsert_crm_from_user(db, u: dict, *, status: str = "prospect", extra_notes: str = "") -> None:
    """Ensure a matching row exists in `crm_customers` for the given user.

    Matches on email (case-insensitive). If the CRM row already exists,
    refresh a small set of contact fields but never downgrade its status.
    """
    email = (u.get("email") or "").lower()
    if not email:
        return
    existing = await db.crm_customers.find_one({"email": email})
    now = _now()
    payload = {
        "name": u.get("name", ""),
        "email": email,
        "phone": u.get("phone", ""),
        "company": u.get("company", ""),
        "updated_at": now,
    }
    if existing:
        # Never downgrade a manually-set status; keep as-is
        payload.pop("email", None)
        if u.get("_id"):
            payload["user_id"] = u["_id"]
        await db.crm_customers.update_one({"_id": existing["_id"]}, {"$set": payload})
        return
    payload.update({
        "position": "",
        "industry": u.get("industry") or "",
        "status": status,
        "notes": extra_notes,
        "user_id": u.get("_id"),
        "source": "self_registration" if status == "prospect" else "admin_registered",
        "created_at": now,
    })
    await db.crm_customers.insert_one(payload)


@router.post("/auth/register", response_model=m.LoginOut)
@_rl_limiter.limit(AUTH_REGISTER_LIMIT)
async def register(payload: m.RegisterIn, request: Request):
    """Public self-registration endpoint.

    Creates a `client` user, mirrors them into `crm_customers` (as a
    `prospect`), and returns a signed JWT so the browser can auto-login.
    """
    db = await _get_db()
    from portal import integrations_v2 as _iv2
    await _iv2.enforce_recaptcha(
        db, payload.recaptcha_token, "register",
        request.client.host if request.client else None,
    )
    email = payload.email.lower().strip()

    if not payload.accepts_tos:
        raise HTTPException(status_code=400, detail="You must accept the Terms of Service to register")

    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="An account already exists for this email")

    doc = {
        "email": email,
        "password_hash": hash_password(payload.password),
        "name": payload.name.strip(),
        "role": "client",
        "company": payload.company,
        "phone": payload.phone,
        "assigned_client_ids": [],
        "billing_emails": [],
        "attention": payload.attention or payload.name.strip(),
        "address_line1": payload.address_line1,
        "address_line2": payload.address_line2,
        "city": payload.city,
        "province": payload.province,
        "postal_code": payload.postal_code,
        "country": payload.country or "Indonesia",
        "npwp": payload.npwp,
        "industry": payload.industry,
        "created_at": _now(),
    }
    r = await db.users.insert_one(doc)
    doc["_id"] = r.inserted_id

    # Mirror into CRM as a prospect
    try:
        await _upsert_crm_from_user(db, doc, status="prospect",
                                    extra_notes="Registered via portal self-signup")
    except Exception:
        # CRM mirroring must never block registration
        pass

    # Fire welcome email (best-effort - never blocks registration)
    try:
        from portal import emails as _em
        await _em.on_user_registered(db, doc)
    except Exception:
        pass

    token = create_access_token(str(doc["_id"]), doc["email"], doc["role"])
    return {"token": token, "user": _user_public(doc)}


@router.get("/auth/me", response_model=m.UserOut)
async def me(user=Depends(get_current_user)):
    return user


# ============================================================
# CLIENT
# ============================================================
@router.get("/client/dashboard")
async def client_dashboard(user=Depends(get_current_user)):
    db = await _get_db()
    await _mark_overdue(db)
    uid = ObjectId(user["id"])
    services_count = await db.services.count_documents({"user_id": uid, "status": "active"})
    unpaid = await db.invoices.count_documents({"user_id": uid, "status": "unpaid"})
    overdue = await db.invoices.count_documents({"user_id": uid, "status": "overdue"})
    open_tickets = await db.tickets.count_documents(
        {"user_id": uid, "status": {"$in": ["open", "awaiting_client", "awaiting_staff"]}}
    )
    # Overdue invoice summary
    overdue_docs = await db.invoices.find({"user_id": uid, "status": "overdue"}).to_list(20)
    overdue_total = sum(d.get("total", 0) for d in overdue_docs)
    return {
        "stats": {
            "active_services": services_count,
            "unpaid_invoices": unpaid,
            "overdue_invoices": overdue,
            "open_tickets": open_tickets,
            "overdue_total": overdue_total,
        },
    }


@router.get("/client/services")
async def client_services(user=Depends(get_current_user)):
    db = await _get_db()
    docs = await db.services.find({"user_id": ObjectId(user["id"])}).sort("created_at", -1).to_list(500)
    result = []
    for d in docs:
        result.append({
            "id": str(d["_id"]),
            "user_id": str(d["user_id"]),
            "product_id": str(d["product_id"]),
            "product_name": d.get("product_name", ""),
            "category": d.get("category", ""),
            "name": d.get("name", ""),
            "status": d.get("status", "active"),
            "start_date": d.get("start_date", ""),
            "next_renewal": d.get("next_renewal", ""),
            "price_monthly": d.get("price_monthly", 0),
            "auto_renew": d.get("auto_renew", True),
            "config": d.get("config", {}),
        })
    return result


@router.get("/client/hosting-accounts")
async def client_hosting_accounts(user=Depends(get_current_user)):
    """Daftar akun hosting (cPanel/Plesk/DirectAdmin) milik klien."""
    db = await _get_db()
    docs = await db.services.find(
        {"user_id": ObjectId(user["id"]), "category": "hosting"}
    ).sort("created_at", -1).to_list(200)
    out = []
    for d in docs:
        cfg = d.get("config") or {}
        out.append({
            "id": str(d["_id"]),
            "product_name": d.get("product_name", ""),
            "name": d.get("name", ""),
            "status": d.get("status", "active"),
            "next_renewal": d.get("next_renewal", ""),
            "price_monthly": d.get("price_monthly", 0),
            "control_panel": cfg.get("control_panel", ""),
            "domain": cfg.get("domain") or cfg.get("hostname", ""),
            "username": cfg.get("username", ""),
            "ip": cfg.get("ip", ""),
            "provision_status": cfg.get("provision_status", "manual"),
        })
    return out


@router.get("/client/services/{sid}")
async def client_service_detail(sid: str, user=Depends(get_current_user)):
    db = await _get_db()
    d = await db.services.find_one({"_id": _oid(sid), "user_id": ObjectId(user["id"])})
    if not d:
        raise HTTPException(status_code=404, detail="Service not found")
    return {
        "id": str(d["_id"]),
        "user_id": str(d["user_id"]),
        "product_id": str(d["product_id"]),
        "product_name": d.get("product_name", ""),
        "category": d.get("category", ""),
        "name": d.get("name", ""),
        "status": d.get("status", "active"),
        "start_date": d.get("start_date", ""),
        "next_renewal": d.get("next_renewal", ""),
        "price_monthly": d.get("price_monthly", 0),
        "auto_renew": d.get("auto_renew", True),
        "config": d.get("config", {}),
        "self_service_log": (d.get("self_service_log") or [])[-10:],
        "pending_upgrade": bool(d.get("pending_upgrade")),
    }


@router.put("/client/services/{sid}/auto-renew")
async def client_service_auto_renew(sid: str, payload: dict, user=Depends(get_current_user)):
    """Klien mengatur auto-renewal per layanan. False = sweep tidak membuat invoice perpanjangan."""
    db = await _get_db()
    d = await db.services.find_one({"_id": _oid(sid), "user_id": ObjectId(user["id"])})
    if not d:
        raise HTTPException(status_code=404, detail="Service not found")
    val = bool(payload.get("auto_renew", True))
    await db.services.update_one({"_id": d["_id"]}, {
        "$set": {"auto_renew": val},
        "$push": {"self_service_log": {"at": _now(), "action": "auto_renew",
                                       "message": f"Auto-renewal {'diaktifkan' if val else 'dimatikan'} oleh klien."}},
    })
    return {"ok": True, "auto_renew": val}


# ---------------- Client self-service VM control ----------------
_CLIENT_VM_ACTIONS = ("start", "stop", "reboot")
_VM_CATEGORIES = ("vps", "cloud", "dedicated")


@router.get("/client/vms")
async def client_vms(user=Depends(get_current_user)):
    """Daftar semua VM milik klien dengan status terkini dari Proxmox."""
    db = await _get_db()
    svcs = await db.services.find(
        {"user_id": ObjectId(user["id"]), "category": {"$in": list(_VM_CATEGORIES)}}
    ).sort("created_at", -1).to_list(100)
    s = await iv2.get_settings(db, "proxmox")
    client = iv2.ProxmoxClient(s) if (s and s.get("enabled")) else None
    out = []
    for svc in svcs:
        cfg = svc.get("config") or {}
        item = {
            "service_id": str(svc["_id"]),
            "name": svc.get("name", ""),
            "product_name": svc.get("product_name", ""),
            "service_status": svc.get("status", ""),
            "node": cfg.get("node"),
            "vmid": cfg.get("vmid"),
            "configured": False,
            "status": "unknown",
        }
        if client and cfg.get("node") and cfg.get("vmid"):
            item["configured"] = True
            try:
                st = await client.vm_status(cfg["node"], int(cfg["vmid"]))
                item["status"] = st.get("status", "unknown")
                item["uptime"] = st.get("uptime")
                item["cpu"] = st.get("cpu")
                item["mem"] = st.get("mem")
            except Exception:
                item["status"] = "unreachable"
        out.append(item)
    return out


@router.get("/client/services/{sid}/vm")
async def client_vm_status(sid: str, user=Depends(get_current_user)):
    """Live VM status for the client's own service (read-only)."""
    db = await _get_db()
    svc = await db.services.find_one({"_id": _oid(sid), "user_id": ObjectId(user["id"])})
    if not svc:
        raise HTTPException(status_code=404, detail="Service not found")
    cfg = svc.get("config") or {}
    s = await iv2.get_settings(db, "proxmox")
    if not (s and s.get("enabled") and cfg.get("node") and cfg.get("vmid")):
        return {"configured": False, "status": "unknown"}
    try:
        st = await iv2.ProxmoxClient(s).vm_status(cfg["node"], int(cfg["vmid"]))
        return {"configured": True, "status": st.get("status", "unknown"),
                "uptime": st.get("uptime"), "cpu": st.get("cpu"),
                "mem": st.get("mem"), "maxmem": st.get("maxmem"),
                "node": cfg["node"], "vmid": int(cfg["vmid"])}
    except Exception as e:
        return {"configured": True, "status": "unreachable", "error": str(e)[:200]}


@router.post("/client/services/{sid}/vm/reset-password")
async def client_vm_reset_password(sid: str, payload: dict, request: Request, user=Depends(get_current_user)):
    """Self-service guest OS password reset via QEMU guest agent (audited)."""
    username = (payload.get("username") or "root").strip()
    password = payload.get("password") or ""
    generated = False
    if not password and payload.get("generate"):
        alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789!@#$%"
        password = "".join(secrets.choice(alphabet) for _ in range(16))
        generated = True
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password minimal 8 karakter")
    db = await _get_db()
    svc = await db.services.find_one({"_id": _oid(sid), "user_id": ObjectId(user["id"])})
    if not svc:
        raise HTTPException(status_code=404, detail="Service not found")
    if svc.get("category") not in _VM_CATEGORIES:
        raise HTTPException(status_code=400, detail="Layanan ini bukan VM")
    if svc.get("status") == "suspended":
        raise HTTPException(status_code=403, detail="Layanan ditangguhkan. Lunasi tagihan untuk mengaktifkan kembali.")
    cfg = svc.get("config") or {}
    node, vmid = cfg.get("node"), cfg.get("vmid")
    if not (node and vmid):
        raise HTTPException(status_code=400, detail="VM belum terhubung ke layanan ini. Hubungi support.")
    s = await iv2.get_settings(db, "proxmox")
    if not (s and s.get("enabled")):
        raise HTTPException(status_code=400, detail="Integrasi Proxmox belum aktif. Hubungi support.")
    client = iv2.ProxmoxClient(s)
    try:
        st = await client.vm_status(node, int(vmid))
        if st.get("status") != "running":
            raise HTTPException(status_code=400, detail="VM harus dalam keadaan running untuk reset password.")
        await client.set_user_password(node, int(vmid), username, password)
    except HTTPException:
        raise
    except Exception as e:
        detail = str(e)[:200]
        if ("agent" in detail.lower() or "500" in detail or "timeout" in detail.lower()
                or "timed out" in detail.lower()):
            detail = "QEMU guest agent tidak aktif atau tidak merespons di VM ini. Hubungi support untuk reset manual."
        raise HTTPException(status_code=502, detail=detail)
    await log_audit(db, actor=user, action="client_vm.reset_password", category="services",
                    target_type="service", target_id=str(svc["_id"]),
                    target_label=svc.get("name", ""),
                    metadata={"node": node, "vmid": int(vmid), "os_username": username},
                    severity="warning", request=request)
    await db.services.update_one(
        {"_id": svc["_id"]},
        {"$push": {"self_service_log": {"at": _now(), "action": "reset_password", "by": user["email"]}}})
    out = {"ok": True, "username": username}
    if generated:
        out["generated_password"] = password
    return out


@router.post("/client/services/{sid}/vm/{action}")
async def client_vm_action(sid: str, action: str, request: Request, user=Depends(get_current_user)):
    """Self-service start/stop/reboot on the client's own VM (audited)."""
    if action not in _CLIENT_VM_ACTIONS:
        raise HTTPException(status_code=400, detail="Aksi tidak didukung")
    db = await _get_db()
    svc = await db.services.find_one({"_id": _oid(sid), "user_id": ObjectId(user["id"])})
    if not svc:
        raise HTTPException(status_code=404, detail="Service not found")
    if svc.get("category") not in _VM_CATEGORIES:
        raise HTTPException(status_code=400, detail="Layanan ini bukan VM")
    if svc.get("status") == "suspended":
        raise HTTPException(status_code=403, detail="Layanan ditangguhkan. Lunasi tagihan untuk mengaktifkan kembali.")
    cfg = svc.get("config") or {}
    node, vmid = cfg.get("node"), cfg.get("vmid")
    if not (node and vmid):
        raise HTTPException(status_code=400, detail="VM belum terhubung ke layanan ini. Hubungi support.")
    s = await iv2.get_settings(db, "proxmox")
    if not (s and s.get("enabled")):
        raise HTTPException(status_code=400, detail="Integrasi Proxmox belum aktif. Hubungi support.")
    try:
        result = await iv2.ProxmoxClient(s).vm_action(node, int(vmid), action)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Proxmox error: {str(e)[:200]}")
    await log_audit(db, actor=user, action=f"client_vm.{action}", category="services",
                    target_type="service", target_id=str(svc["_id"]),
                    target_label=svc.get("name", ""),
                    metadata={"node": node, "vmid": int(vmid)}, request=request)
    await db.services.update_one(
        {"_id": svc["_id"]},
        {"$push": {"self_service_log": {"at": _now(), "action": action, "by": user["email"]}}})
    return {"ok": True, "action": action, "task": result}


# ---------------- Client self-service resource upgrade ----------------
_UPGRADE_PRICING_DEFAULT = {"cpu_per_core": 50000.0, "ram_per_gb": 25000.0, "disk_per_gb": 2000.0}


def _upgrade_quote(svc: dict, pricing: dict, cpu: int, ram_gb: int, disk_gb: int, tax_percent: float) -> dict:
    monthly_delta = (cpu * float(pricing["cpu_per_core"])
                     + ram_gb * float(pricing["ram_per_gb"])
                     + disk_gb * float(pricing["disk_per_gb"]))
    today = datetime.now(timezone.utc).date()
    try:
        renewal = datetime.fromisoformat(svc.get("next_renewal", "")).date()
        days_left = max(0, min(31, (renewal - today).days))
    except Exception:
        days_left = 30
    factor = days_left / 30.0
    prorated = round(monthly_delta * factor, 2)
    tax = round(prorated * tax_percent / 100.0, 2)
    return {
        "cpu": cpu, "ram_gb": ram_gb, "disk_gb": disk_gb,
        "monthly_delta": round(monthly_delta, 2),
        "days_left": days_left,
        "prorated_charge": prorated,
        "tax_percent": tax_percent,
        "tax_amount": tax,
        "total": round(prorated + tax, 2),
    }


async def _upgrade_ctx(db, sid: str, user):
    svc = await db.services.find_one({"_id": _oid(sid), "user_id": ObjectId(user["id"])})
    if not svc:
        raise HTTPException(status_code=404, detail="Service not found")
    if svc.get("category") not in _VM_CATEGORIES:
        raise HTTPException(status_code=400, detail="Upgrade resource hanya tersedia untuk layanan VM")
    if svc.get("status") not in ("active",):
        raise HTTPException(status_code=403, detail="Layanan harus aktif untuk melakukan upgrade")
    pricing = await _get_setting_value(db, "upgrade_pricing", _UPGRADE_PRICING_DEFAULT)
    pricing = {**_UPGRADE_PRICING_DEFAULT, **(pricing or {})}
    tax_percent = float(await _get_setting_value(db, "default_tax_percent", 11.0))
    return svc, pricing, tax_percent


def _upgrade_units(payload: dict) -> tuple:
    try:
        cpu = max(0, min(64, int(payload.get("cpu") or 0)))
        ram_gb = max(0, min(256, int(payload.get("ram_gb") or 0)))
        disk_gb = max(0, min(2000, int(payload.get("disk_gb") or 0)))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Nilai upgrade tidak valid")
    if cpu + ram_gb + disk_gb == 0:
        raise HTTPException(status_code=400, detail="Pilih minimal satu resource untuk di-upgrade")
    return cpu, ram_gb, disk_gb


@router.get("/client/services/{sid}/upgrade/options")
async def client_upgrade_options(sid: str, user=Depends(get_current_user)):
    db = await _get_db()
    svc, pricing, tax_percent = await _upgrade_ctx(db, sid, user)
    cfg = svc.get("config") or {}
    return {
        "pricing": pricing,
        "tax_percent": tax_percent,
        "current": {"cpu": cfg.get("cpu"), "ram_gb": cfg.get("ram_gb"), "disk_gb": cfg.get("disk_gb")},
        "pending_upgrade": bool(svc.get("pending_upgrade")),
    }


@router.post("/client/services/{sid}/upgrade/preview")
async def client_upgrade_preview(sid: str, payload: dict, user=Depends(get_current_user)):
    db = await _get_db()
    svc, pricing, tax_percent = await _upgrade_ctx(db, sid, user)
    cpu, ram_gb, disk_gb = _upgrade_units(payload)
    return _upgrade_quote(svc, pricing, cpu, ram_gb, disk_gb, tax_percent)


@router.post("/client/services/{sid}/upgrade")
async def client_upgrade_request(sid: str, payload: dict, request: Request, user=Depends(get_current_user)):
    """Create the prorated difference invoice for a self-service resource upgrade."""
    db = await _get_db()
    svc, pricing, tax_percent = await _upgrade_ctx(db, sid, user)
    if svc.get("pending_upgrade"):
        raise HTTPException(status_code=400, detail="Masih ada upgrade yang menunggu pembayaran untuk layanan ini")
    cpu, ram_gb, disk_gb = _upgrade_units(payload)
    q = _upgrade_quote(svc, pricing, cpu, ram_gb, disk_gb, tax_percent)
    items = []
    if cpu:
        items.append({"description": f"Upgrade +{cpu} vCPU - {svc.get('product_name','')} (prorata {q['days_left']} hari)",
                      "qty": 1, "price": round(cpu * pricing["cpu_per_core"] * q["days_left"] / 30.0, 2)})
    if ram_gb:
        items.append({"description": f"Upgrade +{ram_gb} GB RAM - {svc.get('product_name','')} (prorata {q['days_left']} hari)",
                      "qty": 1, "price": round(ram_gb * pricing["ram_per_gb"] * q["days_left"] / 30.0, 2)})
    if disk_gb:
        items.append({"description": f"Upgrade +{disk_gb} GB Disk - {svc.get('product_name','')} (prorata {q['days_left']} hari)",
                      "qty": 1, "price": round(disk_gb * pricing["disk_per_gb"] * q["days_left"] / 30.0, 2)})
    for it in items:
        it["total"] = round(it["qty"] * it["price"], 2)
    subtotal = round(sum(i["total"] for i in items), 2)
    tax = round(subtotal * tax_percent / 100.0, 2)
    due = (datetime.now(timezone.utc) + timedelta(days=7)).date().isoformat()
    inv = {
        "user_id": ObjectId(user["id"]),
        "items": items,
        "subtotal": subtotal,
        "tax_percent": tax_percent,
        "tax_amount": tax,
        "total": round(subtotal + tax, 2),
        "due_date": due,
        "status": "unpaid",
        "payment_method": None,
        "paid_at": None,
        "notes": f"Upgrade resource {svc.get('name','')} - berlaku setelah pembayaran.",
        "service_id": str(svc["_id"]),
        "upgrade": {"cpu": cpu, "ram_gb": ram_gb, "disk_gb": disk_gb,
                    "monthly_delta": q["monthly_delta"]},
        "created_at": _now(),
    }
    inv = await _insert_numbered(db, "invoices", "INV", inv)
    await db.services.update_one(
        {"_id": svc["_id"]},
        {"$set": {"pending_upgrade": {
            "cpu": cpu, "ram_gb": ram_gb, "disk_gb": disk_gb,
            "monthly_delta": q["monthly_delta"],
            "invoice_id": str(inv["_id"]),
            "requested_at": _now(),
        }}})
    await log_audit(db, actor=user, action="client_service.upgrade_requested", category="services",
                    target_type="service", target_id=str(svc["_id"]), target_label=svc.get("name", ""),
                    metadata={"cpu": cpu, "ram_gb": ram_gb, "disk_gb": disk_gb,
                              "invoice": inv["number"], "total": inv["total"]},
                    request=request)
    return {"ok": True, "invoice_id": str(inv["_id"]), "number": inv["number"],
            "total": inv["total"], "due_date": due, "quote": q}


async def _serialize_invoice(db, d: dict) -> dict:
    u = await db.users.find_one({"_id": d["user_id"]}) or {}
    return {
        "id": str(d["_id"]),
        "number": d["number"],
        "user_id": str(d["user_id"]),
        "user_name": u.get("name", ""),
        "user_email": u.get("email", ""),
        "items": d.get("items", []),
        "subtotal": d.get("subtotal", 0),
        "tax_percent": d.get("tax_percent"),
        "tax_amount": d.get("tax_amount", 0),
        "total": d.get("total", 0),
        "due_date": d.get("due_date", ""),
        "status": d.get("status", "unpaid"),
        "payment_method": d.get("payment_method"),
        "paid_at": d.get("paid_at"),
        "payment_link": d.get("payment_link"),
        "payment_ref": d.get("payment_ref"),
        "service_id": d.get("service_id"),
        "order_id": d.get("order_id"),
        "renewal_period": d.get("renewal_period"),
        "created_at": _iso(d.get("created_at", "")),
        "notes": d.get("notes", ""),
        "source_quotation_id": d.get("source_quotation_id"),
        "source_quotation_number": d.get("source_quotation_number"),
    }


@router.get("/client/invoices")
async def client_invoices(user=Depends(get_current_user)):
    db = await _get_db()
    await _mark_overdue(db)
    docs = await db.invoices.find({"user_id": ObjectId(user["id"])}).sort("created_at", -1).to_list(500)
    return [await _serialize_invoice(db, d) for d in docs]


@router.get("/client/invoices/{iid}")
async def client_invoice_detail(iid: str, user=Depends(get_current_user)):
    db = await _get_db()
    await _mark_overdue(db)
    d = await db.invoices.find_one({"_id": _oid(iid), "user_id": ObjectId(user["id"])})
    if not d:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return await _serialize_invoice(db, d)


@router.post("/client/orders")
async def create_order(payload: m.OrderIn, user=Depends(get_current_user)):
    db = await _get_db()
    prod = await db.products.find_one({"_id": _oid(payload.product_id)})
    if not prod or prod.get("is_addon"):
        raise HTTPException(status_code=404, detail="Product not found")

    # Price the cart from the selected options + add-ons.
    selections_data = [s.model_dump() for s in (payload.selections or [])]
    cart = await _price_cart(
        db, product=prod,
        selections=selections_data,
        addon_ids=payload.addon_ids or [],
        tax_percent=float(await _get_setting_value(db, "default_tax_percent", 11.0)),
    )

    # 1. Create the order
    doc = {
        "user_id": ObjectId(user["id"]),
        "user_name": user["name"],
        "user_email": user["email"],
        "product_id": prod["_id"],
        "product_name": prod["name"],
        "notes": payload.notes,
        "config": payload.config,
        "selections": selections_data,
        "addon_ids": [ObjectId(x) for x in (payload.addon_ids or [])],
        "cart_snapshot": cart,   # audit - the price shown to the user at confirm time
        "billing_cycle": payload.billing_cycle or prod.get("billing_cycle", "monthly"),
        "status": "pending_payment",
        "assigned_admin_id": None,
        "invoice_id": None,
        "service_id": None,
        "provision_log": [{"at": _now(), "step": "order_created", "message": "Order placed by client."}],
        "created_at": _now(),
    }
    r = await db.orders.insert_one(doc)
    doc["_id"] = r.inserted_id

    # 2. Auto-create an invoice for the order (14-day due window)
    items = []
    # Base line - first billing period
    base_line = cart["base_line"]
    if base_line["monthly"]:
        items.append({
            "description": f"{prod['name']} - first month",
            "qty": 1, "unit_price": base_line["monthly"], "total": base_line["monthly"],
        })
    if base_line["setup"]:
        items.append({
            "description": f"{prod['name']} - setup fee",
            "qty": 1, "unit_price": base_line["setup"], "total": base_line["setup"],
        })
    # Configurable option lines
    for ol in cart["option_lines"]:
        if ol.get("monthly"):
            items.append({
                "description": f"{ol['group_label']}: {ol['choice']} - monthly",
                "qty": 1, "unit_price": ol["monthly"], "total": ol["monthly"],
            })
        if ol.get("setup"):
            items.append({
                "description": f"{ol['group_label']}: {ol['choice']} - setup",
                "qty": 1, "unit_price": ol["setup"], "total": ol["setup"],
            })
    # Add-on lines
    for al in cart["addon_lines"]:
        if al.get("monthly"):
            items.append({
                "description": f"Add-on: {al['name']} - monthly",
                "qty": 1, "unit_price": al["monthly"], "total": al["monthly"],
            })
        if al.get("setup"):
            items.append({
                "description": f"Add-on: {al['name']} - setup",
                "qty": 1, "unit_price": al["setup"], "total": al["setup"],
            })

    if not items:
        # Custom quote / firewall - mark order as needing manual quotation, no auto-invoice
        await db.orders.update_one({"_id": doc["_id"]}, {"$set": {"status": "awaiting_quote"}})
        doc["status"] = "awaiting_quote"
        doc["provision_log"].append({"at": _now(), "step": "awaiting_quote", "message": "Custom-priced product; sales will send a quotation."})
    else:
        line_subtotal = sum(i["total"] for i in items)
        # PPN pre-filled from the admin-editable global default - stored on the
        # invoice and manually overridable afterwards; never recalculated.
        tax_percent = float(await _get_setting_value(db, "default_tax_percent", 11.0))
        tax = round(line_subtotal * tax_percent / 100, 2)
        total = round(line_subtotal + tax, 2)
        due = (datetime.now(timezone.utc) + timedelta(days=14)).date().isoformat()
        number = await _next_number(db, "invoices", "INV")
        inv = {
            "number": number,
            "user_id": ObjectId(user["id"]),
            "items": items,
            "subtotal": line_subtotal,
            "tax_percent": tax_percent,
            "tax_amount": tax,
            "total": total,
            "due_date": due,
            "status": "unpaid",
            "payment_method": None,
            "paid_at": None,
            "notes": f"Auto-generated from order for {prod['name']}.",
            "order_id": str(doc["_id"]),
            "created_at": _now(),
        }
        ir = await db.invoices.insert_one(inv)
        await db.orders.update_one(
            {"_id": doc["_id"]},
            {"$set": {"invoice_id": ir.inserted_id},
             "$push": {"provision_log": {"at": _now(), "step": "invoice_created", "message": f"Invoice {number} generated ({total:,.0f} IDR)."}}},
        )
        doc["invoice_id"] = ir.inserted_id
        doc["provision_log"].append({"at": _now(), "step": "invoice_created", "message": f"Invoice {number} generated ({total:,.0f} IDR)."})

    # Fire order + invoice notification emails (best-effort - never blocks the order)
    try:
        from portal import emails as _em
        user_doc = await db.users.find_one({"_id": ObjectId(user["id"])}) or {"email": user["email"], "name": user["name"]}
        await _em.on_order_created(db, doc, user_doc)
        if doc.get("invoice_id"):
            inv_doc = await db.invoices.find_one({"_id": doc["invoice_id"]})
            if inv_doc:
                await _em.on_invoice_generated(db, inv_doc, user_doc, order_doc=doc)
    except Exception:
        pass

    return _serialize_order(doc)


def _serialize_order(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "user_id": str(d["user_id"]),
        "user_name": d.get("user_name", ""),
        "user_email": d.get("user_email", ""),
        "product_id": str(d["product_id"]),
        "product_name": d.get("product_name", ""),
        "notes": d.get("notes", ""),
        "config": d.get("config", {}),
        "status": d.get("status", "pending"),
        "assigned_admin_id": str(d["assigned_admin_id"]) if d.get("assigned_admin_id") else None,
        "invoice_id": str(d["invoice_id"]) if d.get("invoice_id") else None,
        "service_id": str(d["service_id"]) if d.get("service_id") else None,
        "provision_log": d.get("provision_log", []),
        "created_at": _iso(d.get("created_at", "")),
    }


async def _auto_provision(db, order: dict) -> dict:
    """Actually run auto-provisioning based on product category.
    Returns the created service (or None if manual setup is required).
    Currently uses realistic mocked module calls - swap for real cPanel/Plesk/
    Proxmox API calls once credentials are wired via /admin/integrations.
    """
    prod = await db.products.find_one({"_id": order["product_id"]})
    if not prod:
        return None
    cat = prod.get("category", "other")
    now = datetime.now(timezone.utc)
    cfg = dict(order.get("config", {}))

    # Append a provision log entry
    async def _log(step, msg):
        await db.orders.update_one(
            {"_id": order["_id"]},
            {"$push": {"provision_log": {"at": _now(), "step": step, "message": msg}}},
        )

    await _log("provisioning_started", f"Provisioning started for category '{cat}'.")

    # Spesifikasi dari opsi konfigurasi + add-on ikut dibawa ke service config
    sel_specs = {}
    for s in order.get("selections", []) or []:
        gk = s.get("group_key")
        if not gk:
            continue
        if s.get("option_labels"):
            sel_specs[gk] = s["option_labels"][0] if len(s["option_labels"]) == 1 else s["option_labels"]
        elif s.get("quantity") is not None:
            sel_specs[gk] = s["quantity"]
    if sel_specs:
        cfg.setdefault("selected_options", sel_specs)
    if order.get("addon_ids"):
        addon_docs = await db.products.find({"_id": {"$in": order["addon_ids"]}}).to_list(50)
        if addon_docs:
            cfg.setdefault("addons", [a.get("name", "") for a in addon_docs])
            await _log("addons_attached",
                       "Add-on terpasang: " + ", ".join(a.get("name", "") for a in addon_docs))

    hosting_credentials = None
    if cat in ("hosting",):
        # Live provisioning cPanel/Plesk/DirectAdmin bila integrasi aktif; fallback mock.
        _PANELS = [
            ("cpanel", "cPanel/WHM", iv2.CpanelClient, "package"),
            ("plesk", "Plesk", iv2.PleskClient, "plan"),
            ("directadmin", "DirectAdmin", iv2.DirectAdminClient, "package"),
        ]
        chosen = None
        for key, label, cls, pkg_kw in _PANELS:
            s = await iv2.get_settings(db, key)
            if s and s.get("enabled"):
                chosen = (key, label, cls, pkg_kw, s)
                break
        if chosen:
            key, panel_label, cls, pkg_kw, settings = chosen
            domain = cfg.get("domain") or f"{order['user_email'].split('@')[0]}.icd-cust.net"
            uname = re.sub(r"[^a-z0-9]", "", order["user_email"].split("@")[0].lower())[:8] or "icduser"
            if uname[0].isdigit():
                uname = "u" + uname[:7]
            pw_alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789!@#$%"
            password = "".join(secrets.choice(pw_alphabet) for _ in range(16))
            try:
                await cls(settings).create_account(
                    domain=domain, username=uname, password=password,
                    contact_email=order["user_email"], **{pkg_kw: cfg.get("package") or None})
                cfg.update({"control_panel": panel_label, "domain": domain,
                            "username": uname, "provision_status": "provisioned",
                            "provisioned_at": _now()})
                await _log("panel_account_created",
                           f"{panel_label} account '{uname}' created for {domain} (live).")
                bare_host = re.sub(r"^https?://", "", (settings.get("credentials") or {}).get("host") or "").split(":")[0].split("/")[0]
                panel_port = {"cpanel": 2083, "plesk": 8443, "directadmin": 2222}[key]
                hosting_credentials = {
                    "panel": panel_label, "domain": domain, "username": uname,
                    "password": password,
                    "panel_url": f"https://{bare_host}:{panel_port}" if bare_host else "",
                }
            except Exception as e:
                cfg.update({"control_panel": panel_label, "domain": domain,
                            "username": uname, "provision_status": "failed"})
                await _log("panel_account_failed",
                           f"{panel_label} provisioning gagal: {str(e)[:150]}. Perlu tindak lanjut manual.")
        else:
            module = await db.integrations.find_one({"module": {"$in": ["cpanel", "plesk", "directadmin"]}, "status": "enabled"})
            provider = module["module"] if module else "cpanel"
            cfg.setdefault("control_panel", {"cpanel": "cPanel/WHM", "plesk": "Plesk", "directadmin": "DirectAdmin"}.get(provider, "cPanel/WHM"))
            cfg.setdefault("hostname", f"{order['user_email'].split('@')[0]}.icd-cust.net")
            pool_ip = await _auto_allocate_customer_ip(db, hostname=cfg.get("hostname", ""),
                                                       customer=order.get("user_email", ""),
                                                       ref=f"order {str(order['_id'])[-6:]}")
            cfg.setdefault("ip", pool_ip or ("103.28.14." + str((hash(str(order["_id"])) % 240) + 10)))
            if pool_ip:
                await _log("ip_allocated", f"IP {pool_ip} dialokasikan otomatis dari IP pool (DCIM).")
            cfg.setdefault("provision_status", "manual")
            await _log("panel_account_created", f"{provider.upper()} account provisioned (mock).")
    elif cat in ("vps", "cloud"):
        module = await db.integrations.find_one({"module": "proxmox", "status": "enabled"})
        cfg.setdefault("node", (module or {}).get("config", {}).get("default_node") or "prox-jkt-05")
        cfg.setdefault("os", cfg.get("os") or "Ubuntu 22.04 LTS Server")
        cfg.setdefault("hostname", f"vm-{str(order['_id'])[-6:]}.icd-cust.net")
        pool_ip = await _auto_allocate_customer_ip(db, hostname=cfg.get("hostname", ""),
                                                   customer=order.get("user_email", ""),
                                                   ref=f"order {str(order['_id'])[-6:]}")
        cfg.setdefault("ip", pool_ip or ("103.28.14." + str((hash(str(order["_id"])) % 240) + 10)))
        if pool_ip:
            await _log("ip_allocated", f"IP {pool_ip} dialokasikan otomatis dari IP pool (DCIM).")
        await _log("vm_created", f"Proxmox VM created on {cfg['node']} with {cfg['os']} (mock).")
    elif cat in ("dedicated", "colocation", "interconnect", "firewall", "lease"):
        # These need manual DC/network setup - mark the service as provisioning
        cfg.setdefault("rack", "TBD by NOC")
        await _log("manual_setup_required", "Requires physical / network setup by NOC team.")
    else:
        await _log("manual_setup_required", "Category needs manual handling.")

    svc = {
        "user_id": order["user_id"],
        "product_id": prod["_id"],
        "product_name": prod["name"],
        "category": cat,
        "name": f"{prod['name']} - {order.get('user_name','')}",
        "status": "active" if cat in ("hosting", "vps", "cloud") else "pending",
        "start_date": now.date().isoformat(),
        "next_renewal": (now + timedelta(days=30)).date().isoformat(),
        "price_monthly": prod.get("price_monthly", 0),
        "config": cfg,
        "order_id": str(order["_id"]),
        "created_at": _now(),
    }
    sr = await db.services.insert_one(svc)
    await db.orders.update_one(
        {"_id": order["_id"]},
        {"$set": {"service_id": sr.inserted_id, "status": "active" if svc["status"] == "active" else "provisioning"},
         "$push": {"provision_log": {"at": _now(), "step": "service_handover", "message": "Service delivered to client dashboard."}}},
    )
    if hosting_credentials:
        u = await db.users.find_one({"_id": order["user_id"]})
        if u:
            from portal import emails as _em
            try:
                await _em.on_hosting_provisioned(db, u, svc, hosting_credentials)
                await _log("credentials_emailed",
                           f"Detail akun hosting dikirim via email ke {u.get('email', '')}.")
            except Exception as e:
                await _log("credentials_email_failed", f"Gagal mengirim email detail akun: {str(e)[:120]}")
    return svc


@router.get("/client/orders")
async def client_orders(user=Depends(get_current_user)):
    db = await _get_db()
    docs = await db.orders.find({"user_id": ObjectId(user["id"])}).sort("created_at", -1).to_list(500)
    return [_serialize_order(d) for d in docs]


# Client tickets
def _deny_creative(user: dict) -> None:
    """Creative role is content-scoped: block billing/CRM/sales surfaces."""
    if user.get("role") == "creative":
        raise HTTPException(status_code=403, detail="Content team only - no billing/CRM access")


async def _serialize_ticket(db, d: dict, include_internal: bool = True) -> dict:
    u = await db.users.find_one({"_id": d["user_id"]}) or {}
    dev_name = None
    if d.get("related_device_id"):
        try:
            dev = await db.mikrotik_devices.find_one({"_id": ObjectId(d["related_device_id"])})
            dev_name = (dev or {}).get("name")
        except Exception:
            dev_name = None
    replies = d.get("replies", [])
    if not include_internal:
        replies = [r for r in replies if not r.get("internal")]
    return {
        "id": str(d["_id"]),
        "number": d.get("number", ""),
        "user_id": str(d["user_id"]),
        "user_name": u.get("name", ""),
        "user_email": u.get("email", ""),
        "subject": d.get("subject", ""),
        "department": d.get("department", "technical"),
        "priority": d.get("priority", "medium"),
        "status": d.get("status", "open"),
        "replies": replies,
        "related_device_id": d.get("related_device_id"),
        "related_device_name": dev_name,
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }


@router.get("/client/tickets")
async def client_tickets(user=Depends(get_current_user)):
    db = await _get_db()
    docs = await db.tickets.find({"user_id": ObjectId(user["id"])}).sort("updated_at", -1).to_list(500)
    return [await _serialize_ticket(db, d, include_internal=False) for d in docs]


@router.post("/client/tickets")
async def client_create_ticket(payload: m.TicketIn, user=Depends(get_current_user)):
    db = await _get_db()
    now = _now()
    number = await _next_number(db, "tickets", "TCK")
    doc = {
        "user_id": ObjectId(user["id"]),
        "number": number,
        "subject": payload.subject,
        "department": payload.department,
        "priority": payload.priority,
        "status": "open",
        "related_device_id": payload.related_device_id or None,
        "replies": [{
            "author_id": user["id"],
            "author_name": user["name"],
            "author_role": "client",
            "message": payload.message,
            "created_at": now,
        }],
        "created_at": now,
        "updated_at": now,
    }
    r = await db.tickets.insert_one(doc)
    doc["_id"] = r.inserted_id
    return await _serialize_ticket(db, doc)


@router.post("/client/tickets/{tid}/replies")
async def client_reply_ticket(tid: str, payload: m.TicketReplyIn, user=Depends(get_current_user)):
    db = await _get_db()
    d = await db.tickets.find_one({"_id": _oid(tid), "user_id": ObjectId(user["id"])})
    if not d:
        raise HTTPException(status_code=404, detail="Ticket not found")
    reply = {
        "author_id": user["id"],
        "author_name": user["name"],
        "author_role": "client",
        "message": payload.message,
        "created_at": _now(),
    }
    await db.tickets.update_one(
        {"_id": d["_id"]},
        {"$push": {"replies": reply}, "$set": {"status": "awaiting_staff", "updated_at": _now()}},
    )
    d = await db.tickets.find_one({"_id": d["_id"]})
    return await _serialize_ticket(db, d, include_internal=False)


@router.put("/client/tickets/{tid}/close")
async def client_close_ticket(tid: str, user=Depends(get_current_user)):
    """UAT-011: klien dapat menutup tiketnya sendiri."""
    db = await _get_db()
    d = await db.tickets.find_one({"_id": _oid(tid), "user_id": ObjectId(user["id"])})
    if not d:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if d.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Tiket sudah ditutup")
    now = _now()
    await db.tickets.update_one({"_id": d["_id"]}, {
        "$set": {"status": "closed", "closed_at": now, "closed_by": "client", "updated_at": now},
        "$push": {"replies": {"author_id": user["id"], "author_name": user["name"],
                              "author_role": "system",
                              "message": "Tiket ditutup oleh klien.", "created_at": now}},
    })
    d = await db.tickets.find_one({"_id": d["_id"]})
    return await _serialize_ticket(db, d, include_internal=False)


@router.put("/admin/tickets/{tid}/status")
async def admin_ticket_status(tid: str, payload: m.TicketStatusIn, staff=Depends(get_current_staff)):
    """UAT-011: staf dapat mengubah status tiket (termasuk close/resolve)."""
    db = await _get_db()
    d = await db.tickets.find_one({"_id": _oid(tid)})
    if not d:
        raise HTTPException(status_code=404, detail="Ticket not found")
    now = _now()
    upd = {"status": payload.status, "updated_at": now}
    if payload.status == "closed":
        upd["closed_at"] = now
        upd["closed_by"] = "staff"
    await db.tickets.update_one({"_id": d["_id"]}, {
        "$set": upd,
        "$push": {"replies": {"author_id": staff["id"], "author_name": staff["name"],
                              "author_role": "system",
                              "message": f"Status tiket diubah menjadi '{payload.status}' oleh {staff['name']}.",
                              "created_at": now}},
    })
    d = await db.tickets.find_one({"_id": d["_id"]})
    return await _serialize_ticket(db, d)


# ============================================================
# ADMIN
# ============================================================
async def _build_admin_alerts(db, staff, scope_user_ids=None) -> list:
    """Pusat Notifikasi: kumpulkan peringatan penting lintas modul dengan severity."""
    alerts = []
    if staff["role"] in FINANCE_ROLES or staff["role"] == "sales":
        inv_q = {} if scope_user_ids is None else {"user_id": {"$in": scope_user_ids}}
        overdue_docs = await db.invoices.find(
            {**inv_q, "status": "overdue"}).sort("due_date", 1).to_list(5)
        for d in overdue_docs:
            alerts.append({
                "type": "invoice_overdue",
                "severity": "danger",
                "title": f"Invoice {d.get('number', '')} jatuh tempo",
                "detail": f"Jatuh tempo {d.get('due_date', '')}",
                "link": "/portal/admin/invoices",
            })
        today = datetime.now(timezone.utc).date()
        soon = (today + timedelta(days=3)).isoformat()
        due_soon = await db.invoices.find(
            {**inv_q, "status": "unpaid",
             "due_date": {"$gte": today.isoformat(), "$lte": soon}}).sort("due_date", 1).to_list(5)
        for d in due_soon:
            alerts.append({
                "type": "invoice_due_soon",
                "severity": "warning",
                "title": f"Invoice {d.get('number', '')} segera jatuh tempo",
                "detail": f"Jatuh tempo {d.get('due_date', '')}",
                "link": "/portal/admin/invoices",
            })
    down_states = await db.noc_device_state.find({"status": "down"}).to_list(20)
    dev_map = {m["_id"]: m.get("name", "unnamed") async for m in db.mikrotik_devices.find(
        {"_id": {"$in": [s["device_id"] for s in down_states]}})} if down_states else {}
    for s in down_states:
        if s["device_id"] not in dev_map:
            continue
        alerts.append({
            "type": "device_down",
            "severity": "danger",
            "title": f"Perangkat {dev_map[s['device_id']]} DOWN",
            "detail": s.get("last_message") or "Tidak merespons probe",
            "link": "/portal/admin/noc",
        })
    pending_orders = await db.orders.count_documents({"status": "pending_verification"})
    if pending_orders:
        alerts.append({
            "type": "orders_pending",
            "severity": "warning",
            "title": f"{pending_orders} order menunggu verifikasi",
            "detail": "Verifikasi pembayaran untuk memproses provisioning",
            "link": "/portal/admin/orders",
        })
    failed_services = await db.services.find(
        {"config.provision_status": "failed"}).sort("created_at", -1).to_list(5)
    for s in failed_services:
        addons = (s.get("config") or {}).get("addons") or []
        detail = ("Termasuk add-on: " + ", ".join(addons)) if addons else "Perlu tindak lanjut manual di panel"
        alerts.append({
            "type": "provision_failed",
            "severity": "danger",
            "title": f"Provisioning gagal: {s.get('product_name') or s.get('name', '')}",
            "detail": detail,
            "link": "/portal/admin/services",
        })
    order = {"danger": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda a: order.get(a["severity"], 9))
    return alerts


@router.get("/admin/notifications")
async def admin_notifications(severity: str | None = None, staff=Depends(get_current_staff)):
    """Pusat Notifikasi dengan filter prioritas (?severity=danger|warning)."""
    db = await _get_db()
    scope_user_ids = None
    if staff["role"] == "sales":
        scope_user_ids = [ObjectId(cid) for cid in staff.get("assigned_client_ids") or []]
    alerts = await _build_admin_alerts(db, staff, scope_user_ids)
    if severity:
        if severity not in ("danger", "warning", "info"):
            raise HTTPException(status_code=400, detail="severity harus danger, warning, atau info")
        alerts = [a for a in alerts if a["severity"] == severity]
    return {"alerts": alerts, "count": len(alerts)}


@router.get("/admin/dashboard")
async def admin_dashboard(staff=Depends(get_current_staff)):
    db = await _get_db()
    await _mark_overdue(db)

    # ---- Scope filter: sales sees ONLY their assigned clients ----
    # Everyone else (admin/finance/support) sees the global tenant view.
    scope_user_ids = None
    if staff["role"] == "sales":
        scope_user_ids = [ObjectId(cid) for cid in (staff.get("assigned_client_ids") or [])]
        # Unassigned sales → all counts are zero (no doc matches _id:None).
        if not scope_user_ids:
            scope_user_ids = [ObjectId("000000000000000000000000")]

    if scope_user_ids is None:
        client_q = {"role": "client"}
        svc_q = {"status": "active"}
        tkt_q = {"status": {"$in": ["open", "awaiting_staff"]}}
    else:
        client_q = {"role": "client", "_id": {"$in": scope_user_ids}}
        svc_q = {"status": "active", "user_id": {"$in": scope_user_ids}}
        tkt_q = {"status": {"$in": ["open", "awaiting_staff"]}, "user_id": {"$in": scope_user_ids}}

    total_users = await db.users.count_documents(client_q)
    active_services = await db.services.count_documents(svc_q)
    open_tickets = await db.tickets.count_documents(tkt_q)

    stats = {
        "total_clients": total_users,
        "active_services": active_services,
        "open_tickets": open_tickets,
    }

    # Financial stats - visible to finance/admin OR sales (scoped to their book).
    if staff["role"] in FINANCE_ROLES or staff["role"] == "sales":
        inv_q_base = {} if scope_user_ids is None else {"user_id": {"$in": scope_user_ids}}
        unpaid = await db.invoices.count_documents({**inv_q_base, "status": "unpaid"})
        overdue = await db.invoices.count_documents({**inv_q_base, "status": "overdue"})
        pending_orders = await db.orders.count_documents({**inv_q_base, "status": "pending"})

        now = datetime.now(timezone.utc)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        paid_docs = await db.invoices.find(
            {**inv_q_base, "status": "paid", "paid_at": {"$gte": month_start}}
        ).to_list(1000)
        revenue_month = sum(d.get("total", 0) for d in paid_docs)
        all_paid = await db.invoices.find({**inv_q_base, "status": "paid"}).to_list(5000)
        revenue_total = sum(d.get("total", 0) for d in all_paid)
        overdue_docs = await db.invoices.find({**inv_q_base, "status": "overdue"}).to_list(1000)
        overdue_total = sum(d.get("total", 0) for d in overdue_docs)
        stats.update({
            "unpaid_invoices": unpaid,
            "overdue_invoices": overdue,
            "pending_orders": pending_orders,
            "revenue_month": revenue_month,
            "revenue_total": revenue_total,
            "overdue_total": overdue_total,
        })

    # ---- Tagihan terbaru (Ringkasan Umum) + Pusat Notifikasi ----
    recent_invoices = []
    if staff["role"] in FINANCE_ROLES or staff["role"] == "sales":
        inv_q_base = {} if scope_user_ids is None else {"user_id": {"$in": scope_user_ids}}
        recent_docs = await db.invoices.find(inv_q_base).sort("created_at", -1).to_list(5)
        uid_set = {d["user_id"] for d in recent_docs}
        user_map = {u["_id"]: u.get("name", "") async for u in db.users.find({"_id": {"$in": list(uid_set)}})}
        recent_invoices = [{
            "id": str(d["_id"]),
            "number": d.get("number", ""),
            "user_name": user_map.get(d["user_id"], ""),
            "total": d.get("total", 0),
            "status": d.get("status", "unpaid"),
            "due_date": d.get("due_date", ""),
        } for d in recent_docs]

    alerts = await _build_admin_alerts(db, staff, scope_user_ids)
    down_count = sum(1 for a in alerts if a["type"] == "device_down")

    # ---- System health: status real dari registry integrations ----
    health = [
        {"name": "API Backend", "status": "ok", "detail": "Online"},
        {"name": "MongoDB", "status": "ok", "detail": "Connected"},
    ]
    mikrotik_count = await db.mikrotik_devices.count_documents({})
    if mikrotik_count:
        health.append({
            "name": "MikroTik Ops",
            "status": "warn" if down_count else "ok",
            "detail": f"{mikrotik_count} device ({down_count} down)" if down_count else f"{mikrotik_count} device",
        })
    else:
        health.append({"name": "MikroTik Ops", "status": "off", "detail": "Not configured"})
    allow_extra = bool(await _get_setting_value(db, "enable_extra_payment_gateways", False))
    integ_docs = await db.integrations.find({}).sort("created_at", -1).to_list(100)
    for idoc in integ_docs:
        if not allow_extra and idoc.get("module") in _EXTRA_PAYMENT_MODULES:
            continue
        schema = module_schema(idoc.get("module", ""))
        enabled = idoc.get("status") == "enabled"
        last = idoc.get("last_test_result") or {}
        test_ok = last.get("ok") if isinstance(last, dict) else None
        health.append({
            "name": idoc.get("name") or (schema["label"] if schema else idoc.get("module", "")),
            "status": ("warn" if test_ok is False else "ok") if enabled else "off",
            "detail": ("Test failed" if test_ok is False else "Enabled") if enabled else "Disabled",
        })
    me_doc = await db.users.find_one({"_id": ObjectId(staff["id"])}) or {}
    es = me_doc.get("email_settings") or {}
    smtp_ok = bool(((es.get("smtp") or {}).get("credentials") or {}).get("host"))
    health.append({
        "name": "SMTP (personal)",
        "status": "ok" if smtp_ok else "off",
        "detail": "Configured" if smtp_ok else "Not configured",
    })

    return {"stats": stats, "role": staff["role"],
            "recent_invoices": recent_invoices, "alerts": alerts, "health": health}


# ============================================================
# Per-admin email settings (F1) - every staff member configures their
# own IMAP/SMTP so Admin ▸ Mail shows their personal inbox instead of
# a single shared mailbox. Stored on the user document under
# `email_settings` with the same shape the shared iv2 integration uses.
# ============================================================
def _mask_email_settings(es: dict) -> dict:
    """Return a copy with the password redacted for GET responses."""
    if not es:
        return {}
    def _redact(creds: dict) -> dict:
        c = dict(creds or {})
        if c.get("password"):
            c["password"] = "•" * 8
        return c
    imap_creds_legacy = _redact(es.get("credentials") or {})
    smtp_creds = _redact((es.get("smtp") or {}).get("credentials") or {})
    imap_creds = _redact((es.get("imap") or {}).get("credentials") or imap_creds_legacy)
    return {
        "smtp": {"credentials": smtp_creds,
                  "options": dict((es.get("smtp") or {}).get("options") or {})},
        "imap": {"credentials": imap_creds, "options": dict(es.get("options") or {})},
        "from_name": es.get("from_name") or "",
        "from_email": es.get("from_email") or (imap_creds.get("username") or ""),
        "configured": bool(imap_creds.get("host") and imap_creds.get("username")),
    }


@router.get("/settings/email")
async def get_my_email_settings(user=Depends(get_current_user)):
    """Return the calling staff member's IMAP/SMTP config (password masked)."""
    db = await _get_db()
    doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    return _mask_email_settings((doc or {}).get("email_settings") or {})


def _merge_email_payload(existing: dict, payload: dict) -> dict:
    """Merge an incoming (possibly password-masked) email-settings payload with
    what's already stored on the user document. Shared by the save endpoint and
    the test-connection endpoint so both resolve masked "••••••••" passwords
    back to the stored value."""
    def _merge(kind: str) -> dict:
        old_creds = ((existing.get(kind) or {}).get("credentials") or
                     (existing.get("credentials") if kind == "imap" else {})) or {}
        new = (payload or {}).get(kind) or {}
        pwd = new.get("password") or ""
        if not pwd or set(pwd) == {"•"}:
            pwd = old_creds.get("password") or ""
        return {
            "credentials": {
                "host":     (new.get("host") or old_creds.get("host") or "").strip(),
                "port":     int(new.get("port") or old_creds.get("port") or (993 if kind == "imap" else 465)),
                "username": (new.get("username") or old_creds.get("username") or "").strip(),
                "password": pwd,
            },
            "options": {"use_ssl": bool(new.get("use_ssl", True))},
        }

    imap = _merge("imap")
    return {
        "from_name":  ((payload or {}).get("from_name") or "").strip(),
        "from_email": ((payload or {}).get("from_email") or "").strip(),
        "imap": imap,
        "smtp": _merge("smtp"),
        # Legacy top-level fields kept for backward compat with IMAPClient
        "credentials": imap["credentials"],
        "options":     imap["options"],
    }


@router.post("/settings/email")
async def save_my_email_settings(payload: dict, user=Depends(get_current_user)):
    """Save this staff member's personal cPanel IMAP/SMTP credentials.

    Expected shape:
      {
        "from_name": "Anang Support",
        "from_email": "anang@intercloud-digital.com",
        "imap": {"host":"...", "port":993, "username":"...", "password":"...", "use_ssl":true},
        "smtp": {"host":"...", "port":465, "username":"...", "password":"...", "use_ssl":true},
      }
    Passwords equal to "•••••••" are treated as "unchanged" so operators
    can edit host/port without re-entering their password.
    """
    db = await _get_db()
    doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    existing = (doc or {}).get("email_settings") or {}
    stored = _merge_email_payload(existing, payload)
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$set": {"email_settings": stored}},
    )
    return _mask_email_settings(stored)


@router.post("/settings/email/test")
async def test_my_email_settings(payload: dict = None, user=Depends(get_current_user)):
    """Test IMAP **and** SMTP connectivity for this staff member's webmail.

    Accepts the same payload shape as POST /settings/email (so the setup modal
    can test what's currently typed before saving). Masked passwords fall back
    to the stored value; an empty/missing payload tests the saved settings.
    Returns per-protocol results:
      { "ok": bool, "imap": {"ok":…, "message":…}, "smtp": {"ok":…, "message":…} }
    """
    db = await _get_db()
    doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    existing = (doc or {}).get("email_settings") or {}
    merged = _merge_email_payload(existing, payload or {})

    def _test(kind: str) -> dict:
        cfg = merged[kind]
        c = cfg["credentials"]
        if not (c.get("host") and c.get("username") and c.get("password")):
            return {"ok": False,
                    "message": f"{kind.upper()} belum dikonfigurasi (host/username/password wajib diisi)."}
        try:
            if kind == "imap":
                return iv2.IMAPClient(cfg).test_connection()
            return iv2.SMTPMailer(cfg).test_connection()
        except Exception as e:  # pragma: no cover - test_connection already catches
            return {"ok": False, "message": f"{type(e).__name__}: {e}"}

    imap_res = _test("imap")
    smtp_res = _test("smtp")
    return {"ok": bool(imap_res.get("ok") and smtp_res.get("ok")),
            "imap": imap_res, "smtp": smtp_res}


@router.delete("/settings/email")
async def clear_my_email_settings(user=Depends(get_current_user)):
    db = await _get_db()
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$unset": {"email_settings": ""}},
    )
    return {"ok": True}


# Menu catalog (used by the Admin User Access modal to render per-menu checkboxes).
# The frontend PortalLayout.jsx must import ADMIN_MENU_CATALOG matching these keys.
ADMIN_MENU_CATALOG = [
    {"key": "dashboard",       "label": "Dashboard",        "group": "Overview",       "default_roles": ["admin", "sales", "finance", "support", "ticket_only", "creative"]},
    {"key": "orders",          "label": "Orders",           "group": "Sales & Billing", "default_roles": ["admin", "sales", "finance"]},
    {"key": "invoices",        "label": "Invoices",         "group": "Sales & Billing", "default_roles": ["admin", "finance"]},
    {"key": "quotations",      "label": "Quotations",       "group": "Sales & Billing", "default_roles": ["admin", "sales", "finance"]},
    {"key": "finance",         "label": "Finance",          "group": "Sales & Billing", "default_roles": ["admin", "finance"]},
    {"key": "assets",          "label": "Assets",           "group": "Sales & Billing", "default_roles": ["admin", "finance"]},
    {"key": "products",        "label": "Products",         "group": "Catalog",        "default_roles": ["admin", "support"]},
    {"key": "addons",          "label": "Add-ons",          "group": "Catalog",        "default_roles": ["admin", "support"]},
    {"key": "categories",      "label": "Categories",       "group": "Catalog",        "default_roles": ["admin"]},
    {"key": "services",        "label": "Services",         "group": "Catalog",        "default_roles": ["admin", "finance", "support"]},
    {"key": "users",           "label": "Users / Clients",  "group": "Support & CRM",  "default_roles": ["admin", "sales", "finance", "support"]},
    {"key": "tickets",         "label": "Tickets",          "group": "Support & CRM",  "default_roles": ["admin", "sales", "finance", "support", "ticket_only"]},
    {"key": "mail",            "label": "Webmail",          "group": "Support & CRM",  "default_roles": ["admin", "sales", "finance", "support"]},
    {"key": "email",           "label": "Email Automation", "group": "Support & CRM",  "default_roles": ["admin"]},
    {"key": "articles",        "label": "Articles",         "group": "Support & CRM",  "default_roles": ["admin", "sales", "finance", "support", "creative"]},
    {"key": "provisioning",    "label": "Provisioning",     "group": "Operations",     "default_roles": ["admin", "support"]},
    {"key": "mikrotik",        "label": "MikroTik Ops",     "group": "Operations",     "default_roles": ["admin", "support"]},
    {"key": "dcim",            "label": "DCIM & IPAM",      "group": "Operations",     "default_roles": ["admin", "support"]},
    {"key": "diagnostics",     "label": "Diagnostics",      "group": "Operations",     "default_roles": ["admin", "support"]},
    {"key": "crm",             "label": "Customer DB (CRM)","group": "Business",       "default_roles": ["admin", "sales", "finance", "support"]},
    {"key": "projects",        "label": "Project Tracker",  "group": "Business",       "default_roles": ["admin", "sales", "support"]},
    {"key": "content",         "label": "Content Planner",  "group": "Business",       "default_roles": ["admin", "sales", "finance", "support", "creative"]},
    {"key": "followups",       "label": "Follow-ups",       "group": "Business",       "default_roles": ["admin", "sales", "finance", "support"]},
    {"key": "documents",       "label": "Documents",        "group": "Business",       "default_roles": ["admin", "sales", "finance", "support"]},
    {"key": "integrations",    "label": "Integrations",     "group": "System",         "default_roles": ["admin"]},
    {"key": "security",        "label": "Security",         "group": "System",         "default_roles": ["admin"]},
    {"key": "audit_log",       "label": "Audit Log",        "group": "System",         "default_roles": ["admin"]},
    {"key": "noc",             "label": "NOC Monitor",      "group": "Operations",     "default_roles": ["admin", "support"]},
    {"key": "credit_notes",    "label": "Credit Notes",     "group": "Sales & Billing", "default_roles": ["admin", "finance"]},
    {"key": "owner_dashboard", "label": "Executive Overview","group": "Overview",       "default_roles": ["admin", "owner"]},
    {"key": "media_library",   "label": "Media Library",    "group": "Creative",       "default_roles": ["admin", "creative"]},
    {"key": "content_calendar","label": "Content Calendar", "group": "Creative",       "default_roles": ["admin", "creative"]},
    {"key": "utm_builder",     "label": "UTM Builder",      "group": "Creative",       "default_roles": ["admin", "creative", "sales"]},
    {"key": "branding",        "label": "Branding",         "group": "System",         "default_roles": ["admin"]},
    {"key": "site_content",    "label": "Landing CMS",      "group": "System",         "default_roles": ["admin"]},
    {"key": "backup",          "label": "Backup & Restore", "group": "System",         "default_roles": ["admin"]},
    {"key": "user_settings",   "label": "User Settings",    "group": "System",         "default_roles": ["admin"]},
]


FEATURE_FLAG_CATALOG = [
    # Delete / destructive
    {"key": "can_delete_invoices",   "label": "Can permanently delete invoices"},
    {"key": "can_delete_users",      "label": "Can delete user accounts"},
    {"key": "can_delete_tickets",    "label": "Can delete support tickets"},
    {"key": "can_run_factory_reset", "label": "Can trigger Factory Reset (DANGEROUS)"},
    # Financial
    {"key": "can_view_all_invoices", "label": "Can see ALL invoices (not just own clients)"},
    {"key": "can_edit_pricing",      "label": "Can edit product / add-on pricing"},
    {"key": "can_apply_discounts",   "label": "Can apply discounts on quotations"},
    {"key": "can_view_assets",       "label": "Can view Assets & depreciation reports"},
    {"key": "can_export_data",       "label": "Can export CRM / invoices to CSV"},
    # Technical / Ops
    {"key": "can_edit_dcim_devices", "label": "Can edit DCIM devices (racks / equipment)"},
    {"key": "can_edit_ip_prefixes",  "label": "Can allocate & edit IP prefixes"},
    {"key": "can_run_provisioning",  "label": "Can trigger auto-provisioning manually"},
    {"key": "can_control_mikrotik",  "label": "Can run live MikroTik ops (traffic-ctrl, blackhole)"},
    {"key": "can_run_diagnostics",   "label": "Can run diagnostics (ping / traceroute / BGP lookup)"},
    # System
    {"key": "can_manage_users",      "label": "Can create/edit staff users (assign roles)"},
    {"key": "can_manage_articles",   "label": "Can publish / edit knowledge-base articles"},
    {"key": "can_edit_branding",     "label": "Can update logo / colours / landing CMS"},
    {"key": "can_manage_integrations","label": "Can configure MikroTik / Telegram / SMTP integrations"},
    {"key": "can_run_backups",       "label": "Can trigger backups & restore snapshots"},
    # CRM / Sales
    {"key": "can_view_all_clients",  "label": "Can view ALL client accounts (not just assigned)"},
    {"key": "can_reassign_clients",  "label": "Can reassign clients to different sales staff"},
    {"key": "can_impersonate_client","label": "Can impersonate a client for support"},
    {"key": "can_view_email_automation", "label": "Can access Email Automation module"},
]


@router.get("/admin/user-access-catalog")
async def admin_user_access_catalog(admin=Depends(get_current_admin)):
    """Returns everything the User Access UI needs to render checkboxes.

    - `menu_catalog`: list of menu keys with human labels & the default roles that
      have access to each menu (used to show the default vs. the override).
    - `feature_flags`: list of extra per-user feature toggles.
    """
    return {
        "menu_catalog": ADMIN_MENU_CATALOG,
        "feature_flags": FEATURE_FLAG_CATALOG,
    }


def _paginate(items: list, page: Optional[int], limit: int):
    """Server-side pagination opsional: tanpa `page` kembalikan list penuh (kompatibel lama)."""
    if page is None:
        return items
    limit = max(1, min(int(limit or 25), 200))
    page = max(1, int(page))
    total = len(items)
    start = (page - 1) * limit
    return {"items": items[start:start + limit], "total": total, "page": page,
            "limit": limit, "pages": max(1, (total + limit - 1) // limit)}


@router.get("/admin/users")
async def admin_list_users(staff=Depends(get_current_staff),
                           page: Optional[int] = None, limit: int = 25):
    """Sales sees only their assigned clients; other staff see all."""
    db = await _get_db()
    if staff["role"] == "sales":
        ids = [ObjectId(x) for x in (staff.get("assigned_client_ids") or [])]
        docs = await db.users.find({"_id": {"$in": ids}}).to_list(500) if ids else []
    else:
        docs = await db.users.find({}).sort("created_at", -1).to_list(1000)
    return _paginate([_user_public(u) for u in docs], page, limit)


@router.get("/admin/users/{uid}")
async def admin_get_user(uid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    if staff["role"] == "sales" and not sales_can_access(staff, uid):
        raise HTTPException(status_code=403, detail="Not your client")
    u = await _load_user(db, uid)
    return _user_public(u)


@router.post("/admin/users", response_model=m.UserOut)
async def admin_create_user(payload: m.UserCreateIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    email = payload.email.lower().strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email already registered")
    doc = {
        "email": email,
        "password_hash": hash_password(payload.password),
        "name": payload.name,
        "role": payload.role,
        "company": payload.company,
        "phone": payload.phone,
        "assigned_client_ids": [ObjectId(x) for x in (payload.assigned_client_ids or [])],
        "billing_emails": payload.billing_emails or [],
        "attention": payload.attention,
        "address_line1": payload.address_line1,
        "address_line2": payload.address_line2,
        "city": payload.city,
        "province": payload.province,
        "postal_code": payload.postal_code,
        "country": payload.country or "Indonesia",
        "npwp": payload.npwp,
        "menu_keys": payload.menu_keys,
        "feature_flags": payload.feature_flags or [],
        "is_active": payload.is_active,
        "created_at": _now(),
    }
    r = await db.users.insert_one(doc)
    doc["_id"] = r.inserted_id
    # Mirror this new user into CRM (as "existing" client - admin-created)
    if payload.role == "client":
        try:
            await _upsert_crm_from_user(db, doc, status="existing",
                                        extra_notes="Registered by admin from Users console")
        except Exception:
            pass
        try:
            from portal import emails as _em
            await _em.on_user_registered(db, doc)
        except Exception:
            pass
    return _user_public(doc)


@router.put("/admin/users/{uid}", response_model=m.UserOut)
async def admin_update_user(uid: str, payload: m.UserUpdateIn, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    existing = await _load_user(db, uid)
    upd = {}
    for k in ("name", "role", "company", "phone", "billing_emails",
              "attention", "address_line1", "address_line2", "city",
              "province", "postal_code", "country", "npwp",
              "menu_keys", "feature_flags", "is_active"):
        v = getattr(payload, k, None)
        if v is not None:
            upd[k] = v
    if payload.assigned_client_ids is not None:
        upd["assigned_client_ids"] = [ObjectId(x) for x in payload.assigned_client_ids]
    password_changed = False
    if payload.password:
        upd["password_hash"] = hash_password(payload.password)
        password_changed = True
    if upd:
        await db.users.update_one({"_id": _oid(uid)}, {"$set": upd})
    u = await _load_user(db, uid)
    # ---- Audit: log role changes and password resets separately for clarity ----
    if existing.get("role") != u.get("role"):
        await log_audit(db, actor=admin, action="user.role_change", category="users",
                        target_type="user", target_id=uid, target_label=u.get("email", ""),
                        before={"role": existing.get("role")}, after={"role": u.get("role")},
                        severity="warning", request=request)
    if password_changed:
        await log_audit(db, actor=admin, action="user.password_change", category="security",
                        target_type="user", target_id=uid, target_label=u.get("email", ""),
                        severity="warning", request=request)
    if existing.get("is_active") != u.get("is_active"):
        await log_audit(db, actor=admin, action="user.active_toggle", category="users",
                        target_type="user", target_id=uid, target_label=u.get("email", ""),
                        before={"is_active": existing.get("is_active")},
                        after={"is_active": u.get("is_active")},
                        severity="warning", request=request)
    return _user_public(u)


@router.delete("/admin/users/{uid}")
async def admin_delete_user(uid: str, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    if str(admin["id"]) == uid:
        raise HTTPException(status_code=400, detail="You cannot delete yourself")
    target = await db.users.find_one({"_id": _oid(uid)})
    r = await db.users.delete_one({"_id": _oid(uid)})
    if r.deleted_count and target:
        await log_audit(db, actor=admin, action="user.delete", category="users",
                        target_type="user", target_id=uid,
                        target_label=target.get("email", ""),
                        before={"role": target.get("role"), "email": target.get("email")},
                        severity="critical", request=request)
    return {"deleted": r.deleted_count}


# Client-side billing email preferences
@router.get("/client/billing-emails")
async def client_billing_emails(user=Depends(get_current_user)):
    db = await _get_db()
    u = await db.users.find_one({"_id": ObjectId(user["id"])})
    return {"billing_emails": list(u.get("billing_emails") or [])}


@router.put("/client/billing-emails")
async def client_update_billing_emails(payload: m.BillingEmailsIn, user=Depends(get_current_user)):
    db = await _get_db()
    emails = [str(e).lower() for e in payload.billing_emails]
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$set": {"billing_emails": emails}},
    )
    return {"billing_emails": emails}


# Products
def _serialize_product(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "name": d["name"],
        "category": d.get("category", "other"),
        "description": d.get("description", ""),
        "price_monthly": d.get("price_monthly", 0),
        "setup_fee": d.get("setup_fee", 0),
        "billing_cycle": d.get("billing_cycle", "monthly"),
        "features": d.get("features", []),
        "is_active": d.get("is_active", True),
        "is_addon": d.get("is_addon", False),
        "applies_to_product_ids": [str(x) for x in (d.get("applies_to_product_ids") or [])],
        "applies_to_categories": list(d.get("applies_to_categories") or []),
        "option_groups": list(d.get("option_groups") or []),
        "stock_qty": d.get("stock_qty"),
        "sort_order": d.get("sort_order", 100),
        "created_at": _iso(d.get("created_at", "")),
    }


DEFAULT_CATEGORIES = [
    {"slug": "cloud",        "label": "Cloud",        "icon": "Cloud",       "sort_order": 10},
    {"slug": "vps",          "label": "VPS",          "icon": "Server",      "sort_order": 20},
    {"slug": "hosting",      "label": "Web Hosting",  "icon": "Globe",       "sort_order": 30},
    {"slug": "dedicated",    "label": "Dedicated",    "icon": "HardDrive",   "sort_order": 40},
    {"slug": "colocation",   "label": "Colocation",   "icon": "Building2",   "sort_order": 50},
    {"slug": "firewall",     "label": "Firewall",     "icon": "Shield",      "sort_order": 60},
    {"slug": "interconnect", "label": "Interconnect", "icon": "Network",     "sort_order": 70},
    {"slug": "lease",        "label": "Lease-to-Own", "icon": "Package",     "sort_order": 80},
    {"slug": "domain",       "label": "Domains",      "icon": "Globe2",      "sort_order": 90},
    {"slug": "other",        "label": "Other",        "icon": "Boxes",       "sort_order": 999},
]


async def _ensure_default_categories(db):
    for c in DEFAULT_CATEGORIES:
        if not await db.categories.find_one({"slug": c["slug"]}):
            await db.categories.insert_one({
                **c, "description": "", "is_active": True,
                "created_at": _now(),
            })


def _serialize_category(d: dict, product_count: int = 0) -> dict:
    return {
        "id": str(d["_id"]),
        "slug": d["slug"],
        "label": d.get("label", d["slug"]),
        "description": d.get("description", ""),
        "icon": d.get("icon", ""),
        "sort_order": d.get("sort_order", 100),
        "is_active": d.get("is_active", True),
        "product_count": product_count,
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/categories")
async def admin_list_categories(staff=Depends(get_current_staff)):
    db = await _get_db()
    await _ensure_default_categories(db)
    docs = await db.categories.find({}).sort("sort_order", 1).to_list(500)
    # Product counts
    counts = {}
    async for c in db.products.aggregate([
        {"$group": {"_id": "$category", "n": {"$sum": 1}}},
    ]):
        counts[c["_id"]] = c["n"]
    return [_serialize_category(d, counts.get(d["slug"], 0)) for d in docs]


@router.get("/portal-public/categories")
async def public_categories():
    db = await _get_db()
    await _ensure_default_categories(db)
    docs = await db.categories.find({"is_active": True}).sort("sort_order", 1).to_list(500)
    return [_serialize_category(d) for d in docs]


@router.post("/admin/categories")
async def admin_create_category(payload: m.CategoryIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    slug = payload.slug.lower().strip()
    if await db.categories.find_one({"slug": slug}):
        raise HTTPException(status_code=409, detail=f"Category '{slug}' already exists")
    doc = payload.model_dump()
    doc["slug"] = slug
    doc["created_at"] = _now()
    r = await db.categories.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_category(doc)


@router.put("/admin/categories/{cid}")
async def admin_update_category(cid: str, payload: m.CategoryIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    slug = payload.slug.lower().strip()
    current = await db.categories.find_one({"_id": _oid(cid)})
    if not current:
        raise HTTPException(status_code=404, detail="Category not found")
    # If slug changed, cascade-update all products
    if slug != current["slug"]:
        if await db.categories.find_one({"slug": slug, "_id": {"$ne": _oid(cid)}}):
            raise HTTPException(status_code=409, detail=f"Category '{slug}' already exists")
        await db.products.update_many({"category": current["slug"]}, {"$set": {"category": slug}})
    upd = payload.model_dump()
    upd["slug"] = slug
    await db.categories.update_one({"_id": _oid(cid)}, {"$set": upd})
    d = await db.categories.find_one({"_id": _oid(cid)})
    return _serialize_category(d)


@router.delete("/admin/categories/{cid}")
async def admin_delete_category(cid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await db.categories.find_one({"_id": _oid(cid)})
    if not d:
        raise HTTPException(status_code=404, detail="Category not found")
    if await db.products.count_documents({"category": d["slug"]}) > 0:
        raise HTTPException(status_code=400, detail="Cannot delete: products still use this category. Reassign them first.")
    r = await db.categories.delete_one({"_id": _oid(cid)})
    return {"deleted": r.deleted_count}


@router.get("/admin/products")
async def admin_list_products(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.products.find({}).sort([("sort_order", 1), ("created_at", -1)]).to_list(500)
    return [_serialize_product(d) for d in docs]


@router.get("/portal-public/products")
async def public_products():
    """Products list - public catalog used by client order flow."""
    db = (await _get_db())
    docs = await db.products.find({"is_active": True, "is_addon": {"$ne": True}}).sort([("sort_order", 1), ("category", 1)]).to_list(500)
    return [_serialize_product(d) for d in docs]


@router.get("/portal-public/addons")
async def public_addons(product_id: Optional[str] = None):
    """Add-on products - used by client Order flow to attach to a base product.
    Dengan ?product_id= hasil difilter sesuai kompatibilitas add-on (applies_to)."""
    db = (await _get_db())
    docs = await db.products.find({"is_active": True, "is_addon": True}).sort([("sort_order", 1), ("name", 1)]).to_list(500)
    rows = [_serialize_product(d) for d in docs]
    if product_id:
        prod = await db.products.find_one({"_id": _oid(product_id)})
        if not prod:
            raise HTTPException(status_code=404, detail="Product not found")
        cat = prod.get("category", "")

        def _compatible(a: dict) -> bool:
            pids = a.get("applies_to_product_ids") or []
            cats = a.get("applies_to_categories") or []
            if not pids and not cats:
                return True
            return product_id in pids or cat in cats

        rows = [a for a in rows if _compatible(a)]
    return rows


def _serialize_lead(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "email": d.get("email", ""),
        "phone": d.get("phone", ""),
        "company": d.get("company", ""),
        "need": d.get("need", ""),
        "message": d.get("message", ""),
        "source": d.get("source", "landing"),
        "status": d.get("status", "new"),
        "crm_id": str(d["crm_id"]) if d.get("crm_id") else None,
        "created_at": _iso(d.get("created_at", "")),
    }


@router.post("/portal-public/leads")
async def public_submit_lead(payload: m.LeadIn, request: Request):
    """Terima lead dari form landing page (publik). reCAPTCHA v3 diverifikasi bila aktif,
    lead disimpan dan otomatis disinkronkan ke CRM sebagai prospect."""
    db = await _get_db()
    remote_ip = request.client.host if request.client else None
    await iv2.enforce_recaptcha(db, payload.recaptcha_token, "lead", remote_ip)
    email = payload.email.lower()
    ten_min_ago = (datetime.now(timezone.utc) - timedelta(minutes=10)).isoformat()
    dup = await db.leads.find_one({"email": email, "created_at": {"$gte": ten_min_ago}})
    if dup:
        return {"ok": True, "lead_id": str(dup["_id"]), "duplicate": True}
    # Sinkron CRM: upsert prospect berdasarkan email
    crm = await db.crm_customers.find_one({"email": email})
    if crm:
        crm_id = crm["_id"]
    else:
        r = await db.crm_customers.insert_one({
            "name": payload.name, "email": email, "phone": payload.phone,
            "company": payload.company, "position": "", "industry": "",
            "status": "prospect",
            "notes": f"Lead dari landing page ({payload.need or 'umum'})",
            "created_at": _now(), "updated_at": _now(),
        })
        crm_id = r.inserted_id
    doc = {
        "name": payload.name,
        "email": email,
        "phone": payload.phone,
        "company": payload.company,
        "need": payload.need,
        "message": payload.message,
        "source": payload.source or "landing",
        "status": "new",
        "crm_id": crm_id,
        "ip": remote_ip,
        "created_at": _now(),
    }
    r = await db.leads.insert_one(doc)
    # Auto follow-up: buat tugas follow-up H+1 agar lead cepat dihubungi
    try:
        due = (datetime.now(timezone.utc) + timedelta(days=1)).date().isoformat()
        await db.followups.insert_one({
            "customer_id": crm_id, "customer_name": payload.name,
            "task": f"Follow up lead baru: {payload.name} ({payload.need or 'umum'})",
            "channel": "whatsapp", "due_date": due, "done": False,
            "owner": "auto", "created_at": _now(),
        })
    except Exception:
        pass
    return {"ok": True, "lead_id": str(r.inserted_id), "duplicate": False}


# ---------- Lead Form Builder ----------

_FB_FIELD_TYPES = {"text", "email", "phone", "textarea", "select", "checkbox", "number"}


def _fb_serialize(d: dict) -> dict:
    return {
        "id": str(d["_id"]), "name": d.get("name", ""), "slug": d.get("slug", ""),
        "title": d.get("title", ""), "description": d.get("description", ""),
        "submit_label": d.get("submit_label", "Kirim"),
        "success_message": d.get("success_message", "Terima kasih, kami akan segera menghubungi Anda."),
        "fields": d.get("fields", []), "active": d.get("active", True),
        "submissions": d.get("submissions", 0),
        "created_at": _iso(d.get("created_at", "")), "updated_at": _iso(d.get("updated_at", "")),
    }


def _fb_clean_fields(fields) -> list:
    out = []
    for i, f in enumerate(fields or []):
        ftype = str(f.get("type", "text")).lower()
        if ftype not in _FB_FIELD_TYPES:
            raise HTTPException(status_code=400, detail=f"Tipe field tidak dikenal: {ftype}")
        key = re.sub(r"[^a-z0-9_]+", "_", str(f.get("key") or f.get("label", f"field_{i}")).lower()).strip("_")
        if not key:
            raise HTTPException(status_code=400, detail="Field key kosong")
        out.append({
            "key": key, "label": str(f.get("label", key)), "type": ftype,
            "required": bool(f.get("required")), "placeholder": str(f.get("placeholder", "")),
            "options": [str(o) for o in (f.get("options") or [])],
            "order": int(f.get("order", i)),
        })
    out.sort(key=lambda x: x["order"])
    return out


@router.get("/admin/form-builder")
async def form_builder_list(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.form_configs.find({}).sort("created_at", -1).to_list(100)
    return [_fb_serialize(d) for d in docs]


@router.post("/admin/form-builder")
async def form_builder_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    name = str(payload.get("name", "")).strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nama form wajib diisi")
    slug = re.sub(r"[^a-z0-9-]+", "-", str(payload.get("slug") or name).lower()).strip("-")
    if await db.form_configs.find_one({"slug": slug}):
        raise HTTPException(status_code=409, detail=f"Slug '{slug}' sudah dipakai")
    doc = {
        "name": name, "slug": slug,
        "title": payload.get("title", name), "description": payload.get("description", ""),
        "submit_label": payload.get("submit_label", "Kirim"),
        "success_message": payload.get("success_message", "Terima kasih, kami akan segera menghubungi Anda."),
        "fields": _fb_clean_fields(payload.get("fields") or [
            {"key": "name", "label": "Nama lengkap", "type": "text", "required": True},
            {"key": "email", "label": "Email", "type": "email", "required": True},
            {"key": "phone", "label": "No. WhatsApp", "type": "phone", "required": False},
            {"key": "message", "label": "Pesan", "type": "textarea", "required": False},
        ]),
        "active": bool(payload.get("active", True)), "submissions": 0,
        "created_at": _now(), "updated_at": _now(),
    }
    r = await db.form_configs.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _fb_serialize(doc)


@router.put("/admin/form-builder/{fid}")
async def form_builder_update(fid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    d = await db.form_configs.find_one({"_id": _oid(fid)})
    if not d:
        raise HTTPException(status_code=404, detail="Form not found")
    upd = {"updated_at": _now()}
    for k in ("name", "title", "description", "submit_label", "success_message"):
        if k in payload:
            upd[k] = str(payload[k])
    if "active" in payload:
        upd["active"] = bool(payload["active"])
    if "fields" in payload:
        upd["fields"] = _fb_clean_fields(payload["fields"])
    await db.form_configs.update_one({"_id": d["_id"]}, {"$set": upd})
    d = await db.form_configs.find_one({"_id": d["_id"]})
    return _fb_serialize(d)


@router.delete("/admin/form-builder/{fid}")
async def form_builder_delete(fid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    await db.form_configs.delete_one({"_id": _oid(fid)})
    return {"ok": True}


@router.get("/portal-public/forms/{slug}")
async def public_form_config(slug: str):
    """Konfigurasi form publik untuk dirender di landing page."""
    db = await _get_db()
    d = await db.form_configs.find_one({"slug": slug, "active": True})
    if not d:
        raise HTTPException(status_code=404, detail="Form tidak ditemukan atau nonaktif")
    out = _fb_serialize(d)
    out.pop("submissions", None)
    return out


@router.post("/portal-public/forms/{slug}/submit")
async def public_form_submit(slug: str, payload: dict, request: Request):
    """Terima kiriman form dinamis: validasi server-side per field, simpan sebagai lead + CRM prospect."""
    db = await _get_db()
    d = await db.form_configs.find_one({"slug": slug, "active": True})
    if not d:
        raise HTTPException(status_code=404, detail="Form tidak ditemukan atau nonaktif")
    remote_ip = request.client.host if request.client else None
    await iv2.enforce_recaptcha(db, payload.get("recaptcha_token"), "lead", remote_ip)
    values, errors = {}, {}
    for f in d.get("fields", []):
        raw = payload.get(f["key"])
        val = ("" if raw is None else str(raw)).strip() if f["type"] != "checkbox" else bool(raw)
        if f["required"] and (val == "" or val is False):
            errors[f["key"]] = f"{f['label']} wajib diisi"
            continue
        if val == "" or val is None:
            values[f["key"]] = val
            continue
        if f["type"] == "email" and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", str(val)):
            errors[f["key"]] = "Format email tidak valid"
        elif f["type"] == "phone" and not re.match(r"^[0-9+()\-\s]{7,20}$", str(val)):
            errors[f["key"]] = "Format nomor telepon tidak valid"
        elif f["type"] == "number":
            try:
                float(val)
            except ValueError:
                errors[f["key"]] = "Harus berupa angka"
        elif f["type"] == "select" and f.get("options") and val not in f["options"]:
            errors[f["key"]] = "Pilihan tidak valid"
        values[f["key"]] = val
    if errors:
        raise HTTPException(status_code=422, detail={"errors": errors})
    email = str(values.get("email", "")).lower()
    name = str(values.get("name") or values.get("nama") or email or "Anonim")
    crm_id = None
    if email:
        crm = await db.crm_customers.find_one({"email": email})
        if crm:
            crm_id = crm["_id"]
        else:
            r = await db.crm_customers.insert_one({
                "name": name, "email": email, "phone": str(values.get("phone", "")),
                "company": str(values.get("company", "")), "position": "", "industry": "",
                "status": "prospect", "notes": f"Lead dari form '{d.get('name','')}'",
                "created_at": _now(), "updated_at": _now(),
            })
            crm_id = r.inserted_id
    lead = {
        "name": name, "email": email, "phone": str(values.get("phone", "")),
        "company": str(values.get("company", "")), "need": d.get("name", ""),
        "message": str(values.get("message", "")), "source": f"form:{slug}",
        "form_slug": slug, "form_values": values, "status": "new",
        "crm_id": crm_id, "ip": remote_ip, "created_at": _now(),
    }
    r = await db.leads.insert_one(lead)
    await db.form_configs.update_one({"_id": d["_id"]}, {"$inc": {"submissions": 1}})
    return {"ok": True, "lead_id": str(r.inserted_id),
            "message": d.get("success_message", "Terima kasih!")}


@router.get("/admin/leads")
async def admin_list_leads(status: Optional[str] = None, staff=Depends(get_current_staff)):
    _deny_creative(staff)
    db = await _get_db()
    q = {"status": status} if status else {}
    docs = await db.leads.find(q).sort("created_at", -1).to_list(1000)
    return [_serialize_lead(d) for d in docs]


@router.put("/admin/leads/{lid}/status")
async def admin_update_lead_status(lid: str, payload: m.LeadStatusIn, staff=Depends(get_current_staff)):
    _deny_creative(staff)
    db = await _get_db()
    await db.leads.update_one({"_id": _oid(lid)}, {"$set": {"status": payload.status}})
    d = await db.leads.find_one({"_id": _oid(lid)})
    if not d:
        raise HTTPException(status_code=404, detail="Lead not found")
    return _serialize_lead(d)


@router.post("/admin/products", response_model=m.ProductOut)
async def admin_create_product(payload: m.ProductIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    doc = payload.model_dump()
    doc["created_at"] = _now()
    r = await db.products.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_product(doc)


@router.put("/admin/products/{pid}", response_model=m.ProductOut)
async def admin_update_product(pid: str, payload: m.ProductIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    await db.products.update_one({"_id": _oid(pid)}, {"$set": payload.model_dump()})
    d = await db.products.find_one({"_id": _oid(pid)})
    if not d:
        raise HTTPException(status_code=404, detail="Product not found")
    return _serialize_product(d)


@router.delete("/admin/products/{pid}")
async def admin_delete_product(pid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    r = await db.products.delete_one({"_id": _oid(pid)})
    return {"deleted": r.deleted_count}


# ============================================================
# ORDER PREVIEW - build a WHMCS-style price cart WITHOUT persisting
# ============================================================

async def _price_cart(db, *, product: dict, selections: list, addon_ids: list, tax_percent: float = 11.0) -> dict:
    """Compute a full price breakdown from a product + option selections + add-ons."""
    base_monthly = float(product.get("price_monthly") or 0)
    base_setup = float(product.get("setup_fee") or 0)
    lines_options = []      # [{group_key, group_label, choice, monthly, setup}]
    monthly_options_sum = 0.0
    setup_options_sum = 0.0

    groups_by_key = {g.get("key"): g for g in (product.get("option_groups") or [])}
    for sel in (selections or []):
        gk = sel.get("group_key") if isinstance(sel, dict) else sel.group_key
        grp = groups_by_key.get(gk)
        if not grp:
            continue
        gtype = grp.get("type", "dropdown")
        glabel = grp.get("label", gk)
        if gtype == "quantity":
            qty = int((sel.get("quantity") if isinstance(sel, dict) else sel.quantity) or 0)
            if qty <= 0:
                continue
            unit_m = float(grp.get("unit_price_monthly") or 0)
            unit_s = float(grp.get("unit_price_setup") or 0)
            m_total = qty * unit_m
            s_total = qty * unit_s
            unit = grp.get("unit_label") or ""
            lines_options.append({
                "group_key": gk, "group_label": glabel,
                "choice": f"{qty} {unit}".strip(),
                "monthly": m_total, "setup": s_total,
            })
            monthly_options_sum += m_total
            setup_options_sum += s_total
        else:
            labels = (sel.get("option_labels") if isinstance(sel, dict) else sel.option_labels) or []
            for opt_lbl in labels:
                opt = next((o for o in (grp.get("options") or []) if o.get("label") == opt_lbl), None)
                if not opt:
                    continue
                m_delta = float(opt.get("price_monthly_delta") or 0)
                s_delta = float(opt.get("price_setup_delta") or 0)
                lines_options.append({
                    "group_key": gk, "group_label": glabel,
                    "choice": opt_lbl, "monthly": m_delta, "setup": s_delta,
                })
                monthly_options_sum += m_delta
                setup_options_sum += s_delta

    # Add-ons
    addon_lines = []
    if addon_ids:
        addon_docs = await db.products.find({"_id": {"$in": [_oid(x) for x in addon_ids]}, "is_addon": True}).to_list(50)
        for a in addon_docs:
            addon_lines.append({
                "id": str(a["_id"]),
                "name": a["name"],
                "monthly": float(a.get("price_monthly") or 0),
                "setup": float(a.get("setup_fee") or 0),
            })

    subtotal_monthly = base_monthly + monthly_options_sum + sum(x["monthly"] for x in addon_lines)
    setup_total = base_setup + setup_options_sum + sum(x["setup"] for x in addon_lines)
    # First-invoice basis = first month + setup fees
    first_invoice_subtotal = subtotal_monthly + setup_total
    tax_amount = round(first_invoice_subtotal * (tax_percent / 100.0), 2)
    total = round(first_invoice_subtotal + tax_amount, 2)
    return {
        "base_line": {
            "product_name": product["name"],
            "monthly": base_monthly,
            "setup": base_setup,
            "billing_cycle": product.get("billing_cycle", "monthly"),
        },
        "option_lines": lines_options,
        "addon_lines": addon_lines,
        "subtotal_monthly": round(subtotal_monthly, 2),
        "setup_total": round(setup_total, 2),
        "subtotal": round(first_invoice_subtotal, 2),
        "tax_percent": tax_percent,
        "tax_amount": tax_amount,
        "total": total,
    }


@router.post("/orders/preview")
async def order_preview(payload: m.OrderIn, user=Depends(get_current_user)):
    """Compute a full price breakdown for a client's cart WITHOUT creating an order.

    Used by the client's Order → Review step so users can confirm the total
    before we generate an invoice.
    """
    db = await _get_db()
    prod = await db.products.find_one({"_id": _oid(payload.product_id)})
    if not prod or prod.get("is_addon"):
        raise HTTPException(status_code=404, detail="Product not found")
    cart = await _price_cart(
        db, product=prod,
        selections=[s.model_dump() for s in (payload.selections or [])],
        addon_ids=payload.addon_ids or [],
        tax_percent=float(await _get_setting_value(db, "default_tax_percent", 11.0)),
    )
    return cart


# Orders
@router.get("/admin/orders")
async def admin_list_orders(staff=Depends(get_current_staff),
                            page: Optional[int] = None, limit: int = 25):
    _deny_creative(staff)
    db = await _get_db()
    q = {}
    if staff["role"] == "sales":
        # Sales sees only orders belonging to clients they're assigned to.
        assigned = [ObjectId(cid) for cid in (staff.get("assigned_client_ids") or [])]
        q = {"user_id": {"$in": assigned}} if assigned else {"_id": None}  # empty result if unassigned
    docs = await db.orders.find(q).sort("created_at", -1).to_list(1000)
    return _paginate([_serialize_order(d) for d in docs], page, limit)


@router.put("/admin/orders/{oid}/status")
async def admin_update_order_status(
    oid: str, payload: m.OrderStatusUpdateIn, admin=Depends(get_current_admin)
):
    """Manually move an order between statuses. Auto-provisioning primarily
    happens when the linked invoice is marked paid, but admins can still nudge
    the state machine (e.g. mark rejected)."""
    db = await _get_db()
    upd = {"status": payload.status}
    if payload.status == "assigned":
        upd["assigned_admin_id"] = ObjectId(admin["id"])
    await db.orders.update_one(
        {"_id": _oid(oid)},
        {"$set": upd,
         "$push": {"provision_log": {"at": _now(), "step": f"admin_set_{payload.status}",
                                      "message": f"Admin set status to {payload.status}."}}},
    )
    d = await db.orders.find_one({"_id": _oid(oid)})
    return _serialize_order(d)


# Client can flag a bank-transfer payment as sent; admin still needs to confirm.
@router.post("/client/orders/{oid}/confirm-transfer")
async def client_confirm_transfer(oid: str, payload: dict, user=Depends(get_current_user)):
    db = await _get_db()
    o = await db.orders.find_one({"_id": _oid(oid), "user_id": ObjectId(user["id"])})
    if not o:
        raise HTTPException(status_code=404, detail="Order not found")
    await db.orders.update_one(
        {"_id": o["_id"]},
        {"$set": {"status": "awaiting_verification"},
         "$push": {"provision_log": {"at": _now(), "step": "transfer_declared",
                                      "message": f"Client declared bank transfer. Ref: {payload.get('reference','-')}"}}},
    )
    d = await db.orders.find_one({"_id": o["_id"]})
    return _serialize_order(d)


# Invoices (staff - Sales sees only invoices of their assigned clients)
@router.get("/admin/invoices")
async def admin_list_invoices(staff=Depends(get_current_staff),
                              page: Optional[int] = None, limit: int = 25):
    _deny_creative(staff)
    db = await _get_db()
    await _mark_overdue(db)
    q = _sales_scope_filter(staff, key="user_id")
    docs = await db.invoices.find(q).sort("created_at", -1).to_list(2000)
    return _paginate([await _serialize_invoice(db, d) for d in docs], page, limit)


@router.get("/admin/invoices/{iid}")
async def admin_get_invoice(iid: str, staff=Depends(get_current_staff)):
    _deny_creative(staff)
    db = await _get_db()
    await _mark_overdue(db)
    q = {"_id": _oid(iid), **_sales_scope_filter(staff, key="user_id")}
    d = await db.invoices.find_one(q)
    if not d:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return await _serialize_invoice(db, d)


@router.post("/admin/invoices/{iid}/send")
async def admin_send_invoice(iid: str, staff=Depends(get_current_staff)):
    """Kirim invoice ke klien via email + siapkan link WhatsApp."""
    from portal import emails as _em
    db = await _get_db()
    d = await db.invoices.find_one({"_id": _oid(iid), **_sales_scope_filter(staff, key="user_id")})
    if not d:
        raise HTTPException(status_code=404, detail="Invoice not found")
    u = await db.users.find_one({"_id": d["user_id"]}) or {}
    email_sent = False
    try:
        await _em.on_invoice_generated(db, d, u)
        email_sent = True
    except Exception:
        email_sent = False
    total_str = "Rp " + f"{float(d.get('total') or 0):,.0f}".replace(",", ".")
    text = (f"Halo {u.get('name','')}, invoice {d.get('number','')} sebesar {total_str} "
            f"jatuh tempo {d.get('due_date','')}. Silakan login ke portal untuk membayar: "
            f"{os.environ.get('PORTAL_PUBLIC_URL', '')}/portal/login")
    phone = re.sub(r"[^0-9]", "", u.get("phone") or "")
    if phone.startswith("0"):
        phone = "62" + phone[1:]
    wa_link = f"https://wa.me/{phone}?text={quote(text)}" if phone else None
    return {"ok": True, "email_sent": email_sent, "email": u.get("email", ""),
            "wa_link": wa_link, "message": text}


@router.put("/admin/orders/{oid}")
async def admin_update_order(oid: str, payload: dict, staff=Depends(get_current_staff)):
    """Ubah detail pesanan (catatan + konfigurasi) sebelum/di tengah provisioning."""
    db = await _get_db()
    d = await db.orders.find_one({"_id": _oid(oid)})
    if not d:
        raise HTTPException(status_code=404, detail="Order not found")
    upd = {}
    if "notes" in payload:
        upd["notes"] = str(payload["notes"])
    if isinstance(payload.get("config"), dict):
        cfg = d.get("config") or {}
        cfg.update({k: v for k, v in payload["config"].items() if isinstance(k, str)})
        upd["config"] = cfg
    if not upd:
        raise HTTPException(status_code=400, detail="Tidak ada perubahan")
    upd["updated_at"] = _now()
    await db.orders.update_one({"_id": d["_id"]}, {
        "$set": upd,
        "$push": {"provision_log": {"at": _now(), "step": "order_updated",
                                    "message": f"Pesanan diubah oleh {staff.get('name','staff')}."}},
    })
    d = await db.orders.find_one({"_id": d["_id"]})
    return _serialize_order(d)


@router.post("/admin/invoices", response_model=m.InvoiceOut)
async def admin_create_invoice(payload: m.InvoiceIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    u = await _load_user(db, payload.user_id)
    subtotal = sum(i.total for i in payload.items)
    tax_amount = round(subtotal * payload.tax_percent / 100, 2)
    total = round(subtotal + tax_amount, 2)
    number = await _next_number(db, "invoices", "INV")
    doc = {
        "number": number,
        "user_id": u["_id"],
        "items": [i.model_dump() for i in payload.items],
        "subtotal": subtotal,
        "tax_percent": payload.tax_percent,
        "tax_amount": tax_amount,
        "total": total,
        "due_date": payload.due_date,
        "status": "unpaid",
        "payment_method": None,
        "paid_at": None,
        "notes": payload.notes,
        "created_at": _now(),
    }
    r = await db.invoices.insert_one(doc)
    doc["_id"] = r.inserted_id
    return await _serialize_invoice(db, doc)


async def _apply_pending_upgrade(db, inv: dict) -> bool:
    """Terapkan upgrade resource setelah invoice selisih dibayar (idempotent).

    Dipanggil dari semua jalur invoice -> paid (webhook Duitku, admin mark-paid,
    settle via credit note). Best-effort live resize cores/memory di Proxmox."""
    up = inv.get("upgrade")
    sid = inv.get("service_id")
    if not (up and sid):
        return False
    try:
        svc = await db.services.find_one({"_id": ObjectId(sid)})
    except Exception:
        return False
    pending = (svc or {}).get("pending_upgrade")
    if not svc or not pending or pending.get("invoice_id") != str(inv["_id"]):
        return False

    def _num(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return 0

    cfg = svc.get("config") or {}
    new_cpu = _num(cfg.get("cpu")) + int(up.get("cpu") or 0)
    new_ram = _num(cfg.get("ram_gb")) + int(up.get("ram_gb") or 0)
    new_disk = _num(cfg.get("disk_gb")) + int(up.get("disk_gb") or 0)
    await db.services.update_one(
        {"_id": svc["_id"]},
        {"$set": {"config.cpu": new_cpu, "config.ram_gb": new_ram, "config.disk_gb": new_disk},
         "$inc": {"price_monthly": float(up.get("monthly_delta") or 0)},
         "$unset": {"pending_upgrade": ""},
         "$push": {"self_service_log": {"at": _now(), "action": "upgrade_applied",
                                         "by": f"billing (invoice {inv.get('number', '')})"}}})
    try:
        s = await iv2.get_settings(db, "proxmox")
        if s and s.get("enabled") and cfg.get("node") and cfg.get("vmid"):
            body = {}
            if up.get("cpu"):
                body["cores"] = new_cpu
            if up.get("ram_gb"):
                body["memory"] = new_ram * 1024
            if body:
                await iv2.ProxmoxClient(s)._post(
                    f"/nodes/{cfg['node']}/qemu/{int(cfg['vmid'])}/config", body)
    except Exception:
        pass
    return True


@router.put("/admin/invoices/{iid}/status")
async def admin_update_invoice_status(
    iid: str, payload: m.InvoiceStatusIn, admin=Depends(get_current_admin)
):
    db = await _get_db()
    upd = {"status": payload.status}
    if payload.status == "paid":
        upd["paid_at"] = _now()
        upd["payment_method"] = payload.payment_method or "bank_transfer"
    await db.invoices.update_one({"_id": _oid(iid)}, {"$set": upd})
    d = await db.invoices.find_one({"_id": _oid(iid)})
    if not d:
        raise HTTPException(status_code=404, detail="Invoice not found")

    # If payment just confirmed AND invoice is linked to an order → auto-provision.
    if payload.status == "paid" and d.get("order_id"):
        order = await db.orders.find_one({"_id": _oid(d["order_id"])})
        if order and not order.get("service_id"):
            await db.orders.update_one(
                {"_id": order["_id"]},
                {"$set": {"status": "payment_verified"},
                 "$push": {"provision_log": {"at": _now(), "step": "payment_verified",
                                              "message": f"Payment received for invoice {d['number']}."}}},
            )
            order = await db.orders.find_one({"_id": order["_id"]})
            await _auto_provision(db, order)

    # Eksekusi upgrade resource yang menunggu pembayaran invoice ini.
    if payload.status == "paid":
        await _apply_pending_upgrade(db, d)
        await _auto_register_domain(db, d)
        await _apply_domain_renewal(db, d)

    return await _serialize_invoice(db, d)


# Quotations
async def _serialize_quotation(db, d: dict) -> dict:
    u = await db.users.find_one({"_id": d["user_id"]}) or {}
    return {
        "id": str(d["_id"]),
        "number": d["number"],
        "user_id": str(d["user_id"]),
        "user_name": u.get("name", ""),
        "user_email": u.get("email", ""),
        "items": d.get("items", []),
        "subtotal": d.get("subtotal", 0),
        "tax_amount": d.get("tax_amount", 0),
        "total": d.get("total", 0),
        "valid_until": d.get("valid_until", ""),
        "status": d.get("status", "draft"),
        "created_at": _iso(d.get("created_at", "")),
        "notes": d.get("notes", ""),
        "converted_invoice_id": d.get("converted_invoice_id"),
        "converted_invoice_number": d.get("converted_invoice_number"),
    }


@router.get("/admin/quotations")
async def admin_list_quotations(staff=Depends(get_current_staff)):
    _deny_creative(staff)
    db = await _get_db()
    q = {}
    if staff["role"] == "sales":
        assigned = [ObjectId(cid) for cid in (staff.get("assigned_client_ids") or [])]
        q = {"user_id": {"$in": assigned}} if assigned else {"_id": None}
    docs = await db.quotations.find(q).sort("created_at", -1).to_list(1000)
    return [await _serialize_quotation(db, d) for d in docs]


@router.post("/admin/quotations", response_model=m.QuotationOut)
async def admin_create_quotation(payload: m.QuotationIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    u = await _load_user(db, payload.user_id)
    subtotal = sum(i.total for i in payload.items)
    tax_amount = round(subtotal * payload.tax_percent / 100, 2)
    total = round(subtotal + tax_amount, 2)
    number = await _next_number(db, "quotations", "QTN")
    doc = {
        "number": number,
        "user_id": u["_id"],
        "items": [i.model_dump() for i in payload.items],
        "subtotal": subtotal,
        "tax_percent": payload.tax_percent,
        "tax_amount": tax_amount,
        "total": total,
        "valid_until": payload.valid_until,
        "status": "draft",
        "notes": payload.notes,
        "created_at": _now(),
    }
    r = await db.quotations.insert_one(doc)
    doc["_id"] = r.inserted_id
    return await _serialize_quotation(db, doc)


@router.put("/admin/quotations/{qid}/status")
async def admin_update_quotation_status(
    qid: str, payload: m.QuotationStatusIn, admin=Depends(get_current_admin)
):
    db = await _get_db()
    await db.quotations.update_one({"_id": _oid(qid)}, {"$set": {"status": payload.status}})
    d = await db.quotations.find_one({"_id": _oid(qid)})
    if not d:
        raise HTTPException(status_code=404, detail="Quotation not found")
    return await _serialize_quotation(db, d)


@router.post("/admin/quotations/{qid}/convert-to-invoice", response_model=m.InvoiceOut)
async def admin_convert_quotation_to_invoice(qid: str, payload: m.QuotationConvertIn,
                                             request: Request, admin=Depends(get_current_admin)):
    """Buat invoice dari quotation (menyalin item/pajak). Idempotent: quotation yang
    sudah dikonversi tidak bisa dikonversi ulang."""
    db = await _get_db()
    q = await db.quotations.find_one({"_id": _oid(qid)})
    if not q:
        raise HTTPException(status_code=404, detail="Quotation not found")
    if q.get("converted_invoice_id"):
        raise HTTPException(status_code=400, detail="Quotation ini sudah dikonversi menjadi invoice")
    due = payload.due_date or (datetime.now(timezone.utc) + timedelta(days=7)).date().isoformat()
    items = q.get("items") or []
    subtotal = sum(float(i.get("total") or 0) for i in items)
    tax_percent = float(q.get("tax_percent") or 0)
    tax_amount = round(subtotal * tax_percent / 100, 2)
    doc = {
        "number": await _next_number(db, "invoices", "INV"),
        "user_id": q["user_id"],
        "items": items,
        "subtotal": subtotal,
        "tax_percent": tax_percent,
        "tax_amount": tax_amount,
        "total": round(subtotal + tax_amount, 2),
        "due_date": due,
        "status": "unpaid",
        "payment_method": None,
        "paid_at": None,
        "notes": q.get("notes") or f"Dibuat dari quotation {q.get('number', '')}.",
        "source_quotation_id": str(q["_id"]),
        "source_quotation_number": q.get("number", ""),
        "created_at": _now(),
    }
    r = await db.invoices.insert_one(doc)
    doc["_id"] = r.inserted_id
    await db.quotations.update_one({"_id": q["_id"]}, {"$set": {
        "status": "accepted", "converted_invoice_id": str(r.inserted_id),
        "converted_invoice_number": doc["number"], "converted_at": _now()}})
    await log_audit(db, actor=admin, action="quotation.convert_to_invoice", category="billing",
                    target_type="quotation", target_id=qid, target_label=q.get("number", ""),
                    metadata={"invoice": doc["number"], "total": doc["total"]}, request=request)
    return await _serialize_invoice(db, doc)


# Tickets (staff - any staff role can view/reply)
@router.get("/admin/tickets")
async def admin_list_tickets(staff=Depends(get_current_staff),
                             device_id: Optional[str] = None,
                             view: Optional[str] = None,
                             page: Optional[int] = None, limit: int = 25):
    db = await _get_db()
    query: dict = {}
    if device_id:
        query["related_device_id"] = device_id
    if view == "active":
        query["status"] = {"$ne": "closed"}
    elif view == "archive":
        query["status"] = "closed"
    docs = await db.tickets.find(query).sort("updated_at", -1).to_list(2000)
    return _paginate([await _serialize_ticket(db, d) for d in docs], page, limit)


@router.post("/admin/tickets/{tid}/replies")
async def admin_reply_ticket(tid: str, payload: m.TicketReplyIn, staff=Depends(get_current_staff)):
    db = await _get_db()
    d = await db.tickets.find_one({"_id": _oid(tid)})
    if not d:
        raise HTTPException(status_code=404, detail="Ticket not found")
    reply = {
        "author_id": staff["id"],
        "author_name": staff["name"],
        "author_role": staff["role"],
        "message": payload.message,
        "internal": bool(payload.internal),
        "created_at": _now(),
    }
    upd = {"updated_at": _now()}
    if not payload.internal:
        upd["status"] = "awaiting_client"
    await db.tickets.update_one(
        {"_id": d["_id"]},
        {"$push": {"replies": reply}, "$set": upd},
    )
    d = await db.tickets.find_one({"_id": d["_id"]})
    return await _serialize_ticket(db, d)


@router.get("/admin/tickets/{tid}/timeline")
async def admin_ticket_timeline(tid: str, staff=Depends(get_current_staff)):
    """Timeline gabungan: pembuatan tiket, balasan (internal + publik), dan perubahan status."""
    db = await _get_db()
    d = await db.tickets.find_one({"_id": _oid(tid)})
    if not d:
        raise HTTPException(status_code=404, detail="Ticket not found")
    events = [{"kind": "created", "at": _iso(d.get("created_at", "")),
               "actor": d.get("user_name", ""), "message": f"Tiket {d.get('number','')} dibuat."}]
    for r in d.get("replies", []):
        events.append({
            "kind": "status" if r.get("author_role") == "system" else ("internal_note" if r.get("internal") else "reply"),
            "at": _iso(r.get("created_at", "")), "actor": r.get("author_name", ""),
            "role": r.get("author_role", ""), "message": r.get("message", ""),
            "internal": bool(r.get("internal")),
        })
    if d.get("closed_at"):
        events.append({"kind": "closed", "at": _iso(d["closed_at"]),
                       "actor": d.get("closed_by", ""), "message": "Tiket ditutup."})
    events.sort(key=lambda e: e["at"])
    return {"ticket_id": str(d["_id"]), "number": d.get("number", ""), "events": events}


@router.get("/admin/noc/devices/{did}/tickets")
async def admin_tickets_by_device(did: str, staff=Depends(get_current_staff)):
    """Daftar tiket yang terkait dengan sebuah perangkat NOC."""
    db = await _get_db()
    docs = await db.tickets.find({"related_device_id": did}).sort("updated_at", -1).to_list(200)
    return [await _serialize_ticket(db, d) for d in docs]


# Finance
@router.get("/admin/finance/summary")
async def admin_finance_summary(admin=Depends(get_current_admin)):
    db = await _get_db()
    # Aggregate by month for last 12 months
    paid = await db.invoices.find({"status": "paid"}).to_list(5000)
    by_month = {}
    for inv in paid:
        p = inv.get("paid_at") or inv.get("created_at", "")
        if not p:
            continue
        key = p[:7]  # YYYY-MM
        by_month[key] = by_month.get(key, 0) + inv.get("total", 0)
    series = sorted(
        [{"month": k, "revenue": v} for k, v in by_month.items()],
        key=lambda x: x["month"],
    )
    unpaid = await db.invoices.find({"status": {"$in": ["unpaid", "overdue"]}}).to_list(2000)
    outstanding = sum(d.get("total", 0) for d in unpaid)
    total_revenue = sum(d.get("total", 0) for d in paid)
    return {
        "total_revenue": total_revenue,
        "outstanding": outstanding,
        "paid_invoices": len(paid),
        "monthly_series": series,
    }


# Services (admin)
def _serialize_service(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "user_id": str(d["user_id"]),
        "product_id": str(d["product_id"]),
        "product_name": d.get("product_name", ""),
        "category": d.get("category", ""),
        "name": d.get("name", ""),
        "status": d.get("status", "active"),
        "start_date": d.get("start_date", ""),
        "next_renewal": d.get("next_renewal", ""),
        "price_monthly": d.get("price_monthly", 0),
        "config": d.get("config", {}),
        "order_id": d.get("order_id"),
    }


@router.get("/admin/services")
async def admin_list_services(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.services.find({}).sort("created_at", -1).to_list(2000)
    return [_serialize_service(d) for d in docs]


@router.get("/admin/services/{sid}/detail")
async def admin_service_detail(sid: str, admin=Depends(get_current_admin)):
    """Service + client + provisioning trail for the admin detail modal."""
    db = await _get_db()
    d = await db.services.find_one({"_id": _oid(sid)})
    if not d:
        raise HTTPException(status_code=404, detail="Service not found")
    out = _serialize_service(d)
    u = await db.users.find_one({"_id": d["user_id"]}) or {}
    out["user"] = {"name": u.get("name", ""), "email": u.get("email", ""), "company": u.get("company", "")}
    order = None
    if d.get("order_id"):
        try:
            order = await db.orders.find_one({"_id": ObjectId(d["order_id"])})
        except Exception:
            order = None
    out["provision_log"] = (order or {}).get("provision_log", [])
    out["self_service_log"] = (d.get("self_service_log") or [])[-10:]
    out["pending_upgrade"] = d.get("pending_upgrade")
    return out


@router.post("/admin/users/{uid}/impersonate")
async def admin_impersonate_client(uid: str, request: Request, admin=Depends(get_current_admin)):
    """Login-as-client untuk troubleshooting. Hanya klien aktif, tercatat di audit log."""
    db = await _get_db()
    target = await db.users.find_one({"_id": _oid(uid)})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") != "client":
        raise HTTPException(status_code=400, detail="Hanya akun klien yang bisa diimpersonasi")
    if target.get("status") == "suspended":
        raise HTTPException(status_code=400, detail="Akun klien sedang disuspensi")
    token = create_access_token(str(target["_id"]), target["email"], "client")
    await log_audit(db, actor=admin, action="user.impersonate", category="security",
                    target_type="user", target_id=str(target["_id"]),
                    target_label=target.get("email", ""), severity="warning",
                    metadata={"admin_email": admin.get("email", "")}, request=request)
    return {"token": token,
            "user": {"id": str(target["_id"]), "name": target.get("name", ""),
                     "email": target.get("email", ""), "role": "client"}}


@router.get("/admin/users/{uid}/profile")
async def admin_user_profile(uid: str, admin=Depends(get_current_admin)):
    """Client 360 profile: services (hosting accounts highlighted), billing summary."""
    db = await _get_db()
    u = await db.users.find_one({"_id": _oid(uid)})
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    svc_docs = await db.services.find({"user_id": u["_id"]}).sort("created_at", -1).to_list(200)
    services = [_serialize_service(s) for s in svc_docs]
    inv_docs = await db.invoices.find({"user_id": u["_id"]}).sort("created_at", -1).to_list(500)
    outstanding = sum(i.get("total", 0) for i in inv_docs if i.get("status") in ("unpaid", "overdue"))
    today = datetime.now(timezone.utc).date()
    overdue_list = []
    for i in inv_docs:
        if i.get("status") != "overdue":
            continue
        try:
            days = (today - datetime.strptime((i.get("due_date") or "")[:10], "%Y-%m-%d").date()).days
        except Exception:
            days = 0
        overdue_list.append({"id": str(i["_id"]), "number": i.get("number", ""),
                             "total": i.get("total", 0), "due_date": i.get("due_date", ""),
                             "days_past_due": max(0, days)})
    suspended_list = [{"id": str(s["_id"]), "name": s.get("name") or s.get("product_name", ""),
                       "reason": s.get("suspended_reason", ""), "suspended_at": s.get("suspended_at", "")}
                      for s in svc_docs if s.get("status") == "suspended"]
    max_days = max((o["days_past_due"] for o in overdue_list), default=0)
    if suspended_list:
        dunning_level = "suspended"
    elif max_days >= 7:
        dunning_level = "urgent"
    elif overdue_list:
        dunning_level = "reminder"
    else:
        dunning_level = "clear"
    return {
        "user": {"id": str(u["_id"]), "name": u.get("name", ""), "email": u.get("email", ""),
                 "company": u.get("company", ""), "phone": u.get("phone", ""),
                 "role": u.get("role", "client"), "created_at": _iso(u.get("created_at", ""))},
        "services": services,
        "hosting_accounts": [s for s in services if s.get("category") == "hosting"],
        "stats": {
            "orders": await db.orders.count_documents({"user_id": u["_id"]}),
            "tickets": await db.tickets.count_documents({"user_id": u["_id"]}),
            "invoices": len(inv_docs),
            "outstanding": round(outstanding, 2),
        },
        "recent_invoices": [{
            "id": str(i["_id"]), "number": i.get("number", ""), "total": i.get("total", 0),
            "status": i.get("status", "unpaid"), "due_date": i.get("due_date", ""),
        } for i in inv_docs[:5]],
        "dunning": {
            "level": dunning_level,
            "overdue_invoices": overdue_list,
            "max_days_past_due": max_days,
            "suspended_services": suspended_list,
        },
    }


@router.post("/admin/services")
async def admin_create_service(payload: m.ServiceCreateIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    u = await _load_user(db, payload.user_id)
    prod = await db.products.find_one({"_id": _oid(payload.product_id)})
    if not prod:
        raise HTTPException(status_code=404, detail="Product not found")
    now = datetime.now(timezone.utc)
    doc = {
        "user_id": u["_id"],
        "product_id": prod["_id"],
        "product_name": prod["name"],
        "category": prod.get("category", "other"),
        "name": payload.name,
        "status": payload.status,
        "start_date": now.date().isoformat(),
        "next_renewal": (now + timedelta(days=30)).date().isoformat(),
        "price_monthly": payload.price_monthly or prod.get("price_monthly", 0),
        "config": payload.config,
        "created_at": _now(),
    }
    r = await db.services.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_service(doc)


# Payment gateway config lives inside /admin/integrations (WHMCS-style module hub).


@router.get("/client/payment-info")
async def client_payment_info(user=Depends(get_current_user)):
    """Bank accounts + Duitku availability visible to clients (read from integrations)."""
    db = await _get_db()
    # Read bank accounts from settings (admin-editable)
    bank_doc = await db.settings.find_one({"key": "bank_accounts"}) or {}
    banks = bank_doc.get("value") or [
        {"bank": "MANDIRI", "number": "1240011911816", "holder": "INTERCLOUD DIGITAL INOVASI"},
        {"bank": "BCA", "number": "4730862038", "holder": "ANANG MADIA CUGITA"},
    ]
    # Any enabled duitku integration?
    duitku = await db.integrations.find_one({"module": "duitku", "status": "enabled"})
    if not duitku:
        # iv2-style settings count too (either storage system may hold the creds)
        iv2_duitku = await iv2.get_settings(db, "duitku")
        duitku = iv2_duitku if (iv2_duitku or {}).get("enabled") else None
    return {
        "bank_accounts": banks,
        "duitku_enabled": bool(duitku),
    }


# ============================================================
# BILLING DEFAULTS (admin-editable global settings)
# ============================================================
@router.get("/admin/billing/settings")
async def get_billing_settings(staff=Depends(get_current_staff)):
    """Global billing defaults. `default_tax_percent` is only the *suggested*
    initial PPN % pre-filled into new invoices/quotations and renewal
    auto-invoices - always overridable per document, down to 0."""
    db = await _get_db()
    return {k: await _get_setting_value(db, k, dv)
            for k, dv in BILLING_SETTING_DEFAULTS.items()}


@router.put("/admin/billing/settings")
async def put_billing_settings(payload: dict, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    before = {k: await _get_setting_value(db, k, dv)
              for k, dv in BILLING_SETTING_DEFAULTS.items()}
    if "default_tax_percent" in payload and payload["default_tax_percent"] is not None:
        await _set_setting_value(db, "default_tax_percent",
                                 max(0.0, float(payload["default_tax_percent"])))
    if "renewal_lead_days" in payload and payload["renewal_lead_days"] is not None:
        await _set_setting_value(db, "renewal_lead_days",
                                 max(1, int(payload["renewal_lead_days"])))
    if "enable_extra_payment_gateways" in payload and payload["enable_extra_payment_gateways"] is not None:
        await _set_setting_value(db, "enable_extra_payment_gateways",
                                 bool(payload["enable_extra_payment_gateways"]))
    if "noc_alert_recipients" in payload and payload["noc_alert_recipients"] is not None:
        v = payload["noc_alert_recipients"]
        if isinstance(v, str):
            v = [x for x in v.replace(",", "\n").splitlines()]
        cleaned = [str(x).strip() for x in (v or []) if str(x).strip()]
        await _set_setting_value(db, "noc_alert_recipients", cleaned)
    after = {k: await _get_setting_value(db, k, dv)
             for k, dv in BILLING_SETTING_DEFAULTS.items()}
    if before != after:
        await log_audit(db, actor=admin, action="billing.settings_update", category="billing",
                        target_type="settings", target_label="Billing Defaults",
                        before=before, after=after, severity="info", request=request)
    return after


@router.post("/admin/billing/run-renewal-sweep")
async def run_renewal_sweep_now(admin=Depends(get_current_admin)):
    """Manually trigger the renewal auto-invoice sweep (same job the hourly
    scheduler runs). Idempotent - re-running never duplicates invoices."""
    db = await _get_db()
    from portal import emails as _em
    return await _em.run_renewal_invoice_sweep(db)


# ============================================================
# INTEGRATIONS (WHMCS-style module hub)
# ============================================================
from .integrations_registry import (
    module_list, module_schema, redact, mock_test_connection,
)


@router.get("/admin/integrations/modules")
async def list_integration_modules(admin=Depends(get_current_admin)):
    """Return the module registry (schemas for the Add Server dialog).

    Midtrans/Xendit stay hidden unless `enable_extra_payment_gateways` is on -
    Duitku is the only active payment gateway by policy."""
    db = await _get_db()
    allow_extra = bool(await _get_setting_value(db, "enable_extra_payment_gateways", False))
    return [mdl for mdl in module_list()
            if allow_extra or mdl["key"] not in _EXTRA_PAYMENT_MODULES]


def _serialize_integration(d: dict, hide_secrets: bool = True) -> dict:
    schema = module_schema(d.get("module", ""))
    cfg = d.get("config", {}) or {}
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "module": d.get("module", ""),
        "module_label": schema["label"] if schema else d.get("module", ""),
        "category": schema["category"] if schema else "other",
        "config": redact(cfg, schema) if hide_secrets else cfg,
        "status": d.get("status", "disabled"),
        "last_test_at": d.get("last_test_at"),
        "last_test_result": d.get("last_test_result"),
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }


@router.get("/admin/integrations")
async def list_integrations(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.integrations.find({}).sort("created_at", -1).to_list(500)
    allow_extra = bool(await _get_setting_value(db, "enable_extra_payment_gateways", False))
    if not allow_extra:
        docs = [d for d in docs if d.get("module") not in _EXTRA_PAYMENT_MODULES]
    return [_serialize_integration(d) for d in docs]


@router.post("/admin/integrations")
async def create_integration(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    module = payload.get("module")
    if not module_schema(module):
        raise HTTPException(status_code=400, detail=f"Unknown module: {module}")
    doc = {
        "name": payload.get("name") or f"{module_schema(module)['label']} {int(datetime.now(timezone.utc).timestamp())}",
        "module": module,
        "config": payload.get("config", {}),
        "status": payload.get("status", "disabled"),
        "last_test_at": None,
        "last_test_result": None,
        "created_at": _now(),
        "updated_at": _now(),
    }
    r = await db.integrations.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_integration(doc)


@router.put("/admin/integrations/{iid}")
async def update_integration(iid: str, payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    existing = await db.integrations.find_one({"_id": _oid(iid)})
    if not existing:
        raise HTTPException(status_code=404, detail="Integration not found")
    schema = module_schema(existing["module"])
    # Merge config so masked secret fields aren't wiped
    new_cfg = payload.get("config", {})
    merged = dict(existing.get("config", {}))
    if schema:
        for f in schema["fields"]:
            if f["key"] in new_cfg:
                val = new_cfg[f["key"]]
                # Skip masked placeholder
                if f["type"] == "password" and isinstance(val, str) and val.strip() in ("••••••••", "", "*", None):
                    continue
                merged[f["key"]] = val
    upd = {
        "name": payload.get("name", existing["name"]),
        "config": merged,
        "status": payload.get("status", existing.get("status", "disabled")),
        "updated_at": _now(),
    }
    await db.integrations.update_one({"_id": existing["_id"]}, {"$set": upd})
    d = await db.integrations.find_one({"_id": existing["_id"]})
    return _serialize_integration(d)


@router.delete("/admin/integrations/{iid}")
async def delete_integration(iid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    r = await db.integrations.delete_one({"_id": _oid(iid)})
    return {"deleted": r.deleted_count}


@router.post("/admin/integrations/{iid}/test")
async def test_integration(iid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await db.integrations.find_one({"_id": _oid(iid)})
    if not d:
        raise HTTPException(status_code=404, detail="Integration not found")
    result = mock_test_connection(d["module"], d.get("config", {}))
    await db.integrations.update_one(
        {"_id": d["_id"]},
        {"$set": {"last_test_at": _now(), "last_test_result": result}},
    )
    return result


@router.post("/admin/integrations/test-config")
async def test_integration_draft(payload: dict, admin=Depends(get_current_admin)):
    """Test connection with an unsaved config (used by the Add Server dialog)."""
    return mock_test_connection(payload.get("module", ""), payload.get("config", {}))


# Bank accounts admin CRUD (simple)
@router.get("/admin/bank-accounts")
async def get_bank_accounts(admin=Depends(get_current_admin)):
    db = await _get_db()
    doc = await db.settings.find_one({"key": "bank_accounts"}) or {}
    return doc.get("value") or [
        {"bank": "MANDIRI", "number": "1240011911816", "holder": "INTERCLOUD DIGITAL INOVASI"},
        {"bank": "BCA", "number": "4730862038", "holder": "ANANG MADIA CUGITA"},
    ]


@router.put("/admin/bank-accounts")
async def update_bank_accounts(payload: list, admin=Depends(get_current_admin)):
    db = await _get_db()
    await db.settings.update_one(
        {"key": "bank_accounts"},
        {"$set": {"key": "bank_accounts", "value": payload}},
        upsert=True,
    )
    return payload


# ============================================================
# WEBMAIL (staff-only) - SMTP for sending, IMAP for inbox.
# Currently backed by MongoDB (mock) so the UX works end-to-end.
# When an SMTP + IMAP integration is enabled under /admin/integrations,
# these endpoints can be swapped to real IMAP/SMTP calls.
# ============================================================

@router.get("/admin/mail/inbox")
async def admin_mail_inbox(staff=Depends(get_current_staff)):
    db = await _get_db()
    # ---- F1: Per-admin inbox - use the caller's OWN email_settings ----
    user_doc = await db.users.find_one({"_id": ObjectId(staff["id"])})
    my_settings = (user_doc or {}).get("email_settings") or {}
    my_imap = my_settings.get("imap") or {}
    if my_imap.get("credentials", {}).get("host") and my_imap.get("credentials", {}).get("username"):
        try:
            live = iv2.IMAPClient(my_imap).fetch_recent()
        except iv2.IMAPConnectionError as e:
            # Personal IMAP creds are set but the mailbox can't be reached -
            # tell the operator exactly why so they can fix it.
            return {"not_setup": True, "reason": "connection_failed",
                    "message": f"IMAP tidak bisa terhubung. {e}",
                    "detail": str(e)}
        return [{
            "id": f"imap-{msg['id']}",
            "from_name": msg["from"].split("<")[0].strip(" \""),
            "from_email": (msg["from"].split("<")[-1].rstrip(">") if "<" in msg["from"] else msg["from"]),
            "subject": msg["subject"],
            "preview": msg["preview"],
            "received_at": msg["date"],
            "unread": False,
            "starred": False,
            "_live": True,
        } for msg in live]

    # No personal creds → surface an actionable "click to setup" hint.
    # Frontend AdminMail.jsx renders a big card with a Configure button.
    return {"not_setup": True, "reason": "no_credentials",
            "message": "Email pribadi Anda belum di-setup. Klik untuk konfigurasi IMAP + SMTP cPanel Anda."}


@router.get("/admin/mail/messages/{mid}")
async def admin_mail_message(mid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    # ---- IMAP live message (id prefixed with "imap-") - uses caller's own creds ----
    if mid.startswith("imap-"):
        uid = mid[len("imap-"):]
        user_doc = await db.users.find_one({"_id": ObjectId(staff["id"])})
        my_imap = ((user_doc or {}).get("email_settings") or {}).get("imap") or {}
        if my_imap.get("credentials", {}).get("host"):
            try:
                for msg in iv2.IMAPClient(my_imap).fetch_recent():
                    if str(msg.get("id")) == uid:
                        return {
                            "id": mid,
                            "from_name": msg["from"].split("<")[0].strip(" \""),
                            "from_email": (msg["from"].split("<")[-1].rstrip(">") if "<" in msg["from"] else msg["from"]),
                            "subject": msg.get("subject", ""),
                            "body": msg.get("body") or msg.get("preview") or "",
                            "received_at": msg.get("date"),
                            "starred": False,
                        }
            except iv2.IMAPConnectionError as e:
                raise HTTPException(status_code=502, detail=f"IMAP tidak bisa terhubung: {e}")
        raise HTTPException(status_code=404, detail="IMAP message no longer available (mailbox may have been re-synced)")

    # ---- Mongo-backed message (legacy seeded demo - no per-user creds path) ----
    try:
        oid = _oid(mid)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid message id")
    d = await db.mail_inbox.find_one({"_id": oid})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    if d.get("unread"):
        await db.mail_inbox.update_one({"_id": d["_id"]}, {"$set": {"unread": False}})
    return {
        "id": str(d["_id"]),
        "from_name": d.get("from_name", ""),
        "from_email": d.get("from_email", ""),
        "subject": d.get("subject", ""),
        "body": d.get("body", ""),
        "received_at": d.get("received_at"),
        "starred": bool(d.get("starred", False)),
    }


@router.post("/admin/mail/messages/{mid}/toggle-star")
async def admin_mail_toggle_star(mid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    d = await db.mail_inbox.find_one({"_id": _oid(mid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    await db.mail_inbox.update_one({"_id": d["_id"]}, {"$set": {"starred": not d.get("starred", False)}})
    return {"starred": not d.get("starred", False)}


@router.post("/admin/mail/send")
async def admin_mail_send(payload: dict, staff=Depends(get_current_staff)):
    """Send outgoing email using the *caller's own* SMTP credentials.

    Every staff member configures their personal cPanel SMTP under
    Settings ▸ Email; the outbox uses those creds so replies come back to
    the same mailbox they read from. If no personal SMTP is configured,
    we hard-fail with 400 so the user is nudged to set it up.
    """
    db = await _get_db()
    to = payload.get("to", "")
    subject = payload.get("subject", "")
    body = payload.get("body", "")
    if not to or not subject:
        raise HTTPException(status_code=400, detail="to and subject are required")

    # ---- Load caller's personal SMTP settings ----
    user_doc = await db.users.find_one({"_id": ObjectId(staff["id"])})
    my_settings = (user_doc or {}).get("email_settings") or {}
    my_smtp = my_settings.get("smtp") or {}
    smtp_creds = my_smtp.get("credentials") or {}
    if not (smtp_creds.get("host") and smtp_creds.get("username") and smtp_creds.get("password")):
        raise HTTPException(
            status_code=400,
            detail="Silakan setup SMTP dulu di Settings ▸ Email sebelum mengirim.",
        )

    # Build a settings dict compatible with SMTPMailer, merging in the
    # caller's display name / from-address from the per-user config.
    smtp_settings = {
        "credentials": smtp_creds,
        "options": {
            **(my_smtp.get("options") or {}),
            "from_email": my_settings.get("from_email") or smtp_creds.get("username"),
            "from_name":  my_settings.get("from_name")  or staff.get("name") or "Intercloud",
        },
    }

    delivered = False
    delivered_via = "queued"
    from_email = smtp_settings["options"]["from_email"]
    from_name  = smtp_settings["options"]["from_name"]
    try:
        iv2.SMTPMailer(smtp_settings).send(to=to, subject=subject, html=body or "")
        delivered = True
        delivered_via = "smtp"
    except Exception as e:
        # Surface the underlying reason to the caller so the UI can show a
        # meaningful error instead of a silent "queued".
        raise HTTPException(
            status_code=502,
            detail=f"SMTP kirim gagal ({type(e).__name__}): {e}",
        )

    doc = {
        "from_email": from_email or "no-reply@intercloud-digital.com",
        "from_name": from_name,
        "to": to, "subject": subject, "body": body,
        "sent_at": _now(),
        "sent_by_id": staff["id"], "sent_by_name": staff["name"],
        "delivered": delivered, "delivered_via": delivered_via,
    }
    r = await db.mail_sent.insert_one(doc)
    doc["_id"] = r.inserted_id
    return {
        "id": str(doc["_id"]),
        "delivered": doc["delivered"],
        "delivered_via": doc["delivered_via"],
        "sent_at": doc["sent_at"],
    }


@router.get("/admin/mail/sent")
async def admin_mail_sent(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.mail_sent.find({}).sort("sent_at", -1).to_list(200)
    return [{
        "id": str(d["_id"]),
        "from_email": d.get("from_email"),
        "to": d.get("to"),
        "subject": d.get("subject"),
        "body": d.get("body"),
        "sent_at": d.get("sent_at"),
        "delivered": d.get("delivered", False),
        "delivered_via": d.get("delivered_via", ""),
    } for d in docs]


# ============================================================
# BUSINESS - CRM, Projects, Content Planner, Follow-ups, Documents
# ============================================================

# ---------- CRM (customers/prospects) ----------
def _serialize_crm(d):
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "email": d.get("email", ""),
        "phone": d.get("phone", ""),
        "company": d.get("company", ""),
        "position": d.get("position", ""),
        "industry": d.get("industry", ""),
        "status": d.get("status", "prospect"),
        "notes": d.get("notes", ""),
        "user_id": str(d["user_id"]) if d.get("user_id") else None,
        "source": d.get("source", ""),
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }


# Order statuses that count as "in-progress" (needs attention) vs "won" vs "closed"
ORDER_TERMINAL_LOST = {"rejected", "cancelled"}
ORDER_IN_PROGRESS = {"pending", "pending_payment", "awaiting_verification",
                     "awaiting_quote", "payment_verified", "assigned", "provisioning"}
ORDER_WON = {"active"}


async def _crm_enrichment_by_uid(db, user_ids: list) -> dict:
    """Return {user_id_str: {latest_order, active_orders_count, lifetime_value, in_progress_count}}
    for the given user IDs, in one round-trip per collection."""
    if not user_ids:
        return {}
    result = {}
    # ---- Orders (grouped in-memory: small dataset per tenant) ----
    orders_cur = db.orders.find(
        {"user_id": {"$in": user_ids}},
        {"user_id": 1, "status": 1, "created_at": 1, "product_name": 1,
         "invoice_id": 1, "config": 1},
    ).sort("created_at", -1)
    async for o in orders_cur:
        key = str(o["user_id"])
        bucket = result.setdefault(key, {
            "latest_order": None,
            "active_orders_count": 0,
            "in_progress_count": 0,
            "won_orders_count": 0,
            "lifetime_value": 0.0,
        })
        if bucket["latest_order"] is None:
            bucket["latest_order"] = {
                "id": str(o["_id"]),
                "status": o.get("status", "pending"),
                "product_name": o.get("product_name", ""),
                "created_at": _iso(o.get("created_at", "")),
                "invoice_id": str(o["invoice_id"]) if o.get("invoice_id") else None,
            }
        st = o.get("status", "pending")
        if st not in ORDER_TERMINAL_LOST:
            bucket["active_orders_count"] += 1
        if st in ORDER_IN_PROGRESS:
            bucket["in_progress_count"] += 1
        if st in ORDER_WON:
            bucket["won_orders_count"] += 1
    # ---- Paid invoices → lifetime value ----
    inv_cur = db.invoices.find(
        {"user_id": {"$in": user_ids}, "status": "paid"},
        {"user_id": 1, "total": 1, "number": 1},
    )
    async for inv in inv_cur:
        key = str(inv["user_id"])
        bucket = result.setdefault(key, {
            "latest_order": None,
            "active_orders_count": 0,
            "in_progress_count": 0,
            "won_orders_count": 0,
            "lifetime_value": 0.0,
        })
        try:
            bucket["lifetime_value"] += float(inv.get("total") or 0)
        except Exception:
            pass
    return result


@router.get("/admin/crm")
async def crm_list(staff=Depends(get_current_staff)):
    _deny_creative(staff)
    db = await _get_db()
    q = _sales_scope_filter(staff, key="user_id")
    docs = await db.crm_customers.find(q).sort("updated_at", -1).to_list(2000)
    # Collect user_ids for enrichment
    uid_pairs = [(str(d.get("user_id")), d.get("user_id")) for d in docs if d.get("user_id")]
    uids = [pair[1] for pair in uid_pairs]
    enrich = await _crm_enrichment_by_uid(db, uids)
    out = []
    for d in docs:
        row = _serialize_crm(d)
        e = enrich.get(str(d.get("user_id"))) if d.get("user_id") else None
        row["latest_order"] = (e or {}).get("latest_order")
        row["active_orders_count"] = (e or {}).get("active_orders_count", 0)
        row["in_progress_count"] = (e or {}).get("in_progress_count", 0)
        row["won_orders_count"] = (e or {}).get("won_orders_count", 0)
        row["lifetime_value"] = (e or {}).get("lifetime_value", 0.0)
        # Warm-lead heuristic: any prospect / lead with an in-progress order,
        # OR an existing customer with a fresh in-progress order (upsell signal)
        row["is_warm"] = row["in_progress_count"] > 0
        out.append(row)
    return out


@router.post("/admin/crm")
async def crm_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "name": payload.get("name", ""),
        "email": (payload.get("email") or "").lower(),
        "phone": payload.get("phone", ""),
        "company": payload.get("company", ""),
        "position": payload.get("position", ""),
        "industry": payload.get("industry", ""),
        "status": payload.get("status", "prospect"),
        "notes": payload.get("notes", ""),
        "created_at": _now(),
        "updated_at": _now(),
    }
    r = await db.crm_customers.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_crm(doc)


async def _assert_sales_can_touch_crm(db, staff: dict, cid: str) -> dict:
    """Load a CRM row and 403 if `staff` is a sales user whose assigned
    clients don't include the row's linked user_id."""
    d = await db.crm_customers.find_one({"_id": _oid(cid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    if staff.get("role") == "sales":
        assigned = {str(x) for x in (staff.get("assigned_client_ids") or [])}
        if not (d.get("user_id") and str(d["user_id"]) in assigned):
            raise HTTPException(status_code=403, detail="Not your client")
    return d


@router.put("/admin/crm/{cid}")
async def crm_update(cid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    await _assert_sales_can_touch_crm(db, staff, cid)
    payload = {k: v for k, v in payload.items() if k in {
        "name", "email", "phone", "company", "position", "industry", "status", "notes"
    }}
    payload["updated_at"] = _now()
    if "email" in payload and payload["email"]:
        payload["email"] = payload["email"].lower()
    await db.crm_customers.update_one({"_id": _oid(cid)}, {"$set": payload})
    d = await db.crm_customers.find_one({"_id": _oid(cid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    return _serialize_crm(d)


@router.delete("/admin/crm/{cid}")
async def crm_delete(cid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    await _assert_sales_can_touch_crm(db, staff, cid)
    r = await db.crm_customers.delete_one({"_id": _oid(cid)})
    return {"deleted": r.deleted_count}


# ---------- Projects ----------
def _serialize_project(d):
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "customer_id": str(d.get("customer_id", "")) if d.get("customer_id") else None,
        "customer_name": d.get("customer_name", ""),
        "owner": d.get("owner", ""),
        "status": d.get("status", "planning"),
        "priority": d.get("priority", "medium"),
        "progress": d.get("progress", 0),
        "start_date": d.get("start_date", ""),
        "target_date": d.get("target_date", ""),
        "description": d.get("description", ""),
        "tasks": d.get("tasks", []),
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }


@router.get("/admin/projects")
async def projects_list(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.projects.find({}).sort("updated_at", -1).to_list(1000)
    return [_serialize_project(d) for d in docs]


@router.post("/admin/projects")
async def projects_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "name": payload.get("name", ""),
        "customer_id": _oid(payload["customer_id"]) if payload.get("customer_id") else None,
        "customer_name": payload.get("customer_name", ""),
        "owner": payload.get("owner", ""),
        "status": payload.get("status", "planning"),
        "priority": payload.get("priority", "medium"),
        "progress": int(payload.get("progress", 0)),
        "start_date": payload.get("start_date", ""),
        "target_date": payload.get("target_date", ""),
        "description": payload.get("description", ""),
        "tasks": payload.get("tasks", []),
        "created_at": _now(),
        "updated_at": _now(),
    }
    r = await db.projects.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_project(doc)


@router.put("/admin/projects/{pid}")
async def projects_update(pid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    upd = {k: v for k, v in payload.items() if k in {
        "name", "customer_name", "owner", "status", "priority", "progress",
        "start_date", "target_date", "description", "tasks"
    }}
    if "customer_id" in payload:
        upd["customer_id"] = _oid(payload["customer_id"]) if payload["customer_id"] else None
    upd["updated_at"] = _now()
    await db.projects.update_one({"_id": _oid(pid)}, {"$set": upd})
    d = await db.projects.find_one({"_id": _oid(pid)})
    return _serialize_project(d)


@router.delete("/admin/projects/{pid}")
async def projects_delete(pid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.projects.delete_one({"_id": _oid(pid)})
    return {"deleted": r.deleted_count}


# ---------- Content Planner ----------
def _serialize_content(d):
    return {
        "id": str(d["_id"]),
        "title": d.get("title", ""),
        "channel": d.get("channel", "blog"),
        "type": d.get("type", "post"),
        "status": d.get("status", "idea"),
        "owner": d.get("owner", ""),
        "publish_date": d.get("publish_date", ""),
        "hook": d.get("hook", ""),
        "url": d.get("url", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/content")
async def content_list(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.content_plan.find({}).sort("publish_date", 1).to_list(1000)
    return [_serialize_content(d) for d in docs]


@router.post("/admin/content")
async def content_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "title": payload.get("title", ""),
        "channel": payload.get("channel", "blog"),
        "type": payload.get("type", "post"),
        "status": payload.get("status", "idea"),
        "owner": payload.get("owner", ""),
        "publish_date": payload.get("publish_date", ""),
        "hook": payload.get("hook", ""),
        "url": payload.get("url", ""),
        "created_at": _now(),
    }
    r = await db.content_plan.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_content(doc)


@router.put("/admin/content/{cid}")
async def content_update(cid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    upd = {k: v for k, v in payload.items() if k in {
        "title", "channel", "type", "status", "owner", "publish_date", "hook", "url"
    }}
    await db.content_plan.update_one({"_id": _oid(cid)}, {"$set": upd})
    d = await db.content_plan.find_one({"_id": _oid(cid)})
    return _serialize_content(d)


@router.delete("/admin/content/{cid}")
async def content_delete(cid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.content_plan.delete_one({"_id": _oid(cid)})
    return {"deleted": r.deleted_count}


# ---------- Follow-ups ----------
def _serialize_followup(d):
    return {
        "id": str(d["_id"]),
        "customer_id": str(d.get("customer_id", "")) if d.get("customer_id") else None,
        "customer_name": d.get("customer_name", ""),
        "task": d.get("task", ""),
        "channel": d.get("channel", "whatsapp"),
        "due_date": d.get("due_date", ""),
        "done": bool(d.get("done", False)),
        "owner": d.get("owner", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


async def _sales_followup_filter(db, staff: dict) -> dict | None:
    """Return a Mongo filter that restricts follow-ups to CRM rows the sales
    staff can access. Returns {} for non-sales. Returns None if the caller is
    a sales user with zero visible CRM rows (endpoint should short-circuit)."""
    if staff.get("role") != "sales":
        return {}
    ids = await _sales_visible_crm_ids(db, staff)
    if not ids:
        return None
    return {"customer_id": {"$in": ids}}


async def _assert_sales_can_touch_followup(db, staff: dict, fid: str) -> dict:
    d = await db.followups.find_one({"_id": _oid(fid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    if staff.get("role") == "sales":
        visible = await _sales_visible_crm_ids(db, staff) or []
        cust_id = d.get("customer_id")
        if not (cust_id and any(str(cust_id) == str(x) for x in visible)):
            raise HTTPException(status_code=403, detail="Not your follow-up")
    return d


@router.get("/admin/followups")
async def followups_list(staff=Depends(get_current_staff)):
    _deny_creative(staff)
    db = await _get_db()
    q = await _sales_followup_filter(db, staff)
    if q is None:
        return []
    docs = await db.followups.find(q).sort("due_date", 1).to_list(1000)
    return [_serialize_followup(d) for d in docs]


@router.post("/admin/followups")
async def followups_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    cust_id = _oid(payload["customer_id"]) if payload.get("customer_id") else None
    if staff.get("role") == "sales":
        visible = await _sales_visible_crm_ids(db, staff) or []
        if not (cust_id and any(str(cust_id) == str(x) for x in visible)):
            raise HTTPException(status_code=403, detail="Follow-up harus untuk pelanggan yang di-assign ke Anda")
    doc = {
        "customer_id": cust_id,
        "customer_name": payload.get("customer_name", ""),
        "task": payload.get("task", ""),
        "channel": payload.get("channel", "whatsapp"),
        "due_date": payload.get("due_date", ""),
        "done": False,
        "owner": payload.get("owner", staff["name"]),
        "created_at": _now(),
    }
    r = await db.followups.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_followup(doc)


@router.put("/admin/followups/{fid}")
async def followups_update(fid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    await _assert_sales_can_touch_followup(db, staff, fid)
    upd = {k: v for k, v in payload.items() if k in {"task", "channel", "due_date", "done", "owner", "customer_name"}}
    await db.followups.update_one({"_id": _oid(fid)}, {"$set": upd})
    d = await db.followups.find_one({"_id": _oid(fid)})
    return _serialize_followup(d)


@router.delete("/admin/followups/{fid}")
async def followups_delete(fid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    await _assert_sales_can_touch_followup(db, staff, fid)
    r = await db.followups.delete_one({"_id": _oid(fid)})
    return {"deleted": r.deleted_count}


# ---------- Documents (metadata + file upload lokal) ----------
from pathlib import Path as _DocPath  # noqa: E402

DOCS_DIR = _DocPath(__file__).resolve().parent.parent / "uploads" / "documents"
_DOC_ALLOWED_TYPES = {
    "application/pdf", "image/png", "image/jpeg", "image/webp",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/zip", "application/x-zip-compressed", "text/plain", "text/csv",
}
_DOC_MAX_BYTES = 15 * 1024 * 1024  # 15 MB


def _serialize_doc(d):
    return {
        "id": str(d["_id"]),
        "title": d.get("title", ""),
        "category": d.get("category", "contract"),
        "customer_name": d.get("customer_name", ""),
        "url": d.get("url", ""),
        "notes": d.get("notes", ""),
        "filename": d.get("filename", ""),
        "size_bytes": d.get("size_bytes", 0),
        "has_file": bool(d.get("stored_name")),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/documents")
async def docs_list(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.documents.find({}).sort("created_at", -1).to_list(1000)
    return [_serialize_doc(d) for d in docs]


@router.post("/admin/documents")
async def docs_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "title": payload.get("title", ""),
        "category": payload.get("category", "contract"),
        "customer_name": payload.get("customer_name", ""),
        "url": payload.get("url", ""),
        "notes": payload.get("notes", ""),
        "created_at": _now(),
    }
    r = await db.documents.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_doc(doc)


@router.delete("/admin/documents/{did}")
async def docs_delete(did: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    d = await db.documents.find_one({"_id": _oid(did)})
    if d and d.get("stored_name"):
        try:
            (DOCS_DIR / d["stored_name"]).unlink(missing_ok=True)
        except Exception:
            pass
    r = await db.documents.delete_one({"_id": _oid(did)})
    return {"deleted": r.deleted_count}


@router.get("/documents/file/{did}")
async def docs_file(did: str):
    """Serve dokumen bisnis yang di-upload (URL ber-ObjectId, seperti media)."""
    db = await _get_db()
    d = await db.documents.find_one({"_id": _oid(did)})
    if not d or not d.get("stored_name"):
        raise HTTPException(status_code=404, detail="Document not found")
    fp = DOCS_DIR / d["stored_name"]
    if not fp.exists():
        raise HTTPException(status_code=404, detail="File missing on disk")
    return FileResponse(fp, media_type=d.get("content_type") or "application/octet-stream",
                        filename=d.get("filename") or d["stored_name"])


# ============================================================
# DCIM / IPAM (native, not via NetBox)
# ============================================================
async def _allocate_ip_from_pool(db, prefix_doc: dict, *, hostname: str = "",
                                 customer: str = "", description: str = "") -> str | None:
    """Ambil IP bebas berikutnya dari sebuah prefix DCIM dan catat di dcim_ips."""
    import ipaddress as _ip
    try:
        net = _ip.ip_network(prefix_doc.get("prefix", ""), strict=False)
    except ValueError:
        return None
    used_docs = await db.dcim_ips.find({"prefix_id": prefix_doc["_id"]}).to_list(5000)
    used = {d.get("address", "").split("/")[0] for d in used_docs}
    for host in net.hosts():
        a = str(host)
        if a.endswith(".0") or a.endswith(".1") or a in used:
            continue
        await db.dcim_ips.insert_one({
            "address": a, "prefix_id": prefix_doc["_id"], "status": "allocated",
            "role": "customer", "hostname": hostname, "customer": customer,
            "description": description, "created_at": _now(),
        })
        await db.dcim_prefixes.update_one({"_id": prefix_doc["_id"]}, {"$inc": {"usage": 1}})
        return a
    return None


async def _auto_allocate_customer_ip(db, *, hostname: str, customer: str, ref: str) -> str | None:
    """Pilih prefix IPv4 customer dengan slot tersisa, lalu alokasikan IP."""
    prefixes = await db.dcim_prefixes.find({"family": 4}).to_list(100)
    for p in prefixes:
        if str(p.get("site", "")).lower() == "internal":
            continue
        if int(p.get("usage", 0)) >= int(p.get("capacity", 0)):
            continue
        ip = await _allocate_ip_from_pool(db, p, hostname=hostname, customer=customer,
                                          description=f"Auto-allocated for {ref}")
        if ip:
            return ip
    return None


@router.post("/admin/dcim/prefixes/{pid}/allocate")
async def dcim_prefix_allocate(pid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    p = await db.dcim_prefixes.find_one({"_id": _oid(pid)})
    if not p:
        raise HTTPException(status_code=404, detail="Prefix not found")
    ip = await _allocate_ip_from_pool(db, p, hostname=payload.get("hostname", ""),
                                      customer=payload.get("customer", ""),
                                      description=payload.get("description", ""))
    if not ip:
        raise HTTPException(status_code=409, detail="Tidak ada IP bebas tersisa di prefix ini")
    return {"ok": True, "address": ip, "prefix": p.get("prefix", "")}


@router.get("/admin/dcim/prefixes/{pid}/utilization")
async def dcim_prefix_utilization(pid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    p = await db.dcim_prefixes.find_one({"_id": _oid(pid)})
    if not p:
        raise HTTPException(status_code=404, detail="Prefix not found")
    allocated = await db.dcim_ips.count_documents({"prefix_id": p["_id"]})
    capacity = int(p.get("capacity", 0)) or 1
    usage = int(p.get("usage", 0))
    return {"prefix": p.get("prefix", ""), "capacity": capacity, "usage": usage,
            "allocated_records": allocated,
            "utilization_pct": round(min(100.0, usage / capacity * 100), 2)}


@router.get("/admin/dcim/racks")
async def dcim_racks(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.dcim_racks.find({}).sort("name", 1).to_list(500)
    if not docs:
        # First-load seed for demo
        seed = [
            {"name": "Rack B12", "site": "Cyber 1 - Metta (Lantai 5)", "u_size": 42,
             "occupancy": [{"u_top": 40, "u_bot": 40, "label": "Patch Panel", "customer": ""},
                            {"u_top": 39, "u_bot": 39, "label": "sw-tor-1", "customer": "internal"},
                            {"u_top": 38, "u_bot": 38, "label": "sw-tor-2", "customer": "internal"},
                            {"u_top": 36, "u_bot": 34, "label": "3U Server", "customer": "PT Contoh Digital"}],
             "power_draw_w": 2450, "power_cap_w": 6000},
            {"name": "Rack A05", "site": "Cyber 1 - Omni (Lantai 2)", "u_size": 42,
             "occupancy": [{"u_top": 24, "u_bot": 20, "label": "5U Blade Chassis", "customer": "Bank ABC"}],
             "power_draw_w": 3800, "power_cap_w": 6000},
        ]
        for s in seed:
            await db.dcim_racks.insert_one({**s, "created_at": _now()})
        docs = await db.dcim_racks.find({}).sort("name", 1).to_list(500)
    return [{"id": str(d["_id"]), **{k: v for k, v in d.items() if k != "_id"}} for d in docs]


@router.get("/admin/dcim/prefixes")
async def dcim_prefixes(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.dcim_prefixes.find({}).to_list(500)
    if not docs:
        seed = [
            {"prefix": "103.28.14.0/24", "usage": 148, "capacity": 256, "vlan": "vlan-100", "site": "Cyber 1 - Metta", "family": 4},
            {"prefix": "103.28.15.0/24", "usage": 22, "capacity": 256, "vlan": "vlan-110", "site": "Cyber 1 - Omni", "family": 4},
            {"prefix": "2401:a900:1234::/48", "usage": 3, "capacity": 65536, "vlan": "vlan-100", "site": "Cyber 1 - Metta", "family": 6},
            {"prefix": "10.10.0.0/16", "usage": 1284, "capacity": 65534, "vlan": "mgmt", "site": "Internal", "family": 4},
        ]
        for s in seed:
            await db.dcim_prefixes.insert_one({**s, "created_at": _now()})
        docs = await db.dcim_prefixes.find({}).to_list(500)
    return [{"id": str(d["_id"]), **{k: v for k, v in d.items() if k != "_id"}} for d in docs]


@router.post("/admin/dcim/racks")
async def dcim_rack_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "name": payload.get("name", "Untitled Rack"),
        "site": payload.get("site", ""),
        "u_size": int(payload.get("u_size", 42)),
        "occupancy": payload.get("occupancy", []),
        "power_draw_w": int(payload.get("power_draw_w", 0) or 0),
        "power_cap_w": int(payload.get("power_cap_w", 6000) or 6000),
        "notes": payload.get("notes", ""),
        "created_at": _now(),
    }
    r = await db.dcim_racks.insert_one(doc)
    doc["_id"] = r.inserted_id
    return {"id": str(doc["_id"]), **{k: v for k, v in doc.items() if k != "_id"}}


@router.put("/admin/dcim/racks/{rid}")
async def dcim_rack_update(rid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    upd = {k: v for k, v in payload.items() if k in {"name", "site", "u_size", "occupancy", "power_draw_w", "power_cap_w", "notes"}}
    for k in ("u_size", "power_draw_w", "power_cap_w"):
        if k in upd:
            upd[k] = int(upd[k] or 0)
    await db.dcim_racks.update_one({"_id": _oid(rid)}, {"$set": upd})
    d = await db.dcim_racks.find_one({"_id": _oid(rid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    return {"id": str(d["_id"]), **{k: v for k, v in d.items() if k != "_id"}}


@router.delete("/admin/dcim/racks/{rid}")
async def dcim_rack_delete(rid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.dcim_racks.delete_one({"_id": _oid(rid)})
    return {"deleted": r.deleted_count}


@router.put("/admin/dcim/prefixes/{pid}")
async def dcim_prefix_update(pid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    upd = {k: v for k, v in payload.items() if k in {"prefix", "usage", "capacity", "vlan", "site", "family", "description"}}
    for k in ("usage", "capacity", "family"):
        if k in upd:
            upd[k] = int(upd[k] or 0)
    await db.dcim_prefixes.update_one({"_id": _oid(pid)}, {"$set": upd})
    d = await db.dcim_prefixes.find_one({"_id": _oid(pid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    return {"id": str(d["_id"]), **{k: v for k, v in d.items() if k != "_id"}}


# IP Addresses (within a prefix)
@router.get("/admin/dcim/ips")
async def dcim_ips_list(prefix_id: str | None = None, staff=Depends(get_current_staff)):
    db = await _get_db()
    q = {}
    if prefix_id:
        q["prefix_id"] = _oid(prefix_id)
    docs = await db.dcim_ips.find(q).sort("address", 1).to_list(2000)
    return [{"id": str(d["_id"]), "prefix_id": str(d.get("prefix_id", "")) if d.get("prefix_id") else None,
             **{k: v for k, v in d.items() if k not in ("_id", "prefix_id")}} for d in docs]


@router.post("/admin/dcim/ips")
async def dcim_ip_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "address": payload.get("address", ""),
        "prefix_id": _oid(payload["prefix_id"]) if payload.get("prefix_id") else None,
        "status": payload.get("status", "active"),
        "role": payload.get("role", ""),
        "hostname": payload.get("hostname", ""),
        "customer": payload.get("customer", ""),
        "description": payload.get("description", ""),
        "created_at": _now(),
    }
    r = await db.dcim_ips.insert_one(doc)
    doc["_id"] = r.inserted_id
    return {"id": str(doc["_id"]), "prefix_id": str(doc["prefix_id"]) if doc.get("prefix_id") else None,
            **{k: v for k, v in doc.items() if k not in ("_id", "prefix_id")}}


@router.put("/admin/dcim/ips/{ipid}")
async def dcim_ip_update(ipid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    upd = {k: v for k, v in payload.items() if k in {"address", "status", "role", "hostname", "customer", "description"}}
    if "prefix_id" in payload:
        upd["prefix_id"] = _oid(payload["prefix_id"]) if payload["prefix_id"] else None
    await db.dcim_ips.update_one({"_id": _oid(ipid)}, {"$set": upd})
    d = await db.dcim_ips.find_one({"_id": _oid(ipid)})
    return {"id": str(d["_id"]), "prefix_id": str(d.get("prefix_id", "")) if d.get("prefix_id") else None,
            **{k: v for k, v in d.items() if k not in ("_id", "prefix_id")}}


@router.delete("/admin/dcim/ips/{ipid}")
async def dcim_ip_delete(ipid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.dcim_ips.delete_one({"_id": _oid(ipid)})
    return {"deleted": r.deleted_count}


# Sites
@router.get("/admin/dcim/sites")
async def dcim_sites(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.dcim_sites.find({}).sort("name", 1).to_list(500)
    if not docs:
        for s in [
            {"name": "Cyber 1 - Metta Lantai 5", "code": "JKT-METTA-5F", "address": "Cyber 1 Building, Jakarta"},
            {"name": "Cyber 1 - Omni Lantai 2", "code": "JKT-OMNI-2F", "address": "Cyber 1 Building, Jakarta"},
            {"name": "TIFA Building", "code": "JKT-TIFA", "address": "TIFA Building, Jakarta"},
            {"name": "APJII DC Cyber 1 Lantai 1", "code": "JKT-APJII-1F", "address": "Cyber 1 Building, Jakarta"},
        ]:
            await db.dcim_sites.insert_one({**s, "created_at": _now()})
        docs = await db.dcim_sites.find({}).sort("name", 1).to_list(500)
    return [{"id": str(d["_id"]), **{k: v for k, v in d.items() if k != "_id"}} for d in docs]


@router.post("/admin/dcim/sites")
async def dcim_site_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {**payload, "created_at": _now()}
    r = await db.dcim_sites.insert_one(doc)
    return {"id": str(r.inserted_id), **payload}


@router.delete("/admin/dcim/sites/{sid}")
async def dcim_site_delete(sid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.dcim_sites.delete_one({"_id": _oid(sid)})
    return {"deleted": r.deleted_count}


@router.post("/admin/dcim/prefixes")
async def dcim_prefix_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {**payload, "created_at": _now()}
    r = await db.dcim_prefixes.insert_one(doc)
    doc["_id"] = r.inserted_id
    return {"id": str(doc["_id"]), **{k: v for k, v in doc.items() if k != "_id"}}


@router.delete("/admin/dcim/prefixes/{pid}")
async def dcim_prefix_delete(pid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.dcim_prefixes.delete_one({"_id": _oid(pid)})
    return {"deleted": r.deleted_count}


# ============================================================
# PROXMOX - available OS templates & OS request ticket bridge
# ============================================================
@router.get("/admin/proxmox/os-templates")
async def proxmox_os_templates(staff=Depends(get_current_staff)):
    """Return OS templates as would be reported by Proxmox ISO storage.
    Reads from an admin-editable settings doc; falls back to a common list."""
    db = await _get_db()
    doc = await db.settings.find_one({"key": "proxmox_os_templates"})
    if doc and doc.get("value"):
        return doc["value"]
    return [
        {"name": "Ubuntu 22.04 LTS Server", "family": "ubuntu", "type": "iso"},
        {"name": "Ubuntu 20.04 LTS Server", "family": "ubuntu", "type": "iso"},
        {"name": "Debian 12", "family": "debian", "type": "iso"},
        {"name": "AlmaLinux 9", "family": "rhel", "type": "iso"},
        {"name": "Rocky Linux 9", "family": "rhel", "type": "iso"},
        {"name": "CentOS Stream 9", "family": "rhel", "type": "iso"},
        {"name": "Windows Server 2022 Std", "family": "windows", "type": "iso"},
        {"name": "cloud-init/ubuntu-24.04-noble", "family": "ubuntu", "type": "template"},
        {"name": "cloud-init/debian-12", "family": "debian", "type": "template"},
    ]


@router.put("/admin/proxmox/os-templates")
async def proxmox_os_templates_set(payload: list, admin=Depends(get_current_admin)):
    db = await _get_db()
    await db.settings.update_one(
        {"key": "proxmox_os_templates"},
        {"$set": {"key": "proxmox_os_templates", "value": payload}},
        upsert=True,
    )
    return payload


@router.post("/client/proxmox/os-request")
async def client_request_os(payload: dict, user=Depends(get_current_user)):
    """Client requests an OS that isn't currently in the Proxmox library.
    Creates a ticket in the technical department."""
    db = await _get_db()
    os_name = (payload.get("os_name") or "").strip()
    if not os_name:
        raise HTTPException(status_code=400, detail="os_name is required")
    now = _now()
    number = await _next_number(db, "tickets", "TCK")
    subject = f"OS Provision Request: {os_name}"
    doc = {
        "user_id": ObjectId(user["id"]),
        "number": number,
        "subject": subject,
        "department": "technical",
        "priority": "medium",
        "status": "open",
        "replies": [{
            "author_id": user["id"], "author_name": user["name"], "author_role": "client",
            "message": f"Hi team, I'd like to request that '{os_name}' be added to the Proxmox ISO library. Additional notes: {payload.get('notes','-')}",
            "created_at": now,
        }],
        "created_at": now, "updated_at": now,
    }
    r = await db.tickets.insert_one(doc)
    return {"ticket_number": number, "ticket_id": str(r.inserted_id)}


# ============================================================
# ASSETS (native asset tracking + STRAIGHT-LINE depreciation)
# Formula (Metode Garis Lurus):
#   Penyusutan per Tahun = (Harga Perolehan − Nilai Sisa) / Umur Ekonomis
#   Penyusutan per Bulan = Penyusutan per Tahun / 12
#   Akumulasi Penyusutan = Penyusutan per Bulan × bulan_terpakai
#   Nilai Buku            = max(Harga Perolehan − Akumulasi Penyusutan, Nilai Sisa)
# ============================================================
def _asset_life_years(a: dict) -> int:
    """Effective useful-life in years. Prefer explicit field, fall back to
    legacy fields (`useful_life_months`, `depreciation_percent`) so we stay
    backwards-compatible with data seeded before the straight-line rewrite."""
    life_y = int(a.get("useful_life_years", 0) or 0)
    if life_y > 0:
        return life_y
    life_m = int(a.get("useful_life_months", 0) or 0)
    if life_m > 0:
        return max(1, round(life_m / 12))
    dep_pct = float(a.get("depreciation_percent", 0) or 0)
    if dep_pct > 0:
        return max(1, round(100.0 / dep_pct))
    return 0


def _asset_depreciation(a: dict) -> dict:
    """Compute straight-line depreciation snapshot for an asset document."""
    value = float(a.get("value", 0) or 0)
    salvage = float(a.get("salvage_value", 0) or 0)
    life_y = _asset_life_years(a)
    purchase = a.get("purchase_date", "") or ""

    if life_y <= 0 or not purchase:
        return {
            "life_years": life_y,
            "depreciable_base": max(value - salvage, 0.0),
            "annual_depreciation": 0.0,
            "monthly_depreciation": 0.0,
            "months_elapsed": 0,
            "total_months": life_y * 12,
            "accumulated_depreciation": 0.0,
            "book_value": round(value, 2),
            "is_fully_depreciated": False,
        }

    base = max(value - salvage, 0.0)
    annual = base / life_y
    monthly = annual / 12.0
    total_months = life_y * 12

    try:
        p = datetime.fromisoformat(purchase[:10])
    except Exception:
        return {
            "life_years": life_y, "depreciable_base": base,
            "annual_depreciation": round(annual, 2), "monthly_depreciation": round(monthly, 2),
            "months_elapsed": 0, "total_months": total_months,
            "accumulated_depreciation": 0.0, "book_value": round(value, 2),
            "is_fully_depreciated": False,
        }

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    months = (now.year - p.year) * 12 + (now.month - p.month)
    if now.day >= p.day:
        months += 1
    months = max(0, min(months, total_months))

    accumulated = round(monthly * months, 2)
    book = max(round(value - accumulated, 2), salvage)
    return {
        "life_years": life_y,
        "depreciable_base": round(base, 2),
        "annual_depreciation": round(annual, 2),
        "monthly_depreciation": round(monthly, 2),
        "months_elapsed": months,
        "total_months": total_months,
        "accumulated_depreciation": accumulated,
        "book_value": book,
        "is_fully_depreciated": months >= total_months,
    }


def _asset_book_value(a: dict) -> float:
    """Current book value using the straight-line method (floored at salvage)."""
    return _asset_depreciation(a)["book_value"]


def _asset_schedule(a: dict) -> list:
    """Yearly schedule from purchase year through end of useful life."""
    value = float(a.get("value", 0) or 0)
    salvage = float(a.get("salvage_value", 0) or 0)
    life_y = _asset_life_years(a)
    purchase = a.get("purchase_date", "") or ""
    if life_y <= 0 or not purchase:
        return []
    base = max(value - salvage, 0.0)
    annual = base / life_y
    try:
        start_year = datetime.fromisoformat(purchase[:10]).year
    except Exception:
        return []
    rows, accumulated = [], 0.0
    for i in range(life_y):
        accumulated = min(accumulated + annual, base)
        book = max(value - accumulated, salvage)
        rows.append({
            "period": i + 1,
            "year": start_year + i,
            "depreciation": round(annual, 2),
            "accumulated_depreciation": round(accumulated, 2),
            "book_value": round(book, 2),
        })
    return rows


def _serialize_asset(d):
    dep = _asset_depreciation(d)
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "category": d.get("category", "server"),
        "serial_number": d.get("serial_number", ""),
        "location": d.get("location", ""),
        "vendor": d.get("vendor", ""),
        "value": float(d.get("value", 0)),
        "salvage_value": float(d.get("salvage_value", 0) or 0),
        "useful_life_years": dep["life_years"],
        # legacy fields kept for backward compat with UI/tests still referencing them
        "depreciation_percent": float(d.get("depreciation_percent", 0) or 0),
        "useful_life_months": int(d.get("useful_life_months", 0) or 0),
        "purchase_date": d.get("purchase_date", ""),
        "annual_depreciation": dep["annual_depreciation"],
        "monthly_depreciation": dep["monthly_depreciation"],
        "accumulated_depreciation": dep["accumulated_depreciation"],
        "book_value": dep["book_value"],
        # kept for compat with old frontend field name
        "depreciated_amount": dep["accumulated_depreciation"],
        "months_elapsed": dep["months_elapsed"],
        "total_months": dep["total_months"],
        "is_fully_depreciated": dep["is_fully_depreciated"],
        "status": d.get("status", "active"),
        "disposed_at": d.get("disposed_at", ""),
        "notes": d.get("notes", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/assets")
async def assets_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.assets.find({}).sort("created_at", -1).to_list(2000)
    return [_serialize_asset(d) for d in docs]


@router.get("/admin/assets/{aid}")
async def assets_get(aid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await db.assets.find_one({"_id": _oid(aid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    payload = _serialize_asset(d)
    payload["schedule"] = _asset_schedule(d)
    return payload


def _coerce_asset_payload(payload: dict) -> dict:
    """Normalize incoming asset payload. Falls back to legacy fields when
    salvage/useful_life_years are omitted."""
    life_y = payload.get("useful_life_years")
    if life_y in (None, "", 0, "0"):
        # derive from legacy fields if provided in same payload
        dep_pct = float(payload.get("depreciation_percent", 0) or 0)
        life_m = int(payload.get("useful_life_months", 0) or 0)
        if life_m > 0:
            life_y = max(1, round(life_m / 12))
        elif dep_pct > 0:
            life_y = max(1, round(100.0 / dep_pct))
        else:
            life_y = 0
    return {
        "salvage_value": float(payload.get("salvage_value", 0) or 0),
        "useful_life_years": int(life_y or 0),
    }


@router.post("/admin/assets")
async def assets_create(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    coerced = _coerce_asset_payload(payload)
    doc = {
        "name": payload.get("name", ""),
        "category": payload.get("category", "server"),
        "serial_number": payload.get("serial_number", ""),
        "location": payload.get("location", ""),
        "vendor": payload.get("vendor", ""),
        "value": float(payload.get("value", 0) or 0),
        "salvage_value": coerced["salvage_value"],
        "useful_life_years": coerced["useful_life_years"],
        # legacy fields retained if the client still sends them
        "depreciation_percent": float(payload.get("depreciation_percent", 0) or 0),
        "useful_life_months": int(payload.get("useful_life_months", 0) or 0),
        "purchase_date": payload.get("purchase_date", ""),
        "status": payload.get("status") if payload.get("status") in ("active", "disposed") else "active",
        "notes": payload.get("notes", ""),
        "created_at": _now(),
    }
    if doc["status"] == "disposed":
        doc["disposed_at"] = payload.get("disposed_at") or datetime.now(timezone.utc).date().isoformat()
    r = await db.assets.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_asset(doc)


@router.put("/admin/assets/{aid}")
async def assets_update(aid: str, payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    allowed = {
        "name", "category", "serial_number", "location", "vendor", "value",
        "salvage_value", "useful_life_years",
        "depreciation_percent", "useful_life_months",
        "purchase_date", "notes", "status", "disposed_at",
    }
    upd = {k: v for k, v in payload.items() if k in allowed}
    if "status" in upd:
        if upd["status"] not in ("active", "disposed"):
            upd.pop("status")
        elif upd["status"] == "disposed" and not upd.get("disposed_at"):
            upd["disposed_at"] = datetime.now(timezone.utc).date().isoformat()
        elif upd["status"] == "active":
            upd["disposed_at"] = ""
    coerced = _coerce_asset_payload({**upd})
    upd["salvage_value"] = coerced["salvage_value"]
    if coerced["useful_life_years"] > 0:
        upd["useful_life_years"] = coerced["useful_life_years"]
    for k in ("value", "depreciation_percent"):
        if k in upd:
            upd[k] = float(upd[k] or 0)
    if "useful_life_months" in upd:
        upd["useful_life_months"] = int(upd["useful_life_months"] or 0)
    await db.assets.update_one({"_id": _oid(aid)}, {"$set": upd})
    d = await db.assets.find_one({"_id": _oid(aid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    return _serialize_asset(d)


@router.delete("/admin/assets/{aid}")
async def assets_delete(aid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    r = await db.assets.delete_one({"_id": _oid(aid)})
    return {"deleted": r.deleted_count}


# ============================================================
# EXPENSES (manual bookkeeping)
# ============================================================
def _serialize_expense(d):
    return {
        "id": str(d["_id"]),
        "date": d.get("date", ""),
        "category": d.get("category", "other"),
        "vendor": d.get("vendor", ""),
        "amount": float(d.get("amount", 0)),
        "description": d.get("description", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/expenses")
async def expenses_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.expenses.find({}).sort("date", -1).to_list(5000)
    return [_serialize_expense(d) for d in docs]


@router.post("/admin/expenses")
async def expenses_create(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    doc = {
        "date": payload.get("date", datetime.now(timezone.utc).date().isoformat()),
        "category": payload.get("category", "other"),
        "vendor": payload.get("vendor", ""),
        "amount": float(payload.get("amount", 0) or 0),
        "description": payload.get("description", ""),
        "created_at": _now(),
    }
    r = await db.expenses.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_expense(doc)


@router.delete("/admin/expenses/{eid}")
async def expenses_delete(eid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    r = await db.expenses.delete_one({"_id": _oid(eid)})
    return {"deleted": r.deleted_count}


# Extended finance report (revenue + expenses + assets)
@router.get("/admin/assets/report/depreciation")
async def assets_depreciation_report(months: int = 12, admin=Depends(get_current_admin)):
    """Laporan total beban depresiasi aset aktif per periode (bulanan, N bulan terakhir)."""
    db = await _get_db()
    months = max(1, min(months, 60))
    assets = await db.assets.find({}).to_list(2000)
    now = datetime.now(timezone.utc)
    periods = []
    for i in range(months - 1, -1, -1):
        y = now.year + (now.month - 1 - i) // 12
        mo = (now.month - 1 - i) % 12 + 1
        periods.append(f"{y:04d}-{mo:02d}")
    rows = []
    for pkey in periods:
        cur_idx = int(pkey[:4]) * 12 + int(pkey[5:7]) - 1
        total, count = 0.0, 0
        for a in assets:
            dep = _asset_depreciation(a)
            monthly = dep["monthly_depreciation"]
            purchase = (a.get("purchase_date") or "")[:10]
            if monthly <= 0 or not purchase or pkey < purchase[:7]:
                continue
            try:
                pd_ = datetime.strptime(purchase, "%Y-%m-%d")
            except Exception:
                continue
            end_idx = pd_.year * 12 + (pd_.month - 1) + dep["total_months"]
            if cur_idx >= end_idx:
                continue
            disposed = (a.get("disposed_at") or "")[:7]
            if a.get("status") == "disposed" and disposed and pkey > disposed:
                continue
            total += monthly
            count += 1
        rows.append({"period": pkey, "depreciation": round(total, 2), "active_assets": count})
    return {"months": months, "rows": rows,
            "total_depreciation": round(sum(r["depreciation"] for r in rows), 2)}


@router.get("/admin/finance/report")
async def admin_finance_report(admin=Depends(get_current_admin)):
    db = await _get_db()
    paid = await db.invoices.find({"status": "paid"}).to_list(5000)
    total_revenue = sum(d.get("total", 0) for d in paid)
    expenses = await db.expenses.find({}).to_list(5000)
    total_expenses = sum(d.get("amount", 0) for d in expenses)
    net_profit = total_revenue - total_expenses

    assets = await db.assets.find({}).to_list(2000)
    total_assets_value = sum(float(a.get("value", 0)) for a in assets)
    net_assets_value = sum(_asset_book_value(a) for a in assets)
    total_depreciation = round(total_assets_value - net_assets_value, 2)

    # Revenue & expenses by month (last 12 months)
    by_month_rev, by_month_exp = {}, {}
    for inv in paid:
        p = inv.get("paid_at") or inv.get("created_at", "")
        if not p:
            continue
        key = p[:7]
        by_month_rev[key] = by_month_rev.get(key, 0) + inv.get("total", 0)
    for e in expenses:
        key = (e.get("date") or "")[:7]
        if not key:
            continue
        by_month_exp[key] = by_month_exp.get(key, 0) + e.get("amount", 0)

    all_keys = sorted(set(by_month_rev.keys()) | set(by_month_exp.keys()))
    monthly = [
        {"month": k, "revenue": by_month_rev.get(k, 0), "expenses": by_month_exp.get(k, 0),
         "profit": by_month_rev.get(k, 0) - by_month_exp.get(k, 0)}
        for k in all_keys
    ]

    return {
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "total_assets_value": total_assets_value,
        "net_assets_value": net_assets_value,
        "total_depreciation": total_depreciation,
        "asset_count": len(assets),
        "monthly": monthly,
    }


# ============================================================
# PDF (HTML/PDF) documents - Invoice & Quotation
# Rendered as an HTML preview by default; add ?format=pdf for a real
# WeasyPrint-rendered downloadable .pdf that matches the WHMCS-style layout.
# ============================================================
from fastapi.responses import HTMLResponse, Response


# Long-form English/Indonesian date used inside the document
def _long_date(iso_or_ymd: str) -> str:
    if not iso_or_ymd:
        return "-"
    try:
        s = iso_or_ymd[:10]
        dt = datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        try:
            dt = datetime.fromisoformat(iso_or_ymd.replace("Z", "+00:00"))
        except Exception:
            return iso_or_ymd
    # e.g. "Thursday, June 18th, 2026"
    day = dt.day
    suffix = "th" if 10 <= day % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return dt.strftime(f"%A, %B {day}{suffix}, %Y")


def _idr(v) -> str:
    """Format IDR as 'Rp3,300,000.00' (WHMCS-style)."""
    try:
        f = float(v or 0)
    except Exception:
        f = 0.0
    return "Rp" + f"{f:,.2f}"


def _period_label(item: dict) -> str:
    """If item has period_start / period_end (YYYY-MM-DD), append ' (dd/mm/yyyy - dd/mm/yyyy)'."""
    ps, pe = item.get("period_start"), item.get("period_end")
    if not (ps and pe):
        return ""
    try:
        s = datetime.strptime(ps[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        e = datetime.strptime(pe[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        return f" ({s} - {e})"
    except Exception:
        return ""


def _addressed_to_block(u: dict) -> str:
    company = u.get("company") or u.get("name") or ""
    attn = u.get("attention") or u.get("name") or ""
    lines = []
    if company:
        lines.append(f"<div style='font-weight:700;color:#111'>{company}</div>")
    if attn:
        lines.append(f"<div>ATTN: {attn}</div>")
    if u.get("address_line1"):
        lines.append(f"<div>{u['address_line1']}</div>")
    if u.get("address_line2"):
        lines.append(f"<div>{u['address_line2']}</div>")
    city_line = ", ".join([x for x in [u.get("city"), u.get("province"), u.get("postal_code")] if x])
    if city_line:
        lines.append(f"<div>{city_line}</div>")
    if u.get("country"):
        lines.append(f"<div>{u['country']}</div>")
    return "\n".join(lines)


# Diagonal corner ribbon (top-right), color depends on status
def _corner_ribbon(status: str) -> str:
    s = (status or "").lower()
    if s == "paid":
        color = "#22c55e"  # green
        label = "PAID"
    elif s == "overdue":
        color = "#dc2626"  # red
        label = "OVERDUE"
    elif s == "cancelled":
        color = "#64748b"
        label = "CANCELLED"
    elif s == "unpaid":
        color = "#f59e0b"  # amber
        label = "UNPAID"
    elif s in ("draft", "sent"):
        color = "#0a2350"  # navy
        label = s.upper()
    elif s == "accepted":
        color = "#22c55e"
        label = "ACCEPTED"
    elif s == "rejected":
        color = "#dc2626"
        label = "REJECTED"
    elif s == "expired":
        color = "#64748b"
        label = "EXPIRED"
    else:
        color = "#f5b120"
        label = (status or "").upper() or "&nbsp;"
    return f"""
    <div class="ribbon-wrap">
      <div class="ribbon" style="background:{color}">{label}</div>
    </div>
    """


COMPANY_HEADER_HTML = """
<div class="company-block">
  <div style="font-weight:800;letter-spacing:.02em;color:#0a2350">PT. INTERCLOUD DIGITAL INOVASI</div>
  <div>Menara Cakrawala Lt 12, Unit 1205A</div>
  <div>Jl. M.H. Thamrin No.9, RT.2/RW.1,</div>
  <div>Kb. Sirih, Kec. Menteng Kota Jakarta Pusat,</div>
  <div>Daerah Khusus Ibukota Jakarta,</div>
  <div>10340</div>
  <div style="margin-top:6px">NPWP : 62.573.806.7-021.000</div>
</div>
"""


from portal.branding import get_branding as _get_branding_dict, DEFAULTS as _BRANDING_DEFAULTS, BRANDING_KEYS as _BRANDING_KEYS
LOGO_URL = _BRANDING_DEFAULTS["logo_dark"]


# ============================================================
# Branding endpoints (Admin ▸ Branding)
# ============================================================
@router.get("/branding")
async def branding_get():
    """Public read - landing/emails/frontend fetch the current branding."""
    db = await _get_db()
    return await _get_branding_dict(db)


@router.post("/admin/branding")
async def branding_set(payload: dict, admin=Depends(get_current_admin)):
    """Update one or more branding fields. Payload example:
        { "logo_dark": "data:image/png;base64,....",
          "favicon":   "https://cdn.example.com/favicon.png" }
    Only the four known keys (logo_light, logo_dark, favicon, email_banner)
    are stored; unknown keys are dropped for safety.
    """
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Body must be a JSON object")
    incoming = {k: v for k, v in payload.items() if k in _BRANDING_KEYS and isinstance(v, str)}
    # data-URI size sanity: refuse anything over 4 MB to keep the settings doc small
    for k, v in list(incoming.items()):
        if v.startswith("data:") and len(v) > 4 * 1024 * 1024:
            raise HTTPException(status_code=413, detail=f"{k}: image is larger than 4 MB")
    if not incoming:
        raise HTTPException(status_code=400, detail="No valid branding fields provided")
    db = await _get_db()
    existing = await db.settings.find_one({"key": "branding"}) or {}
    merged = dict(existing.get("value") or {})
    merged.update(incoming)
    await db.settings.update_one(
        {"key": "branding"},
        {"$set": {"value": merged, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return await _get_branding_dict(db)


@router.delete("/admin/branding/{key}")
async def branding_reset(key: str, admin=Depends(get_current_admin)):
    """Reset one field to its hardcoded default."""
    if key not in _BRANDING_KEYS:
        raise HTTPException(status_code=400, detail=f"Unknown branding key: {key}")
    db = await _get_db()
    existing = await db.settings.find_one({"key": "branding"}) or {}
    merged = dict(existing.get("value") or {})
    merged.pop(key, None)
    await db.settings.update_one(
        {"key": "branding"},
        {"$set": {"value": merged, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return await _get_branding_dict(db)


# ============================================================
# Landing-page CMS
# ============================================================
from portal.backups import (
    get_landing_content as _get_landing_content,
    LANDING_CONTENT_DEFAULT as _LANDING_DEFAULT,
    run_mongodump as _run_mongodump,
    run_mongorestore as _run_mongorestore,
)


@router.get("/landing-content")
async def landing_content_get():
    """Public - Landing page fetches on mount and merges overrides on top of
    the shipped i18n dict + hardcoded FAQ list."""
    db = await _get_db()
    return await _get_landing_content(db)


@router.post("/admin/landing-content")
async def landing_content_set(payload: dict, admin=Depends(get_current_admin)):
    """Replace the landing-content JSON. Body:
        {
          "overrides": {"hero.h1a": {"id": "...", "en": "..."}},
          "faqs":      [{"q": {"id": "...", "en": "..."},
                          "a": {"id": "...", "en": "..."}}, ...],
          "contact":   {"phone": "...", "email": "...", "address_id": "...", ...}
        }
    Unknown top-level keys are ignored. Any missing top-level key is set to
    an empty dict/list so the page never crashes on missing shape."""
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Body must be a JSON object")
    clean = {
        "overrides": payload.get("overrides") if isinstance(payload.get("overrides"), dict) else {},
        "faqs":      payload.get("faqs")      if isinstance(payload.get("faqs"), list)      else [],
        "contact":   payload.get("contact")   if isinstance(payload.get("contact"), dict)   else {},
    }
    # 128 KB cap on the whole doc - plenty for a landing page's worth of text.
    approx = len(str(clean))
    if approx > 128 * 1024:
        raise HTTPException(status_code=413,
                            detail=f"landing-content is {approx // 1024} KB; cap is 128 KB")
    db = await _get_db()
    await db.settings.update_one(
        {"key": "landing_content"},
        {"$set": {"value": clean, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return await _get_landing_content(db)


@router.delete("/admin/landing-content")
async def landing_content_reset(admin=Depends(get_current_admin)):
    """Wipe all landing overrides - Landing renders the shipped defaults."""
    db = await _get_db()
    await db.settings.update_one(
        {"key": "landing_content"},
        {"$set": {"value": {"overrides": {}, "faqs": [], "contact": {}},
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return dict(_LANDING_DEFAULT)


# ============================================================
# Backup / Restore
# ============================================================
# ---------- UTM link persistence ----------

@router.get("/admin/utm-links")
async def utm_links_list(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.utm_links.find({}).sort("created_at", -1).to_list(200)
    return [{"id": str(d["_id"]), "url": d.get("url", ""), "base": d.get("base", ""),
             "params": d.get("params", {}), "label": d.get("label", ""),
             "created_by": d.get("created_by", ""), "created_at": _iso(d.get("created_at", ""))}
            for d in docs]


@router.post("/admin/utm-links")
async def utm_links_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    url = str(payload.get("url", "")).strip()
    if not url.startswith("http"):
        raise HTTPException(status_code=400, detail="URL tidak valid")
    doc = {"url": url, "base": payload.get("base", ""), "params": payload.get("params", {}),
           "label": payload.get("label", ""), "created_by": staff.get("email", ""),
           "created_at": _now()}
    res = await db.utm_links.insert_one(doc)
    return {"ok": True, "id": str(res.inserted_id)}


@router.delete("/admin/utm-links/{lid}")
async def utm_links_delete(lid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    await db.utm_links.delete_one({"_id": _oid(lid)})
    return {"ok": True}


# ---------- Content calendar: hari libur nasional ----------

ID_HOLIDAYS_2026 = [
    {"date": "2026-01-01", "name": "Tahun Baru Masehi"},
    {"date": "2026-01-16", "name": "Isra Mikraj Nabi Muhammad SAW"},
    {"date": "2026-02-17", "name": "Tahun Baru Imlek 2577"},
    {"date": "2026-03-19", "name": "Hari Suci Nyepi"},
    {"date": "2026-03-20", "name": "Idul Fitri 1447 H (perkiraan)"},
    {"date": "2026-03-21", "name": "Idul Fitri 1447 H (hari kedua)"},
    {"date": "2026-04-03", "name": "Wafat Isa Almasih"},
    {"date": "2026-05-01", "name": "Hari Buruh Internasional"},
    {"date": "2026-05-14", "name": "Kenaikan Isa Almasih"},
    {"date": "2026-05-27", "name": "Idul Adha 1447 H (perkiraan)"},
    {"date": "2026-05-31", "name": "Hari Raya Waisak"},
    {"date": "2026-06-01", "name": "Hari Lahir Pancasila"},
    {"date": "2026-06-16", "name": "Tahun Baru Islam 1448 H"},
    {"date": "2026-08-17", "name": "Hari Kemerdekaan RI"},
    {"date": "2026-08-25", "name": "Maulid Nabi Muhammad SAW"},
    {"date": "2026-12-25", "name": "Hari Raya Natal"},
]


@router.get("/admin/content-calendar/holidays")
async def content_calendar_holidays(year: int = 2026, staff=Depends(get_current_staff)):
    return {"year": year, "holidays": [h for h in ID_HOLIDAYS_2026 if h["date"].startswith(str(year))]}


# ---------- Backup history ----------

BACKUP_DIR = "/app/backups"


@router.post("/admin/backup/trigger")
async def backup_trigger(request: Request, admin=Depends(get_current_admin)):
    """Backup manual: mongodump ke file + catat di riwayat."""
    import pathlib
    blob, filename = await _run_mongodump()
    pathlib.Path(BACKUP_DIR).mkdir(parents=True, exist_ok=True)
    path = f"{BACKUP_DIR}/{filename}"
    with open(path, "wb") as f:
        f.write(blob)
    db = await _get_db()
    res = await db.backup_history.insert_one({
        "filename": filename, "path": path, "size_bytes": len(blob),
        "kind": "manual", "by": admin.get("email", ""), "created_at": _now(),
    })
    return {"ok": True, "id": str(res.inserted_id), "filename": filename, "size_bytes": len(blob)}


@router.get("/admin/backup/history")
async def backup_history_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.backup_history.find({}).sort("created_at", -1).to_list(100)
    return [{"id": str(d["_id"]), "filename": d.get("filename", ""),
             "size_bytes": d.get("size_bytes", 0), "kind": d.get("kind", "manual"),
             "by": d.get("by", ""), "created_at": _iso(d.get("created_at", ""))}
            for d in docs]


@router.get("/admin/backup/history/{bid}/download")
async def backup_history_download(bid: str, admin=Depends(get_current_admin)):
    from fastapi.responses import Response as _R
    db = await _get_db()
    d = await db.backup_history.find_one({"_id": _oid(bid)})
    if not d:
        raise HTTPException(status_code=404, detail="Backup not found")
    try:
        with open(d["path"], "rb") as f:
            blob = f.read()
    except FileNotFoundError:
        raise HTTPException(status_code=410, detail="File backup sudah tidak tersedia di disk")
    return _R(content=blob, media_type="application/gzip",
              headers={"Content-Disposition": f'attachment; filename="{d["filename"]}"'})


@router.get("/admin/backup/download")
async def backup_download(admin=Depends(get_current_admin)):
    """Download a full gzipped BSON archive of every collection.
    Streams via a plain `bytes` response - the archive is small enough
    for the ~1000-row datasets this portal carries."""
    from fastapi.responses import Response as _R
    try:
        blob, filename = await _run_mongodump()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup failed: {e}")
    return _R(
        content=blob,
        media_type="application/gzip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Backup-Size":       str(len(blob)),
            "Cache-Control":       "no-store",
        },
    )


@router.post("/admin/system/factory-reset")
async def system_factory_reset(payload: m.FactoryResetIn,
                               request: Request,
                               admin=Depends(get_current_admin)):
    """DANGER: wipe the database back to a fresh-install state.

    Behaviour (see /app/memory/PRD.md - user-approved scope):
      • Preserves the entire `settings` collection (branding + landing CMS).
      • Preserves ALL users whose `role == "admin"` (multiple admins survive).
      • Deletes every other document in every other collection.
      • System collections (`system.*`) and any index metadata are left alone.

    Guards (both required):
      1. `admin_password` must match the calling admin's current password.
      2. `confirm` body field must be the exact string "FACTORY RESET".

    Returns a per-collection summary of documents removed so the operator
    can see exactly what was purged. The admin's session token stays valid
    because the admin user document itself is preserved.
    """
    if payload.confirm != "FACTORY RESET":
        raise HTTPException(
            status_code=400,
            detail='Confirmation phrase mismatch. Type "FACTORY RESET" exactly.',
        )

    db = await _get_db()

    # ---- Verify admin password against the fresh DB record ----
    admin_doc = await db.users.find_one({"_id": ObjectId(admin["id"])})
    if not admin_doc or admin_doc.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin account not found")
    if not verify_password(payload.admin_password, admin_doc.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Admin password is incorrect")

    # ---- Take a safety snapshot BEFORE wiping (best-effort) ----
    safety_backup_path = None
    try:
        blob, filename = await _run_mongodump()
        import os as _os
        backup_dir = "/var/backups/intercloud"
        try:
            _os.makedirs(backup_dir, exist_ok=True)
            safety_backup_path = _os.path.join(backup_dir, f"pre-factory-reset-{filename}")
            with open(safety_backup_path, "wb") as fh:
                fh.write(blob)
        except PermissionError:
            # Fall back to /tmp when running without root (e.g., preview env)
            backup_dir = "/tmp"
            safety_backup_path = _os.path.join(backup_dir, f"pre-factory-reset-{filename}")
            with open(safety_backup_path, "wb") as fh:
                fh.write(blob)
    except Exception as e:  # nosec
        # Never block the reset on a backup failure, but surface it in logs.
        safety_backup_path = f"(snapshot skipped: {e})"

    # ---- Wipe ----
    PRESERVE_COLLECTIONS = {"settings"}
    summary: dict = {}

    all_names = await db.list_collection_names()
    for name in all_names:
        if name.startswith("system."):
            continue
        if name in PRESERVE_COLLECTIONS:
            continue
        if name == "users":
            # Purge everything EXCEPT admins.
            res = await db.users.delete_many({"role": {"$ne": "admin"}})
            summary[name] = {"deleted": res.deleted_count, "kept": "role==admin"}
            continue
        # Drop the whole collection (much faster than delete_many on big sets).
        try:
            count = await db[name].estimated_document_count()
        except Exception:
            count = None
        await db[name].drop()
        summary[name] = {"deleted": count, "dropped": True}

    # ---- Immutable audit trail: insert a single reset marker into the
    # freshly-wiped audit_logs collection so the operator can see WHO wiped
    # the system and WHEN, even after the reset dropped historical logs.
    await log_audit(db, actor=admin, action="system.factory_reset", category="system",
                    target_type="system", target_label="Factory Reset",
                    metadata={"safety_backup": str(safety_backup_path)[:400],
                              "collections_affected": len(summary)},
                    severity="critical", request=request)

    return {
        "ok": True,
        "message": "Factory reset complete. Admin account and settings preserved.",
        "safety_backup": safety_backup_path,
        "collections": summary,
    }


async def _log_factory_reset_after(db, admin, request, summary, backup_path):
    """Reserved (unused): factory-reset dropped audit_logs; a marker row is
    inserted inline before returning so the operator sees the trigger."""
    await log_audit(db, actor=admin, action="system.factory_reset", category="system",
                    target_type="system", target_label="Factory Reset",
                    metadata={"safety_backup": str(backup_path)[:400],
                              "collections_affected": len(summary or {})},
                    severity="critical", request=request)


@router.post("/admin/backup/restore")
async def backup_restore(request: Request, admin=Depends(get_current_admin), confirm: str = ""):
    """Restore a full snapshot. **Wipes every collection contained in the
    archive** (mongorestore --drop) and reinstates the uploaded content.

    Expects the archive as the raw request body (`Content-Type` is
    ignored; slugs like `application/gzip` or `application/octet-stream`
    both work). Requires `?confirm=REPLACE` as a safety guard."""
    if confirm != "REPLACE":
        raise HTTPException(status_code=400,
                            detail="Confirmation required: pass ?confirm=REPLACE")
    blob = await request.body()
    if not blob or len(blob) < 32:
        raise HTTPException(status_code=400, detail="Empty or too-small upload")
    try:
        log = await _run_mongorestore(blob, drop=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore failed: {e}")
    # Audit - restore also drops audit_logs, so the marker is the first row after.
    db = await _get_db()
    await log_audit(db, actor=admin, action="system.backup_restore", category="system",
                    target_type="system", target_label="Backup Restore",
                    metadata={"bytes_received": len(blob)},
                    severity="critical", request=request)
    return {
        "ok": True,
        "bytes_received": len(blob),
        "log_tail": log[-1200:],
    }


# ============================================================
# System update - runs scripts/update.sh which git-pulls, installs deps,
# rebuilds the frontend, and restarts supervisor. Auto-backs up first.
# ============================================================
@router.get("/admin/system/version")
async def system_version(admin=Depends(get_current_admin)):
    """Return current git SHA + short version info for the update UI."""
    import asyncio as _asyncio, os as _os
    repo_root = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), "..", ".."))
    async def _run(cmd):
        p = await _asyncio.create_subprocess_exec(
            *cmd, cwd=repo_root,
            stdout=_asyncio.subprocess.PIPE, stderr=_asyncio.subprocess.PIPE,
        )
        out, err = await p.communicate()
        return (out.decode(errors="replace").strip() if p.returncode == 0 else "")
    sha    = await _run(["git", "rev-parse", "HEAD"])
    short  = await _run(["git", "rev-parse", "--short", "HEAD"])
    branch = await _run(["git", "rev-parse", "--abbrev-ref", "HEAD"])
    subject= await _run(["git", "log", "-1", "--pretty=%s"])
    date   = await _run(["git", "log", "-1", "--pretty=%cI"])
    return {
        "sha": sha or None,
        "short": short or None,
        "branch": branch or None,
        "subject": subject or None,
        "date": date or None,
        "repo_root": repo_root,
    }


@router.post("/admin/system/update")
async def system_update(admin=Depends(get_current_admin), confirm: str = ""):
    """Run `scripts/update.sh` in the checkout - auto-backs up first, then
    `git pull`, `pip install`, `yarn install && yarn build`, and restarts
    supervisor. **Preserves both .env files and the live database**.

    Guarded by `?confirm=UPDATE` so a stray click cannot trigger an update.
    Returns the STATUS line (`STATUS=ok OLD=<sha> NEW=<sha> BACKUP=<path>`)
    and the last ~2 KB of the script's log for diagnostics.

    Uses a filesystem lock at `/tmp/intercloud-update.lock` so two concurrent
    clicks return 409 instead of racing two `bash update.sh` invocations."""
    if confirm != "UPDATE":
        raise HTTPException(status_code=400,
                            detail="Confirmation required: pass ?confirm=UPDATE")
    import asyncio as _asyncio, os as _os, fcntl as _fcntl
    repo_root = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), "..", ".."))
    script = _os.path.join(repo_root, "scripts", "update.sh")
    if not _os.path.isfile(script):
        raise HTTPException(status_code=500,
                            detail=f"update.sh not found at {script}")

    lock_path = "/tmp/intercloud-update.lock"
    lock_fd = _os.open(lock_path, _os.O_CREAT | _os.O_RDWR, 0o644)
    try:
        try:
            _fcntl.flock(lock_fd, _fcntl.LOCK_EX | _fcntl.LOCK_NB)
        except BlockingIOError:
            raise HTTPException(status_code=409,
                                detail="Another update is already running.")

        proc = await _asyncio.create_subprocess_exec(
            "/bin/bash", script,
            cwd=repo_root,
            stdout=_asyncio.subprocess.PIPE,
            stderr=_asyncio.subprocess.STDOUT,
        )
        try:
            stdout_b, _ = await _asyncio.wait_for(proc.communicate(), timeout=600)
        except _asyncio.TimeoutError:
            proc.kill()
            raise HTTPException(status_code=504,
                                detail="Update timed out after 10 min")
        text = stdout_b.decode(errors="replace")
        status_line = next((l for l in text.splitlines() if l.startswith("STATUS=")), "")
        rc = proc.returncode

        # Map well-known exit codes to distinct 4xx statuses so the UI can
        # render a helpful message instead of a raw stderr traceback:
        # 0 = ok/noop | 2 = backup failed | 3 = dirty tree | 4 = no remote
        if rc == 0:
            return {"ok": True, "status": status_line, "return_code": 0,
                    "log_tail": text[-2400:]}
        if rc == 3:
            raise HTTPException(status_code=409,
                                detail="Working tree has uncommitted changes; "
                                       "commit or reset before updating. "
                                       f"Log: {text[-800:]}")
        if rc == 4:
            raise HTTPException(status_code=422,
                                detail="This checkout has no git remote - "
                                       "cannot update. Deploy a proper git "
                                       "clone (see docs/production.md).")
        raise HTTPException(status_code=500,
                            detail=f"update.sh exited {rc}: {text[-800:]}")
    finally:
        try: _fcntl.flock(lock_fd, _fcntl.LOCK_UN)
        except Exception: pass
        try: _os.close(lock_fd)
        except Exception: pass


def _pdf_template(
    *,
    doc_kind: str,           # "invoice" or "quotation"
    number: str,
    issued_date: str,        # YYYY-MM-DD or ISO
    due_or_valid_date: str,  # YYYY-MM-DD
    due_or_valid_label: str, # "Due Date" or "Valid Until"
    items: list,
    subtotal: float,
    tax_amount: float,    total: float,
    tax_percent: float,
    status: str,
    billed_to: dict,
    transactions: list = None,
    balance: float = None,
    notes: str = "",
    banks: list = None,
    extra_footer: str = "",
    for_pdf: bool = False,
    logo_url: str = LOGO_URL,
) -> str:
    """Renders the invoice/quotation HTML matching the reference layout."""
    transactions = transactions or []
    title = "Invoice" if doc_kind == "invoice" else "Quotation"
    header_title = f"{title} #{number}"

    # ---- items table (Description | Total) ----
    item_rows = "".join(
        f"<tr>"
        f"<td class='desc'>{i.get('description','')}{_period_label(i)}</td>"
        f"<td class='amt'>{_idr(i.get('total', (i.get('qty',1) * i.get('unit_price',0))))}</td>"
        f"</tr>"
        for i in items
    )

    # ---- transactions table (only if there are any) ----
    tx_rows = "".join(
        f"<tr>"
        f"<td>{_long_date(t.get('date',''))}</td>"
        f"<td>{t.get('gateway','')}</td>"
        f"<td>{t.get('transaction_id','') or '-'}</td>"
        f"<td class='amt'>{_idr(t.get('amount',0))}</td>"
        f"</tr>"
        for t in transactions
    )
    if transactions:
        bal = balance if balance is not None else max(0.0, float(total or 0) - sum(float(x.get("amount") or 0) for x in transactions))
        tx_block = f"""
        <div class="section-title">Transactions</div>
        <table class="tx">
          <thead>
            <tr>
              <th style="width:28%">Transaction Date</th>
              <th style="width:22%">Gateway</th>
              <th style="width:28%">Transaction ID</th>
              <th style="width:22%;text-align:right">Amount</th>
            </tr>
          </thead>
          <tbody>{tx_rows}</tbody>
          <tfoot>
            <tr><td colspan="3" class="bal-lbl">Balance</td><td class="amt bal-val">{_idr(bal)}</td></tr>
          </tfoot>
        </table>
        """
    else:
        tx_block = ""

    # ---- banks (unpaid only) ----
    if banks:
        bank_rows = "".join(
            f"<div class='bank-row'><span class='bank-name'>{b['bank']}</span>"
            f"<span class='bank-num'>{b['number']}</span>"
            f"<span class='bank-holder'>A/N {b['holder']}</span></div>"
            for b in banks
        )
        bank_block = f"""
        <div class="bank-panel">
          <div class="section-title" style="margin-top:0">Payment - Bank Transfer</div>
          {bank_rows}
          <div class="bank-note">Please include invoice number <b>{number}</b> in the transfer memo. Confirmation via WhatsApp speeds up reconciliation.</div>
        </div>
        """
    else:
        bank_block = ""

    ribbon = _corner_ribbon(status)
    generated_on = _long_date(datetime.now(timezone.utc).strftime("%Y-%m-%d"))

    # Actions bar only for browser (HTML) view
    actions_bar = "" if for_pdf else (
        f'<div class="actions">'
        f'<button onclick="window.print()">Print</button>'
        f'<a class="dl" href="?format=pdf&token={{TOKEN_PLACEHOLDER}}">Download PDF</a>'
        f'</div>'
    )

    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>{title} #{number}</title>
<style>
  @page {{ size: A4; margin: 14mm 14mm 16mm 14mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
         color:#334155; margin:0; padding:0; background:#f1f5f9; font-size:12px; line-height:1.45; }}
  .paper {{ background:#fff; padding:34px 40px 30px; max-width:800px; margin:20px auto; position:relative; box-shadow:0 6px 30px rgba(2,6,23,.08); }}

  /* Corner ribbon top-right */
  .ribbon-wrap {{ position:absolute; top:0; right:0; width:170px; height:170px; overflow:hidden; pointer-events:none; }}
  .ribbon {{ position:absolute; top:24px; right:-52px; transform:rotate(45deg); width:220px; text-align:center;
             color:#fff; font-weight:800; letter-spacing:.2em; padding:8px 0; font-size:14px;
             box-shadow:0 2px 6px rgba(0,0,0,.15); }}

  /* Header - logo sized generously so a wordmark reads cleanly on print */
  .head {{ display:flex; justify-content:space-between; align-items:center; gap:24px; }}
  .head .logo {{ flex:0 0 auto; max-width:55%; }}
  .head .logo img {{ height:130px; max-height:130px; width:auto; max-width:100%; object-fit:contain; display:block; }}
  .company-block {{ text-align:right; font-size:11.5px; color:#334155; line-height:1.55; }}

  /* Invoice title strip */
  .titlebar {{ margin-top:28px; background:#e5edf5; padding:14px 18px; }}
  .titlebar h1 {{ margin:0 0 6px 0; font-size:20px; color:#334155; font-weight:800; }}
  .titlebar .meta-line {{ font-size:12px; color:#475569; }}
  .titlebar .meta-line b {{ color:#0f172a; font-weight:600; }}

  /* Invoiced To */
  .to {{ margin-top:22px; }}
  .to .lbl {{ font-weight:700; font-size:12px; color:#111; margin-bottom:6px; }}
  .to .body {{ font-size:11.5px; color:#475569; line-height:1.6; }}

  /* Items table */
  table.items {{ width:100%; border-collapse:collapse; margin-top:22px; font-size:12px; }}
  table.items thead th {{ background:#e5edf5; color:#334155; font-weight:700; padding:9px 12px; text-align:center; border:1px solid #cbd5e1; }}
  table.items tbody td {{ padding:11px 12px; border:1px solid #e2e8f0; vertical-align:top; }}
  table.items td.desc {{ background:#fff; }}
  table.items td.amt {{ text-align:right; white-space:nowrap; font-variant-numeric:tabular-nums; }}

  /* Totals block */
  .totals {{ margin-top:6px; }}
  .totals table {{ margin-left:auto; border-collapse:collapse; font-size:12px; }}
  .totals td {{ padding:7px 14px; border:1px solid #e2e8f0; }}
  .totals td.lbl {{ background:#f1f5f9; text-align:right; font-weight:700; color:#0f172a; width:170px; }}
  .totals td.val {{ text-align:right; width:170px; font-variant-numeric:tabular-nums; }}
  .totals tr.grand td.lbl,
  .totals tr.grand td.val {{ background:#f1f5f9; font-weight:800; color:#0f172a; }}

  /* Transactions */
  .section-title {{ margin-top:28px; font-size:15px; font-weight:800; color:#0f172a; }}
  table.tx {{ width:100%; border-collapse:collapse; margin-top:10px; font-size:12px; }}
  table.tx thead th {{ background:#e5edf5; padding:9px 12px; border:1px solid #cbd5e1; color:#334155; text-align:center; font-weight:700; }}
  table.tx tbody td {{ padding:9px 12px; border:1px solid #e2e8f0; text-align:center; }}
  table.tx td.amt {{ text-align:right; font-variant-numeric:tabular-nums; }}
  table.tx tfoot td {{ padding:9px 12px; border:1px solid #e2e8f0; }}
  table.tx tfoot td.bal-lbl {{ text-align:right; font-weight:800; color:#0f172a; background:#f1f5f9; }}
  table.tx tfoot td.bal-val {{ background:#f1f5f9; font-weight:800; color:#0f172a; }}

  /* Bank panel */
  .bank-panel {{ margin-top:26px; background:#fffbeb; border:1px solid #fde68a; padding:14px 16px; }}
  .bank-row {{ display:flex; gap:16px; padding:4px 0; font-size:12px; }}
  .bank-name {{ font-weight:800; color:#0a2350; min-width:80px; }}
  .bank-num  {{ font-family: "SFMono-Regular", ui-monospace, Menlo, Consolas, monospace; color:#111; min-width:180px; }}
  .bank-holder {{ color:#78350f; }}
  .bank-note {{ font-size:11px; color:#78350f; margin-top:6px; }}

  /* Notes */
  .notes {{ margin-top:22px; font-size:11.5px; color:#475569; }}

  /* Footer */
  .foot {{ margin-top:26px; text-align:center; font-size:11px; color:#94a3b8; }}

  /* Print actions bar (HTML view only) */
  .actions {{ text-align:center; padding:16px 0 0 0; }}
  .actions button, .actions a.dl {{ display:inline-block; background:#0a2350; color:#fff; border:0; border-radius:99px; padding:8px 22px; font-weight:700; font-size:12px; cursor:pointer; text-decoration:none; margin: 0 6px; }}
  .actions a.dl {{ background:#f5b120; color:#0a2350; }}
  @media print {{ body {{ background:#fff }} .paper {{ box-shadow:none; margin:0 }} .actions {{ display:none }} }}
</style></head>
<body>
{actions_bar}
<div class="paper">
  {ribbon}

  <div class="head">
    <div class="logo"><img src="{logo_url}" alt="Intercloud Digital Inovasi"/></div>
    {COMPANY_HEADER_HTML}
  </div>

  <div class="titlebar">
    <h1>{header_title}</h1>
    <div class="meta-line">{title} Date: <b>{_long_date(issued_date)}</b></div>
    <div class="meta-line">{due_or_valid_label}: <b>{_long_date(due_or_valid_date)}</b></div>
  </div>

  <div class="to">
    <div class="lbl">Invoiced To</div>
    <div class="body">
      {_addressed_to_block(billed_to)}
    </div>
  </div>

  <table class="items">
    <thead>
      <tr><th style="text-align:center">Description</th><th style="text-align:center;width:180px">Total</th></tr>
    </thead>
    <tbody>{item_rows}</tbody>
  </table>

  <div class="totals">
    <table>
      <tr><td class="lbl">Sub Total</td><td class="val">{_idr(subtotal)}</td></tr>
      <tr><td class="lbl">Tax ({tax_percent:g}%)</td><td class="val">{_idr(tax_amount)}</td></tr>
      <tr><td class="lbl">Credit</td><td class="val">{_idr(0)}</td></tr>
      <tr class="grand"><td class="lbl">Total</td><td class="val">{_idr(total)}</td></tr>
    </table>
  </div>

  {tx_block}

  {bank_block}

  {("<div class='notes'>" + notes + "</div>") if notes else ""}

  {extra_footer}

  <div class="foot">PDF Generated on {generated_on}</div>
</div>
</body></html>
"""


def _render_pdf_bytes(html: str) -> bytes:
    from weasyprint import HTML
    return HTML(string=html).write_pdf()


@router.get("/documents/invoice/{iid}")
async def render_invoice_pdf(iid: str, format: str = "html", user=Depends(get_current_user)):
    db = await _get_db()
    d = await db.invoices.find_one({"_id": _oid(iid)})
    if not d:
        raise HTTPException(status_code=404, detail="Invoice not found")
    # Access: owner or staff
    if user["role"] == "client" and str(d["user_id"]) != str(user["id"]):
        raise HTTPException(status_code=403, detail="Not your invoice")
    u = await db.users.find_one({"_id": d["user_id"]}) or {}

    bank_doc = await db.settings.find_one({"key": "bank_accounts"}) or {}
    banks = bank_doc.get("value") or [
        {"bank": "MANDIRI", "number": "1240011911816", "holder": "INTERCLOUD DIGITAL INOVASI"},
        {"bank": "BCA", "number": "4730862038", "holder": "ANANG MADIA CUGITA"},
    ]

    status = (d.get("status") or "unpaid").lower()

    # Synthesize transactions from paid_at + payment_method when invoice is paid
    tx_list = list(d.get("transactions") or [])
    if not tx_list and status == "paid" and d.get("paid_at"):
        tx_list = [{
            "date": d.get("paid_at"),
            "gateway": (d.get("payment_method") or "Bank Transfer").replace("_", " ").title(),
            "transaction_id": d.get("payment_ref") or "",
            "amount": d.get("total", 0),
        }]

    html = _pdf_template(
        doc_kind="invoice",
        number=d.get("number", ""),
        issued_date=(d.get("created_at") or "")[:10],
        due_or_valid_date=d.get("due_date", ""),
        due_or_valid_label="Due Date",
        items=d.get("items", []),
        subtotal=d.get("subtotal", 0),
        tax_amount=d.get("tax_amount", 0),
        total=d.get("total", 0),
        tax_percent=d.get("tax_percent", 11.0),
        status=status,
        billed_to=u,
        transactions=tx_list,
        banks=banks if status in ("unpaid", "overdue") else None,
        notes=d.get("notes", ""),
        for_pdf=(format == "pdf"),
        logo_url=(await _get_branding_dict(db))["logo_dark"],
    )

    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        filename = f"Invoice-{d.get('number','invoice')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # For HTML view, inject token so the "Download PDF" button in-page can carry auth
    token = user.get("_token", "")
    html = html.replace("{TOKEN_PLACEHOLDER}", token)
    return HTMLResponse(content=html)


@router.get("/documents/quotation/{qid}")
async def render_quotation_pdf(qid: str, format: str = "html", staff=Depends(get_current_staff)):
    db = await _get_db()
    d = await db.quotations.find_one({"_id": _oid(qid)})
    if not d:
        raise HTTPException(status_code=404, detail="Quotation not found")
    u = await db.users.find_one({"_id": d["user_id"]}) or {}
    status = (d.get("status") or "draft").lower()

    html = _pdf_template(
        doc_kind="quotation",
        number=d.get("number", ""),
        issued_date=(d.get("created_at") or "")[:10],
        due_or_valid_date=d.get("valid_until", ""),
        due_or_valid_label="Valid Until",
        items=d.get("items", []),
        subtotal=d.get("subtotal", 0),
        tax_amount=d.get("tax_amount", 0),
        total=d.get("total", 0),
        tax_percent=d.get("tax_percent", 11.0),
        status=status,
        billed_to=u,
        transactions=[],
        banks=None,
        notes=d.get("notes", ""),
        extra_footer=(
            "<div style='margin-top:22px;font-size:11px;color:#64748b;line-height:1.7'>"
            "This quotation is valid until the date shown above. Prices are in Indonesian Rupiah (IDR) and exclude any applicable "
            "withholding tax. To accept this quotation, reply via email or WhatsApp - an invoice will be issued upon acceptance."
            "</div>"
        ),
        for_pdf=(format == "pdf"),
        logo_url=(await _get_branding_dict(db))["logo_dark"],
    )

    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        filename = f"Quotation-{d.get('number','quotation')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="{filename}"'},
        )

    token = staff.get("_token", "")
    html = html.replace("{TOKEN_PLACEHOLDER}", token)
    return HTMLResponse(content=html)


# Traffic Report (mocked realistic time series)
@router.get("/client/services/{sid}/traffic")
async def client_service_traffic(sid: str, user=Depends(get_current_user)):
    db = await _get_db()
    d = await db.services.find_one({"_id": _oid(sid), "user_id": ObjectId(user["id"])})
    if not d:
        raise HTTPException(status_code=404, detail="Service not found")
    # Deterministic mocked data based on service id hash
    import random
    seed = sum(ord(c) for c in sid)
    r = random.Random(seed)
    now = datetime.now(timezone.utc)
    points = []
    for i in range(24):
        h = now - timedelta(hours=23 - i)
        base_in = r.uniform(150, 850)
        base_out = r.uniform(120, 700)
        points.append({
            "t": h.strftime("%H:00"),
            "in_mbps": round(base_in, 1),
            "out_mbps": round(base_out, 1),
        })
    total_in = round(sum(p["in_mbps"] for p in points) * 60 / 8 / 1024, 2)  # GB
    total_out = round(sum(p["out_mbps"] for p in points) * 60 / 8 / 1024, 2)
    return {
        "service_id": sid,
        "service_name": d.get("name", ""),
        "range": "24h",
        "points": points,
        "totals": {"in_gb": total_in, "out_gb": total_out},
        "peak_in_mbps": max(p["in_mbps"] for p in points),
        "peak_out_mbps": max(p["out_mbps"] for p in points),
    }


# ============================================================
# Password lifecycle - change / admin-reset / forgot / reset
# ============================================================
import hashlib  # noqa: E402
import secrets as _secrets  # noqa: E402
from fastapi import Request  # noqa: E402
from portal import integrations_v2 as iv2  # noqa: E402


@router.post("/auth/change-password")
async def auth_change_password(payload: m.ChangePasswordIn, user=Depends(get_current_user)):
    """Any authenticated user (client or staff) can rotate their OWN password."""
    db = await _get_db()
    u = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not u or not verify_password(payload.current_password, u["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    if payload.current_password == payload.new_password:
        raise HTTPException(status_code=400, detail="New password must differ from the current one")
    await db.users.update_one(
        {"_id": u["_id"]},
        {"$set": {"password_hash": hash_password(payload.new_password),
                  "password_changed_at": _now(),
                  "must_change_password": False}},
    )
    # Invalidate any outstanding reset tokens for this user
    await db.password_resets.update_many({"user_id": u["_id"], "used": False}, {"$set": {"used": True}})
    return {"ok": True, "message": "Password updated"}


@router.post("/admin/users/{uid}/reset-password")
async def admin_reset_user_password(uid: str, payload: m.AdminResetPasswordIn,
                                    request: Request,
                                    admin=Depends(get_current_admin)):
    """Admin sets a new password for another user. Optionally emails them."""
    db = await _get_db()
    target = await db.users.find_one({"_id": _oid(uid)})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    await db.users.update_one(
        {"_id": target["_id"]},
        {"$set": {"password_hash": hash_password(payload.new_password),
                  "password_changed_at": _now(),
                  "password_reset_by": ObjectId(admin["id"])}},
    )
    await db.password_resets.update_many({"user_id": target["_id"], "used": False},
                                         {"$set": {"used": True}})
    await log_audit(db, actor=admin, action="user.password_reset", category="security",
                    target_type="user", target_id=uid,
                    target_label=target.get("email", ""),
                    metadata={"notify_user": bool(payload.notify_user)},
                    severity="warning", request=request)
    # Optional email
    sent = False
    if payload.notify_user:
        try:
            await _send_password_notice(db, target, kind="admin_reset")
            sent = True
        except Exception:
            sent = False
    return {"ok": True, "message": f"Password reset for {target['email']}", "email_sent": sent}


@router.post("/auth/forgot-password")
@_rl_limiter.limit(AUTH_FORGOT_LIMIT)
async def auth_forgot_password(payload: m.ForgotPasswordIn, request: Request):
    """Public. Always returns 200 to avoid email enumeration.

    Generates a signed one-time token, stores it (hashed) in `password_resets`,
    and emails the user a link. If SMTP isn't configured we log the link so
    the admin can still recover the user manually while awaiting SMTP setup.
    """
    db = await _get_db()
    from portal import integrations_v2 as _iv2
    await _iv2.enforce_recaptcha(
        db, payload.recaptcha_token, "forgot",
        request.client.host if request.client else None,
    )
    email = payload.email.lower().strip()
    u = await db.users.find_one({"email": email})
    if u:
        raw = _secrets.token_urlsafe(48)
        token_hash = hashlib.sha256(raw.encode()).hexdigest()
        expires = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
        await db.password_resets.insert_one({
            "user_id": u["_id"], "email": email, "token_hash": token_hash,
            "expires_at": expires, "used": False, "created_at": _now(),
            "requester_ip": (request.client.host if request.client else "unknown"),
        })
        # Best-effort email
        origin = os.environ.get("REACT_APP_BACKEND_URL", "")
        reset_url = f"{origin}/portal/reset-password?token={raw}"
        try:
            await _send_password_notice(db, u, kind="forgot", reset_url=reset_url)
        except Exception as e:
            # SMTP down or not configured → log the link to backend log so admin can share it.
            import logging
            logging.getLogger("portal.password_reset").warning(
                f"[password-reset] SMTP unavailable ({e}) - reset link for {email}: {reset_url}"
            )
    # Always the same response
    return {"ok": True, "message": "If an account exists for that email, a reset link has been sent."}


@router.post("/auth/reset-password")
@_rl_limiter.limit(AUTH_RESET_LIMIT)
async def auth_reset_password(payload: m.ResetPasswordIn, request: Request):
    db = await _get_db()
    token_hash = hashlib.sha256(payload.token.encode()).hexdigest()
    row = await db.password_resets.find_one({"token_hash": token_hash, "used": False})
    if not row:
        raise HTTPException(status_code=400, detail="This reset link is invalid or has already been used.")
    if row.get("expires_at", "") < datetime.now(timezone.utc).isoformat():
        raise HTTPException(status_code=400, detail="This reset link has expired. Request a new one.")
    await db.users.update_one(
        {"_id": row["user_id"]},
        {"$set": {"password_hash": hash_password(payload.new_password),
                  "password_changed_at": _now()}},
    )
    await db.password_resets.update_one({"_id": row["_id"]},
                                        {"$set": {"used": True, "used_at": _now()}})
    return {"ok": True, "message": "Password updated. You may now log in."}


async def _send_password_notice(db, user: dict, *, kind: str, reset_url: str = "") -> None:
    """Compose + send transactional email via SMTP integration.

    Raises on any failure - caller decides whether to swallow.
    """
    smtp = await iv2.get_settings(db, "smtp") if False else None  # avoid circular; done below
    # Late import (routes.py appended block hasn't imported iv2 up here in this hunk)
    from portal import integrations_v2 as _iv2
    smtp = await _iv2.get_settings(db, "smtp")
    if not smtp or not smtp.get("enabled"):
        raise RuntimeError("SMTP not configured")
    if kind == "forgot":
        subject = "Reset your Intercloud portal password"
        html = (
            f"<p>Hi {user.get('name','there')},</p>"
            f"<p>We received a password-reset request for your Intercloud portal account. "
            f"Click the button below within the next 60 minutes to set a new password:</p>"
            f"<p><a href='{reset_url}' style='display:inline-block;padding:10px 22px;"
            f"background:#0a2350;color:#fff;text-decoration:none;border-radius:99px;"
            f"font-weight:700;letter-spacing:.05em'>Reset password</a></p>"
            f"<p style='color:#64748b;font-size:12px'>If the button doesn't work, copy and paste this link:<br>{reset_url}</p>"
            f"<p style='color:#64748b;font-size:12px'>Didn't request this? You can ignore this email - your password wasn't changed.</p>"
        )
    else:
        subject = "Your Intercloud portal password was reset"
        html = (
            f"<p>Hi {user.get('name','there')},</p>"
            f"<p>An administrator has reset the password for your Intercloud portal account. "
            f"Please contact your account manager for the new password, "
            f"or use the &lsquo;Forgot password&rsquo; link on the portal login page.</p>"
        )
    _iv2.SMTPMailer(smtp).send(to=user["email"], subject=subject, html=html)



@router.get("/admin/integrations-v2/schema")
async def integrations_v2_schema(admin=Depends(get_current_admin)):
    """Returns the field schema the admin UI uses to render each integration's settings form."""
    db = await _get_db()
    allow_extra = bool(await _get_setting_value(db, "enable_extra_payment_gateways", False))
    return {k: v for k, v in iv2.INTEGRATION_SCHEMA.items()
            if allow_extra or k not in _EXTRA_PAYMENT_MODULES}


@router.get("/admin/integrations-v2")
async def integrations_v2_list(admin=Depends(get_current_admin)):
    """Return all persisted integration settings (secrets masked)."""
    db = await _get_db()
    allow_extra = bool(await _get_setting_value(db, "enable_extra_payment_gateways", False))
    out = {}
    for provider in iv2.INTEGRATION_SCHEMA.keys():
        if not allow_extra and provider in _EXTRA_PAYMENT_MODULES:
            continue
        d = await iv2.get_settings(db, provider)
        out[provider] = iv2.redact(d) or {"provider": provider, "enabled": False, "credentials": {}, "options": {}}
    return out


@router.put("/admin/integrations-v2/{provider}")
async def integrations_v2_upsert(provider: str, payload: dict, request: Request, admin=Depends(get_current_admin)):
    if provider not in iv2.INTEGRATION_SCHEMA:
        raise HTTPException(status_code=404, detail="Unknown provider")
    db = await _get_db()
    if provider in _EXTRA_PAYMENT_MODULES and not bool(
            await _get_setting_value(db, "enable_extra_payment_gateways", False)):
        raise HTTPException(status_code=400,
                            detail="Gateway ini dinonaktifkan - Duitku adalah satu-satunya payment gateway aktif.")
    # Merge - never drop existing secrets if the incoming value is empty
    existing = await iv2.get_settings(db, provider) or {}
    creds_in = payload.get("credentials") or {}
    merged_creds = {**(existing.get("credentials") or {})}
    for k, v in creds_in.items():
        if v not in ("", None):
            merged_creds[k] = v
    doc = {
        "enabled": bool(payload.get("enabled")),
        "sandbox": payload.get("sandbox", existing.get("sandbox", True)),
        "channel": payload.get("channel", existing.get("channel")),
        "credentials": merged_creds,
        "options": payload.get("options", existing.get("options") or {}),
    }
    saved = await iv2.upsert_settings(db, provider, doc)
    # Redact both snapshots before writing audit (credentials → ••••)
    await log_audit(db, actor=admin, action="integration.update", category="integrations",
                    target_type="integration", target_id=provider, target_label=provider,
                    before=iv2.redact(existing) if existing else None,
                    after=iv2.redact(saved),
                    severity="warning", request=request)
    return iv2.redact(saved)


@router.delete("/admin/integrations-v2/{provider}")
async def integrations_v2_delete(provider: str, request: Request, admin=Depends(get_current_admin)):
    """Wipe all persisted settings for a provider (credentials + options + enabled).

    Useful for rotating credentials cleanly - the PUT endpoint merges by design,
    so it cannot clear a stored secret on its own.
    """
    if provider not in iv2.INTEGRATION_SCHEMA:
        raise HTTPException(status_code=404, detail="Unknown provider")
    db = await _get_db()
    r = await db.integration_settings.delete_one({"provider": provider})
    if r.deleted_count:
        await log_audit(db, actor=admin, action="integration.delete", category="integrations",
                        target_type="integration", target_id=provider, target_label=provider,
                        severity="warning", request=request)
    return {"deleted": r.deleted_count}


@router.post("/admin/integrations-v2/{provider}/test")
async def integrations_v2_test(provider: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    settings = await iv2.get_settings(db, provider)
    if not settings:
        return {"ok": False, "message": "Integration is not configured yet."}
    if provider == "proxmox":
        return await iv2.ProxmoxClient(settings).test_connection()
    if provider == "mikrotik":
        return iv2.MikrotikClient(settings).test_connection()
    if provider in iv2.PAYMENT_PROVIDERS:
        gw = iv2.payment_gateway(provider, settings)
        return await gw.test_connection()
    if provider == "smtp":
        return iv2.SMTPMailer(settings).test_connection()
    if provider == "imap":
        return iv2.IMAPClient(settings).test_connection()
    if provider == "cpanel":
        c = settings.get("credentials") or {}
        missing = [k for k in ("host", "username") if not c.get(k)]
        secret_ok = bool(c.get("api_token") or c.get("password"))
        if missing or not secret_ok:
            return {"ok": False, "message": f"Missing credentials: {', '.join(missing) or 'api_token/password'}"}
        return await iv2.CpanelClient(settings).test_connection()
    if provider == "plesk":
        c = settings.get("credentials") or {}
        missing = [k for k in ("host", "username") if not c.get(k)]
        secret_ok = bool(c.get("api_token") or c.get("password"))
        if missing or not secret_ok:
            return {"ok": False, "message": f"Missing credentials: {', '.join(missing) or 'api_token/password'}"}
        return await iv2.PleskClient(settings).test_connection()
    if provider == "directadmin":
        c = settings.get("credentials") or {}
        missing = [k for k in ("host", "username") if not c.get(k)]
        secret_ok = bool(c.get("api_token") or c.get("password"))
        if missing or not secret_ok:
            return {"ok": False, "message": f"Missing credentials: {', '.join(missing) or 'api_token/password'}"}
        return await iv2.DirectAdminClient(settings).test_connection()
    if provider == "rna":
        c = settings.get("credentials") or {}
        missing = [k for k in ("reseller_id", "api_key") if not c.get(k)]
        if missing:
            return {"ok": False, "message": f"Missing credentials: {', '.join(missing)}"}
        return await iv2.RdashClient(settings).test_connection()
    return {"ok": False, "message": "No test method"}


# ---------------- Client domains (RNA.id / RDASH) ----------------
_DOMAIN_NAME_RE = re.compile(r"^(?=.{4,253}$)([a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,}$")


async def _rna_client(db):
    s = await iv2.get_settings(db, "rna")
    if s and s.get("enabled"):
        return iv2.RdashClient(s)
    return None


async def _rdap_lookup(domain: str) -> dict:
    """Public RDAP fallback (rdap.org) so WHOIS is live even without reseller creds."""
    import httpx
    async with httpx.AsyncClient(timeout=20, follow_redirects=True) as c:
        r = await c.get(f"https://rdap.org/domain/{domain}",
                        headers={"Accept": "application/rdap+json"})
    if r.status_code == 404:
        return {"registered": False, "registrar": "", "status": [], "created": "",
                "updated": "", "expiry": "", "nameservers": [], "dnssec": "unsigned"}
    r.raise_for_status()
    j = r.json()
    events = {e.get("eventAction"): (e.get("eventDate") or "")[:10] for e in j.get("events", [])}
    registrar = ""
    for ent in j.get("entities", []):
        if "registrar" in (ent.get("roles") or []):
            for item in (ent.get("vcardArray") or [None, []])[1]:
                if item and item[0] == "fn" and len(item) > 3:
                    registrar = item[3]
    return {
        "registered": True,
        "registrar": registrar,
        "status": j.get("status") or [],
        "created": events.get("registration", ""),
        "updated": events.get("last changed", ""),
        "expiry": events.get("expiration", ""),
        "nameservers": [n.get("ldhName", "").lower() for n in j.get("nameservers", []) if n.get("ldhName")],
        "dnssec": "signed" if (j.get("secureDNS") or {}).get("delegationSigned") else "unsigned",
    }


@router.get("/client/domains/whois")
async def client_domain_whois(domain: str, user=Depends(get_current_user)):
    db = await _get_db()
    name = domain.strip().lower()
    if not _DOMAIN_NAME_RE.match(name):
        raise HTTPException(status_code=400, detail="Nama domain tidak valid")
    rna = await _rna_client(db)
    if rna:
        try:
            w = await rna.whois(name)
            return {
                "live": True, "source": "rna",
                "domain": w.get("name") or name,
                "registered": not bool(w.get("available")),
                "registrar": w.get("registrar", ""),
                "status": w.get("status") or [],
                "created": (w.get("created_at") or "")[:10],
                "updated": (w.get("updated_at") or "")[:10],
                "expiry": (w.get("expired_at") or "")[:10],
                "registrant": "REDACTED FOR PRIVACY",
                "nameservers": w.get("nameserver") or [],
                "dnssec": w.get("dnssec") or "unsigned",
            }
        except Exception as e:
            import logging
            logging.getLogger("portal.domains").warning("RNA whois gagal untuk %s: %s", name, e)
    try:
        data = await _rdap_lookup(name)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"WHOIS lookup gagal: {str(e)[:120]}")
    return {"live": True, "source": "rdap", "domain": name,
            "registrant": "REDACTED FOR PRIVACY", **data}


_TLD_PRICES_IDR = {".com": 165000, ".id": 250000, ".co.id": 300000, ".net": 185000,
                   ".org": 175000, ".my.id": 25000, ".web.id": 55000, ".biz.id": 55000}


def _tld_price(name: str) -> int:
    tld = next((t for t in sorted(_TLD_PRICES_IDR, key=len, reverse=True) if name.endswith(t)), None)
    return _TLD_PRICES_IDR.get(tld, 95000)


async def _dns_domain_taken(name: str) -> bool:
    """DNS fallback: NXDOMAIN on the NS query means the domain is very likely available."""
    import dns.asyncresolver
    import dns.resolver
    resolver = dns.asyncresolver.Resolver()
    resolver.lifetime = 4
    try:
        await resolver.resolve(name, "NS")
        return True
    except dns.resolver.NXDOMAIN:
        return False
    except Exception:
        return True


async def _check_domains_availability(db, names: list) -> list:
    """Availability via RNA.id bila aktif; fallback live DNS check."""
    import asyncio
    rna = await _rna_client(db)

    async def _one(n: str) -> dict:
        if rna:
            try:
                res = await rna.availability(n)
                item = res[0] if res else {}
                return {"domain": n, "available": bool(item.get("available")), "source": "rna"}
            except Exception:
                pass
        try:
            taken = await _dns_domain_taken(n)
            return {"domain": n, "available": not taken, "source": "dns"}
        except Exception:
            return {"domain": n, "available": None, "source": "dns"}

    out = list(await asyncio.gather(*(_one(n) for n in names)))
    for r in out:
        r["price"] = _tld_price(r["domain"])
    return out


@router.get("/client/domains/suggest")
async def client_domain_suggest(q: str, user=Depends(get_current_user)):
    db = await _get_db()
    base = re.sub(r"[^a-z0-9-]", "", q.strip().lower().split(".")[0])[:40].strip("-")
    if len(base) < 2:
        raise HTTPException(status_code=400, detail="Kata kunci minimal 2 karakter")
    variants = [f"get{base}.com", f"{base}online.com", f"{base}-id.com", f"my{base}.id",
                f"{base}store.com", f"{base}.web.id", f"{base}hq.com", f"{base}.biz.id"]
    results = await _check_domains_availability(db, variants)
    return {"live": True, "query": base, "suggestions": results}


@router.post("/client/domains/order")
async def client_domain_order(payload: m.DomainOrderIn, request: Request, user=Depends(get_current_user)):
    """Order registrasi domain: buat record + invoice; registrasi otomatis jalan saat lunas."""
    db = await _get_db()
    name = payload.domain.strip().lower()
    if not _DOMAIN_NAME_RE.match(name):
        raise HTTPException(status_code=400, detail="Nama domain tidak valid")
    existing = await db.domains.find_one({"domain": name, "status": {"$in": ["pending", "active", "expiring"]}})
    if existing:
        raise HTTPException(status_code=400, detail="Domain sudah terdaftar dalam sistem")
    chk = (await _check_domains_availability(db, [name]))[0]
    if chk["available"] is False:
        raise HTTPException(status_code=400, detail="Domain tidak tersedia untuk registrasi")
    price = _tld_price(name) * payload.years
    tax_percent = float(await _get_setting_value(db, "default_tax_percent", 11.0))
    tax = round(price * tax_percent / 100.0, 2)
    due = (datetime.now(timezone.utc) + timedelta(days=3)).date().isoformat()
    dom = {
        "user_id": ObjectId(user["id"]),
        "domain": name,
        "tld": "." + name.split(".", 1)[1],
        "status": "pending",
        "registrar": "rna",
        "years": payload.years,
        "auto_renew": payload.auto_renew,
        "registered_at": None,
        "expires_at": None,
        "nameservers": [],
        "price": price,
        "invoice_id": None,
        "order_ref": None,
        "created_at": _now(),
    }
    dr = await db.domains.insert_one(dom)
    inv = {
        "user_id": ObjectId(user["id"]),
        "items": [{"description": f"Registrasi domain {name} ({payload.years} tahun)",
                   "qty": 1, "price": price, "total": price}],
        "subtotal": price,
        "tax_percent": tax_percent,
        "tax_amount": tax,
        "total": round(price + tax, 2),
        "due_date": due,
        "status": "unpaid",
        "payment_method": None,
        "paid_at": None,
        "notes": f"Registrasi domain {name} - diproses otomatis setelah pembayaran.",
        "domain_id": str(dr.inserted_id),
        "created_at": _now(),
    }
    inv = await _insert_numbered(db, "invoices", "INV", inv)
    await db.domains.update_one({"_id": dr.inserted_id}, {"$set": {"invoice_id": str(inv["_id"])}})
    await log_audit(db, actor=user, action="client_domain.order_created", category="domains",
                    target_type="domain", target_id=str(dr.inserted_id), target_label=name,
                    metadata={"invoice": inv["number"], "total": inv["total"], "years": payload.years},
                    request=request)
    return {"ok": True, "domain_id": str(dr.inserted_id), "invoice_id": str(inv["_id"]),
            "number": inv["number"], "total": inv["total"], "due_date": due}


async def _auto_register_domain(db, inv: dict) -> bool:
    """Registrasi domain otomatis di RNA.id setelah invoice registrasi lunas. Idempotent
    (hanya memproses domain berstatus pending)."""
    if not inv.get("domain_id"):
        return False
    try:
        dom = await db.domains.find_one({"_id": _oid(inv["domain_id"]), "status": "pending"})
    except Exception:
        return False
    if not dom:
        return False
    now = datetime.now(timezone.utc)
    fallback_expiry = (now + timedelta(days=365 * int(dom.get("years", 1)))).date().isoformat()
    rna = await _rna_client(db)
    if rna:
        try:
            res = await rna.register(dom["domain"], int(dom.get("years", 1)))
            ns = [res.get(f"nameserver_{i}") for i in range(1, 6) if res.get(f"nameserver_{i}")]
            await db.domains.update_one({"_id": dom["_id"]}, {"$set": {
                "status": "active",
                "registered_at": now.date().isoformat(),
                "expires_at": (res.get("expired_at") or "")[:10] or fallback_expiry,
                "nameservers": ns or rna.default_ns,
                "order_ref": str(res.get("id") or ""),
                "provision_note": "Registered live via RNA.id (RDASH).",
            }})
            return True
        except Exception as e:
            await db.domains.update_one({"_id": dom["_id"]}, {"$set": {
                "provision_note": f"Registrasi RNA.id gagal: {str(e)[:150]}. Perlu tindak lanjut manual.",
            }})
            return False
    await db.domains.update_one({"_id": dom["_id"]}, {"$set": {
        "status": "active",
        "registered_at": now.date().isoformat(),
        "expires_at": fallback_expiry,
        "nameservers": ["ns1.intercloud-digital.com", "ns2.intercloud-digital.com"],
        "provision_note": "Integrasi RNA.id belum aktif - dicatat internal, submit manual ke registrar.",
    }})
    return True


@router.post("/client/domains/{did}/renew")
async def client_domain_renew(did: str, payload: m.DomainRenewIn, request: Request,
                              user=Depends(get_current_user)):
    """Order perpanjangan domain: buat invoice; perpanjangan otomatis jalan saat lunas."""
    db = await _get_db()
    dom = await db.domains.find_one({"_id": _oid(did), "user_id": ObjectId(user["id"])})
    if not dom:
        raise HTTPException(status_code=404, detail="Domain tidak ditemukan")
    if dom.get("status") not in ("active", "expiring", "expired"):
        raise HTTPException(status_code=400, detail="Domain belum bisa diperpanjang (masih pending)")
    if dom.get("pending_renewal"):
        raise HTTPException(status_code=400, detail="Masih ada perpanjangan yang menunggu pembayaran")
    name = dom["domain"]
    price = _tld_price(name) * payload.years
    tax_percent = float(await _get_setting_value(db, "default_tax_percent", 11.0))
    tax = round(price * tax_percent / 100.0, 2)
    due = (datetime.now(timezone.utc) + timedelta(days=7)).date().isoformat()
    inv = {
        "user_id": ObjectId(user["id"]),
        "items": [{"description": f"Perpanjangan domain {name} ({payload.years} tahun)",
                   "qty": 1, "price": price, "total": price}],
        "subtotal": price,
        "tax_percent": tax_percent,
        "tax_amount": tax,
        "total": round(price + tax, 2),
        "due_date": due,
        "status": "unpaid",
        "payment_method": None,
        "paid_at": None,
        "notes": f"Perpanjangan domain {name} - diproses otomatis setelah pembayaran.",
        "domain_renewal": {"domain_id": str(dom["_id"]), "years": payload.years},
        "created_at": _now(),
    }
    inv = await _insert_numbered(db, "invoices", "INV", inv)
    await db.domains.update_one({"_id": dom["_id"]}, {"$set": {
        "pending_renewal": {"years": payload.years, "invoice_id": str(inv["_id"]),
                            "requested_at": _now()}}})
    await log_audit(db, actor=user, action="client_domain.renew_requested", category="domains",
                    target_type="domain", target_id=str(dom["_id"]), target_label=name,
                    metadata={"invoice": inv["number"], "total": inv["total"], "years": payload.years},
                    request=request)
    return {"ok": True, "domain_id": str(dom["_id"]), "invoice_id": str(inv["_id"]),
            "number": inv["number"], "total": inv["total"], "due_date": due}


def _add_years(date_str: str, years: int) -> str:
    base = None
    try:
        base = datetime.strptime((date_str or "")[:10], "%Y-%m-%d")
    except Exception:
        base = datetime.now(timezone.utc)
    if base.replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
        base = datetime.now(timezone.utc)
    try:
        return base.replace(year=base.year + years).date().isoformat()
    except ValueError:
        return base.replace(year=base.year + years, day=28).date().isoformat()


async def _apply_domain_renewal(db, inv: dict) -> bool:
    """Perpanjangan domain otomatis (RNA.id) setelah invoice lunas. Idempotent
    (hanya memproses domain yang pending_renewal-nya menunjuk ke invoice ini)."""
    ren = inv.get("domain_renewal")
    if not ren:
        return False
    try:
        dom = await db.domains.find_one({"_id": _oid(ren["domain_id"]),
                                         "pending_renewal.invoice_id": str(inv["_id"])})
    except Exception:
        return False
    if not dom:
        return False
    years = int(ren.get("years", 1))
    new_expiry = _add_years(dom.get("expires_at") or "", years)
    note = "Integrasi RNA.id belum aktif - perpanjangan dicatat internal, submit manual ke registrar."
    rna = await _rna_client(db)
    if rna and dom.get("order_ref"):
        try:
            res = await rna.renew(dom["order_ref"], years, (dom.get("expires_at") or "")[:10])
            new_expiry = (res.get("expired_at") or "")[:10] or new_expiry
            note = "Renewed live via RNA.id (RDASH)."
        except Exception as e:
            await db.domains.update_one({"_id": dom["_id"]}, {"$set": {
                "provision_note": f"Perpanjangan RNA.id gagal: {str(e)[:150]}. Perlu tindak lanjut manual."}})
            return False
    await db.domains.update_one({"_id": dom["_id"]}, {
        "$set": {"status": "active", "expires_at": new_expiry, "provision_note": note},
        "$unset": {"pending_renewal": ""},
    })
    return True


def _serialize_domain(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "domain": d.get("domain", ""),
        "tld": d.get("tld", ""),
        "status": d.get("status", "pending"),
        "registrar": d.get("registrar", "rna"),
        "years": d.get("years", 1),
        "registered_at": d.get("registered_at"),
        "expires_at": d.get("expires_at"),
        "auto_renew": d.get("auto_renew", True),
        "nameservers": d.get("nameservers", []),
        "price": d.get("price", 0),
        "invoice_id": d.get("invoice_id"),
        "pending_renewal": bool(d.get("pending_renewal")),
        "renewal_invoice_id": (d.get("pending_renewal") or {}).get("invoice_id"),
        "provision_note": d.get("provision_note", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/client/domains")
async def client_domains_list(user=Depends(get_current_user)):
    db = await _get_db()
    docs = await db.domains.find({"user_id": ObjectId(user["id"])}).sort("created_at", -1).to_list(500)
    return [_serialize_domain(d) for d in docs]


@router.get("/client/domains/check")
async def client_domain_check(domain: str, user=Depends(get_current_user)):
    """Cek ketersediaan live: nama polos → semua TLD populer; nama ber-TLD → exact match."""
    db = await _get_db()
    raw = domain.strip().lower()
    base = re.sub(r"[^a-z0-9-]", "", raw.split(".")[0])[:63].strip("-")
    if len(base) < 2:
        raise HTTPException(status_code=400, detail="Nama domain minimal 2 karakter")
    if "." in raw and _DOMAIN_NAME_RE.match(raw):
        names = [raw]
    else:
        names = [f"{base}{tld}" for tld in _TLD_PRICES_IDR]
    results = await _check_domains_availability(db, names)
    return {"live": True, "query": raw, "results": results}


# ---------------- Proxmox live actions ----------------
@router.get("/admin/proxmox/nodes")
async def proxmox_nodes(admin=Depends(get_current_admin)):
    db = await _get_db()
    s = await iv2.get_settings(db, "proxmox")
    if not s or not s.get("enabled"):
        raise HTTPException(status_code=400, detail="Proxmox not configured")
    return await iv2.ProxmoxClient(s).list_nodes()


@router.get("/admin/proxmox/vms")
async def proxmox_vms(node: Optional[str] = None, admin=Depends(get_current_admin)):
    db = await _get_db()
    s = await iv2.get_settings(db, "proxmox")
    if not s or not s.get("enabled"):
        raise HTTPException(status_code=400, detail="Proxmox not configured")
    return await iv2.ProxmoxClient(s).list_vms(node)


@router.post("/admin/proxmox/vms/{node}/{vmid}/{action}")
async def proxmox_vm_action(node: str, vmid: int, action: str, admin=Depends(get_current_admin)):
    if action not in ("start", "stop", "reboot", "shutdown", "suspend", "resume"):
        raise HTTPException(status_code=400, detail="Unsupported action")
    db = await _get_db()
    s = await iv2.get_settings(db, "proxmox")
    if not s or not s.get("enabled"):
        raise HTTPException(status_code=400, detail="Proxmox not configured")
    return await iv2.ProxmoxClient(s).vm_action(node, vmid, action)


@router.get("/admin/proxmox/vnc/{node}/{vmid}")
async def proxmox_vnc(node: str, vmid: int, admin=Depends(get_current_admin)):
    db = await _get_db()
    s = await iv2.get_settings(db, "proxmox")
    if not s or not s.get("enabled"):
        raise HTTPException(status_code=400, detail="Proxmox not configured")
    ticket = await iv2.ProxmoxClient(s).vnc_ticket(node, vmid)
    return {"ticket": ticket, "wss": f"{iv2.ProxmoxClient(s).host}/?console=kvm&novnc=1&vmid={vmid}&node={node}"}


# ---------------- Mikrotik multi-device management ----------------
async def _get_mikrotik_device(db, device_id: str | None):
    """Resolve a Mikrotik device by id; if id is missing return the legacy
    single-device credentials from `integration_settings.mikrotik` as a
    seamless fallback."""
    if device_id:
        try:
            doc = await db.mikrotik_devices.find_one({"_id": _oid(device_id)})
        except Exception:
            doc = None
        if not doc:
            raise HTTPException(status_code=404, detail="Mikrotik device not found")
        return doc
    # Fallback: use the integration_settings.mikrotik credentials
    s = await iv2.get_settings(db, "mikrotik")
    if not s or not s.get("enabled"):
        raise HTTPException(status_code=400, detail="Mikrotik not configured. Add a device in Admin ▸ MikroTik Ops ▸ Devices, or enable the legacy MikroTik integration.")
    return {"_id": None, "name": "default", **(s.get("credentials") or {})}


def _serialize_device(d: dict) -> dict:
    return {
        "id": str(d["_id"]) if d.get("_id") else None,
        "name": d.get("name", ""),
        "host": d.get("host", ""),
        "port": int(d.get("port") or 8728),
        "username": d.get("username", ""),
        "use_tls": bool(d.get("use_tls", False)),
        "site": d.get("site", ""),
        "notes": d.get("notes", ""),
        "created_at": d.get("created_at"),
    }


@router.get("/admin/mikrotik/devices")
async def mikrotik_devices_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.mikrotik_devices.find({}).sort("created_at", 1).to_list(500)
    out = [_serialize_device(d) for d in docs]
    # Attach legacy single-device fallback marker so the UI can show it too
    legacy = await iv2.get_settings(db, "mikrotik")
    if legacy and legacy.get("enabled") and (legacy.get("credentials") or {}).get("host"):
        c = legacy["credentials"]
        out.insert(0, {
            "id": None, "name": "Legacy (Integrations)", "host": c.get("host"),
            "port": int(c.get("port") or 8728), "username": c.get("username"),
            "use_tls": bool(c.get("use_tls", False)), "site": "", "notes": "",
            "legacy": True,
        })
    return out


@router.post("/admin/mikrotik/devices")
async def mikrotik_devices_create(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    doc = {
        "name": (payload.get("name") or "").strip() or "unnamed",
        "host": (payload.get("host") or "").strip(),
        "port": int(payload.get("port") or 8728),
        "username": (payload.get("username") or "").strip(),
        "password": payload.get("password") or "",
        "use_tls": bool(payload.get("use_tls", False)),
        "site": payload.get("site") or "",
        "notes": payload.get("notes") or "",
        "created_at": _now(),
    }
    if not doc["host"] or not doc["username"]:
        raise HTTPException(status_code=400, detail="host and username are required")
    r = await db.mikrotik_devices.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_device(doc)


@router.put("/admin/mikrotik/devices/{did}")
async def mikrotik_devices_update(did: str, payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    allowed = {"name", "host", "port", "username", "password", "use_tls", "site", "notes"}
    upd = {k: v for k, v in (payload or {}).items() if k in allowed}
    if "port" in upd: upd["port"] = int(upd["port"] or 8728)
    if "use_tls" in upd: upd["use_tls"] = bool(upd["use_tls"])
    if "password" in upd and not upd["password"]:
        upd.pop("password")  # don't blank out password when field is empty
    r = await db.mikrotik_devices.update_one({"_id": _oid(did)}, {"$set": upd})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Device not found")
    d = await db.mikrotik_devices.find_one({"_id": _oid(did)})
    return _serialize_device(d)


@router.delete("/admin/mikrotik/devices/{did}")
async def mikrotik_devices_delete(did: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    r = await db.mikrotik_devices.delete_one({"_id": _oid(did)})
    return {"ok": True, "deleted": r.deleted_count}


@router.post("/admin/mikrotik/devices/{did}/test")
async def mikrotik_devices_test(did: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await _get_mikrotik_device(db, did)
    import asyncio as _a
    r = await _a.get_event_loop().run_in_executor(None, lambda: iv2.MikrotikClient(d).test_connection())
    return r


async def _run_mikrotik(db, device_id: str | None, fn_name: str, *args, **kwargs):
    """Helper - resolve the device, dispatch a MikrotikClient method in a
    threadpool (librouteros is sync), return the result."""
    d = await _get_mikrotik_device(db, device_id)
    client = iv2.MikrotikClient(d)
    fn = getattr(client, fn_name)
    import asyncio as _a
    return await _a.get_event_loop().run_in_executor(None, lambda: fn(*args, **kwargs))


# ---------------- Mikrotik live views ----------------
@router.get("/admin/mikrotik/interfaces")
async def mikrotik_interfaces(admin=Depends(get_current_admin), device_id: str | None = None):
    db = await _get_db()
    return await _run_mikrotik(db, device_id, "list_interfaces")


@router.get("/admin/mikrotik/bgp")
async def mikrotik_bgp(admin=Depends(get_current_admin), device_id: str | None = None):
    db = await _get_db()
    return await _run_mikrotik(db, device_id, "list_bgp_peers")


@router.get("/admin/mikrotik/traffic")
async def mikrotik_traffic(interface: str, admin=Depends(get_current_admin), device_id: str | None = None):
    db = await _get_db()
    return await _run_mikrotik(db, device_id, "traffic_monitor", interface)


@router.get("/admin/mikrotik/system")
async def mikrotik_system(admin=Depends(get_current_admin), device_id: str | None = None):
    db = await _get_db()
    return await _run_mikrotik(db, device_id, "system_resource")


# ---------- Looking Glass ----------
@router.post("/admin/mikrotik/looking-glass")
async def mikrotik_looking_glass(payload: dict, admin=Depends(get_current_admin)):
    """Execute a Looking-Glass style lookup from the router:
    payload = {device_id?, tool: 'ping'|'traceroute'|'bgp_route', target: str}"""
    db = await _get_db()
    tool = (payload.get("tool") or "ping").lower()
    if tool not in ("ping", "traceroute", "bgp_route"):
        raise HTTPException(status_code=400, detail="tool must be ping/traceroute/bgp_route")
    target = (payload.get("target") or "").strip()
    if not target:
        raise HTTPException(status_code=400, detail="target is required")
    src_address = (payload.get("src_address") or payload.get("src-address") or "").strip() or None
    return await _run_mikrotik(db, payload.get("device_id"),
                               "looking_glass", tool=tool, target=target,
                               src_address=src_address)


# ---------- Blackhole ----------
@router.get("/admin/mikrotik/blackhole")
async def mikrotik_blackhole_list(admin=Depends(get_current_admin),
                                  device_id: str | None = None,
                                  prefix_filter: str | None = None):
    db = await _get_db()
    return await _run_mikrotik(db, device_id, "blackhole_list", prefix_filter=prefix_filter)


@router.post("/admin/mikrotik/blackhole")
async def mikrotik_blackhole_add(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    prefix = (payload.get("prefix") or "").strip()
    if not prefix:
        raise HTTPException(status_code=400, detail="prefix required")
    dev = await _get_mikrotik_device(db, payload.get("device_id"))
    result = await _run_mikrotik(db, payload.get("device_id"),
                                 "blackhole_add", prefix,
                                 comment=payload.get("comment") or "portal-blackhole")
    await db.blackhole_log.insert_one({
        "prefix": prefix, "action": "add", "by": admin.get("email", ""),
        "source": "manual", "device": dev.get("name", ""),
        "ok": not (isinstance(result, dict) and result.get("error")),
        "detail": str(result)[:300], "at": _now(),
    })
    return result


@router.delete("/admin/mikrotik/blackhole/{route_id}")
async def mikrotik_blackhole_remove(route_id: str, admin=Depends(get_current_admin),
                                    device_id: str | None = None):
    db = await _get_db()
    dev = await _get_mikrotik_device(db, device_id)
    result = await _run_mikrotik(db, device_id, "blackhole_remove", route_id)
    prefix = (result or {}).get("prefix") if isinstance(result, dict) else ""
    await db.blackhole_log.insert_one({
        "prefix": prefix or route_id, "action": "remove", "by": admin.get("email", ""),
        "source": "manual", "device": dev.get("name", ""),
        "ok": not (isinstance(result, dict) and result.get("error")),
        "detail": str(result)[:300], "at": _now(),
    })
    return result


@router.get("/admin/noc/blackhole-log")
async def noc_blackhole_log(q: Optional[str] = None, limit: int = 100,
                            admin=Depends(get_current_admin)):
    """Riwayat announce/remove blackhole (auto + manual) dengan pencarian prefix/aktor."""
    db = await _get_db()
    limit = max(1, min(limit, 500))
    query = {}
    if q:
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        query = {"$or": [{"prefix": rx}, {"by": rx}, {"device": rx}]}
    docs = await db.blackhole_log.find(query).sort("at", -1).to_list(limit)
    return [{
        "id": str(d["_id"]),
        "prefix": d.get("prefix", ""),
        "action": d.get("action", "add"),
        "by": d.get("by", ""),
        "source": d.get("source", "manual"),
        "device": d.get("device", ""),
        "ok": d.get("ok", True),
        "at": d.get("at", ""),
    } for d in docs]


@router.get("/admin/noc/netflow/sankey")
async def noc_netflow_sankey(device_id: Optional[str] = None, limit: int = 12,
                             admin=Depends(get_current_admin)):
    """Data agregat arus trafik (torch MikroTik live) untuk Diagram Sankey:
    flows = [{src, dst, gbps}] top-N berdasarkan rate."""
    db = await _get_db()
    limit = max(3, min(limit, 50))
    devices = ([await _get_mikrotik_device(db, device_id)] if device_id
               else await db.mikrotik_devices.find({}).to_list(50))
    agg: dict = {}
    sampled = 0
    for d in devices:
        if not d:
            continue
        iface = d.get("main_interface") or d.get("interface") or "ether1"
        try:
            import asyncio as _a
            client = iv2.MikrotikClient(d)
            res = await _a.get_event_loop().run_in_executor(
                None, lambda c=client, i=iface: c.torch(interface=i, duration=2))
        except Exception:
            continue
        if not res.get("ok"):
            continue
        sampled += 1
        for f in res.get("rows", []):
            src = (f.get("src_address") or "").split("/")[0]
            dst = (f.get("dst_address") or "").split("/")[0]
            if not src or not dst or src == dst:
                continue
            agg[(src, dst)] = agg.get((src, dst), 0) + f.get("rx_rate", 0) + f.get("tx_rate", 0)
    flows = sorted(({"src": k[0], "dst": k[1], "gbps": round(v / 1e9, 3)}
                    for k, v in agg.items() if v > 0),
                   key=lambda x: x["gbps"], reverse=True)[:limit]
    return {"live": sampled > 0, "devices_sampled": sampled,
            "flows": flows, "sampled_at": _now()}


# ---------- Backup ----------
@router.get("/admin/mikrotik/backups")
async def mikrotik_backups_list(admin=Depends(get_current_admin), device_id: str | None = None):
    db = await _get_db()
    return await _run_mikrotik(db, device_id, "backup_list")


@router.post("/admin/mikrotik/backups")
async def mikrotik_backups_create(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    return await _run_mikrotik(db, payload.get("device_id"),
                               "backup_create", name=payload.get("name"))


@router.delete("/admin/mikrotik/backups/{filename}")
async def mikrotik_backups_delete(filename: str, admin=Depends(get_current_admin),
                                  device_id: str | None = None):
    db = await _get_db()
    return await _run_mikrotik(db, device_id, "backup_delete", filename)


# ---------- Reboot ----------
@router.post("/admin/mikrotik/reboot")
async def mikrotik_reboot(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    confirm = (payload.get("confirm") or "").strip().upper()
    if confirm != "REBOOT":
        raise HTTPException(status_code=400, detail="Confirm by sending {confirm: 'REBOOT'}")
    return await _run_mikrotik(db, payload.get("device_id"), "reboot")


# ---------------- NOC: DDoS threshold rules (CRUD) ----------------
def _serialize_threshold_rule(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "metric": d.get("metric", "pps"),
        "threshold": d.get("threshold", 0),
        "window_s": d.get("window_s", 60),
        "action": d.get("action", "alert"),
        "enabled": d.get("enabled", True),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/noc/threshold-rules")
async def noc_threshold_rules_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.ddos_threshold_rules.find({}).sort("created_at", -1).to_list(200)
    return [_serialize_threshold_rule(d) for d in docs]


@router.post("/admin/noc/threshold-rules")
async def noc_threshold_rules_create(payload: m.ThresholdRuleIn, request: Request,
                                     admin=Depends(get_current_admin)):
    db = await _get_db()
    doc = payload.model_dump()
    doc["created_at"] = _now()
    r = await db.ddos_threshold_rules.insert_one(doc)
    doc["_id"] = r.inserted_id
    await log_audit(db, actor=admin, action="noc.threshold_rule_created", category="noc",
                    target_type="threshold_rule", target_id=str(r.inserted_id),
                    target_label=payload.name, metadata=payload.model_dump(), request=request)
    return _serialize_threshold_rule(doc)


@router.put("/admin/noc/threshold-rules/{rid}")
async def noc_threshold_rules_update(rid: str, payload: m.ThresholdRuleIn, request: Request,
                                     admin=Depends(get_current_admin)):
    db = await _get_db()
    upd = payload.model_dump()
    res = await db.ddos_threshold_rules.update_one({"_id": _oid(rid)}, {"$set": upd})
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Rule not found")
    d = await db.ddos_threshold_rules.find_one({"_id": _oid(rid)})
    await log_audit(db, actor=admin, action="noc.threshold_rule_updated", category="noc",
                    target_type="threshold_rule", target_id=rid,
                    target_label=payload.name, metadata=upd, request=request)
    return _serialize_threshold_rule(d)


@router.delete("/admin/noc/threshold-rules/{rid}")
async def noc_threshold_rules_delete(rid: str, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await db.ddos_threshold_rules.find_one({"_id": _oid(rid)})
    r = await db.ddos_threshold_rules.delete_one({"_id": _oid(rid)})
    if d:
        await log_audit(db, actor=admin, action="noc.threshold_rule_deleted", category="noc",
                        target_type="threshold_rule", target_id=rid,
                        target_label=d.get("name", ""), request=request)
    return {"deleted": r.deleted_count}


@router.post("/admin/noc/ddos/run-detect")
async def noc_ddos_run_detect(admin=Depends(get_current_admin)):
    """Jalankan evaluasi threshold + deteksi insiden DDoS sekarang (manual trigger)."""
    db = await _get_db()
    from portal import emails as _em
    return await _em.run_ddos_detection_sweep(db)


def _serialize_ddos_incident(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "target": d.get("target", ""),
        "attack_type": d.get("attack_type", ""),
        "pps": d.get("pps", 0),
        "bps": d.get("bps", 0),
        "severity": d.get("severity", "medium"),
        "status": d.get("status", "active"),
        "action": d.get("action", "alert"),
        "rule_id": d.get("rule_id"),
        "rule_name": d.get("rule_name", ""),
        "device": d.get("device", ""),
        "started_at": d.get("started_at", ""),
        "ended_at": d.get("ended_at"),
        "notified": d.get("notified", []),
    }


@router.get("/admin/noc/ddos/incidents")
async def noc_ddos_incidents(status: Optional[str] = None, limit: int = 100,
                             admin=Depends(get_current_admin)):
    db = await _get_db()
    q = {"status": status} if status else {}
    limit = max(1, min(limit, 500))
    docs = await db.ddos_incidents.find(q).sort("started_at", -1).to_list(limit)
    return [_serialize_ddos_incident(d) for d in docs]


@router.put("/admin/noc/ddos/incidents/{iid}/status")
async def noc_ddos_incident_status(iid: str, payload: m.DDoSIncidentStatusIn, request: Request,
                                   admin=Depends(get_current_admin)):
    db = await _get_db()
    upd = {"status": payload.status}
    if payload.status in ("resolved", "false_positive"):
        upd["ended_at"] = _now()
    res = await db.ddos_incidents.update_one({"_id": _oid(iid)}, {"$set": upd})
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Incident not found")
    d = await db.ddos_incidents.find_one({"_id": _oid(iid)})
    await log_audit(db, actor=admin, action="noc.ddos_incident_status", category="noc",
                    target_type="ddos_incident", target_id=iid,
                    target_label=d.get("target", ""), metadata=upd, request=request)
    return _serialize_ddos_incident(d)


# ---------------- NOC: saluran notifikasi insiden (CRUD) + log pengiriman ----------------
def _serialize_notif_channel(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "type": d.get("type", "email"),
        "target": d.get("target", ""),
        "events": d.get("events", []),
        "enabled": d.get("enabled", True),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/noc/notif-channels")
async def noc_notif_channels_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.notif_channels.find({}).sort("created_at", -1).to_list(100)
    return [_serialize_notif_channel(d) for d in docs]


@router.post("/admin/noc/notif-channels")
async def noc_notif_channels_create(payload: m.NotifChannelIn, request: Request,
                                    admin=Depends(get_current_admin)):
    db = await _get_db()
    doc = payload.model_dump()
    doc["created_at"] = _now()
    r = await db.notif_channels.insert_one(doc)
    doc["_id"] = r.inserted_id
    await log_audit(db, actor=admin, action="noc.notif_channel_created", category="noc",
                    target_type="notif_channel", target_id=str(r.inserted_id),
                    target_label=f"{payload.type}:{payload.target}", request=request)
    return _serialize_notif_channel(doc)


@router.put("/admin/noc/notif-channels/{cid}")
async def noc_notif_channels_update(cid: str, payload: m.NotifChannelIn, request: Request,
                                    admin=Depends(get_current_admin)):
    db = await _get_db()
    res = await db.notif_channels.update_one({"_id": _oid(cid)}, {"$set": payload.model_dump()})
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Channel not found")
    d = await db.notif_channels.find_one({"_id": _oid(cid)})
    return _serialize_notif_channel(d)


@router.delete("/admin/noc/notif-channels/{cid}")
async def noc_notif_channels_delete(cid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    r = await db.notif_channels.delete_one({"_id": _oid(cid)})
    return {"deleted": r.deleted_count}


@router.get("/admin/noc/ddos/notify-log")
async def noc_ddos_notify_log(limit: int = 100, admin=Depends(get_current_admin)):
    db = await _get_db()
    limit = max(1, min(limit, 500))
    docs = await db.ddos_notify_log.find({}).sort("at", -1).to_list(limit)
    return [{
        "id": str(d["_id"]),
        "incident_id": d.get("incident_id", ""),
        "target": d.get("target", ""),
        "channel_type": d.get("channel_type", ""),
        "channel_target": d.get("channel_target", ""),
        "status": d.get("status", ""),
        "at": d.get("at", ""),
    } for d in docs]


# ---------------- Payment gateway - create + webhook ----------------
async def _payment_settings(db, provider: str) -> Optional[dict]:
    """Resolve gateway settings from either storage system:
    1. `integration_settings` (iv2, provider-keyed) - preferred;
    2. the module-hub `integrations` row (admin UI "Add Server" dialog),
       mapped into the iv2 shape. Duitku is the only mapped provider.
    Returns an iv2-shaped dict or None when not configured/enabled."""
    s = await iv2.get_settings(db, provider)
    if s and s.get("enabled") and (s.get("credentials") or {}):
        return s
    row = await db.integrations.find_one({"module": provider, "status": "enabled"})
    if row and provider == "duitku":
        cfg = row.get("config") or {}
        if cfg.get("merchant_code") and cfg.get("api_key"):
            return {
                "provider": "duitku",
                "enabled": True,
                "sandbox": (cfg.get("environment") or "sandbox") != "production",
                "credentials": {"merchant_code": cfg["merchant_code"],
                                 "api_key": cfg["api_key"]},
                "options": {"callback_url": cfg.get("callback_url") or "",
                            "return_url": cfg.get("return_url") or ""},
            }
    return None


@router.post("/client/invoices/{iid}/pay-online")
async def client_pay_online(iid: str, request: Request, provider: str = "duitku",
                            user=Depends(get_current_user)):
    """Create a hosted payment link for the given invoice (Duitku-only policy)."""
    db = await _get_db()
    if provider not in iv2.PAYMENT_PROVIDERS:
        raise HTTPException(status_code=400, detail="Unknown payment provider")
    if provider in _EXTRA_PAYMENT_MODULES and not bool(
            await _get_setting_value(db, "enable_extra_payment_gateways", False)):
        raise HTTPException(status_code=400,
                            detail="Hanya Duitku yang tersedia sebagai payment gateway.")
    inv = await db.invoices.find_one({"_id": _oid(iid), "user_id": ObjectId(user["id"])})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.get("status") == "paid":
        raise HTTPException(status_code=400, detail="Invoice already paid")
    s = await _payment_settings(db, provider)
    if not s:
        raise HTTPException(status_code=400, detail=f"{provider} not configured")
    gw = iv2.payment_gateway(provider, s)
    # Public base URL resolution: env → request headers (behind ingress).
    base = (os.environ.get("REACT_APP_BACKEND_URL") or "").strip().rstrip("/")
    if not base:
        fwd_host = request.headers.get("x-forwarded-host") or request.headers.get("host") or ""
        fwd_proto = request.headers.get("x-forwarded-proto") or "https"
        base = f"{fwd_proto}://{fwd_host}" if fwd_host else ""
    opts = s.get("options") or {}
    # Callback: explicit integration config wins, else derived from base.
    callback = (opts.get("callback_url") or "").strip()
    if not callback:
        if not base:
            raise HTTPException(status_code=500,
                                detail="Cannot determine public callback URL - set it in the Duitku integration config.")
        callback = f"{base}/api/portal/webhooks/{provider}"
    kwargs = dict(
        invoice_id=inv["number"] or str(inv["_id"]),
        amount_idr=int(inv["total"]),
        customer_email=user["email"],
        callback_url=callback,
    )
    if provider == "duitku":
        # returnUrl is REQUIRED by the POP docs - send the client back to their
        # invoices page after payment (config override supported).
        return_url = (opts.get("return_url") or "").strip()
        if not return_url and base:
            return_url = f"{base}/portal/client/invoices"
        kwargs.update(return_url=return_url,
                      customer_name=user.get("name") or "",
                      expiry_minutes=int(opts.get("expiry_minutes") or 1440))
    try:
        result = await gw.create_payment(**kwargs)
    except Exception as e:
        raise HTTPException(status_code=502,
                            detail=f"Gagal membuat transaksi {provider}: {type(e).__name__}: {e}")
    await db.invoices.update_one(
        {"_id": inv["_id"]},
        {"$set": {"payment_provider": provider, "payment_external_id": result.get("external_id"),
                  "payment_link": result.get("payment_url")}},
    )
    return result


@router.post("/webhooks/{provider}")
async def payment_webhook(provider: str, request: Request):
    """Public webhook. Verifies the gateway signature before touching any invoice.

    Idempotent: the paid-transition only happens once (`status != paid` filter),
    so duplicate callback deliveries never double-fire emails, provisioning or
    service reactivation."""
    db = await _get_db()
    s = await _payment_settings(db, provider)
    if not s:
        raise HTTPException(status_code=404, detail="Unknown gateway")
    raw = await request.body()
    gw = iv2.payment_gateway(provider, s)
    try:
        if provider == "xendit":
            verified = gw.verify_webhook({k.lower(): v for k, v in request.headers.items()}, raw)
        else:
            verified = gw.verify_webhook(raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid webhook: {e}")

    if verified["status"] != "paid":
        return {"received": True, "status": verified["status"]}

    inv_no = verified["invoice_id"]
    # ---- Idempotent paid transition (survives duplicate callback delivery) ----
    r = await db.invoices.update_one(
        {"number": inv_no, "status": {"$ne": "paid"}},
        {"$set": {"status": "paid", "paid_at": _now(), "payment_method": provider,
                  "payment_ref": verified.get("external_id")}},
    )
    if not r.modified_count:
        exists = await db.invoices.find_one({"number": inv_no})
        return {"received": True, "status": "paid",
                "duplicate": bool(exists),
                "note": "already processed" if exists else "invoice not found"}

    inv = await db.invoices.find_one({"number": inv_no})
    user = await db.users.find_one({"_id": inv["user_id"]}) if inv else None

    # 1) Payment-received notification email (best-effort, never blocks the ACK)
    if inv and user:
        try:
            from portal import emails as _em
            await _em.on_invoice_paid(db, inv, user)
        except Exception:
            pass

    # 2) Auto-provision the linked order (same hook as manual admin mark-paid)
    if inv and inv.get("order_id"):
        order = await db.orders.find_one({"_id": _oid(inv["order_id"])})
        if order and not order.get("service_id"):
            try:
                await _auto_provision(db, order)
            except Exception:
                pass

    # 2b) Eksekusi upgrade resource yang menunggu pembayaran invoice ini
    if inv:
        try:
            await _apply_pending_upgrade(db, inv)
        except Exception:
            pass

    # 2c) Registrasi domain otomatis (RNA.id) untuk invoice registrasi domain
    if inv:
        try:
            await _auto_register_domain(db, inv)
            await _apply_domain_renewal(db, inv)
        except Exception:
            pass

    # 3) Reactivate services suspended for non-payment of THIS invoice
    reactivated = 0
    if inv:
        import re as _re
        ors = [{"user_id": inv["user_id"],
                "suspended_reason": {"$regex": f"invoice {_re.escape(inv_no)} overdue",
                                      "$options": "i"}}]
        if inv.get("service_id"):
            try:
                ors.append({"_id": _oid(inv["service_id"])})
            except Exception:
                pass
        res = await db.services.update_many(
            {"status": "suspended", "$or": ors},
            {"$set": {"status": "active", "reactivated_at": _now(),
                      "reactivated_reason": f"invoice {inv_no} paid via {provider}"},
             "$unset": {"suspended_at": "", "suspended_reason": ""}},
        )
        reactivated = res.modified_count

    return {"received": True, "status": "paid", "reactivated_services": reactivated}


# ============================================================
# SECURITY - Login Attempt Analytics
# ============================================================
@router.get("/admin/security/login-analytics")
async def login_analytics(
    admin=Depends(get_current_admin),
    window: str = "24h",  # 24h | 7d | 30d
    limit: int = 100,
):
    """Aggregate login attempts for the Admin Security dashboard.
    Powered by the `login_attempts` collection populated by `/auth/login`."""
    db = await _get_db()
    now = datetime.now(timezone.utc)
    windows = {"24h": timedelta(hours=24), "7d": timedelta(days=7), "30d": timedelta(days=30)}
    delta = windows.get(window, timedelta(hours=24))
    since = (now - delta).isoformat()

    cursor = db.login_attempts.find({"created_at": {"$gte": since}}).sort("created_at", -1)
    rows = await cursor.to_list(20000)

    total = len(rows)
    successes = sum(1 for r in rows if r.get("success"))
    failures = total - successes
    success_rate = round((successes / total) * 100, 2) if total else 0.0
    recap_blocks = sum(1 for r in rows if r.get("reason", "").startswith("recaptcha"))

    # Reason breakdown
    reason_counts: dict[str, int] = {}
    for r in rows:
        k = r.get("reason", "unknown")
        reason_counts[k] = reason_counts.get(k, 0) + 1

    # Top offending IPs (failures only)
    ip_counts: dict[str, int] = {}
    for r in rows:
        if not r.get("success"):
            ip = r.get("ip", "unknown")
            ip_counts[ip] = ip_counts.get(ip, 0) + 1
    top_ips = sorted(ip_counts.items(), key=lambda kv: kv[1], reverse=True)[:10]

    # Top targeted emails (failures only)
    email_counts: dict[str, int] = {}
    for r in rows:
        if not r.get("success"):
            em = r.get("email") or "(empty)"
            email_counts[em] = email_counts.get(em, 0) + 1
    top_emails = sorted(email_counts.items(), key=lambda kv: kv[1], reverse=True)[:10]

    # Time series buckets
    if window == "24h":
        # hourly buckets
        buckets: dict[str, dict] = {}
        for h in range(24, -1, -1):
            t = now - timedelta(hours=h)
            key = t.strftime("%Y-%m-%d %H:00")
            buckets[key] = {"bucket": key, "success": 0, "failed": 0, "recaptcha_block": 0}
        for r in rows:
            ts = r.get("created_at", "")
            key = ts[:13] + ":00"
            if key in buckets:
                if r.get("success"):
                    buckets[key]["success"] += 1
                else:
                    buckets[key]["failed"] += 1
                    if r.get("reason", "").startswith("recaptcha"):
                        buckets[key]["recaptcha_block"] += 1
        series = list(buckets.values())
    else:
        days = 7 if window == "7d" else 30
        buckets = {}
        for d in range(days, -1, -1):
            t = now - timedelta(days=d)
            key = t.strftime("%Y-%m-%d")
            buckets[key] = {"bucket": key, "success": 0, "failed": 0, "recaptcha_block": 0}
        for r in rows:
            key = (r.get("created_at", "") or "")[:10]
            if key in buckets:
                if r.get("success"):
                    buckets[key]["success"] += 1
                else:
                    buckets[key]["failed"] += 1
                    if r.get("reason", "").startswith("recaptcha"):
                        buckets[key]["recaptcha_block"] += 1
        series = list(buckets.values())

    # reCAPTCHA score distribution (buckets of 0.1)
    score_buckets = [{"bucket": f"{b/10:.1f}", "count": 0} for b in range(0, 11)]
    for r in rows:
        s = r.get("recaptcha_score")
        if s is None:
            continue
        idx = min(int(float(s) * 10), 10)
        score_buckets[idx]["count"] += 1
    scored_rows = sum(sb["count"] for sb in score_buckets)

    # Recent attempts
    recent = [{
        "id": str(r.get("_id", "")),
        "email": r.get("email", ""),
        "action": r.get("action", ""),
        "success": bool(r.get("success")),
        "reason": r.get("reason", ""),
        "ip": r.get("ip", ""),
        "user_agent": (r.get("user_agent") or "")[:120],
        "recaptcha_enabled": bool(r.get("recaptcha_enabled")),
        "recaptcha_score": r.get("recaptcha_score"),
        "created_at": r.get("created_at", ""),
    } for r in rows[:max(1, min(limit, 500))]]

    return {
        "window": window,
        "since": since,
        "totals": {
            "attempts": total,
            "successes": successes,
            "failures": failures,
            "success_rate": success_rate,
            "recaptcha_blocks": recap_blocks,
        },
        "reason_breakdown": [{"reason": k, "count": v} for k, v in
                             sorted(reason_counts.items(), key=lambda kv: kv[1], reverse=True)],
        "top_ips": [{"ip": k, "count": v} for k, v in top_ips],
        "top_emails": [{"email": k, "count": v} for k, v in top_emails],
        "series": series,
        "score_distribution": {"buckets": score_buckets, "total_scored": scored_rows},
        "recent": recent,
    }


# ---------- Security Settings & Blocked IPs ----------
@router.get("/admin/security/settings")
async def security_settings_get(admin=Depends(get_current_admin)):
    db = await _get_db()
    s = await _get_security_settings(db)
    return s


@router.put("/admin/security/settings")
async def security_settings_put(payload: dict, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    before = await _get_security_settings(db)
    allowed = {"auto_block_enabled", "fail_threshold", "window_minutes", "ban_minutes",
               "notify_emails", "whitelist_ips",
               "email_notify_enabled", "telegram_notify_enabled"}
    upd = {k: v for k, v in (payload or {}).items() if k in allowed}
    # type coercion
    if "fail_threshold" in upd: upd["fail_threshold"] = max(1, int(upd["fail_threshold"]))
    if "window_minutes" in upd: upd["window_minutes"] = max(1, int(upd["window_minutes"]))
    if "ban_minutes" in upd:    upd["ban_minutes"]    = max(1, int(upd["ban_minutes"]))
    for boolkey in ("auto_block_enabled", "email_notify_enabled", "telegram_notify_enabled"):
        if boolkey in upd: upd[boolkey] = bool(upd[boolkey])
    if "notify_emails" in upd:
        v = upd["notify_emails"]
        if isinstance(v, str):
            v = [x for x in v.replace(",", "\n").splitlines()]
        upd["notify_emails"] = [str(x).strip() for x in (v or []) if str(x).strip()]
    if "whitelist_ips" in upd:
        v = upd["whitelist_ips"]
        if isinstance(v, str):
            v = [x for x in v.replace(",", "\n").splitlines()]
        upd["whitelist_ips"] = [str(x).strip() for x in (v or []) if str(x).strip()]
    await db.settings.update_one({"_id": "security"}, {"$set": upd}, upsert=True)
    after = await _get_security_settings(db)
    await log_audit(db, actor=admin, action="security.settings_update", category="security",
                    target_type="settings", target_label="Security Settings",
                    before=before, after=after, severity="warning", request=request)
    return after


@router.get("/admin/security/blocked-ips")
async def blocked_ips_list(admin=Depends(get_current_admin), active_only: bool = False):
    db = await _get_db()
    now_dt = datetime.now(timezone.utc)
    docs = await db.blocked_ips.find({}).sort("blocked_at", -1).to_list(500)
    out = []
    for d in docs:
        exp = d.get("expires_at")
        if isinstance(exp, str):
            try: exp_dt = datetime.fromisoformat(exp.replace("Z", "+00:00"))
            except Exception: exp_dt = None
        else:
            exp_dt = exp
        # Normalize naive datetimes (MongoDB returns tz-naive UTC)
        if isinstance(exp_dt, datetime) and exp_dt.tzinfo is None:
            exp_dt = exp_dt.replace(tzinfo=timezone.utc)
        is_active = (exp_dt is not None and exp_dt > now_dt and not d.get("unblocked_at"))
        if active_only and not is_active:
            continue
        out.append({
            "ip": d.get("ip"),
            "blocked_at": d.get("blocked_at"),
            "expires_at": (exp_dt.isoformat() if exp_dt else None),
            "reason": d.get("reason"),
            "hits": d.get("hits", 0),
            "unblocked_at": d.get("unblocked_at"),
            "active": bool(is_active),
        })
    return out


@router.delete("/admin/security/blocked-ips/{ip}")
async def blocked_ips_unblock(ip: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    await db.blocked_ips.update_one(
        {"ip": ip},
        {"$set": {"unblocked_at": _now(), "expires_at": _now()}},
    )
    return {"ok": True, "ip": ip}


@router.post("/admin/security/blocked-ips")
async def blocked_ips_add(payload: dict, admin=Depends(get_current_admin)):
    ip = (payload.get("ip") or "").strip()
    ban_minutes = max(1, int(payload.get("ban_minutes") or 30))
    if not ip:
        raise HTTPException(status_code=400, detail="ip required")
    db = await _get_db()
    now_dt = datetime.now(timezone.utc)
    await db.blocked_ips.update_one(
        {"ip": ip},
        {"$set": {
            "ip": ip,
            "blocked_at": now_dt.isoformat(),
            "expires_at": now_dt + timedelta(minutes=ban_minutes),
            "reason": payload.get("reason") or "manual_block",
            "hits": int(payload.get("hits", 0)),
            "unblocked_at": None,
        }},
        upsert=True,
    )
    return {"ok": True, "ip": ip}


@router.get("/admin/security/notifications")
async def security_notifications_list(admin=Depends(get_current_admin), limit: int = 50):
    db = await _get_db()
    limit = max(1, min(limit, 200))
    docs = await db.security_notifications.find({}).sort("created_at", -1).to_list(limit)
    return [{**{k: v for k, v in d.items() if k != "_id"}, "id": str(d["_id"])} for d in docs]


@router.post("/admin/security/notifications/mark-read")
async def security_notifications_mark_read(payload: dict, admin=Depends(get_current_admin)):
    db = await _get_db()
    ids = payload.get("ids") or []
    if ids:
        from bson import ObjectId as _OID
        await db.security_notifications.update_many(
            {"_id": {"$in": [_OID(i) for i in ids]}},
            {"$set": {"read": True}},
        )
    else:
        await db.security_notifications.update_many({}, {"$set": {"read": True}})
    return {"ok": True}


@router.post("/admin/security/notifications/test")
async def security_notifications_test(payload: dict, admin=Depends(get_current_admin)):
    """Fire a sample notification through email + Telegram so the admin can
    verify their SMTP / Telegram integrations from the Security dashboard."""
    db = await _get_db()
    s = await _get_security_settings(db)
    from portal import integrations_v2 as _iv2
    result = {"email": {"attempted": False}, "telegram": {"attempted": False}}

    # Email
    recipients = payload.get("emails") or s.get("notify_emails") or []
    recipients = [r.strip() for r in recipients if r and r.strip()]
    smtp_doc = await _iv2.get_settings(db, "smtp")
    if smtp_doc and smtp_doc.get("enabled") and recipients:
        result["email"]["attempted"] = True
        try:
            mailer = _iv2.SMTPMailer(smtp_doc)
            import asyncio as _asyncio
            loop = _asyncio.get_event_loop()
            errs = []
            for to in recipients:
                try:
                    await loop.run_in_executor(None, lambda t=to: mailer.send(
                        to=t, subject="[Intercloud Security] Test alert",
                        html="<p>This is a <b>test</b> alert from the Intercloud Portal Security dashboard.</p>",
                        text="Test alert from Intercloud Portal Security dashboard."))
                except Exception as e:
                    errs.append(f"{to}: {e}")
            result["email"]["ok"] = not errs
            result["email"]["errors"] = errs
            result["email"]["sent_to"] = recipients
        except Exception as e:
            result["email"]["ok"] = False
            result["email"]["errors"] = [str(e)]
    else:
        result["email"]["reason"] = "SMTP integration not enabled or no recipients"

    # Telegram
    tg_doc = await _iv2.get_telegram_settings(db)
    if tg_doc:
        result["telegram"]["attempted"] = True
        try:
            tg = _iv2.TelegramNotifier(tg_doc)
            r = await tg.send("*🔔 Intercloud test alert*\nThis is a test message from Security dashboard.")
            result["telegram"]["ok"] = bool(r.get("ok"))
            result["telegram"]["details"] = r
        except Exception as e:
            result["telegram"]["ok"] = False
            result["telegram"]["errors"] = [str(e)]
    else:
        result["telegram"]["reason"] = "Telegram integration not enabled"

    return result


# ---------- Real Diagnostic Tools ----------
@router.post("/admin/diagnostics/run")
async def diagnostics_run(payload: dict, admin=Depends(get_current_admin)):
    from portal import diagnostics as _diag
    tool = (payload.get("tool") or "").strip().lower()
    target = (payload.get("target") or "").strip()
    extras: dict = {}
    for key in ("count", "max_hops", "record",
                "interface", "src_address", "dst_address",
                "protocol", "port", "duration", "ip_version"):
        if key in payload and payload[key] not in (None, ""):
            extras[key] = payload[key]
    db = await _get_db()
    try:
        result = await _diag.dispatch(tool, target, db=db, **extras)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Diagnostic failed: {type(e).__name__}: {e}")
    return result


@router.get("/admin/diagnostics/tools")
async def diagnostics_tools_list(admin=Depends(get_current_admin)):
    """Advertise which tools are available on this host so the UI can grey out
    any missing binaries (e.g. traceroute) without a round-trip."""
    from portal import diagnostics as _diag
    from portal import integrations_v2 as _iv2
    db = await _get_db()
    mikrotik_settings = await _iv2.get_settings(db, "mikrotik")
    mikrotik_ready = bool(mikrotik_settings and mikrotik_settings.get("enabled"))
    tools_meta = {
        "ping":       {"label": "Ping",       "requires": "ping3 (python)",       "extras": ["count"]},
        "traceroute": {"label": "Traceroute", "requires": "traceroute",           "extras": ["max_hops"]},
        "dns":        {"label": "DNS Lookup", "requires": "dig",                  "extras": ["record"]},
        "whois":      {"label": "WHOIS",      "requires": "whois",                "extras": []},
        "blacklist":  {"label": "DNSBL",      "requires": "dns",                  "extras": []},
        "portscan":   {"label": "Port Scan",  "requires": "tcp sockets",          "extras": []},
        "http":       {"label": "HTTP Check", "requires": "httpx",                "extras": []},
        "torch":      {"label": "MikroTik Torch",
                       "requires": f"mikrotik integration ({'ready' if mikrotik_ready else 'not configured'})",
                       "available": mikrotik_ready,
                       "extras": ["interface", "src_address", "dst_address", "protocol", "port", "duration"]},
    }
    return {"tools": list(_diag.TOOLS.keys()), "meta": tools_meta,
            "mikrotik_ready": mikrotik_ready}


# ============================================================
# FINANCE V2 - Kas Kecil / Salaries / Sales Fees / Excel reports
# ============================================================
from fastapi.responses import StreamingResponse  # noqa: E402
import io as _io  # noqa: E402


def _generic_ledger_serialize(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "date": d.get("date", ""),
        "amount": float(d.get("amount") or 0),
        "category": d.get("category", ""),
        "notes": d.get("notes", ""),
        "vendor": d.get("vendor", ""),
        "employee": d.get("employee", ""),
        "sales_person": d.get("sales_person", ""),
        "invoice_number": d.get("invoice_number", ""),
        "period_yyyy_mm": d.get("period_yyyy_mm") or (d.get("date", "")[:7]),
        "created_at": _iso(d.get("created_at", "")),
    }


def _month_locked(period_yyyy_mm: str) -> bool:
    """A month is 'locked' once its data is frozen for reporting.

    Rule: a month M/Y is locked as soon as we're in month M+1/Y or later.
    Additionally, once the calendar year Y completes, ALL months of Y stay
    locked until January 5th of Y+1 (audit window). Only strictly future
    months are always mutable.
    """
    today = datetime.now(timezone.utc).date()
    try:
        y, m = int(period_yyyy_mm[:4]), int(period_yyyy_mm[5:7])
    except Exception:
        return False
    # Future months → not locked
    if (y, m) > (today.year, today.month):
        return False
    # Prior calendar year: locked until Jan 5 of Y+1
    if y < today.year:
        if today.year == y + 1 and today.month == 1 and today.day <= 5:
            return False   # 5-day amendment window
        return True
    # Same year, prior month: locked
    if y == today.year and m < today.month:
        return True
    return False


def _mk_ledger_router(*, collection: str, label: str, extra_fields: list):
    """Factory that creates a set of endpoints for a simple ledger table.

    Each ledger has: date (YYYY-MM-DD), amount, notes, plus `extra_fields`.
    We register 3 endpoints per ledger: list / create / delete.
    """

    async def _list(admin=Depends(get_current_admin)):
        db = await _get_db()
        docs = await db[collection].find({}).sort("date", -1).to_list(5000)
        return [_generic_ledger_serialize(d) for d in docs]

    async def _create(payload: dict, admin=Depends(get_current_admin)):
        db = await _get_db()
        date_str = payload.get("date") or datetime.now(timezone.utc).date().isoformat()
        period = date_str[:7]
        if _month_locked(period):
            raise HTTPException(status_code=403, detail=f"Cannot add {label} for locked month {period}. Contact finance to unlock.")
        doc = {"date": date_str, "amount": float(payload.get("amount", 0) or 0),
               "notes": payload.get("notes", ""), "period_yyyy_mm": period,
               "created_at": _now()}
        for k in extra_fields:
            doc[k] = payload.get(k, "")
        r = await db[collection].insert_one(doc)
        doc["_id"] = r.inserted_id
        return _generic_ledger_serialize(doc)

    async def _delete(item_id: str, admin=Depends(get_current_admin)):
        db = await _get_db()
        d = await db[collection].find_one({"_id": _oid(item_id)})
        if not d:
            raise HTTPException(status_code=404, detail="Not found")
        if _month_locked(d.get("period_yyyy_mm") or d.get("date", "")[:7]):
            raise HTTPException(status_code=403, detail=f"Cannot delete {label} from a locked month")
        r = await db[collection].delete_one({"_id": _oid(item_id)})
        return {"deleted": r.deleted_count}

    return _list, _create, _delete


# --- kas kecil (petty cash) ---
_kk_list, _kk_create, _kk_delete = _mk_ledger_router(
    collection="kas_kecil", label="petty cash", extra_fields=["category", "vendor"],
)
router.get("/admin/kas-kecil")(_kk_list)
router.post("/admin/kas-kecil")(_kk_create)
router.delete("/admin/kas-kecil/{item_id}")(_kk_delete)

# --- salaries ---
_sal_list, _sal_create, _sal_delete = _mk_ledger_router(
    collection="salaries", label="salary", extra_fields=["employee", "category"],
)
router.get("/admin/salaries")(_sal_list)
router.post("/admin/salaries")(_sal_create)
router.delete("/admin/salaries/{item_id}")(_sal_delete)

# --- sales fees ---
_sf_list, _sf_create, _sf_delete = _mk_ledger_router(
    collection="sales_fees", label="sales fee", extra_fields=["sales_person", "invoice_number"],
)
router.get("/admin/sales-fees")(_sf_list)
router.post("/admin/sales-fees")(_sf_create)
router.delete("/admin/sales-fees/{item_id}")(_sf_delete)


@router.get("/documents/salary-slip/{sid}")
async def render_salary_slip(sid: str, format: str = "pdf", admin=Depends(get_current_admin)):
    """UAT-034: slip gaji PDF per entri salary (WeasyPrint)."""
    db = await _get_db()
    d = await db.salaries.find_one({"_id": _oid(sid)})
    if not d:
        raise HTTPException(status_code=404, detail="Salary entry not found")
    period = d.get("period_yyyy_mm") or (d.get("date") or "")[:7]
    amount = float(d.get("amount") or 0)
    amount_str = "Rp " + f"{amount:,.0f}".replace(",", ".")
    employee = d.get("employee") or "-"
    category = d.get("category") or "Gaji pokok"
    issued = datetime.now(timezone.utc).date().isoformat()
    html = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 24mm 18mm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; color: #0f172a; font-size: 13px; }}
      .head {{ display: flex; justify-content: space-between; border-bottom: 3px solid #0a2350; padding-bottom: 14px; }}
      .co {{ font-size: 19px; font-weight: 800; color: #0a2350; }}
      .co small {{ display:block; font-size: 10px; color:#64748b; font-weight: 400; margin-top: 3px; }}
      h1 {{ font-size: 15px; letter-spacing: 2px; color: #0a2350; margin: 26px 0 4px; text-transform: uppercase; }}
      .meta {{ color: #64748b; font-size: 11px; margin-bottom: 20px; }}
      table {{ width: 100%; border-collapse: collapse; }}
      td, th {{ padding: 10px 12px; border: 1px solid #e2e8f0; text-align: left; }}
      th {{ background: #f8fafc; font-size: 10px; text-transform: uppercase; letter-spacing: 1px; color: #64748b; width: 34%; }}
      .amt {{ font-size: 17px; font-weight: 800; color: #0a2350; }}
      .foot {{ margin-top: 44px; display: flex; justify-content: space-between; font-size: 11px; color: #64748b; }}
      .sig {{ text-align: center; }}
      .sig .line {{ margin-top: 56px; border-top: 1px solid #94a3b8; padding-top: 5px; width: 190px; }}
      .conf {{ margin-top: 26px; font-size: 9.5px; color: #94a3b8; }}
    </style></head><body>
      <div class="head">
        <div class="co">INTERCLOUD DIGITAL INOVASI<small>Jakarta, Indonesia · support@intercloud-digital.com · +62 878-1239-7187</small></div>
        <div style="text-align:right"><div style="font-size:11px;color:#64748b">SLIP GAJI</div>
          <div style="font-weight:800;color:#0a2350">{period}</div></div>
      </div>
      <h1>Slip Gaji Karyawan</h1>
      <div class="meta">Diterbitkan {issued} · No. ref {str(d["_id"])[-8:].upper()}</div>
      <table>
        <tr><th>Nama karyawan</th><td style="font-weight:700">{employee}</td></tr>
        <tr><th>Periode</th><td>{period}</td></tr>
        <tr><th>Kategori</th><td>{category}</td></tr>
        <tr><th>Tanggal pembayaran</th><td>{(d.get("date") or "")[:10]}</td></tr>
        <tr><th>Jumlah diterima (net)</th><td class="amt">{amount_str}</td></tr>
        <tr><th>Catatan</th><td>{d.get("notes") or "-"}</td></tr>
      </table>
      <div class="foot">
        <div class="sig">Diserahkan oleh,<div class="line">Finance - Intercloud</div></div>
        <div class="sig">Diterima oleh,<div class="line">{employee}</div></div>
      </div>
      <div class="conf">Dokumen ini bersifat rahasia dan dihasilkan otomatis oleh Intercloud Portal.</div>
    </body></html>"""
    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        emp_slug = re.sub(r"[^A-Za-z0-9_-]+", "-", employee).strip("-") or "karyawan"
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="SlipGaji-{emp_slug}-{period}.pdf"'})
    return HTMLResponse(content=html)


@router.get("/documents/sales-fee-slip/{sid}")
async def render_sales_fee_slip(sid: str, format: str = "pdf", admin=Depends(get_current_admin)):
    """Slip fee sales PDF per entri sales_fees (WeasyPrint)."""
    db = await _get_db()
    d = await db.sales_fees.find_one({"_id": _oid(sid)})
    if not d:
        raise HTTPException(status_code=404, detail="Sales fee entry not found")
    period = d.get("period_yyyy_mm") or (d.get("date") or "")[:7]
    amount = float(d.get("amount") or 0)
    amount_str = "Rp " + f"{amount:,.0f}".replace(",", ".")
    person = d.get("sales_person") or "-"
    invoice_no = d.get("invoice_number") or "-"
    issued = datetime.now(timezone.utc).date().isoformat()
    html = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 24mm 18mm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; color: #0f172a; font-size: 13px; }}
      .head {{ display: flex; justify-content: space-between; border-bottom: 3px solid #0a2350; padding-bottom: 14px; }}
      .co {{ font-size: 19px; font-weight: 800; color: #0a2350; }}
      .co small {{ display:block; font-size: 10px; color:#64748b; font-weight: 400; margin-top: 3px; }}
      h1 {{ font-size: 15px; letter-spacing: 2px; color: #0a2350; margin: 26px 0 4px; text-transform: uppercase; }}
      .meta {{ color: #64748b; font-size: 11px; margin-bottom: 20px; }}
      table {{ width: 100%; border-collapse: collapse; }}
      td, th {{ padding: 10px 12px; border: 1px solid #e2e8f0; text-align: left; }}
      th {{ background: #f8fafc; font-size: 10px; text-transform: uppercase; letter-spacing: 1px; color: #64748b; width: 34%; }}
      .amt {{ font-size: 17px; font-weight: 800; color: #0a2350; }}
      .foot {{ margin-top: 44px; display: flex; justify-content: space-between; font-size: 11px; color: #64748b; }}
      .sig {{ text-align: center; }}
      .sig .line {{ margin-top: 56px; border-top: 1px solid #94a3b8; padding-top: 5px; width: 190px; }}
      .conf {{ margin-top: 26px; font-size: 9.5px; color: #94a3b8; }}
    </style></head><body>
      <div class="head">
        <div class="co">INTERCLOUD DIGITAL INOVASI<small>Jakarta, Indonesia · support@intercloud-digital.com · +62 878-1239-7187</small></div>
        <div style="text-align:right"><div style="font-size:11px;color:#64748b">SLIP FEE SALES</div>
          <div style="font-weight:800;color:#0a2350">{period}</div></div>
      </div>
      <h1>Slip Fee Penjualan</h1>
      <div class="meta">Diterbitkan {issued} · No. ref {str(d["_id"])[-8:].upper()}</div>
      <table>
        <tr><th>Nama sales</th><td style="font-weight:700">{person}</td></tr>
        <tr><th>Periode</th><td>{period}</td></tr>
        <tr><th>Invoice terkait</th><td>{invoice_no}</td></tr>
        <tr><th>Tanggal pembayaran</th><td>{(d.get("date") or "")[:10]}</td></tr>
        <tr><th>Jumlah fee (net)</th><td class="amt">{amount_str}</td></tr>
        <tr><th>Catatan</th><td>{d.get("notes") or "-"}</td></tr>
      </table>
      <div class="foot">
        <div class="sig">Diserahkan oleh,<div class="line">Finance - Intercloud</div></div>
        <div class="sig">Diterima oleh,<div class="line">{person}</div></div>
      </div>
      <div class="conf">Dokumen ini bersifat rahasia dan dihasilkan otomatis oleh Intercloud Portal.</div>
    </body></html>"""
    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        slug = re.sub(r"[^A-Za-z0-9_-]+", "-", person).strip("-") or "sales"
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="SlipFeeSales-{slug}-{period}.pdf"'})
    return HTMLResponse(content=html)


# ---------------- Finance detailed report ----------------
@router.get("/admin/finance/detailed")
async def finance_detailed(admin=Depends(get_current_admin)):
    """Returns paid-invoice detail + all four expense ledgers + assets + depreciation.

    The frontend Finance page uses this to render tabbed detailed tables.
    """
    db = await _get_db()
    paid = await db.invoices.find({"status": "paid"}).sort("paid_at", -1).to_list(5000)
    revenue_rows = [{
        "id": str(inv["_id"]),
        "number": inv.get("number", ""),
        "paid_at": inv.get("paid_at", "")[:10],
        "customer": inv.get("customer_name") or "",
        "total": float(inv.get("total") or 0),
        "period_yyyy_mm": (inv.get("paid_at") or inv.get("created_at", ""))[:7],
    } for inv in paid]

    async def _fetch(coll):
        docs = await db[coll].find({}).sort("date", -1).to_list(5000)
        return [_generic_ledger_serialize(d) for d in docs]

    expenses_rows = []
    async for d in db.expenses.find({}).sort("date", -1):
        expenses_rows.append({
            "id": str(d["_id"]),
            "date": d.get("date", ""),
            "category": d.get("category", ""),
            "vendor": d.get("vendor", ""),
            "amount": float(d.get("amount", 0)),
            "description": d.get("description", ""),
            "period_yyyy_mm": (d.get("date") or "")[:7],
        })
    kk_rows = await _fetch("kas_kecil")
    sal_rows = await _fetch("salaries")
    sf_rows = await _fetch("sales_fees")

    assets_rows = []
    async for a in db.assets.find({}):
        dep = _asset_depreciation(a)
        assets_rows.append({
            "id": str(a["_id"]),
            "name": a.get("name", ""),
            "category": a.get("category", ""),
            "purchase_date": a.get("purchase_date", ""),
            "value": float(a.get("value", 0)),
            "salvage_value": float(a.get("salvage_value", 0) or 0),
            "useful_life_years": dep["life_years"],
            "useful_life_months": int(a.get("useful_life_months", 0) or 0),
            "annual_depreciation": dep["annual_depreciation"],
            "monthly_depreciation": dep["monthly_depreciation"],
            "book_value": dep["book_value"],
            "accumulated_depreciation": dep["accumulated_depreciation"],
        })

    total_revenue = sum(r["total"] for r in revenue_rows)
    total_expenses = sum(e["amount"] for e in expenses_rows)
    total_kas_kecil = sum(r["amount"] for r in kk_rows)
    total_salaries = sum(r["amount"] for r in sal_rows)
    total_sales_fees = sum(r["amount"] for r in sf_rows)
    total_all_expenses = total_expenses + total_kas_kecil + total_salaries + total_sales_fees
    total_depreciation = sum(a["accumulated_depreciation"] for a in assets_rows)
    return {
        "revenue_rows": revenue_rows,
        "expenses_rows": expenses_rows,
        "kas_kecil_rows": kk_rows,
        "salaries_rows": sal_rows,
        "sales_fees_rows": sf_rows,
        "assets_rows": assets_rows,
        "totals": {
            "revenue": total_revenue,
            "expenses_recurring": total_expenses,
            "kas_kecil": total_kas_kecil,
            "salaries": total_salaries,
            "sales_fees": total_sales_fees,
            "expenses_all": total_all_expenses,
            "depreciation_accumulated": total_depreciation,
            "net_profit": total_revenue - total_all_expenses - total_depreciation,
        },
    }


# ---------------- Excel report generation ----------------
def _idr_fmt(v):
    return f"Rp {float(v or 0):,.0f}".replace(",", ".")


def _write_xlsx(sheets: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF0A2350")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="FFFEF3C7")
    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, cell in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=cell)
                if isinstance(cell, (int, float)) or (isinstance(cell, str) and cell.startswith("=")):
                    c.number_format = "#,##0"
                if r_idx == 1:
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal="center")
                elif r_idx == len(rows) and str(row[0]).lower().startswith(("total", "net ")):
                    c.font = total_font
                    c.fill = total_fill
        # auto-width
        for col in ws.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, 40)
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _gather_period_data(db, *, year: int, month: Optional[int] = None) -> dict:
    def in_period(dt: str) -> bool:
        if not dt or len(dt) < 7:
            return False
        y = int(dt[:4])
        if y != year:
            return False
        if month is not None:
            return int(dt[5:7]) == month
        return True

    paid = await db.invoices.find({"status": "paid"}).to_list(10000)
    revenue = [i for i in paid if in_period(i.get("paid_at") or i.get("created_at", ""))]
    expenses = [e for e in await db.expenses.find({}).to_list(10000)
                if in_period(e.get("date", ""))]
    kk = [e for e in await db.kas_kecil.find({}).to_list(10000) if in_period(e.get("date", ""))]
    sal = [e for e in await db.salaries.find({}).to_list(10000) if in_period(e.get("date", ""))]
    sf = [e for e in await db.sales_fees.find({}).to_list(10000) if in_period(e.get("date", ""))]
    assets = await db.assets.find({}).to_list(10000)
    return {"revenue": revenue, "expenses": expenses, "kk": kk, "sal": sal, "sf": sf, "assets": assets}


@router.get("/admin/finance/report/monthly/{period}")
async def finance_monthly_xlsx(period: str, admin=Depends(get_current_admin)):
    """`period` is YYYY-MM. Returns an .xlsx with 6 sheets:
    Summary / Revenue / Expenses / Kas Kecil / Salaries / Sales Fees.
    Also freezes the month into `finalized_reports` so it becomes read-only.
    """
    try:
        y, m = int(period[:4]), int(period[5:7])
    except Exception:
        raise HTTPException(status_code=400, detail="Bad period, use YYYY-MM")
    db = await _get_db()
    d = await _gather_period_data(db, year=y, month=m)
    rev_total = sum(float(i.get("total") or 0) for i in d["revenue"])
    exp_total = sum(float(e.get("amount") or 0) for e in d["expenses"])
    kk_total = sum(float(e.get("amount") or 0) for e in d["kk"])
    sal_total = sum(float(e.get("amount") or 0) for e in d["sal"])
    sf_total = sum(float(e.get("amount") or 0) for e in d["sf"])
    all_exp = exp_total + kk_total + sal_total + sf_total
    net_profit = rev_total - all_exp
    summary = [
        ["Line", "Amount (IDR)"],
        ["Revenue (paid invoices)", float(rev_total)],
        ["Expenses (recurring)", float(exp_total)],
        ["Kas Kecil (petty cash)", float(kk_total)],
        ["Salaries", float(sal_total)],
        ["Sales Fees", float(sf_total)],
        ["Total expenses", "=SUM(B3:B6)"],
        ["Net profit (before depreciation)", "=B2-B7"],
    ]
    rev_rows = [["Paid at", "Invoice #", "Customer", "Amount"]] + [
        [i.get("paid_at", "")[:10], i.get("number", ""), i.get("customer_name") or "",
         float(i.get("total") or 0)] for i in d["revenue"]
    ] + [["TOTAL", "", "", f"=SUM(D2:D{len(d['revenue']) + 1})"]]
    exp_rows = [["Date", "Category", "Vendor", "Description", "Amount"]] + [
        [e.get("date", ""), e.get("category", ""), e.get("vendor", ""),
         e.get("description", ""), float(e.get("amount") or 0)] for e in d["expenses"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['expenses']) + 1})"]]
    kk_rows = [["Date", "Category", "Vendor", "Notes", "Amount"]] + [
        [e.get("date", ""), e.get("category", ""), e.get("vendor", ""),
         e.get("notes", ""), float(e.get("amount") or 0)] for e in d["kk"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['kk']) + 1})"]]
    sal_rows = [["Date", "Employee", "Category", "Notes", "Amount"]] + [
        [e.get("date", ""), e.get("employee", ""), e.get("category", ""),
         e.get("notes", ""), float(e.get("amount") or 0)] for e in d["sal"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['sal']) + 1})"]]
    sf_rows = [["Date", "Sales person", "Invoice #", "Notes", "Amount"]] + [
        [e.get("date", ""), e.get("sales_person", ""), e.get("invoice_number", ""),
         e.get("notes", ""), float(e.get("amount") or 0)] for e in d["sf"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['sf']) + 1})"]]

    xlsx = _write_xlsx([
        (f"Summary {period}", summary),
        ("Revenue", rev_rows),
        ("Expenses", exp_rows),
        ("Kas Kecil", kk_rows),
        ("Salaries", sal_rows),
        ("Sales Fees", sf_rows),
    ])

    # Freeze: save into finalized_reports for the month (idempotent)
    await db.finalized_reports.update_one(
        {"period": period, "kind": "monthly"},
        {"$set": {"period": period, "kind": "monthly", "totals": {
            "revenue": rev_total, "expenses_all": all_exp, "net_profit": net_profit,
        }, "generated_at": _now()}},
        upsert=True,
    )
    filename = f"Intercloud_Finance_{period}.xlsx"
    return StreamingResponse(
        _io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/admin/finance/report/annual/{year}")
async def finance_annual_xlsx(year: int, admin=Depends(get_current_admin)):
    """One workbook with per-month AND cumulative Jan-Dec P&L + assets."""
    db = await _get_db()
    d = await _gather_period_data(db, year=year)

    # Per-month buckets
    months = [f"{year}-{m:02d}" for m in range(1, 13)]
    buckets = {mm: {"rev": 0.0, "exp": 0.0, "kk": 0.0, "sal": 0.0, "sf": 0.0} for mm in months}
    for i in d["revenue"]:
        k = (i.get("paid_at") or i.get("created_at", ""))[:7]
        if k in buckets: buckets[k]["rev"] += float(i.get("total") or 0)
    for e in d["expenses"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["exp"] += float(e.get("amount") or 0)
    for e in d["kk"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["kk"] += float(e.get("amount") or 0)
    for e in d["sal"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["sal"] += float(e.get("amount") or 0)
    for e in d["sf"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["sf"] += float(e.get("amount") or 0)

    # Monthly + cumulative sheet
    monthly_rows = [["Month", "Revenue", "Recurring exp.", "Kas Kecil", "Salaries",
                     "Sales Fees", "Total expenses", "Net profit",
                     "Cumulative revenue", "Cumulative net"]]
    cum_rev = cum_net = 0.0
    for idx, mm in enumerate(months):
        b = buckets[mm]
        exps = b["exp"] + b["kk"] + b["sal"] + b["sf"]
        net = b["rev"] - exps
        cum_rev += b["rev"]; cum_net += net
        rn = idx + 2
        monthly_rows.append([mm, float(b["rev"]), float(b["exp"]), float(b["kk"]),
                             float(b["sal"]), float(b["sf"]),
                             f"=SUM(C{rn}:F{rn})", f"=B{rn}-G{rn}",
                             float(cum_rev), float(cum_net)])
    total_rev = sum(b["rev"] for b in buckets.values())
    total_exp_all = sum(b["exp"] + b["kk"] + b["sal"] + b["sf"] for b in buckets.values())
    monthly_rows.append(["TOTAL", "=SUM(B2:B13)", "", "", "", "",
                         "=SUM(G2:G13)", "=B14-G14",
                         float(total_rev), float(total_rev - total_exp_all)])

    # Assets sheet (straight-line depreciation)
    asset_rows = [["Asset", "Category", "Purchased", "Cost", "Salvage",
                   "Useful life (yr)", "Annual depreciation",
                   "Book value", "Accumulated depreciation"]]
    total_cost = total_book = 0.0
    for a in d["assets"]:
        dep = _asset_depreciation(a)
        cost = float(a.get("value", 0)); book = dep["book_value"]
        total_cost += cost; total_book += book
        asset_rows.append([a.get("name", ""), a.get("category", ""),
                           a.get("purchase_date", ""), _idr_fmt(cost),
                           _idr_fmt(float(a.get("salvage_value", 0) or 0)),
                           dep["life_years"],
                           _idr_fmt(dep["annual_depreciation"]),
                           _idr_fmt(book),
                           _idr_fmt(dep["accumulated_depreciation"])])
    asset_rows.append(["TOTAL", "", "", _idr_fmt(total_cost), "", "", "",
                       _idr_fmt(total_book), _idr_fmt(total_cost - total_book)])

    # Details
    rev_rows = [["Paid at", "Invoice #", "Customer", "Amount"]] + [
        [i.get("paid_at", "")[:10], i.get("number", ""), i.get("customer_name") or "",
         _idr_fmt(i.get("total"))] for i in d["revenue"]] + [["TOTAL", "", "", _idr_fmt(total_rev)]]

    xlsx = _write_xlsx([
        (f"P&L {year}", monthly_rows),
        (f"Assets {year}", asset_rows),
        (f"Revenue {year}", rev_rows),
    ])

    await db.finalized_reports.update_one(
        {"period": str(year), "kind": "annual"},
        {"$set": {"period": str(year), "kind": "annual",
                  "totals": {"revenue": total_rev, "expenses_all": total_exp_all,
                             "net_profit": total_rev - total_exp_all},
                  "generated_at": _now()}},
        upsert=True,
    )
    return StreamingResponse(
        _io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Intercloud_Finance_Annual_{year}.xlsx"'},
    )


@router.get("/admin/finance/cashflow-forecast")
async def finance_cashflow_forecast(admin=Depends(get_current_admin)):
    """Proyeksi arus kas 30/60/90 hari: inflow dari invoice unpaid/overdue + renewal
    layanan aktif; outflow dari run-rate 3 bulan terakhir keempat buku beban."""
    db = await _get_db()
    return await _compute_cashflow_forecast(db)


async def _compute_cashflow_forecast(db) -> dict:
    today = datetime.now(timezone.utc).date()
    horizon = today + timedelta(days=91)

    events = []  # (date, amount, kind)
    async for inv in db.invoices.find({"status": {"$in": ["unpaid", "overdue"]}}):
        try:
            due = datetime.strptime((inv.get("due_date") or "")[:10], "%Y-%m-%d").date()
        except Exception:
            continue
        due = max(due, today)
        if due < horizon:
            events.append((due, float(inv.get("total") or 0), "invoice"))
    async for svc in db.services.find({"status": "active"}):
        try:
            due = datetime.strptime((svc.get("next_renewal") or "")[:10], "%Y-%m-%d").date()
        except Exception:
            continue
        if today <= due < horizon:
            events.append((due, float(svc.get("price_monthly") or 0), "renewal"))

    three_months_ago = (today - timedelta(days=90)).isoformat()
    monthly_expense = 0.0
    for coll in ("expenses", "kas_kecil", "salaries", "sales_fees"):
        total = 0.0
        async for e in db[coll].find({"date": {"$gte": three_months_ago}}):
            total += float(e.get("amount") or 0)
        monthly_expense += total / 3.0
    daily_expense = monthly_expense / 30.0

    weekly = []
    cumulative = 0.0
    for w in range(13):
        ws = today + timedelta(days=w * 7)
        we = ws + timedelta(days=7)
        inflow = sum(a for (dt, a, _k) in events if ws <= dt < we)
        outflow = daily_expense * 7
        net = inflow - outflow
        cumulative += net
        weekly.append({"week_start": ws.isoformat(), "inflow": round(inflow, 2),
                       "outflow": round(outflow, 2), "net": round(net, 2),
                       "cumulative": round(cumulative, 2)})

    def _bucket(days: int) -> dict:
        end = today + timedelta(days=days)
        inflow = sum(a for (dt, a, _k) in events if dt < end)
        outflow = daily_expense * days
        return {"inflow": round(inflow, 2), "outflow": round(outflow, 2),
                "net": round(inflow - outflow, 2)}

    return {
        "as_of": today.isoformat(),
        "monthly_expense_run_rate": round(monthly_expense, 2),
        "buckets": {"d30": _bucket(30), "d60": _bucket(60), "d90": _bucket(90)},
        "weekly": weekly,
        "sources": {
            "unpaid_invoices": sum(1 for e in events if e[2] == "invoice"),
            "upcoming_renewals": sum(1 for e in events if e[2] == "renewal"),
        },
    }


def _rp(v) -> str:
    return "Rp " + f"{float(v or 0):,.0f}".replace(",", ".")


@router.get("/admin/finance/cashflow-forecast/export")
async def finance_cashflow_export(format: str = "pdf", token: str = "",
                                  admin=Depends(get_current_admin)):
    """Unduh proyeksi arus kas 30/60/90 hari sebagai PDF atau Excel (xlsx)."""
    db = await _get_db()
    f = await _compute_cashflow_forecast(db)
    b = f["buckets"]
    stamp = f["as_of"]
    fname = f"Proyeksi-Arus-Kas-{stamp}"

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyeksi Arus Kas"
        navy = PatternFill("solid", fgColor="0A2350")
        white_bold = Font(color="FFFFFF", bold=True)
        ws["A1"] = "Intercloud Digital - Proyeksi Arus Kas"
        ws["A1"].font = Font(bold=True, size=14, color="0A2350")
        ws["A2"] = f"Per {stamp} · run-rate beban {_rp(f['monthly_expense_run_rate'])}/bulan"
        ws["A2"].font = Font(italic=True, size=9, color="64748B")
        # Buckets summary
        ws["A4"] = "Ringkasan"
        ws["A4"].font = Font(bold=True, color="0A2350")
        hdr = ["Horizon", "Perkiraan Masuk", "Perkiraan Keluar", "Net"]
        for i, h in enumerate(hdr):
            c = ws.cell(row=5, column=1 + i, value=h)
            c.fill = navy
            c.font = white_bold
            c.alignment = Alignment(horizontal="center")
        for r, (lbl, key) in enumerate([("30 hari", "d30"), ("60 hari", "d60"), ("90 hari", "d90")], start=6):
            ws.cell(row=r, column=1, value=lbl)
            ws.cell(row=r, column=2, value=b[key]["inflow"])
            ws.cell(row=r, column=3, value=b[key]["outflow"])
            ws.cell(row=r, column=4, value=b[key]["net"])
        # Weekly detail
        ws["A10"] = "Rincian Mingguan (90 hari)"
        ws["A10"].font = Font(bold=True, color="0A2350")
        whdr = ["Minggu (mulai)", "Masuk", "Keluar", "Net", "Kumulatif"]
        for i, h in enumerate(whdr):
            c = ws.cell(row=11, column=1 + i, value=h)
            c.fill = navy
            c.font = white_bold
        for r, wk in enumerate(f["weekly"], start=12):
            ws.cell(row=r, column=1, value=wk["week_start"])
            ws.cell(row=r, column=2, value=wk["inflow"])
            ws.cell(row=r, column=3, value=wk["outflow"])
            ws.cell(row=r, column=4, value=wk["net"])
            ws.cell(row=r, column=5, value=wk["cumulative"])
        for col, w in zip("ABCDE", (18, 18, 18, 18, 18)):
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=6, max_row=11 + len(f["weekly"]), min_col=2, max_col=5):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        import io
        buf = io.BytesIO()
        wb.save(buf)
        return Response(content=buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})

    # default: PDF
    def _card(lbl, bk):
        net_color = "#059669" if bk["net"] >= 0 else "#dc2626"
        return (f"<td style='padding:14px;border:1px solid #e2e8f0;border-radius:10px;width:33%'>"
                f"<div style='font-size:10px;text-transform:uppercase;letter-spacing:1px;color:#64748b'>Proyeksi {lbl}</div>"
                f"<div style='font-size:20px;font-weight:800;color:{net_color};margin:4px 0'>{_rp(bk['net'])}</div>"
                f"<div style='font-size:11px;color:#334155'>Masuk {_rp(bk['inflow'])}<br/>Keluar {_rp(bk['outflow'])}</div></td>")
    rows = "".join(
        f"<tr><td style='padding:6px 8px;border-bottom:1px solid #eef2f7'>{w['week_start']}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;color:#059669'>{_rp(w['inflow'])}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;color:#dc2626'>{_rp(w['outflow'])}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;font-weight:700'>{_rp(w['net'])}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;font-weight:700;color:#0a2350'>{_rp(w['cumulative'])}</td></tr>"
        for w in f["weekly"])
    html = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 20mm 16mm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; color:#0f172a; font-size:12px; }}
      h1 {{ color:#0a2350; font-size:18px; margin:0; }}
      .sub {{ color:#64748b; font-size:11px; margin:2px 0 18px; }}
      table {{ width:100%; border-collapse:collapse; }}
      th {{ background:#0a2350; color:#fff; padding:7px 8px; font-size:10px; text-transform:uppercase; letter-spacing:1px; text-align:right; }}
      th:first-child {{ text-align:left; }}
    </style></head><body>
      <h1>Proyeksi Arus Kas - Intercloud Digital</h1>
      <div class="sub">Per {stamp} · run-rate beban {_rp(f['monthly_expense_run_rate'])}/bulan ·
        {f['sources']['unpaid_invoices']} invoice belum lunas · {f['sources']['upcoming_renewals']} perpanjangan akan datang</div>
      <table style="margin-bottom:18px"><tr>{_card('30 hari', b['d30'])}{_card('60 hari', b['d60'])}{_card('90 hari', b['d90'])}</tr></table>
      <div style="font-weight:800;color:#0a2350;margin-bottom:6px">Rincian Mingguan (90 hari ke depan)</div>
      <table><thead><tr><th>Minggu (mulai)</th><th>Masuk</th><th>Keluar</th><th>Net</th><th>Kumulatif</th></tr></thead>
      <tbody>{rows}</tbody></table>
      <div style="margin-top:22px;font-size:9px;color:#94a3b8">Estimasi otomatis dari data invoice, jadwal perpanjangan layanan, dan rata-rata beban 3 bulan terakhir. Bukan angka final akuntansi.</div>
    </body></html>"""
    return Response(content=_render_pdf_bytes(html), media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'})


@router.get("/admin/finance/reports")
async def finance_finalized_reports(admin=Depends(get_current_admin)):
    """List all previously-generated monthly/annual reports (audit trail)."""
    db = await _get_db()
    docs = await db.finalized_reports.find({}).sort("period", -1).to_list(500)
    return [{
        "id": str(r["_id"]), "period": r["period"], "kind": r["kind"],
        "totals": r.get("totals", {}), "generated_at": _iso(r.get("generated_at", "")),
        "locked": _month_locked(r["period"]) if r["kind"] == "monthly" else True,
    } for r in docs]


# ============================================================
# Email Automation - templates, preview, logs, blasts
# ============================================================
from portal import emails as _emails  # noqa: E402


def _serialize_template(t: dict) -> dict:
    return {
        "id": str(t["_id"]),
        "event_key": t.get("event_key", ""),
        "name": t.get("name", ""),
        "subject": t.get("subject", ""),
        "body_html": t.get("body_html", ""),
        "offset_days": t.get("offset_days"),
        "send_time": t.get("send_time"),
        "is_active": t.get("is_active", True),
        "notes": t.get("notes", ""),
        "is_system": t.get("is_system", False),
        "last_sent_at": t.get("last_sent_at"),
        "send_count": t.get("send_count", 0),
        "created_at": _iso(t.get("created_at", "")),
        "updated_at": _iso(t.get("updated_at", "")),
    }


@router.get("/admin/email-templates")
async def admin_email_templates_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.email_templates.find({}).sort("event_key", 1).to_list(500)
    return [_serialize_template(d) for d in docs]


@router.get("/admin/email-templates/{tid}")
async def admin_email_template_get(tid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await db.email_templates.find_one({"_id": _oid(tid)})
    if not d:
        raise HTTPException(status_code=404, detail="Template not found")
    return _serialize_template(d)


@router.post("/admin/email-templates")
async def admin_email_template_create(payload: m.EmailTemplateIn, admin=Depends(get_current_admin)):
    db = await _get_db()
    if await db.email_templates.find_one({"event_key": payload.event_key}):
        raise HTTPException(status_code=409, detail="A template with this event_key already exists")
    now = _now()
    doc = {**payload.model_dump(), "is_system": False,
           "created_at": now, "updated_at": now,
           "last_sent_at": None, "send_count": 0}
    r = await db.email_templates.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_template(doc)


@router.put("/admin/email-templates/{tid}")
async def admin_email_template_update(tid: str, payload: m.EmailTemplateIn,
                                      admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await db.email_templates.find_one({"_id": _oid(tid)})
    if not d:
        raise HTTPException(status_code=404, detail="Template not found")
    upd = {**payload.model_dump(), "updated_at": _now()}
    # event_key on system templates is immutable
    if d.get("is_system"):
        upd["event_key"] = d["event_key"]
    await db.email_templates.update_one({"_id": d["_id"]}, {"$set": upd})
    d2 = await db.email_templates.find_one({"_id": d["_id"]})
    return _serialize_template(d2)


@router.delete("/admin/email-templates/{tid}")
async def admin_email_template_delete(tid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    d = await db.email_templates.find_one({"_id": _oid(tid)})
    if not d:
        raise HTTPException(status_code=404, detail="Template not found")
    if d.get("is_system"):
        raise HTTPException(status_code=400,
                            detail="System templates cannot be deleted - pause them via is_active=false instead")
    await db.email_templates.delete_one({"_id": d["_id"]})
    return {"ok": True}


@router.post("/admin/email-templates/preview")
async def admin_email_template_preview(payload: m.EmailPreviewIn,
                                       admin=Depends(get_current_admin)):
    """Render subject + body against a sample user/invoice/order.

    Priority: use raw subject/body_html from payload if provided; else fall
    back to the referenced template. Returns wrapped HTML ready for iframe.
    """
    db = await _get_db()
    subject = payload.subject or ""
    body = payload.body_html or ""
    if not subject and not body and payload.template_id:
        t = await db.email_templates.find_one({"_id": _oid(payload.template_id)})
        if not t:
            raise HTTPException(status_code=404, detail="Template not found")
        subject = t["subject"]
        body = t["body_html"]

    # Build sample context
    user_doc = None
    inv_doc = None
    order_doc = None
    if payload.sample_user_id:
        user_doc = await db.users.find_one({"_id": _oid(payload.sample_user_id)})
    if payload.sample_invoice_id:
        inv_doc = await db.invoices.find_one({"_id": _oid(payload.sample_invoice_id)})
        if inv_doc and not user_doc:
            user_doc = await db.users.find_one({"_id": inv_doc["user_id"]})
    if payload.sample_order_id:
        order_doc = await db.orders.find_one({"_id": _oid(payload.sample_order_id)})
        if order_doc and not user_doc:
            user_doc = await db.users.find_one({"_id": order_doc["user_id"]})
    if not user_doc:
        # Fall back to a demo shape so template preview always renders.
        user_doc = {"name": "Sample Client", "email": "sample@example.com", "company": "Sample Co"}
    if not inv_doc:
        inv_doc = {"number": "INV-2026-00042", "total": 4500000,
                   "due_date": (datetime.now(timezone.utc) + timedelta(days=3)).date().isoformat(),
                   "status": "unpaid"}
    if not order_doc:
        order_doc = {"product_name": "VPS 4 vCPU / 8 GB RAM", "status": "pending_payment"}
    extra = {
        "reset_url": os.environ.get("REACT_APP_BACKEND_URL", "") + "/portal/reset-password?token=SAMPLE",
        "maintenance": {"title": "Emergency network upgrade",
                        "window": "Sabtu, 15 Feb 2026, 02:00-04:00 WIB",
                        "impact": "Kemungkinan latensi meningkat 5-10 menit."},
        "month": {"name": datetime.now(timezone.utc).strftime("%B %Y")},
    }
    ctx = _emails.build_context(user=user_doc, invoice=inv_doc, order=order_doc, extra=extra)
    rendered_subject = _emails.render(subject, ctx)
    rendered_body = _emails.wrap_html(_emails.render(body, ctx))
    return {"subject": rendered_subject, "body_html": rendered_body}


@router.post("/admin/email-templates/send-test")
async def admin_email_template_send_test(payload: m.EmailSendTestIn,
                                         admin=Depends(get_current_admin)):
    db = await _get_db()
    t = await db.email_templates.find_one({"_id": _oid(payload.template_id)})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    # Render with admin's own context so it looks realistic
    admin_doc = await db.users.find_one({"_id": ObjectId(admin["id"])}) or admin
    inv_doc = {"number": "INV-2026-TEST",
               "total": 1500000,
               "due_date": (datetime.now(timezone.utc) + timedelta(days=3)).date().isoformat(),
               "status": "unpaid",
               "_id": ObjectId()}
    ctx = _emails.build_context(user=admin_doc, invoice=inv_doc, extra={
        "reset_url": os.environ.get("REACT_APP_BACKEND_URL", "") + "/portal/reset-password?token=TEST",
        "maintenance": {"title": "Test maintenance", "window": "Test window", "impact": "None."},
        "month": {"name": datetime.now(timezone.utc).strftime("%B %Y")},
    })
    subject = _emails.render(t["subject"], ctx)
    body = _emails.wrap_html(_emails.render(t["body_html"], ctx))
    res = await _emails.deliver(db, to_email=payload.to_email, subject=subject,
                                body_html=body, event_key=f"test:{t['event_key']}",
                                template_id=str(t["_id"]),
                                user_id=str(admin_doc.get("_id") or ""))
    return {"ok": res.get("status") == "sent", **res, "subject": subject}


@router.post("/admin/email/broadcast")
async def admin_email_broadcast(payload: m.EmailNewsletterIn,
                                admin=Depends(get_current_admin)):
    """One-off broadcast - newsletter / maintenance / arbitrary."""
    db = await _get_db()
    recipients: List[dict] = []
    if payload.audience == "all_clients":
        recipients = await db.users.find({"role": "client", "is_active": {"$ne": False}}).to_list(5000)
    elif payload.audience == "all_users":
        recipients = await db.users.find({"is_active": {"$ne": False}}).to_list(5000)
    elif payload.audience == "custom":
        if not payload.to_emails:
            raise HTTPException(status_code=400, detail="Custom audience requires to_emails[]")
        recipients = [{"email": e, "name": e.split("@")[0], "company": ""} for e in payload.to_emails]

    sent = 0
    failed = 0
    skipped = 0
    for u in recipients:
        ctx = _emails.build_context(user=u, extra={
            "month": {"name": datetime.now(timezone.utc).strftime("%B %Y")},
        })
        subject = _emails.render(payload.subject, ctx)
        body = _emails.wrap_html(_emails.render(payload.body_html, ctx))
        res = await _emails.deliver(db, to_email=u["email"], subject=subject,
                                    body_html=body, event_key="broadcast",
                                    user_id=str(u.get("_id") or "") or None)
        s = res.get("status")
        if s == "sent":
            sent += 1
        elif s == "failed":
            failed += 1
        else:
            skipped += 1
    return {"recipients": len(recipients), "sent": sent, "failed": failed, "skipped": skipped}


@router.get("/admin/email-logs")
async def admin_email_logs(limit: int = 200, admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.email_logs.find({}).sort("created_at", -1).to_list(max(1, min(limit, 1000)))
    out = []
    for d in docs:
        out.append({
            "id": str(d["_id"]),
            "event_key": d.get("event_key", ""),
            "template_id": d.get("template_id"),
            "to_email": d.get("to_email", ""),
            "subject": d.get("subject", ""),
            "status": d.get("status", ""),
            "delivered_via": d.get("delivered_via", ""),
            "error": d.get("error"),
            "sent_at": d.get("sent_at"),
            "invoice_id": d.get("invoice_id"),
            "order_id": d.get("order_id"),
            "user_id": d.get("user_id"),
            "created_at": _iso(d.get("created_at", "")),
        })
    return out


@router.post("/admin/email/run-scheduler-now")
async def admin_email_run_scheduler_now(admin=Depends(get_current_admin)):
    """Fire the invoice-reminder sweep on demand (used by admin UI + tests)."""
    db = await _get_db()
    summary = await _emails.run_invoice_reminder_sweep(db)
    return summary


@router.get("/admin/email/event-catalog")
async def admin_email_event_catalog(admin=Depends(get_current_admin)):
    """Return the canonical list of event keys the frontend can reference."""
    return {
        "events": [
            {"key": "welcome", "label": "Welcome (on registration)", "trigger": "instant"},
            {"key": "order_confirmation", "label": "Order confirmation", "trigger": "instant"},
            {"key": "invoice_generated", "label": "Invoice generated (D-14)", "trigger": "instant"},
            {"key": "invoice_reminder_d3", "label": "Payment reminder - D-3", "trigger": "scheduled",
             "offset_days": -3},
            {"key": "invoice_due", "label": "Payment due today", "trigger": "scheduled", "offset_days": 0},
            {"key": "invoice_overdue_d1", "label": "Overdue - D+1", "trigger": "scheduled", "offset_days": 1},
            {"key": "invoice_overdue_d3", "label": "Overdue - D+3", "trigger": "scheduled", "offset_days": 3},
            {"key": "invoice_overdue_d7", "label": "Overdue - D+7 (final)", "trigger": "scheduled", "offset_days": 7},
            {"key": "service_suspension", "label": "Service suspension - D+8",
             "trigger": "scheduled", "offset_days": 8},
            {"key": "password_reset", "label": "Password reset link", "trigger": "instant"},
            {"key": "maintenance", "label": "Maintenance / downtime", "trigger": "on_demand"},
            {"key": "newsletter", "label": "Newsletter (blast)", "trigger": "on_demand"},
        ],
        "variables": [
            "user.name", "user.email", "user.company",
            "invoice.number", "invoice.total_fmt", "invoice.due_date", "invoice.status",
            "order.id_short", "order.product_name", "order.status",
            "portal.login_url", "portal.invoice_url",
            "reset_url", "maintenance.title", "maintenance.window", "maintenance.impact",
            "month.name",
        ],
    }


# ============================================================
# Articles / CMS - admin editor + public listing + search
# ============================================================
import re as _re_slug  # noqa: E402


def _slugify(text: str) -> str:
    s = (text or "").lower().strip()
    s = _re_slug.sub(r"[^a-z0-9]+", "-", s)
    s = _re_slug.sub(r"-+", "-", s).strip("-")
    return s[:80] or "article"


def _norm_tags(tags):
    out = []
    seen = set()
    for t in tags or []:
        s = _slugify(str(t))
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return out


def _serialize_article(d: dict, *, include_body: bool = True) -> dict:
    out = {
        "id": str(d["_id"]),
        "title": d.get("title", ""),
        "slug": d.get("slug", ""),
        "excerpt": d.get("excerpt", ""),
        "cover_image_url": d.get("cover_image_url", ""),
        "cover_image_alt": d.get("cover_image_alt", ""),
        "video_url": d.get("video_url", ""),
        "author_name": d.get("author_name", ""),
        "tags": d.get("tags", []),
        "category": d.get("category", ""),
        "status": d.get("status", "draft"),
        "published_at": d.get("published_at"),
        "meta_title": d.get("meta_title", ""),
        "meta_description": d.get("meta_description", ""),
        "meta_keywords": d.get("meta_keywords", []),
        "og_image_url": d.get("og_image_url", ""),
        "is_featured": bool(d.get("is_featured", False)),
        "view_count": int(d.get("view_count", 0)),
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }
    if include_body:
        out["body_html"] = d.get("body_html", "")
    return out


async def _ensure_article_indexes(db):
    try:
        await db.articles.create_index("slug", unique=True)
    except Exception:
        pass
    try:
        # Text index for search (title, excerpt, body, tags)
        await db.articles.create_index([
            ("title", "text"), ("excerpt", "text"),
            ("body_html", "text"), ("tags", "text"),
        ], default_language="english", name="articles_text_idx")
    except Exception:
        pass


async def _unique_slug(db, base: str, ignore_id: Optional[str] = None) -> str:
    slug = _slugify(base)
    i = 1
    candidate = slug
    while True:
        q = {"slug": candidate}
        if ignore_id:
            q["_id"] = {"$ne": _oid(ignore_id)}
        exists = await db.articles.find_one(q)
        if not exists:
            return candidate
        i += 1
        candidate = f"{slug}-{i}"


# ---- Admin CRUD ----
@router.get("/admin/articles")
async def admin_articles_list(status: str = "", q: str = "", tag: str = "",
                              staff=Depends(get_current_staff)):
    db = await _get_db()
    await _ensure_article_indexes(db)
    filt: dict = {}
    if status in ("draft", "published", "archived"):
        filt["status"] = status
    if tag:
        filt["tags"] = _slugify(tag)
    if q:
        filt["$text"] = {"$search": q}
    docs = await db.articles.find(filt).sort("updated_at", -1).to_list(500)
    return [_serialize_article(d, include_body=False) for d in docs]


@router.get("/admin/articles/{aid}")
async def admin_article_get(aid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    d = await db.articles.find_one({"_id": _oid(aid)})
    if not d:
        raise HTTPException(status_code=404, detail="Article not found")
    return _serialize_article(d)


@router.post("/admin/articles")
async def admin_article_create(payload: m.ArticleIn, admin=Depends(get_current_content)):
    db = await _get_db()
    await _ensure_article_indexes(db)
    now = _now()
    slug_base = payload.slug or payload.title
    slug = await _unique_slug(db, slug_base)
    doc = payload.model_dump()
    doc.update({
        "slug": slug,
        "tags": _norm_tags(payload.tags),
        "meta_keywords": _norm_tags(payload.meta_keywords),
        "author_name": payload.author_name or admin["name"],
        "created_at": now,
        "updated_at": now,
        "view_count": 0,
    })
    if payload.status == "published" and not payload.published_at:
        doc["published_at"] = now
    r = await db.articles.insert_one(doc)
    doc["_id"] = r.inserted_id
    await _sync_article_calendar(db, doc, admin)
    return _serialize_article(doc)


@router.put("/admin/articles/{aid}")
async def admin_article_update(aid: str, payload: m.ArticleIn,
                               admin=Depends(get_current_content)):
    db = await _get_db()
    existing = await db.articles.find_one({"_id": _oid(aid)})
    if not existing:
        raise HTTPException(status_code=404, detail="Article not found")
    upd = payload.model_dump()
    upd["tags"] = _norm_tags(payload.tags)
    upd["meta_keywords"] = _norm_tags(payload.meta_keywords)
    upd["updated_at"] = _now()
    # slug: only regenerate if changed or blank
    incoming_slug = payload.slug or payload.title
    if _slugify(incoming_slug) != existing.get("slug"):
        upd["slug"] = await _unique_slug(db, incoming_slug, ignore_id=aid)
    else:
        upd["slug"] = existing["slug"]
    # First-publish → stamp published_at
    if payload.status == "published" and not existing.get("published_at") and not payload.published_at:
        upd["published_at"] = _now()
    await db.articles.update_one({"_id": _oid(aid)}, {"$set": upd})
    d2 = await db.articles.find_one({"_id": _oid(aid)})
    await _sync_article_calendar(db, d2, admin)
    return _serialize_article(d2)


@router.delete("/admin/articles/{aid}")
async def admin_article_delete(aid: str, admin=Depends(get_current_content)):
    db = await _get_db()
    r = await db.articles.delete_one({"_id": _oid(aid)})
    return {"deleted": r.deleted_count}


@router.get("/admin/articles-tags")
async def admin_articles_tags(staff=Depends(get_current_staff)):
    """Return all tags used across articles with a count (for suggestions)."""
    db = await _get_db()
    pipeline = [
        {"$unwind": "$tags"},
        {"$group": {"_id": "$tags", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    rows = await db.articles.aggregate(pipeline).to_list(500)
    return [{"tag": r["_id"], "count": r["count"]} for r in rows]


# ---- Public endpoints (unauthenticated) ----
@router.get("/public/articles")
async def public_articles_list(q: str = "", tag: str = "",
                               limit: int = 24, skip: int = 0):
    db = await _get_db()
    await _ensure_article_indexes(db)
    filt: dict = {"status": "published"}
    if tag:
        filt["tags"] = _slugify(tag)
    projection = None
    sort = [("published_at", -1)]
    if q:
        filt["$text"] = {"$search": q}
        projection = {"score": {"$meta": "textScore"}}
        sort = [("score", {"$meta": "textScore"}), ("published_at", -1)]
    cursor = db.articles.find(filt, projection).sort(sort).skip(max(0, skip)).limit(max(1, min(limit, 100)))
    docs = await cursor.to_list(200)
    total = await db.articles.count_documents(filt)
    return {
        "total": total,
        "count": len(docs),
        "results": [_serialize_article(d, include_body=False) for d in docs],
    }


@router.get("/public/articles/tags")
async def public_articles_tags():
    """Return every tag that appears on at least one published article."""
    db = await _get_db()
    pipeline = [
        {"$match": {"status": "published"}},
        {"$unwind": "$tags"},
        {"$group": {"_id": "$tags", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    rows = await db.articles.aggregate(pipeline).to_list(500)
    return [{"tag": r["_id"], "count": r["count"]} for r in rows]


@router.get("/public/articles/{slug}")
async def public_article_detail(slug: str):
    db = await _get_db()
    d = await db.articles.find_one({"slug": slug, "status": "published"})
    if not d:
        raise HTTPException(status_code=404, detail="Article not found")
    # Track a view; ignore if it fails.
    try:
        await db.articles.update_one({"_id": d["_id"]}, {"$inc": {"view_count": 1}})
    except Exception:
        pass
    d["view_count"] = int(d.get("view_count", 0)) + 1
    # Sibling: 3 most recent published, excluding this one.
    related_cursor = db.articles.find(
        {"status": "published", "_id": {"$ne": d["_id"]},
         **({"tags": {"$in": d.get("tags", [])}} if d.get("tags") else {})},
    ).sort("published_at", -1).limit(3)
    related = [_serialize_article(x, include_body=False) for x in await related_cursor.to_list(3)]
    return {"article": _serialize_article(d), "related": related}


# ============================================================
# Sitemap - dynamic XML for search engines
# ============================================================
_SITEMAP_STATIC_ROUTES = [
    ("", "1.0", "daily"),                # /
    ("articles", "0.9", "daily"),         # /articles
    ("status", "0.5", "hourly"),          # /status - public uptime page
    ("legal/terms", "0.3", "yearly"),
    ("legal/aup", "0.3", "yearly"),
    ("legal/sla", "0.3", "yearly"),
]

_SITEMAP_ORIGINS = ("https://intercloud-digital.com",)


@router.get("/sitemap.xml", include_in_schema=False)
@_rl_limiter.limit(PUBLIC_STATUS_LIMIT)
async def sitemap_xml(request: Request):
    """Serve a Google-friendly sitemap covering static routes + all
    published articles. Cache-friendly (5-min public cache)."""
    from fastapi.responses import Response as _R
    db = await _get_db()
    origin = _SITEMAP_ORIGINS[0]
    urls: list[str] = []

    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for path, prio, freq in _SITEMAP_STATIC_ROUTES:
        loc = f"{origin}/{path}" if path else f"{origin}/"
        urls.append(
            "  <url>\n"
            f"    <loc>{loc}</loc>\n"
            f"    <lastmod>{now_iso}</lastmod>\n"
            f"    <changefreq>{freq}</changefreq>\n"
            f"    <priority>{prio}</priority>\n"
            "  </url>"
        )

    # Published articles
    try:
        cur = db.articles.find({"status": "published"},
                               {"slug": 1, "updated_at": 1, "published_at": 1}
                               ).sort("published_at", -1).limit(5000)
        async for row in cur:
            slug = row.get("slug")
            if not slug:
                continue
            lm = row.get("updated_at") or row.get("published_at") or ""
            lm = (str(lm)[:10]) or now_iso
            urls.append(
                "  <url>\n"
                f"    <loc>{origin}/articles/{slug}</loc>\n"
                f"    <lastmod>{lm}</lastmod>\n"
                "    <changefreq>weekly</changefreq>\n"
                "    <priority>0.7</priority>\n"
                "  </url>"
            )
    except Exception as e:
        import logging
        logging.getLogger("portal.sitemap").warning(f"sitemap articles fetch failed: {e}")

    body = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        + "\n".join(urls) +
        "\n</urlset>\n"
    )
    return _R(content=body, media_type="application/xml",
              headers={"Cache-Control": "public, max-age=300"})



# ============================================================
# SEO - dynamic rendering for crawlers / link-preview bots
# ============================================================
# Non-JS crawlers (WhatsApp/Telegram/Facebook/Twitter/Slack/Discord link
# unfurlers and some search bots) never execute the SPA, so per-article
# meta tags set client-side are invisible to them. nginx rewrites bot
# requests for /articles/<slug> to this endpoint (see install.sh).
@router.get("/seo/render/articles/{slug}", include_in_schema=False)
@_rl_limiter.limit(PUBLIC_STATUS_LIMIT)
async def seo_render_article(slug: str, request: Request):
    import html as _html
    import json as _json
    db = await _get_db()
    a = await db.articles.find_one({"slug": slug, "status": "published"})
    if not a:
        raise HTTPException(status_code=404, detail="Article not found")
    origin = _SITEMAP_ORIGINS[0]
    title = (a.get("meta_title") or a.get("title") or "").strip()
    desc = (a.get("meta_description") or a.get("excerpt") or "").strip()[:300]
    image = a.get("og_image_url") or a.get("cover_image_url") or f"{origin}/og-image.png"
    canonical = f"{origin}/articles/{slug}"
    ld = _json.dumps({
        "@context": "https://schema.org",
        "@type": "Article",
        "headline": a.get("title") or title,
        "description": desc,
        "image": [image],
        "datePublished": a.get("published_at") or "",
        "dateModified": a.get("updated_at") or "",
        "author": {"@type": "Organization",
                    "name": a.get("author_name") or "PT Intercloud Digital Inovasi"},
        "publisher": {"@type": "Organization",
                       "name": "PT Intercloud Digital Inovasi",
                       "logo": {"@type": "ImageObject", "url": f"{origin}/og-logo.png"}},
        "mainEntityOfPage": canonical,
    }, ensure_ascii=False)
    e = _html.escape
    body_html = f"""<!doctype html>
<html lang="id"><head>
<meta charset="utf-8">
<title>{e(title)} - Intercloud</title>
<meta name="description" content="{e(desc)}">
<link rel="canonical" href="{e(canonical)}">
<meta property="og:type" content="article">
<meta property="og:title" content="{e(title)}">
<meta property="og:description" content="{e(desc)}">
<meta property="og:image" content="{e(image)}">
<meta property="og:url" content="{e(canonical)}">
<meta property="og:site_name" content="PT. Intercloud Digital Inovasi">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="{e(title)}">
<meta name="twitter:description" content="{e(desc)}">
<meta name="twitter:image" content="{e(image)}">
<script type="application/ld+json">{ld}</script>
</head><body>
<h1>{e(a.get('title') or title)}</h1>
<p>{e(desc)}</p>
<a href="{e(canonical)}">Baca artikel lengkap di intercloud-digital.com</a>
</body></html>"""
    return HTMLResponse(content=body_html,
                        headers={"Cache-Control": "public, max-age=300"})


# ============================================================
# AUDIT LOGS - read-only history of sensitive admin actions
# ============================================================
@router.get("/admin/audit-logs")
async def admin_audit_logs_list(
    admin=Depends(get_current_admin),
    limit: int = 200,
    skip: int = 0,
    category: Optional[str] = None,
    action: Optional[str] = None,
    actor_id: Optional[str] = None,
    severity: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Paginated list of audit rows, newest first.

    Filters (all optional):
      • `category` - one of security/billing/integrations/users/system/noc
      • `action`   - exact action key (e.g. "user.role_change")
      • `actor_id` - filter to a specific admin's actions
      • `severity` - info/warning/critical
      • `q`        - case-insensitive substring on actor_email/target_label
      • `date_from`, `date_to` - ISO date strings (inclusive)
    """
    db = await _get_db()
    limit = max(1, min(int(limit or 200), 500))
    skip = max(0, int(skip or 0))
    query: dict = {}
    if category:  query["category"] = category
    if action:    query["action"] = action
    if severity:  query["severity"] = severity
    if actor_id:
        try:
            query["actor_id"] = ObjectId(actor_id)
        except Exception:
            query["actor_id"] = None
    if q:
        needle = q.strip()
        if needle:
            query["$or"] = [
                {"actor_email": {"$regex": needle, "$options": "i"}},
                {"target_label": {"$regex": needle, "$options": "i"}},
                {"action": {"$regex": needle, "$options": "i"}},
            ]
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to + "T23:59:59"
        query["created_at"] = rng
    total = await db.audit_logs.count_documents(query)
    cur = db.audit_logs.find(query).sort("created_at", -1).skip(skip).limit(limit)
    docs = [d async for d in cur]
    return {
        "total": total,
        "limit": limit,
        "skip": skip,
        "items": [_serialize_audit(d) for d in docs],
    }


@router.get("/admin/audit-logs/facets")
async def admin_audit_logs_facets(admin=Depends(get_current_admin)):
    """Distinct values available to power the filter dropdowns."""
    db = await _get_db()
    return {
        "categories": await db.audit_logs.distinct("category"),
        "actions": await db.audit_logs.distinct("action"),
        "severities": ["info", "warning", "critical"],
    }


# ============================================================
# NOC - proactive MikroTik reachability polling
# ============================================================
async def _noc_uptime_window(db, dev_ids, days: int):
    """Uptime % over the last `days` days, combining `noc_daily_uptime`
    rollups (full past days) with raw `noc_probes` for days not yet rolled
    up (today + any recent gap). Survives raw-probe retention deletion."""
    now = datetime.now(timezone.utc)
    day_from = (now - timedelta(days=days)).strftime("%Y-%m-%d")
    today = now.strftime("%Y-%m-%d")
    base: dict = {}
    if dev_ids is not None:
        base["device_id"] = {"$in": dev_ids}
    total = 0
    up = 0.0
    rolled: set = set()
    async for r in db.noc_daily_uptime.find({**base, "date": {"$gte": day_from, "$lt": today}}):
        sc = int(r.get("sample_count") or 0)
        total += sc
        up += sc * float(r.get("uptime_pct") or 0) / 100.0
        rolled.add((r.get("device_id"), r.get("date")))
    async for p in db.noc_probes.find({**base, "at": {"$gte": day_from}},
                                      {"device_id": 1, "at": 1, "ok": 1}):
        if (p.get("device_id"), (p.get("at") or "")[:10]) in rolled:
            continue
        total += 1
        up += 1 if p.get("ok") else 0
    return round(up / total * 100, 2) if total else None


@router.get("/admin/noc/devices")
async def noc_devices_list(admin=Depends(get_current_admin)):
    """Current uptime state for every MikroTik device.

    Aggregates `noc_device_state` (last known status) so the frontend can
    render a device grid without paginating through the full event log."""
    db = await _get_db()
    devices = await db.mikrotik_devices.find({}).to_list(500)
    out = []
    for d in devices:
        state = await db.noc_device_state.find_one({"device_id": d["_id"]}) or {}
        # 24h uptime %: count up-samples / total samples in the last 24h
        since = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        samples = await db.noc_probes.count_documents({
            "device_id": d["_id"], "at": {"$gte": since},
        })
        up_samples = await db.noc_probes.count_documents({
            "device_id": d["_id"], "at": {"$gte": since}, "ok": True,
        })
        uptime_pct = round((up_samples / samples) * 100, 2) if samples else None
        uptime_30d = await _noc_uptime_window(db, [d["_id"]], 30)
        out.append({
            "id": str(d["_id"]),
            "name": d.get("name") or "unnamed",
            "host": d.get("host") or "",
            "site": d.get("site") or "",
            "status": state.get("status") or "unknown",   # up / down / unknown
            "last_probe_at": state.get("last_probe_at"),
            "last_change_at": state.get("last_change_at"),
            "last_message": state.get("last_message") or "",
            "uptime_24h_pct": uptime_pct,
            "uptime_30d_pct": uptime_30d,
            "samples_24h": samples,
        })
    return out


@router.get("/admin/noc/events")
async def noc_events_list(admin=Depends(get_current_admin),
                          limit: int = 200,
                          device_id: Optional[str] = None,
                          type: Optional[str] = None):
    """Chronological device_up / device_down events (newest first)."""
    db = await _get_db()
    limit = max(1, min(int(limit or 200), 500))
    query: dict = {}
    if device_id:
        try:
            query["device_id"] = ObjectId(device_id)
        except Exception:
            query["device_id"] = None
    if type:
        query["type"] = type
    cur = db.noc_events.find(query).sort("at", -1).limit(limit)
    docs = [d async for d in cur]
    return [{
        "id": str(d["_id"]),
        "device_id": str(d.get("device_id")) if d.get("device_id") else None,
        "device_name": d.get("device_name") or "",
        "device_host": d.get("device_host") or "",
        "type": d.get("type") or "",
        "message": d.get("message") or "",
        "at": d.get("at") or "",
        "email_notified": bool(d.get("email_notified")),
    } for d in docs]


@router.post("/admin/noc/run-poll")
async def noc_run_poll_now(admin=Depends(get_current_admin)):
    """Manually trigger a single NOC probe sweep (same code the scheduler runs)."""
    db = await _get_db()
    from portal import emails as _em
    return await _em.run_noc_probe_sweep(db)


@router.post("/admin/noc/run-retention")
async def noc_run_retention_now(admin=Depends(get_current_admin)):
    """Manually trigger the daily probe rollup + retention job."""
    db = await _get_db()
    from portal import emails as _em
    return await _em.run_noc_probe_retention(db)


# ============================================================
# CREDIT NOTES - refunds & adjustments applied to invoices
# ============================================================
def _credit_note_serialize(d: dict, invoice: dict | None = None, user: dict | None = None) -> dict:
    return {
        "id": str(d["_id"]),
        "number": d.get("number", ""),
        "invoice_id": str(d["invoice_id"]) if d.get("invoice_id") else None,
        "invoice_number": (invoice or {}).get("number", d.get("invoice_number", "")),
        "user_id": str(d["user_id"]) if d.get("user_id") else None,
        "user_name": (user or {}).get("name", d.get("user_name", "")),
        "user_email": (user or {}).get("email", d.get("user_email", "")),
        "amount": float(d.get("amount") or 0),
        "reason": d.get("reason") or "",
        "notes": d.get("notes") or "",
        "status": d.get("status") or "draft",   # draft / applied / cancelled
        "applied_at": d.get("applied_at"),
        "applied_by": str(d["applied_by"]) if d.get("applied_by") else None,
        "created_at": d.get("created_at", ""),
    }


async def _sum_applied_credit(db, invoice_id: ObjectId) -> float:
    """Sum of `amount` from applied credit notes for a given invoice."""
    cur = db.credit_notes.find({"invoice_id": invoice_id, "status": "applied"},
                               {"amount": 1})
    total = 0.0
    async for d in cur:
        total += float(d.get("amount") or 0)
    return total


async def _settle_invoice_from_credit(db, invoice: dict, request, admin) -> int:
    """If the total applied credit meets/exceeds the invoice total, flip the
    invoice to paid and reactivate suspended services - same effect as a
    Duitku webhook, but source=credit_note. Returns number of services
    reactivated. Idempotent: filters status != paid."""
    total_credit = await _sum_applied_credit(db, invoice["_id"])
    inv_total = float(invoice.get("total") or 0)
    if total_credit + 0.001 < inv_total:
        return 0
    r = await db.invoices.update_one(
        {"_id": invoice["_id"], "status": {"$ne": "paid"}},
        {"$set": {"status": "paid", "paid_at": _now(),
                  "payment_method": "credit_note",
                  "payment_ref": f"CN>={inv_total:.0f}"}},
    )
    if not r.modified_count:
        return 0
    # Reactivate services suspended for THIS invoice, mirroring the webhook logic
    inv_no = invoice.get("number", "")
    reactivated = 0
    import re as _re
    ors = [{"user_id": invoice["user_id"],
            "suspended_reason": {"$regex": f"invoice {_re.escape(inv_no)} overdue",
                                  "$options": "i"}}]
    if invoice.get("service_id"):
        try:
            ors.append({"_id": _oid(invoice["service_id"])})
        except Exception:
            pass
    res = await db.services.update_many(
        {"status": "suspended", "$or": ors},
        {"$set": {"status": "active", "reactivated_at": _now(),
                  "reactivated_reason": f"invoice {inv_no} settled via credit note"},
         "$unset": {"suspended_at": "", "suspended_reason": ""}},
    )
    reactivated = res.modified_count
    try:
        await _apply_pending_upgrade(db, invoice)
    except Exception:
        pass
    await log_audit(db, actor=admin, action="invoice.settled_by_credit", category="billing",
                    target_type="invoice", target_id=str(invoice["_id"]),
                    target_label=inv_no,
                    metadata={"total_credit": total_credit,
                              "invoice_total": inv_total,
                              "reactivated_services": reactivated},
                    severity="warning", request=request)
    return reactivated


@router.get("/admin/credit-notes")
async def credit_notes_list(admin=Depends(get_current_admin),
                            invoice_id: Optional[str] = None,
                            user_id: Optional[str] = None,
                            status: Optional[str] = None):
    db = await _get_db()
    query: dict = {}
    if invoice_id:
        try: query["invoice_id"] = ObjectId(invoice_id)
        except Exception: query["invoice_id"] = None
    if user_id:
        try: query["user_id"] = ObjectId(user_id)
        except Exception: query["user_id"] = None
    if status:
        query["status"] = status
    docs = await db.credit_notes.find(query).sort("created_at", -1).to_list(500)
    # bulk-fetch related invoices + users
    inv_ids = [d["invoice_id"] for d in docs if d.get("invoice_id")]
    user_ids = [d["user_id"] for d in docs if d.get("user_id")]
    invs = {i["_id"]: i for i in await db.invoices.find({"_id": {"$in": inv_ids}}).to_list(500)}
    users = {u["_id"]: u for u in await db.users.find({"_id": {"$in": user_ids}}).to_list(500)}
    return [_credit_note_serialize(d, invs.get(d.get("invoice_id")), users.get(d.get("user_id")))
            for d in docs]


@router.post("/admin/credit-notes")
async def credit_notes_create(payload: dict, request: Request, admin=Depends(get_current_admin)):
    """Issue a credit note against an invoice.

    Body: `{invoice_id, amount, reason, notes?, auto_apply?: bool}`.
    If `auto_apply` is true, immediately applies (which may settle the invoice)."""
    db = await _get_db()
    inv_id_raw = payload.get("invoice_id") or ""
    try:
        inv_oid = ObjectId(inv_id_raw)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid invoice_id")
    invoice = await db.invoices.find_one({"_id": inv_oid})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be > 0")
    reason = (payload.get("reason") or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Reason is required")
    inv_total = float(invoice.get("total") or 0)
    already_applied = await _sum_applied_credit(db, inv_oid)
    if amount + already_applied > inv_total + 0.001:
        raise HTTPException(status_code=400,
                            detail=f"Credit ({amount:.0f}) + already applied ({already_applied:.0f}) exceeds invoice total ({inv_total:.0f}).")
    user = await db.users.find_one({"_id": invoice["user_id"]}) or {}
    number = await _next_number(db, "credit_notes", "CN")
    doc = {
        "number": number,
        "invoice_id": inv_oid,
        "invoice_number": invoice.get("number", ""),
        "user_id": invoice["user_id"],
        "user_name": user.get("name", ""),
        "user_email": user.get("email", ""),
        "amount": amount,
        "reason": reason,
        "notes": payload.get("notes") or "",
        "status": "draft",
        "created_at": _now(),
        "created_by": ObjectId(admin["id"]),
    }
    r = await db.credit_notes.insert_one(doc)
    doc["_id"] = r.inserted_id
    await log_audit(db, actor=admin, action="credit_note.create", category="billing",
                    target_type="credit_note", target_id=str(r.inserted_id),
                    target_label=number,
                    after={"invoice_number": invoice.get("number"), "amount": amount,
                           "reason": reason},
                    severity="warning", request=request)
    # Auto-apply if requested (single POST for the "quick refund" workflow)
    if bool(payload.get("auto_apply")):
        await _apply_credit_note_inner(db, doc, admin, request)
        doc = await db.credit_notes.find_one({"_id": r.inserted_id})
    return _credit_note_serialize(doc, invoice, user)


async def _apply_credit_note_inner(db, cn: dict, admin, request):
    """Mark the credit note applied, then settle the invoice if fully covered."""
    if cn.get("status") == "applied":
        return {"already_applied": True}
    await db.credit_notes.update_one(
        {"_id": cn["_id"], "status": {"$ne": "applied"}},
        {"$set": {"status": "applied", "applied_at": _now(),
                  "applied_by": ObjectId(admin["id"])}},
    )
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]})
    reactivated = 0
    if invoice:
        reactivated = await _settle_invoice_from_credit(db, invoice, request, admin)
    await log_audit(db, actor=admin, action="credit_note.apply", category="billing",
                    target_type="credit_note", target_id=str(cn["_id"]),
                    target_label=cn.get("number", ""),
                    metadata={"invoice_number": (invoice or {}).get("number"),
                              "amount": cn.get("amount"),
                              "reactivated_services": reactivated},
                    severity="warning", request=request)
    return {"applied": True, "reactivated_services": reactivated}


@router.post("/admin/credit-notes/{cid}/apply")
async def credit_notes_apply(cid: str, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    return await _apply_credit_note_inner(db, cn, admin, request)


@router.post("/admin/credit-notes/{cid}/cancel")
async def credit_notes_cancel(cid: str, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    if cn.get("status") == "applied":
        raise HTTPException(status_code=400,
                            detail="Cannot cancel an already-applied credit note (invoice may be paid).")
    await db.credit_notes.update_one({"_id": cn["_id"]},
                                     {"$set": {"status": "cancelled",
                                               "cancelled_at": _now()}})
    await log_audit(db, actor=admin, action="credit_note.cancel", category="billing",
                    target_type="credit_note", target_id=cid,
                    target_label=cn.get("number", ""),
                    severity="info", request=request)
    return {"ok": True}


@router.get("/admin/credit-notes/{cid}")
async def credit_notes_detail(cid: str, admin=Depends(get_current_admin)):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]}) if cn.get("invoice_id") else None
    user = await db.users.find_one({"_id": cn["user_id"]}) if cn.get("user_id") else None
    return _credit_note_serialize(cn, invoice, user)


# ---- Client-visible credit notes (read-only, own only) ----
@router.get("/client/credit-notes")
async def client_credit_notes(user=Depends(get_current_user)):
    db = await _get_db()
    docs = await db.credit_notes.find({"user_id": ObjectId(user["id"])}).sort("created_at", -1).to_list(200)
    inv_ids = [d["invoice_id"] for d in docs if d.get("invoice_id")]
    invs = {i["_id"]: i for i in await db.invoices.find({"_id": {"$in": inv_ids}}).to_list(500)}
    return [_credit_note_serialize(d, invs.get(d.get("invoice_id"))) for d in docs]


# ---- PDF render (HTML or PDF, mirrors invoice endpoint) ----
@router.get("/documents/credit-note/{cid}")
async def render_credit_note_pdf(cid: str, format: str = "html", user=Depends(get_current_user)):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    # Access: owner or staff
    if user["role"] == "client" and str(cn["user_id"]) != str(user["id"]):
        raise HTTPException(status_code=403, detail="Not your credit note")
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]}) if cn.get("invoice_id") else None
    u = await db.users.find_one({"_id": cn["user_id"]}) or {}
    branding = await _get_branding_dict(db)
    html = _credit_note_html(
        cn=cn,
        invoice=invoice,
        billed_to=u,
        for_pdf=(format == "pdf"),
        logo_url=branding["logo_dark"],
    )
    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        filename = f"CreditNote-{cn.get('number','credit-note')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    token = user.get("_token", "")
    html = html.replace("{TOKEN_PLACEHOLDER}", token)
    return HTMLResponse(content=html)


def _credit_note_html(*, cn: dict, invoice: dict | None, billed_to: dict,
                       for_pdf: bool, logo_url: str) -> str:
    number = cn.get("number", "")
    amount = float(cn.get("amount") or 0)
    status = (cn.get("status") or "draft").lower()
    status_colors = {"draft": "#94a3b8", "applied": "#059669", "cancelled": "#dc2626"}
    status_color = status_colors.get(status, "#64748b")
    inv_no = (invoice or {}).get("number", cn.get("invoice_number", ""))
    inv_total = float((invoice or {}).get("total") or 0)
    generated_on = _long_date(datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    actions_bar = "" if for_pdf else (
        f'<div class="actions">'
        f'<button onclick="window.print()">Print</button>'
        f'<a class="dl" href="?format=pdf&token={{TOKEN_PLACEHOLDER}}">Download PDF</a>'
        f'</div>'
    )
    bill_to = _billing_block(billed_to)
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Credit Note #{number}</title>
<style>
  @page {{ size: A4; margin: 14mm 14mm 16mm 14mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
         color:#334155; margin:0; padding:0; background:#f1f5f9; font-size:12px; line-height:1.5; }}
  .paper {{ background:#fff; padding:34px 40px 30px; max-width:800px; margin:20px auto;
           position:relative; box-shadow:0 6px 30px rgba(2,6,23,.08); }}
  .header {{ display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:22px;
             padding-bottom:18px; border-bottom:1px solid #e2e8f0; }}
  .logo {{ height:64px; width:auto; }}
  h1 {{ margin:0; font-size:26px; letter-spacing:.02em; color:#0a2540; }}
  .sub {{ font-size:11px; color:#64748b; margin-top:4px; }}
  .status {{ display:inline-block; margin-top:8px; padding:6px 14px; border-radius:999px;
             color:#fff; font-weight:700; letter-spacing:.08em; font-size:10px;
             text-transform:uppercase; background:{status_color}; }}
  .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:24px; margin-bottom:22px; }}
  .card {{ background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:14px 16px; }}
  .card h3 {{ margin:0 0 8px 0; font-size:11px; color:#475569; letter-spacing:.14em;
              text-transform:uppercase; font-weight:700; }}
  .amount-box {{ background:#0a2540; color:#fff; padding:18px 22px; border-radius:12px;
                 display:flex; justify-content:space-between; align-items:center; margin-bottom:22px; }}
  .amount-box .lbl {{ font-size:11px; letter-spacing:.14em; text-transform:uppercase;
                       color:#cbd5e1; margin-bottom:4px; }}
  .amount-box .val {{ font-size:26px; font-weight:800; color:#f5b120; }}
  .notes {{ background:#fef3c7; border:1px solid #fbbf24; padding:12px 14px; border-radius:8px;
            margin-top:16px; color:#78350f; font-size:12px; line-height:1.6; }}
  .footer {{ margin-top:28px; padding-top:14px; border-top:1px dashed #cbd5e1;
             font-size:10.5px; color:#64748b; line-height:1.6; }}
  .actions {{ text-align:center; margin:16px 0; }}
  .actions button, .actions .dl {{ display:inline-block; margin:0 6px; padding:8px 18px;
       background:#0a2540; color:#fff; border:0; border-radius:6px; font-size:12px;
       text-decoration:none; cursor:pointer; }}
  .actions .dl {{ background:#f5b120; color:#0a2540; }}
  @media print {{ body{{background:#fff}} .paper{{box-shadow:none;margin:0;max-width:100%}} .actions{{display:none}} }}
</style></head><body>
{actions_bar}
<div class="paper">
  <div class="header">
    <div>
      <img src="{logo_url}" alt="Intercloud" class="logo" onerror="this.style.display='none'"/>
      <div class="sub">PT Intercloud Digital Inovasi<br>Cyber 1 Building, Kuningan · Jakarta 12950</div>
    </div>
    <div style="text-align:right">
      <h1>Credit Note</h1>
      <div class="sub"><b>#{number}</b></div>
      <div class="status">{status}</div>
      <div class="sub" style="margin-top:6px">Issued {_long_date((cn.get('created_at') or '')[:10])}</div>
    </div>
  </div>
  <div class="grid">
    <div class="card">
      <h3>Issued To</h3>
      {bill_to}
    </div>
    <div class="card">
      <h3>Reference Invoice</h3>
      <div style="font-weight:700; color:#0a2540; font-size:14px">#{inv_no or '-'}</div>
      <div class="sub">Invoice total: {_idr(inv_total)}</div>
      {'<div class="sub">Applied on: ' + _long_date((cn.get('applied_at') or '')[:10]) + '</div>' if cn.get('applied_at') else ''}
    </div>
  </div>
  <div class="amount-box">
    <div>
      <div class="lbl">Credit Amount</div>
      <div class="sub" style="color:#cbd5e1; margin-top:2px">{cn.get('reason', '')}</div>
    </div>
    <div class="val">{_idr(amount)}</div>
  </div>
  {'<div class="notes"><b>Notes.</b> ' + (cn.get('notes') or '').replace(chr(10), '<br>') + '</div>' if cn.get('notes') else ''}
  <div class="footer">
    <b>Reason</b><br>{cn.get('reason', '')}
    <br><br>
    Generated on {generated_on}. Credit notes reduce the outstanding amount of the referenced invoice.
    If the total applied credit meets or exceeds the invoice total, the invoice is automatically
    marked as paid and any suspended services are reactivated.
  </div>
</div>
</body></html>"""


def _billing_block(u: dict) -> str:
    """Small helper - same shape used by invoice PDF sidebar."""
    lines = []
    if u.get("attention"): lines.append(f"<b>{u['attention']}</b>")
    elif u.get("name"):    lines.append(f"<b>{u['name']}</b>")
    if u.get("company"):   lines.append(u["company"])
    if u.get("address_line1"): lines.append(u["address_line1"])
    if u.get("address_line2"): lines.append(u["address_line2"])
    city_line = " ".join(x for x in [u.get("city"), u.get("province"), u.get("postal_code")] if x)
    if city_line: lines.append(city_line)
    if u.get("country"): lines.append(u["country"])
    if u.get("email"):   lines.append(f"<span style='color:#64748b'>{u['email']}</span>")
    if u.get("phone"):   lines.append(f"<span style='color:#64748b'>{u['phone']}</span>")
    return "<br>".join(lines) if lines else "<i>No billing address on file</i>"


# ============================================================
# Outstanding-amount helper for InvoiceOut serialization
# ============================================================
async def _invoice_outstanding(db, invoice: dict) -> float:
    """Invoice total minus applied credit notes. Never negative."""
    if invoice.get("status") == "paid":
        return 0.0
    total = float(invoice.get("total") or 0)
    applied = await _sum_applied_credit(db, invoice["_id"])
    return max(0.0, total - applied)



# ============================================================
# EXECUTIVE OVERVIEW - read-only for owner/admin
# ============================================================
async def _get_current_owner(user=Depends(get_current_user)):
    """Access gate: owner or admin only. Owner is READ-ONLY globally."""
    if user.get("role") not in ("admin", "owner"):
        raise HTTPException(status_code=403, detail="Executive access only")
    return user


def _relative_months(n: int):
    """Return a timedelta-like approximation of N months (30d each)."""
    return timedelta(days=30 * n)


@router.get("/admin/owner/overview")
async def owner_overview(owner=Depends(_get_current_owner)):
    """Aggregate MRR/ARPU/churn/uptime/SLA into a single dashboard payload."""
    db = await _get_db()
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # ---- MRR: sum of active-service monthly prices ----
    services = await db.services.find({"status": "active"}).to_list(5000)
    mrr = sum(float(s.get("price_monthly") or 0) for s in services)
    active_services_count = len(services)
    total_clients = await db.users.count_documents({"role": "client"})
    clients_with_active = len({str(s["user_id"]) for s in services if s.get("user_id")})
    arpu = round(mrr / clients_with_active, 2) if clients_with_active else 0.0

    # ---- Churn: services terminated in last 30 days ----
    thirty_ago = (now - timedelta(days=30)).isoformat()
    churned = await db.services.count_documents({
        "status": "terminated", "terminated_at": {"$gte": thirty_ago},
    })
    churn_pct = round((churned / (churned + active_services_count)) * 100, 2) \
                if (churned + active_services_count) else 0.0

    # ---- Revenue: MTD + last 12 months trend ----
    paid_month = await db.invoices.find({"status": "paid",
                                         "paid_at": {"$gte": month_start.isoformat()}}).to_list(5000)
    revenue_month = sum(float(d.get("total") or 0) for d in paid_month)
    trend = []
    for i in range(11, -1, -1):
        start = month_start - _relative_months(i)
        end = start + _relative_months(1)
        docs = await db.invoices.find({
            "status": "paid",
            "paid_at": {"$gte": start.isoformat(), "$lt": end.isoformat()},
        }).to_list(5000)
        trend.append({"period": start.strftime("%Y-%m"),
                      "revenue": sum(float(d.get("total") or 0) for d in docs),
                      "invoices": len(docs)})

    # ---- Outstanding & overdue ----
    overdue_docs = await db.invoices.find({"status": "overdue"}).to_list(2000)
    unpaid_count = await db.invoices.count_documents({"status": {"$in": ["unpaid", "overdue"]}})
    overdue_total = sum(float(d.get("total") or 0) for d in overdue_docs)

    # ---- NOC uptime: fleet-wide 24h / 7d ----
    since_24h = (now - timedelta(hours=24)).isoformat()
    total_samples = await db.noc_probes.count_documents({"at": {"$gte": since_24h}})
    up_samples = await db.noc_probes.count_documents({"at": {"$gte": since_24h}, "ok": True})
    uptime_24h_pct = round((up_samples / total_samples) * 100, 2) if total_samples else None
    since_7d = (now - timedelta(days=7)).isoformat()
    samples_7d = await db.noc_probes.count_documents({"at": {"$gte": since_7d}})
    up_7d = await db.noc_probes.count_documents({"at": {"$gte": since_7d}, "ok": True})
    uptime_7d_pct = round((up_7d / samples_7d) * 100, 2) if samples_7d else None
    devices_down = await db.noc_device_state.count_documents({"status": "down"})
    devices_total = await db.mikrotik_devices.count_documents({})
    # SLA: rough outage minutes = down samples * 5-min cadence
    down_samples_30d = await db.noc_probes.count_documents({
        "at": {"$gte": (now - timedelta(days=30)).isoformat()}, "ok": False,
    })
    outage_minutes_30d = down_samples_30d * 5

    # ---- Ticket load ----
    open_tickets = await db.tickets.count_documents({"status": {"$in": ["open", "awaiting_staff"]}})
    critical_tickets = await db.tickets.count_documents({"status": {"$nin": ["resolved", "closed"]},
                                                          "priority": "critical"})

    # ---- Top clients by lifetime revenue ----
    paid_all = await db.invoices.find({"status": "paid"}).to_list(20000)
    per_client: dict = {}
    for d in paid_all:
        uid = str(d.get("user_id"))
        if not uid or uid == "None":
            continue
        per_client[uid] = per_client.get(uid, 0.0) + float(d.get("total") or 0)
    top_pairs = sorted(per_client.items(), key=lambda x: x[1], reverse=True)[:5]
    top_uids = []
    for uid, _rev in top_pairs:
        try: top_uids.append(ObjectId(uid))
        except Exception: pass
    top_users = {str(u["_id"]): u for u in await db.users.find(
        {"_id": {"$in": top_uids}}).to_list(20)} if top_uids else {}
    top_clients = [{
        "user_id": uid,
        "name": (top_users.get(uid) or {}).get("name", ""),
        "email": (top_users.get(uid) or {}).get("email", ""),
        "lifetime_revenue": rev,
    } for uid, rev in top_pairs]

    return {
        "generated_at": _now(),
        "mrr": mrr,
        "arr": mrr * 12,
        "arpu": arpu,
        "churn_pct_30d": churn_pct,
        "clients_total": total_clients,
        "clients_with_active_service": clients_with_active,
        "active_services": active_services_count,
        "revenue_month_to_date": revenue_month,
        "revenue_trend_12m": trend,
        "unpaid_invoices": unpaid_count,
        "overdue_total": overdue_total,
        "noc": {
            "uptime_24h_pct": uptime_24h_pct,
            "uptime_7d_pct": uptime_7d_pct,
            "devices_total": devices_total,
            "devices_down": devices_down,
            "outage_minutes_30d": outage_minutes_30d,
            "samples_24h": total_samples,
        },
        "support": {
            "open_tickets": open_tickets,
            "critical_open": critical_tickets,
        },
        "top_clients": top_clients,
    }


# ============================================================
# PUBLIC STATUS PAGE
# ============================================================
_DEFAULT_STATUS_GROUPS = [
    {"key": "core_network",  "label": "Core Network"},
    {"key": "customer_edge", "label": "Customer Edge"},
    {"key": "peering",       "label": "Peering & Transit"},
]


@router.get("/public/status")
@_rl_limiter.limit(PUBLIC_STATUS_LIMIT)
async def public_status_page(request: Request):
    """Customer-friendly uptime snapshot with NO device names/IPs leaked.

    Devices are bucketed by their `status_group` field (default:
    customer_edge). Which groups are visible + their display labels come
    from `settings.status_page.groups`; falls back to defaults."""
    db = await _get_db()
    doc = await db.settings.find_one({"key": "status_page"}) or {}
    cfg = doc.get("value") or {}
    groups = cfg.get("groups") or _DEFAULT_STATUS_GROUPS
    company = cfg.get("company") or "Intercloud Digital Inovasi"
    incident_note = cfg.get("incident_note") or ""

    devices = await db.mikrotik_devices.find({}, {"status_group": 1}).to_list(1000)
    dev_group: dict = {d["_id"]: (d.get("status_group") or "customer_edge") for d in devices}

    now = datetime.now(timezone.utc)
    since_24h = (now - timedelta(hours=24)).isoformat()

    out_groups = []
    any_degraded = False
    any_operational = False
    for grp in groups:
        gkey = grp["key"]
        dev_ids = [did for did, gk in dev_group.items() if gk == gkey]
        base = {"device_id": {"$in": dev_ids}} if dev_ids else {"device_id": None}
        total = await db.noc_probes.count_documents({**base, "at": {"$gte": since_24h}})
        up = await db.noc_probes.count_documents({**base, "at": {"$gte": since_24h}, "ok": True})
        uptime_24h = round((up / total) * 100, 2) if total else None
        # 30d window uses daily rollups + recent raw samples (retention-safe)
        uptime_30d = await _noc_uptime_window(db, dev_ids if dev_ids else [None], 30)
        down_now = await db.noc_device_state.count_documents({
            "device_id": {"$in": dev_ids}, "status": "down",
        }) if dev_ids else 0
        if uptime_24h is None:
            status = "unknown"
        elif down_now > 0:
            status = "degraded"; any_degraded = True
        elif uptime_24h >= 99.0:
            status = "operational"; any_operational = True
        else:
            status = "degraded"; any_degraded = True
        out_groups.append({
            "key": gkey, "label": grp["label"],
            "status": status,
            "uptime_24h_pct": uptime_24h,
            "uptime_30d_pct": uptime_30d,
            "devices_count": len(dev_ids),
        })
    # Overall: degraded takes precedence; otherwise operational only if at
    # least one group is genuinely operational; else unknown.
    if any_degraded:
        overall_status = "degraded"
    elif any_operational:
        overall_status = "operational"
    else:
        overall_status = "unknown"
    return {
        "company": company,
        "generated_at": now.isoformat(),
        "overall_status": overall_status,
        "groups": out_groups,
        "incident_note": incident_note,
    }


@router.get("/admin/status-page/config")
async def status_page_config_get(admin=Depends(get_current_admin)):
    db = await _get_db()
    doc = await db.settings.find_one({"key": "status_page"}) or {}
    return doc.get("value") or {"groups": _DEFAULT_STATUS_GROUPS,
                                 "company": "Intercloud Digital Inovasi",
                                 "incident_note": ""}


@router.put("/admin/status-page/config")
async def status_page_config_put(payload: dict, request: Request, admin=Depends(get_current_admin)):
    db = await _get_db()
    before_doc = await db.settings.find_one({"key": "status_page"}) or {}
    value = {
        "groups": payload.get("groups") or _DEFAULT_STATUS_GROUPS,
        "company": (payload.get("company") or "Intercloud Digital Inovasi").strip(),
        "incident_note": (payload.get("incident_note") or "").strip(),
    }
    await db.settings.update_one({"key": "status_page"},
                                 {"$set": {"key": "status_page", "value": value,
                                           "updated_at": _now()}}, upsert=True)
    await log_audit(db, actor=admin, action="status_page.config_update", category="system",
                    target_type="settings", target_label="Status Page",
                    before=before_doc.get("value"), after=value,
                    severity="info", request=request)
    return value


# ============================================================
# MEDIA LIBRARY - shared assets for the Digital Creative team
# ============================================================
from fastapi import UploadFile, File, Form  # noqa: E402
from fastapi.responses import FileResponse  # noqa: E402
import uuid as _uuid  # noqa: E402
from pathlib import Path as _Path  # noqa: E402

MEDIA_DIR = _Path(__file__).resolve().parent.parent / "uploads" / "media"
_MEDIA_ALLOWED_TYPES = {"image/png", "image/jpeg", "image/webp", "image/gif", "image/svg+xml"}
_MEDIA_MAX_BYTES = 8 * 1024 * 1024  # 8 MB


async def _media_usage(db, media_id: str) -> list:
    """Where is this asset referenced? Scans articles (cover/OG/body) and
    branding/landing settings. Computed live so it never goes stale."""
    needle = f"/media/file/{media_id}"
    used = []
    cur = db.articles.find({"$or": [
        {"cover_image_url": {"$regex": needle}},
        {"og_image_url": {"$regex": needle}},
        {"body_html": {"$regex": needle}},
    ]}, {"title": 1, "slug": 1})
    async for a in cur:
        used.append({"type": "article", "id": str(a["_id"]),
                     "label": a.get("title") or a.get("slug") or "article"})
    async for s in db.settings.find({"key": {"$in": ["branding", "landing_content"]}}):
        if needle in str(s.get("value", "")):
            used.append({"type": "settings", "id": s.get("key"),
                         "label": f"Settings: {s.get('key')}"})
    return used


def _serialize_media(d: dict, used_in=None) -> dict:
    return {
        "id": str(d["_id"]),
        "filename": d.get("filename", ""),
        "url": d.get("url", ""),
        "content_type": d.get("content_type", ""),
        "size_bytes": int(d.get("size_bytes") or 0),
        "alt_text": d.get("alt_text", ""),
        "tags": d.get("tags", []),
        "uploaded_by": d.get("uploaded_by", ""),
        "created_at": d.get("created_at", ""),
        "used_in": used_in if used_in is not None else d.get("used_in", []),
    }


@router.get("/admin/media")
async def media_list(staff=Depends(get_current_staff),
                     tag: Optional[str] = None, q: Optional[str] = None):
    db = await _get_db()
    query: dict = {}
    if tag:
        query["tags"] = tag.strip().lower()
    if q:
        query["$or"] = [
            {"filename": {"$regex": q.strip(), "$options": "i"}},
            {"alt_text": {"$regex": q.strip(), "$options": "i"}},
        ]
    docs = await db.media_assets.find(query).sort("created_at", -1).to_list(500)
    out = []
    for d in docs:
        used = await _media_usage(db, str(d["_id"]))
        if used != d.get("used_in"):
            await db.media_assets.update_one({"_id": d["_id"]}, {"$set": {"used_in": used}})
        out.append(_serialize_media(d, used))
    return out


@router.post("/admin/media")
async def media_upload(file: UploadFile = File(...),
                       alt_text: str = Form(""),
                       tags: str = Form(""),
                       staff=Depends(get_current_content)):
    db = await _get_db()
    if file.content_type not in _MEDIA_ALLOWED_TYPES:
        raise HTTPException(status_code=400,
                            detail=f"Unsupported type {file.content_type}. Allowed: PNG, JPEG, WebP, GIF, SVG.")
    raw = await file.read()
    if len(raw) > _MEDIA_MAX_BYTES:
        raise HTTPException(status_code=400, detail="File exceeds 8 MB limit")
    MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    ext = _Path(file.filename or "upload.bin").suffix.lower() or ".bin"
    mid = ObjectId()
    stored_name = f"{mid}{ext}"
    (MEDIA_DIR / stored_name).write_bytes(raw)
    doc = {
        "_id": mid,
        "filename": file.filename or stored_name,
        "stored_name": stored_name,
        "url": f"/api/portal/media/file/{mid}",
        "content_type": file.content_type,
        "size_bytes": len(raw),
        "alt_text": (alt_text or "").strip(),
        "tags": sorted({t.strip().lower() for t in (tags or "").split(",") if t.strip()}),
        "uploaded_by": staff["email"],
        "used_in": [],
        "created_at": _now(),
    }
    await db.media_assets.insert_one(doc)
    return _serialize_media(doc)


@router.post("/admin/documents/upload")
async def docs_upload(file: UploadFile = File(...), title: str = Form(""),
                      category: str = Form("contract"), customer_name: str = Form(""),
                      notes: str = Form(""), staff=Depends(get_current_staff)):
    """UAT-003: upload dokumen lokal (drag & drop) selain link URL."""
    db = await _get_db()
    ctype = file.content_type or "application/octet-stream"
    if ctype not in _DOC_ALLOWED_TYPES:
        raise HTTPException(status_code=400,
                            detail=f"Tipe file {ctype} tidak didukung. Gunakan PDF, Word, Excel, gambar, ZIP, atau teks.")
    raw = await file.read()
    if len(raw) > _DOC_MAX_BYTES:
        raise HTTPException(status_code=400, detail="Ukuran file melebihi 15 MB")
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    did = ObjectId()
    ext = _DocPath(file.filename or "dokumen.bin").suffix.lower() or ".bin"
    stored_name = f"{did}{ext}"
    (DOCS_DIR / stored_name).write_bytes(raw)
    doc = {
        "_id": did,
        "title": (title or "").strip() or (file.filename or "Dokumen"),
        "category": category or "contract",
        "customer_name": customer_name or "",
        "url": f"/api/portal/documents/file/{did}",
        "notes": notes or "",
        "filename": file.filename or stored_name,
        "stored_name": stored_name,
        "content_type": ctype,
        "size_bytes": len(raw),
        "uploaded_by": staff["email"],
        "created_at": _now(),
    }
    await db.documents.insert_one(doc)
    return _serialize_doc(doc)


@router.put("/admin/media/{mid}")
async def media_update(mid: str, payload: dict, staff=Depends(get_current_content)):
    db = await _get_db()
    d = await db.media_assets.find_one({"_id": _oid(mid)})
    if not d:
        raise HTTPException(status_code=404, detail="Media not found")
    upd: dict = {}
    if "alt_text" in payload:
        upd["alt_text"] = (payload.get("alt_text") or "").strip()
    if "tags" in payload:
        raw_tags = payload.get("tags") or []
        if isinstance(raw_tags, str):
            raw_tags = raw_tags.split(",")
        upd["tags"] = sorted({str(t).strip().lower() for t in raw_tags if str(t).strip()})
    if upd:
        await db.media_assets.update_one({"_id": d["_id"]}, {"$set": upd})
    d = await db.media_assets.find_one({"_id": d["_id"]})
    return _serialize_media(d)


@router.delete("/admin/media/{mid}")
async def media_delete(mid: str, staff=Depends(get_current_content)):
    db = await _get_db()
    d = await db.media_assets.find_one({"_id": _oid(mid)})
    if not d:
        raise HTTPException(status_code=404, detail="Media not found")
    used = await _media_usage(db, mid)
    if used:
        raise HTTPException(status_code=409, detail={
            "message": "Asset is still in use - detach it first.",
            "used_in": used,
        })
    try:
        (MEDIA_DIR / d.get("stored_name", "")).unlink(missing_ok=True)
    except Exception:
        pass
    await db.media_assets.delete_one({"_id": d["_id"]})
    return {"deleted": 1}


@router.get("/media/file/{mid}", include_in_schema=False)
async def media_file(mid: str):
    """Public file serve - media is referenced from public articles."""
    db = await _get_db()
    d = await db.media_assets.find_one({"_id": _oid(mid)})
    if not d:
        raise HTTPException(status_code=404, detail="Media not found")
    fp = MEDIA_DIR / d.get("stored_name", "")
    if not fp.exists():
        raise HTTPException(status_code=404, detail="File missing on disk")
    return FileResponse(fp, media_type=d.get("content_type") or "application/octet-stream",
                        headers={"Cache-Control": "public, max-age=86400"})


# ============================================================
# CONTENT CALENDAR - plan articles / campaigns / social posts
# ============================================================
def _serialize_calendar(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "title": d.get("title", ""),
        "type": d.get("type", "article"),
        "scheduled_at": d.get("scheduled_at", ""),
        "status": d.get("status", "draft"),
        "linked_article_id": d.get("linked_article_id"),
        "owner_id": d.get("owner_id"),
        "notes": d.get("notes", ""),
        "created_at": d.get("created_at", ""),
    }


_CAL_TYPES = {"article", "campaign", "social_post"}
_CAL_STATUSES = {"draft", "scheduled", "published"}


@router.get("/admin/content-calendar")
async def calendar_list(staff=Depends(get_current_staff),
                        date_from: Optional[str] = None,
                        date_to: Optional[str] = None):
    db = await _get_db()
    query: dict = {}
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to + "T23:59:59"
        query["scheduled_at"] = rng
    docs = await db.content_calendar.find(query).sort("scheduled_at", 1).to_list(1000)
    return [_serialize_calendar(d) for d in docs]


@router.post("/admin/content-calendar")
async def calendar_create(payload: dict, staff=Depends(get_current_content)):
    db = await _get_db()
    title = (payload.get("title") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title is required")
    ctype = payload.get("type") or "article"
    if ctype not in _CAL_TYPES:
        raise HTTPException(status_code=400, detail=f"type must be one of {sorted(_CAL_TYPES)}")
    status = payload.get("status") or "draft"
    if status not in _CAL_STATUSES:
        raise HTTPException(status_code=400, detail=f"status must be one of {sorted(_CAL_STATUSES)}")
    doc = {
        "title": title,
        "type": ctype,
        "scheduled_at": payload.get("scheduled_at") or _now(),
        "status": status,
        "linked_article_id": payload.get("linked_article_id"),
        "owner_id": staff["id"],
        "notes": payload.get("notes") or "",
        "created_at": _now(),
    }
    r = await db.content_calendar.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_calendar(doc)


@router.put("/admin/content-calendar/{cid}")
async def calendar_update(cid: str, payload: dict, staff=Depends(get_current_content)):
    db = await _get_db()
    d = await db.content_calendar.find_one({"_id": _oid(cid)})
    if not d:
        raise HTTPException(status_code=404, detail="Calendar entry not found")
    upd: dict = {}
    for k in ("title", "scheduled_at", "linked_article_id", "notes"):
        if k in payload:
            upd[k] = payload[k]
    if "type" in payload:
        if payload["type"] not in _CAL_TYPES:
            raise HTTPException(status_code=400, detail=f"type must be one of {sorted(_CAL_TYPES)}")
        upd["type"] = payload["type"]
    if "status" in payload:
        if payload["status"] not in _CAL_STATUSES:
            raise HTTPException(status_code=400, detail=f"status must be one of {sorted(_CAL_STATUSES)}")
        upd["status"] = payload["status"]
    if upd:
        await db.content_calendar.update_one({"_id": d["_id"]}, {"$set": upd})
    d = await db.content_calendar.find_one({"_id": d["_id"]})
    return _serialize_calendar(d)


@router.delete("/admin/content-calendar/{cid}")
async def calendar_delete(cid: str, staff=Depends(get_current_content)):
    db = await _get_db()
    r = await db.content_calendar.delete_one({"_id": _oid(cid)})
    if not r.deleted_count:
        raise HTTPException(status_code=404, detail="Calendar entry not found")
    return {"deleted": 1}


async def _sync_article_calendar(db, article: dict, staff) -> None:
    """When an article is published, upsert its calendar entry to published.
    Fire-and-forget: never blocks the article save."""
    try:
        if not article or article.get("status") != "published":
            return
        aid = str(article["_id"])
        await db.content_calendar.update_one(
            {"linked_article_id": aid},
            {"$set": {"title": article.get("title", ""),
                      "type": "article",
                      "status": "published",
                      "scheduled_at": article.get("published_at") or _now(),
                      "linked_article_id": aid},
             "$setOnInsert": {"owner_id": staff.get("id"), "notes": "",
                               "created_at": _now()}},
            upsert=True,
        )
    except Exception:
        import logging
        logging.getLogger("portal.calendar").exception("calendar sync failed")


# ============================================================
# TICKET ↔ DEVICE linking - minimal device options for dropdowns
# ============================================================
@router.get("/tickets/device-options")
async def ticket_device_options(user=Depends(get_current_user)):
    """Names only (no hosts/IPs) so clients can point a ticket at a device."""
    db = await _get_db()
    docs = await db.mikrotik_devices.find({}, {"name": 1}).sort("name", 1).to_list(500)
    return [{"id": str(d["_id"]), "name": d.get("name") or "unnamed"} for d in docs]

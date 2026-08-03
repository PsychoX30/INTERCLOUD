"""Admin core: dashboard, notifications, reports, backup/restore, system update/health, audit logs, owner overview.

Split from the former monolithic routes.py - behavior preserved 1:1.
"""
import os
import asyncio
import logging
import secrets
import re
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Request, WebSocket
from bson import ObjectId
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from .. import models as m
from ..auth import (
    verify_password, hash_password, create_access_token,
    get_current_user, get_current_admin, get_current_staff, get_current_content,
    require_roles, sales_can_access,
    STAFF_ROLES, FINANCE_ROLES, BILLING_ROLES, CATALOG_ROLES,
    OPS_ROLES, USER_MGMT_ROLES, TICKET_ROLES, CONTENT_ROLES,
)
from ..audit import log_audit, serialize as _serialize_audit
from ..secretbox import (dec_value as _sb_dec, enc_value as _sb_enc,
                         decrypt_config as _sb_dec_config)
from .. import integrations_v2 as iv2
from .documents import _render_pdf_bytes  # noqa: E402
from .shared import _EXTRA_PAYMENT_MODULES, _get_db, _get_setting_value, _iso, _mark_overdue, _now, _oid  # noqa: E402
from ..integrations_registry import module_schema  # noqa: E402
from fastapi.responses import Response  # noqa: E402
from portal.backups import run_mongodump as _run_mongodump  # noqa: E402
from portal.backups import run_mongorestore as _run_mongorestore  # noqa: E402
import io as _io  # noqa: E402

router = APIRouter()


# ============================================================
# ADMIN
# ============================================================
async def _build_admin_alerts(db, staff, scope_user_ids=None) -> list:
    """Pusat Notifikasi: kumpulkan peringatan penting lintas modul dengan severity."""
    alerts = []
    if staff["role"] in FINANCE_ROLES or staff["role"] == "sales":
        inv_q = {} if scope_user_ids is None else {"user_id": {"$in": scope_user_ids}}
        overdue_docs = await db.invoices.find(
            {**inv_q, "status": "overdue"}).sort("due_date", 1).to_list(5)
        for d in overdue_docs:
            alerts.append({
                "type": "invoice_overdue",
                "severity": "danger",
                "title": f"Invoice {d.get('number', '')} jatuh tempo",
                "detail": f"Jatuh tempo {d.get('due_date', '')}",
                "link": "/portal/admin/invoices",
            })
        today = datetime.now(timezone.utc).date()
        soon = (today + timedelta(days=3)).isoformat()
        due_soon = await db.invoices.find(
            {**inv_q, "status": "unpaid",
             "due_date": {"$gte": today.isoformat(), "$lte": soon}}).sort("due_date", 1).to_list(5)
        for d in due_soon:
            alerts.append({
                "type": "invoice_due_soon",
                "severity": "warning",
                "title": f"Invoice {d.get('number', '')} segera jatuh tempo",
                "detail": f"Jatuh tempo {d.get('due_date', '')}",
                "link": "/portal/admin/invoices",
            })
    down_states = await db.noc_device_state.find({"status": "down"}).to_list(20)
    dev_map = {m["_id"]: m.get("name", "unnamed") async for m in db.mikrotik_devices.find(
        {"_id": {"$in": [s["device_id"] for s in down_states]}})} if down_states else {}
    for s in down_states:
        if s["device_id"] not in dev_map:
            continue
        alerts.append({
            "type": "device_down",
            "severity": "danger",
            "title": f"Perangkat {dev_map[s['device_id']]} DOWN",
            "detail": s.get("last_message") or "Tidak merespons probe",
            "link": "/portal/admin/noc",
        })
    pending_orders = await db.orders.count_documents({"status": "pending_verification"})
    if pending_orders:
        alerts.append({
            "type": "orders_pending",
            "severity": "warning",
            "title": f"{pending_orders} order menunggu verifikasi",
            "detail": "Verifikasi pembayaran untuk memproses provisioning",
            "link": "/portal/admin/orders",
        })
    failed_services = await db.services.find(
        {"config.provision_status": "failed"}).sort("created_at", -1).to_list(5)
    for s in failed_services:
        addons = (s.get("config") or {}).get("addons") or []
        detail = ("Termasuk add-on: " + ", ".join(addons)) if addons else "Perlu tindak lanjut manual di panel"
        alerts.append({
            "type": "provision_failed",
            "severity": "danger",
            "title": f"Provisioning gagal: {s.get('product_name') or s.get('name', '')}",
            "detail": detail,
            "link": "/portal/admin/services",
        })
    pending_provision = await db.services.find(
        {"config.provision_status": "pending"}).sort("created_at", -1).to_list(5)
    for s in pending_provision:
        alerts.append({
            "type": "provision_pending",
            "severity": "warning",
            "title": f"Menunggu provisioning manual: {s.get('product_name') or s.get('name', '')}",
            "detail": "Integrasi tidak aktif saat order dibayar - selesaikan provisioning lalu aktifkan layanan",
            "link": "/portal/admin/services",
        })
    # Kesehatan server: disk (cek live) + SSL (dari sweep harian health_alert_log)
    if staff["role"] in ("admin", "support"):
        try:
            import shutil as _sh
            du = _sh.disk_usage("/")
            disk_pct = round(du.used / du.total * 100, 1)
            if disk_pct >= 85.0:
                free_gb = round(du.free / 1024 ** 3, 1)
                alerts.append({
                    "type": "disk_full",
                    "severity": "danger",
                    "title": f"Disk server {disk_pct}% terpakai",
                    "detail": f"Sisa {free_gb} GB - bersihkan log/backup lama atau tambah kapasitas",
                    "link": "/portal/admin/diagnostics",
                })
        except Exception:
            pass
        cutoff = (datetime.now(timezone.utc).date() - timedelta(days=2)).isoformat()
        hdocs = await db.health_alert_log.find(
            {"date": {"$gte": cutoff}}).sort("date", -1).to_list(3)
        seen_issues = set()
        for h in hdocs:
            for issue in h.get("issues", []):
                low = issue.lower()
                if low.startswith("disk") or issue in seen_issues:
                    continue  # disk sudah dicek live di atas
                seen_issues.add(issue)
                alerts.append({
                    "type": "ssl_expiry" if "ssl" in low else "server_health",
                    "severity": "danger",
                    "title": "Sertifikat SSL akan kedaluwarsa" if "ssl" in low
                             else "Peringatan kesehatan server",
                    "detail": issue,
                    "link": "/portal/admin/diagnostics",
                })
    order = {"danger": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda a: order.get(a["severity"], 9))
    return alerts


def _alert_key(a: dict) -> str:
    import hashlib
    raw = f"{a.get('type', '')}|{a.get('title', '')}|{a.get('detail', '')}"
    return hashlib.sha1(raw.encode()).hexdigest()[:16]


@router.get("/admin/notifications")
async def admin_notifications(severity: str | None = None, staff=Depends(get_current_staff)):
    """Pusat Notifikasi dengan filter prioritas (?severity=danger|warning) +
    status baca per staf (koleksi notification_reads)."""
    db = await _get_db()
    scope_user_ids = None
    if staff["role"] == "sales":
        scope_user_ids = [ObjectId(cid) for cid in staff.get("assigned_client_ids") or []]
    alerts = await _build_admin_alerts(db, staff, scope_user_ids)
    for a in alerts:
        a["key"] = _alert_key(a)
    read_set = {r["key"] for r in await db.notification_reads.find(
        {"staff_id": staff["id"]}, {"key": 1}).to_list(2000)}
    for a in alerts:
        a["read"] = a["key"] in read_set
    if severity:
        if severity not in ("danger", "warning", "info"):
            raise HTTPException(status_code=400, detail="severity harus danger, warning, atau info")
        alerts = [a for a in alerts if a["severity"] == severity]
    return {"alerts": alerts, "count": len(alerts),
            "unread": sum(1 for a in alerts if not a["read"])}


@router.post("/admin/notifications/mark-read")
async def admin_notifications_mark_read(payload: dict, staff=Depends(get_current_staff)):
    """Tandai notifikasi sudah dibaca. Body: {keys: [..]}"""
    keys = payload.get("keys") or []
    if not isinstance(keys, list) or not keys:
        raise HTTPException(status_code=400, detail="keys (list) wajib diisi")
    db = await _get_db()
    now = _now()
    for k in keys[:200]:
        await db.notification_reads.update_one(
            {"staff_id": staff["id"], "key": str(k)},
            {"$set": {"read_at": now}}, upsert=True)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    await db.notification_reads.delete_many({"read_at": {"$lt": cutoff}})
    return {"ok": True, "marked": len(keys[:200])}


@router.post("/admin/reports/monthly/send")
async def admin_send_monthly_report(payload: dict = None, admin=Depends(require_roles("admin", "finance"))):
    """Kirim laporan bulanan (tagihan + trafik) ke email support untuk dokumentasi.
    Body opsional: {month: 'YYYY-MM'} - default bulan lalu."""
    db = await _get_db()
    from portal import emails as _em
    month = ((payload or {}).get("month") or "").strip() or None
    return await _em.run_monthly_report(db, month=month)


@router.get("/admin/reports/monthly")
async def admin_list_monthly_reports(admin=Depends(require_roles("admin", "finance"))):
    """Arsip laporan bulanan (terbaru dulu) untuk menu Finance."""
    db = await _get_db()
    docs = await db.monthly_reports.find({}).sort("month", -1).to_list(60)
    return [{"id": str(d["_id"]),
             "month": d.get("month", ""),
             "generated_at": d.get("generated_at", ""),
             "to_email": d.get("to_email", ""),
             "delivery_status": (d.get("last_delivery") or {}).get("status", ""),
             "summary": d.get("summary", {})} for d in docs]


@router.get("/admin/reports/monthly/{month}/pdf")
async def admin_monthly_report_pdf(month: str, admin=Depends(require_roles("admin", "finance"))):
    """Unduh ulang arsip laporan bulanan sebagai PDF."""
    db = await _get_db()
    d = await db.monthly_reports.find_one({"month": month})
    if not d:
        raise HTTPException(status_code=404, detail="Arsip laporan tidak ditemukan")
    from portal import emails as _em
    html = _em.wrap_html(d.get("body_html", ""))
    return Response(content=_render_pdf_bytes(html),
                    media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="Laporan-Bulanan-{month}.pdf"'})


def _monthly_report_workbook(month: str, summary: dict, invoices: list,
                             user_map: dict, traffic: list, svc_map: dict) -> bytes:
    """Excel arsip bulanan siap olah: formula SUM/referensi antar-sheet,
    format Rupiah & persen, penomoran baris, freeze header, border."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    RP = '"Rp" #,##0'
    NUM = "#,##0"
    GB = "#,##0.00"
    PCT = "0.0%"
    thin = Side(style="thin", color="FFD8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF0A2350")
    tot_font = Font(bold=True)
    tot_fill = PatternFill("solid", fgColor="FFFEF3C7")

    wb = Workbook()

    def _style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        ws.freeze_panes = "A2"

    def _autofit(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 42)

    # ---- Sheet 2 data first (Invoice) so Ringkasan can reference its totals ----
    ws_inv = wb.active
    ws_inv.title = "Invoice"
    inv_headers = ["No", "Nomor Invoice", "Klien", "Tanggal Terbit", "Jatuh Tempo",
                   "Status", "Subtotal", "PPN", "Total"]
    ws_inv.append(inv_headers)
    for idx, inv in enumerate(invoices, start=1):
        ws_inv.append([
            idx,
            inv.get("number", ""),
            user_map.get(str(inv.get("user_id") or ""), ""),
            (inv.get("created_at") or "")[:10],
            (inv.get("due_date") or "")[:10],
            inv.get("status", ""),
            float(inv.get("subtotal") or 0),
            float(inv.get("tax_amount") or 0),
            float(inv.get("total") or 0),
        ])
    n_inv = len(invoices)
    total_row = n_inv + 2
    if n_inv:
        ws_inv.append(["", "TOTAL", "", "", "", "",
                       f"=SUM(G2:G{total_row - 1})",
                       f"=SUM(H2:H{total_row - 1})",
                       f"=SUM(I2:I{total_row - 1})"])
    else:
        ws_inv.append(["", "TOTAL", "", "", "", "", 0, 0, 0])
    for r in range(2, total_row + 1):
        for c in range(1, 10):
            cell = ws_inv.cell(row=r, column=c)
            cell.border = border
            if c == 1:
                cell.number_format = NUM
            elif c >= 7:
                cell.number_format = RP
            if r == total_row:
                cell.font = tot_font
                cell.fill = tot_fill
    _style_header(ws_inv, 9)
    _autofit(ws_inv)

    # ---- Sheet 3: Trafik ----
    ws_tr = wb.create_sheet("Trafik")
    ws_tr.append(["No", "Layanan", "Inbound (GB)", "Outbound (GB)", "Total (GB)"])
    if traffic:
        for idx, t in enumerate(traffic, start=1):
            r = idx + 1
            ws_tr.append([idx,
                          svc_map.get(t.get("service_id", ""), t.get("service_id", "")),
                          round(float(t.get("in_gb") or 0), 3),
                          round(float(t.get("out_gb") or 0), 3),
                          f"=C{r}+D{r}"])
        tr_total = len(traffic) + 2
        ws_tr.append(["", "TOTAL", f"=SUM(C2:C{tr_total - 1})",
                      f"=SUM(D2:D{tr_total - 1})", f"=SUM(E2:E{tr_total - 1})"])
        for r in range(2, tr_total + 1):
            for c in range(1, 6):
                cell = ws_tr.cell(row=r, column=c)
                cell.border = border
                if c == 1:
                    cell.number_format = NUM
                elif c >= 3:
                    cell.number_format = GB
                if r == tr_total:
                    cell.font = tot_font
                    cell.fill = tot_fill
    else:
        ws_tr.append(["", "Belum ada data trafik untuk bulan ini", "", "", ""])
    _style_header(ws_tr, 5)
    _autofit(ws_tr)

    # ---- Sheet 1: Ringkasan (referensi formula ke sheet Invoice/Trafik) ----
    ws_sum = wb.create_sheet("Ringkasan", 0)
    ws_sum.append(["Metrik", "Jumlah", "Nilai"])
    paid_count = int(summary.get("invoices_paid") or 0)
    rows = [
        ("Periode", month, None, None),
        ("Invoice diterbitkan", n_inv, f"=Invoice!I{total_row}", RP),
        ("Invoice dibayar", paid_count, float(summary.get("invoices_paid_total") or 0), RP),
        ("Collection rate (dibayar / terbit)", None, "=IF(C3=0,0,C4/C3)", PCT),
        ("Outstanding saat ini (unpaid + overdue)", None, float(summary.get("outstanding_total") or 0), RP),
        ("Credit note diterapkan", int(summary.get("credit_notes_applied") or 0),
         float(summary.get("credit_notes_total") or 0), RP),
        ("Klien baru", int(summary.get("new_clients") or 0), None, None),
        ("Order baru", int(summary.get("new_orders") or 0), None, None),
        ("Trafik inbound (GB)", None,
         (f"=Trafik!C{len(traffic) + 2}" if traffic else 0), GB),
        ("Trafik outbound (GB)", None,
         (f"=Trafik!D{len(traffic) + 2}" if traffic else 0), GB),
    ]
    for r_off, (label, qty, val, fmt) in enumerate(rows, start=2):
        ws_sum.cell(row=r_off, column=1, value=label).border = border
        qcell = ws_sum.cell(row=r_off, column=2, value=qty)
        qcell.border = border
        qcell.number_format = NUM
        vcell = ws_sum.cell(row=r_off, column=3, value=val)
        vcell.border = border
        if fmt:
            vcell.number_format = fmt
    _style_header(ws_sum, 3)
    _autofit(ws_sum)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/admin/reports/monthly/{month}/xlsx")
async def admin_monthly_report_xlsx(month: str, admin=Depends(require_roles("admin", "finance"))):
    """Arsip laporan bulanan sebagai Excel (formula, format Rupiah/persen,
    penomoran) untuk diolah tim finance."""
    db = await _get_db()
    d = await db.monthly_reports.find_one({"month": month})
    if not d:
        raise HTTPException(status_code=404, detail="Arsip laporan tidak ditemukan")
    q_month = {"$regex": f"^{month}"}
    invoices = await db.invoices.find({"created_at": q_month}).sort("created_at", 1).to_list(3000)
    uids = list({inv["user_id"] for inv in invoices if inv.get("user_id")})
    user_map = {}
    if uids:
        async for u in db.users.find({"_id": {"$in": uids}}):
            user_map[str(u["_id"])] = u.get("name") or u.get("email", "")
    traffic = await db.traffic_monthly.find({"month": month}).sort("in_gb", -1).to_list(500)
    svc_ids = []
    for t in traffic:
        try:
            svc_ids.append(ObjectId(t["service_id"]))
        except Exception:
            pass
    svc_map = {}
    if svc_ids:
        async for s in db.services.find({"_id": {"$in": svc_ids}}):
            svc_map[str(s["_id"])] = s.get("name") or s.get("product_name", "")
    data = _monthly_report_workbook(month, d.get("summary", {}), invoices,
                                    user_map, traffic, svc_map)
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Laporan-Bulanan-{month}.xlsx"'})


@router.post("/admin/reports/weekly/send")
async def admin_send_weekly_summary(admin=Depends(require_roles("admin", "finance"))):
    """Kirim ringkasan mingguan (order baru, tiket terbuka, invoice jatuh tempo) sekarang."""
    db = await _get_db()
    from portal import emails as _em
    return await _em.run_weekly_summary(db)


@router.post("/admin/integrations-v2/smtp/send-test")
async def integrations_v2_smtp_send_test(payload: dict = None, admin=Depends(get_current_admin)):
    """Kirim email percobaan via SMTP tersimpan - verifikasi cepat produksi."""
    db = await _get_db()
    smtp = await iv2.get_settings(db, "smtp")
    if not smtp or not smtp.get("enabled"):
        raise HTTPException(status_code=400,
                            detail="Integrasi SMTP belum aktif. Simpan & aktifkan konfigurasi SMTP terlebih dahulu.")
    to_email = ((payload or {}).get("to") or "").strip() or admin.get("email", "")
    from portal import emails as _em
    inner = ("<p>Email percobaan dari <b>Intercloud Portal</b>.</p>"
             f"<p>Dikirim oleh <b>{admin.get('name', '')}</b> ({admin.get('email', '')}) pada "
             f"{datetime.now(timezone.utc).isoformat()[:19]} UTC.</p>"
             "<p>Jika Anda menerima email ini, konfigurasi SMTP produksi Anda bekerja dengan baik.</p>")
    result = await _em.deliver(db, to_email=to_email, subject="Uji SMTP - Intercloud Portal",
                               body_html=_em.wrap_html(inner), event_key="smtp_test")
    return {"ok": result.get("status") == "sent", "to": to_email, **result}


@router.get("/admin/dashboard")
async def admin_dashboard(staff=Depends(get_current_staff)):
    db = await _get_db()
    await _mark_overdue(db)

    # ---- Scope filter: sales sees ONLY their assigned clients ----
    # Everyone else (admin/finance/support) sees the global tenant view.
    scope_user_ids = None
    if staff["role"] == "sales":
        scope_user_ids = [ObjectId(cid) for cid in (staff.get("assigned_client_ids") or [])]
        # Unassigned sales → all counts are zero (no doc matches _id:None).
        if not scope_user_ids:
            scope_user_ids = [ObjectId("000000000000000000000000")]

    if scope_user_ids is None:
        client_q = {"role": "client"}
        svc_q = {"status": "active"}
        tkt_q = {"status": {"$in": ["open", "awaiting_staff"]}}
    else:
        client_q = {"role": "client", "_id": {"$in": scope_user_ids}}
        svc_q = {"status": "active", "user_id": {"$in": scope_user_ids}}
        tkt_q = {"status": {"$in": ["open", "awaiting_staff"]}, "user_id": {"$in": scope_user_ids}}

    total_users = await db.users.count_documents(client_q)
    active_services = await db.services.count_documents(svc_q)
    open_tickets = await db.tickets.count_documents(tkt_q)

    stats = {
        "total_clients": total_users,
        "active_services": active_services,
        "open_tickets": open_tickets,
    }

    # Financial stats - visible to finance/admin OR sales (scoped to their book).
    if staff["role"] in FINANCE_ROLES or staff["role"] == "sales":
        inv_q_base = {} if scope_user_ids is None else {"user_id": {"$in": scope_user_ids}}
        unpaid = await db.invoices.count_documents({**inv_q_base, "status": "unpaid"})
        overdue = await db.invoices.count_documents({**inv_q_base, "status": "overdue"})
        pending_orders = await db.orders.count_documents({**inv_q_base, "status": "pending"})

        now = datetime.now(timezone.utc)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        paid_docs = await db.invoices.find(
            {**inv_q_base, "status": "paid", "paid_at": {"$gte": month_start}}
        ).to_list(1000)
        revenue_month = sum(d.get("total", 0) for d in paid_docs)
        all_paid = await db.invoices.find({**inv_q_base, "status": "paid"}).to_list(5000)
        revenue_total = sum(d.get("total", 0) for d in all_paid)
        overdue_docs = await db.invoices.find({**inv_q_base, "status": "overdue"}).to_list(1000)
        overdue_total = sum(d.get("total", 0) for d in overdue_docs)
        stats.update({
            "unpaid_invoices": unpaid,
            "overdue_invoices": overdue,
            "pending_orders": pending_orders,
            "revenue_month": revenue_month,
            "revenue_total": revenue_total,
            "overdue_total": overdue_total,
        })

    # ---- Tagihan terbaru (Ringkasan Umum) + Pusat Notifikasi ----
    recent_invoices = []
    if staff["role"] in FINANCE_ROLES or staff["role"] == "sales":
        inv_q_base = {} if scope_user_ids is None else {"user_id": {"$in": scope_user_ids}}
        recent_docs = await db.invoices.find(inv_q_base).sort("created_at", -1).to_list(5)
        uid_set = {d["user_id"] for d in recent_docs}
        user_map = {u["_id"]: u.get("name", "") async for u in db.users.find({"_id": {"$in": list(uid_set)}})}
        recent_invoices = [{
            "id": str(d["_id"]),
            "number": d.get("number", ""),
            "user_name": user_map.get(d["user_id"], ""),
            "total": d.get("total", 0),
            "status": d.get("status", "unpaid"),
            "due_date": d.get("due_date", ""),
        } for d in recent_docs]

    alerts = await _build_admin_alerts(db, staff, scope_user_ids)
    down_count = sum(1 for a in alerts if a["type"] == "device_down")

    # ---- System health: status real dari registry integrations ----
    health = [
        {"name": "API Backend", "status": "ok", "detail": "Online"},
        {"name": "MongoDB", "status": "ok", "detail": "Connected"},
    ]
    mikrotik_count = await db.mikrotik_devices.count_documents({})
    if mikrotik_count:
        health.append({
            "name": "MikroTik Ops",
            "status": "warn" if down_count else "ok",
            "detail": f"{mikrotik_count} device ({down_count} down)" if down_count else f"{mikrotik_count} device",
        })
    else:
        health.append({"name": "MikroTik Ops", "status": "off", "detail": "Not configured"})
    allow_extra = bool(await _get_setting_value(db, "enable_extra_payment_gateways", False))
    integ_docs = await db.integrations.find({}).sort("created_at", -1).to_list(100)
    for idoc in integ_docs:
        if not allow_extra and idoc.get("module") in _EXTRA_PAYMENT_MODULES:
            continue
        schema = module_schema(idoc.get("module", ""))
        enabled = idoc.get("status") == "enabled"
        last = idoc.get("last_test_result") or {}
        test_ok = last.get("ok") if isinstance(last, dict) else None
        health.append({
            "name": idoc.get("name") or (schema["label"] if schema else idoc.get("module", "")),
            "status": ("warn" if test_ok is False else "ok") if enabled else "off",
            "detail": ("Test failed" if test_ok is False else "Enabled") if enabled else "Disabled",
        })
    me_doc = await db.users.find_one({"_id": ObjectId(staff["id"])}) or {}
    es = me_doc.get("email_settings") or {}
    smtp_ok = bool(((es.get("smtp") or {}).get("credentials") or {}).get("host"))
    health.append({
        "name": "SMTP (personal)",
        "status": "ok" if smtp_ok else "off",
        "detail": "Configured" if smtp_ok else "Not configured",
    })

    return {"stats": stats, "role": staff["role"],
            "recent_invoices": recent_invoices, "alerts": alerts, "health": health}


# ---------- Backup history ----------

BACKUP_DIR = "/app/backups"


@router.post("/admin/backup/trigger")
async def backup_trigger(request: Request, admin=Depends(get_current_admin)):
    """Backup manual: mongodump ke file + catat di riwayat."""
    import pathlib
    blob, filename = await _run_mongodump()
    pathlib.Path(BACKUP_DIR).mkdir(parents=True, exist_ok=True)
    path = f"{BACKUP_DIR}/{filename}"
    with open(path, "wb") as f:
        f.write(blob)
    db = await _get_db()
    res = await db.backup_history.insert_one({
        "filename": filename, "path": path, "size_bytes": len(blob),
        "kind": "manual", "by": admin.get("email", ""), "created_at": _now(),
    })
    return {"ok": True, "id": str(res.inserted_id), "filename": filename, "size_bytes": len(blob)}


@router.get("/admin/backup/history")
async def backup_history_list(admin=Depends(get_current_admin)):
    db = await _get_db()
    docs = await db.backup_history.find({}).sort("created_at", -1).to_list(100)
    return [{"id": str(d["_id"]), "filename": d.get("filename", ""),
             "size_bytes": d.get("size_bytes", 0), "kind": d.get("kind", "manual"),
             "by": d.get("by", ""), "created_at": _iso(d.get("created_at", ""))}
            for d in docs]


@router.get("/admin/backup/history/{bid}/download")
async def backup_history_download(bid: str, admin=Depends(get_current_admin)):
    from fastapi.responses import Response as _R
    db = await _get_db()
    d = await db.backup_history.find_one({"_id": _oid(bid)})
    if not d:
        raise HTTPException(status_code=404, detail="Backup not found")
    try:
        with open(d["path"], "rb") as f:
            blob = f.read()
    except FileNotFoundError:
        raise HTTPException(status_code=410, detail="File backup sudah tidak tersedia di disk")
    return _R(content=blob, media_type="application/gzip",
              headers={"Content-Disposition": f'attachment; filename="{d["filename"]}"'})


@router.get("/admin/backup/download")
async def backup_download(admin=Depends(get_current_admin)):
    """Download a full gzipped BSON archive of every collection.
    Streams via a plain `bytes` response - the archive is small enough
    for the ~1000-row datasets this portal carries."""
    from fastapi.responses import Response as _R
    try:
        blob, filename = await _run_mongodump()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup failed: {e}")
    return _R(
        content=blob,
        media_type="application/gzip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Backup-Size":       str(len(blob)),
            "Cache-Control":       "no-store",
        },
    )


@router.post("/admin/system/factory-reset")
async def system_factory_reset(payload: m.FactoryResetIn,
                               request: Request,
                               admin=Depends(get_current_admin)):
    """DANGER: wipe the database back to a fresh-install state.

    Behaviour (see /app/memory/PRD.md - user-approved scope):
      • Preserves the entire `settings` collection (branding + landing CMS).
      • Preserves ALL users whose `role == "admin"` (multiple admins survive).
      • Deletes every other document in every other collection.
      • System collections (`system.*`) and any index metadata are left alone.

    Guards (both required):
      1. `admin_password` must match the calling admin's current password.
      2. `confirm` body field must be the exact string "FACTORY RESET".

    Returns a per-collection summary of documents removed so the operator
    can see exactly what was purged. The admin's session token stays valid
    because the admin user document itself is preserved.
    """
    if payload.confirm != "FACTORY RESET":
        raise HTTPException(
            status_code=400,
            detail='Confirmation phrase mismatch. Type "FACTORY RESET" exactly.',
        )

    db = await _get_db()

    # ---- Verify admin password against the fresh DB record ----
    admin_doc = await db.users.find_one({"_id": ObjectId(admin["id"])})
    if not admin_doc or admin_doc.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin account not found")
    if not verify_password(payload.admin_password, admin_doc.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Admin password is incorrect")

    # ---- Take a safety snapshot BEFORE wiping (best-effort) ----
    safety_backup_path = None
    try:
        blob, filename = await _run_mongodump()
        import os as _os
        backup_dir = "/var/backups/intercloud"
        try:
            _os.makedirs(backup_dir, exist_ok=True)
            safety_backup_path = _os.path.join(backup_dir, f"pre-factory-reset-{filename}")
            with open(safety_backup_path, "wb") as fh:
                fh.write(blob)
        except PermissionError:
            # Fall back to /tmp when running without root (e.g., preview env)
            backup_dir = "/tmp"
            safety_backup_path = _os.path.join(backup_dir, f"pre-factory-reset-{filename}")
            with open(safety_backup_path, "wb") as fh:
                fh.write(blob)
    except Exception as e:  # nosec
        # Never block the reset on a backup failure, but surface it in logs.
        safety_backup_path = f"(snapshot skipped: {e})"

    # ---- Wipe ----
    PRESERVE_COLLECTIONS = {"settings"}
    summary: dict = {}

    all_names = await db.list_collection_names()
    for name in all_names:
        if name.startswith("system."):
            continue
        if name in PRESERVE_COLLECTIONS:
            continue
        if name == "users":
            # Purge everything EXCEPT admins.
            res = await db.users.delete_many({"role": {"$ne": "admin"}})
            summary[name] = {"deleted": res.deleted_count, "kept": "role==admin"}
            continue
        # Drop the whole collection (much faster than delete_many on big sets).
        try:
            count = await db[name].estimated_document_count()
        except Exception:
            count = None
        await db[name].drop()
        summary[name] = {"deleted": count, "dropped": True}

    # ---- Immutable audit trail: insert a single reset marker into the
    # freshly-wiped audit_logs collection so the operator can see WHO wiped
    # the system and WHEN, even after the reset dropped historical logs.
    await log_audit(db, actor=admin, action="system.factory_reset", category="system",
                    target_type="system", target_label="Factory Reset",
                    metadata={"safety_backup": str(safety_backup_path)[:400],
                              "collections_affected": len(summary)},
                    severity="critical", request=request)

    return {
        "ok": True,
        "message": "Factory reset complete. Admin account and settings preserved.",
        "safety_backup": safety_backup_path,
        "collections": summary,
    }


async def _log_factory_reset_after(db, admin, request, summary, backup_path):
    """Reserved (unused): factory-reset dropped audit_logs; a marker row is
    inserted inline before returning so the operator sees the trigger."""
    await log_audit(db, actor=admin, action="system.factory_reset", category="system",
                    target_type="system", target_label="Factory Reset",
                    metadata={"safety_backup": str(backup_path)[:400],
                              "collections_affected": len(summary or {})},
                    severity="critical", request=request)


@router.post("/admin/backup/restore")
async def backup_restore(request: Request, admin=Depends(get_current_admin), confirm: str = ""):
    """Restore a full snapshot. **Wipes every collection contained in the
    archive** (mongorestore --drop) and reinstates the uploaded content.

    Expects the archive as the raw request body (`Content-Type` is
    ignored; slugs like `application/gzip` or `application/octet-stream`
    both work). Requires `?confirm=REPLACE` as a safety guard."""
    if confirm != "REPLACE":
        raise HTTPException(status_code=400,
                            detail="Confirmation required: pass ?confirm=REPLACE")
    blob = await request.body()
    if not blob or len(blob) < 32:
        raise HTTPException(status_code=400, detail="Empty or too-small upload")
    try:
        log = await _run_mongorestore(blob, drop=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore failed: {e}")
    # Audit - restore also drops audit_logs, so the marker is the first row after.
    db = await _get_db()
    await log_audit(db, actor=admin, action="system.backup_restore", category="system",
                    target_type="system", target_label="Backup Restore",
                    metadata={"bytes_received": len(blob)},
                    severity="critical", request=request)
    return {
        "ok": True,
        "bytes_received": len(blob),
        "log_tail": log[-1200:],
    }


# ============================================================
# System update - runs scripts/update.sh which git-pulls, installs deps,
# rebuilds the frontend, and restarts supervisor. Auto-backs up first.
# ============================================================
@router.get("/admin/system/version")
async def system_version(admin=Depends(get_current_admin)):
    """Return current git SHA + short version info for the update UI."""
    import asyncio as _asyncio, os as _os
    repo_root = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), "..", "..", ".."))
    async def _run(cmd):
        p = await _asyncio.create_subprocess_exec(
            *cmd, cwd=repo_root,
            stdout=_asyncio.subprocess.PIPE, stderr=_asyncio.subprocess.PIPE,
        )
        out, err = await p.communicate()
        return (out.decode(errors="replace").strip() if p.returncode == 0 else "")
    sha    = await _run(["git", "rev-parse", "HEAD"])
    short  = await _run(["git", "rev-parse", "--short", "HEAD"])
    branch = await _run(["git", "rev-parse", "--abbrev-ref", "HEAD"])
    subject= await _run(["git", "log", "-1", "--pretty=%s"])
    date   = await _run(["git", "log", "-1", "--pretty=%cI"])
    return {
        "sha": sha or None,
        "short": short or None,
        "branch": branch or None,
        "subject": subject or None,
        "date": date or None,
        "repo_root": repo_root,
    }


@router.post("/admin/system/update")
async def system_update(admin=Depends(get_current_admin), confirm: str = ""):
    """Run `scripts/update.sh` in the checkout - auto-backs up first, then
    `git pull`, `pip install`, `yarn install && yarn build`, and restarts
    supervisor. **Preserves both .env files and the live database**.

    Guarded by `?confirm=UPDATE` so a stray click cannot trigger an update.
    Returns the STATUS line (`STATUS=ok OLD=<sha> NEW=<sha> BACKUP=<path>`)
    and the last ~2 KB of the script's log for diagnostics.

    Uses a filesystem lock at `/tmp/intercloud-update.lock` so two concurrent
    clicks return 409 instead of racing two `bash update.sh` invocations."""
    if confirm != "UPDATE":
        raise HTTPException(status_code=400,
                            detail="Confirmation required: pass ?confirm=UPDATE")
    import asyncio as _asyncio, os as _os, fcntl as _fcntl
    repo_root = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), "..", "..", ".."))
    script = _os.path.join(repo_root, "scripts", "update.sh")
    if not _os.path.isfile(script):
        raise HTTPException(status_code=500,
                            detail=f"update.sh not found at {script}")

    lock_path = "/tmp/intercloud-update.lock"
    lock_fd = _os.open(lock_path, _os.O_CREAT | _os.O_RDWR, 0o644)
    try:
        try:
            _fcntl.flock(lock_fd, _fcntl.LOCK_EX | _fcntl.LOCK_NB)
        except BlockingIOError:
            raise HTTPException(status_code=409,
                                detail="Another update is already running.")

        proc = await _asyncio.create_subprocess_exec(
            "/bin/bash", script,
            cwd=repo_root,
            stdout=_asyncio.subprocess.PIPE,
            stderr=_asyncio.subprocess.STDOUT,
        )
        try:
            stdout_b, _ = await _asyncio.wait_for(proc.communicate(), timeout=600)
        except _asyncio.TimeoutError:
            proc.kill()
            raise HTTPException(status_code=504,
                                detail="Update timed out after 10 min")
        text = stdout_b.decode(errors="replace")
        status_line = next((l for l in text.splitlines() if l.startswith("STATUS=")), "")
        rc = proc.returncode

        # Map well-known exit codes to distinct 4xx statuses so the UI can
        # render a helpful message instead of a raw stderr traceback:
        # 0 = ok/noop | 2 = backup failed | 3 = dirty tree | 4 = no remote
        if rc == 0:
            return {"ok": True, "status": status_line, "return_code": 0,
                    "log_tail": text[-2400:]}
        if rc == 3:
            raise HTTPException(status_code=409,
                                detail="Working tree has uncommitted changes; "
                                       "commit or reset before updating. "
                                       f"Log: {text[-800:]}")
        if rc == 4:
            raise HTTPException(status_code=422,
                                detail="This checkout has no git remote - "
                                       "cannot update. Deploy a proper git "
                                       "clone (see docs/production.md).")
        raise HTTPException(status_code=500,
                            detail=f"update.sh exited {rc}: {text[-800:]}")
    finally:
        try: _fcntl.flock(lock_fd, _fcntl.LOCK_UN)
        except Exception: pass
        try: _os.close(lock_fd)
        except Exception: pass


# ============================================================
# AUDIT LOGS - read-only history of sensitive admin actions
# ============================================================
@router.get("/admin/audit-logs")
async def admin_audit_logs_list(
    admin=Depends(get_current_admin),
    limit: int = 200,
    skip: int = 0,
    category: Optional[str] = None,
    action: Optional[str] = None,
    actor_id: Optional[str] = None,
    severity: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Paginated list of audit rows, newest first.

    Filters (all optional):
      • `category` - one of security/billing/integrations/users/system/noc
      • `action`   - exact action key (e.g. "user.role_change")
      • `actor_id` - filter to a specific admin's actions
      • `severity` - info/warning/critical
      • `q`        - case-insensitive substring on actor_email/target_label
      • `date_from`, `date_to` - ISO date strings (inclusive)
    """
    db = await _get_db()
    limit = max(1, min(int(limit or 200), 500))
    skip = max(0, int(skip or 0))
    query: dict = {}
    if category:  query["category"] = category
    if action:    query["action"] = action
    if severity:  query["severity"] = severity
    if actor_id:
        try:
            query["actor_id"] = ObjectId(actor_id)
        except Exception:
            query["actor_id"] = None
    if q:
        needle = q.strip()
        if needle:
            query["$or"] = [
                {"actor_email": {"$regex": needle, "$options": "i"}},
                {"target_label": {"$regex": needle, "$options": "i"}},
                {"action": {"$regex": needle, "$options": "i"}},
            ]
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to + "T23:59:59"
        query["created_at"] = rng
    total = await db.audit_logs.count_documents(query)
    cur = db.audit_logs.find(query).sort("created_at", -1).skip(skip).limit(limit)
    docs = [d async for d in cur]
    return {
        "total": total,
        "limit": limit,
        "skip": skip,
        "items": [_serialize_audit(d) for d in docs],
    }


@router.get("/admin/audit-logs/facets")
async def admin_audit_logs_facets(admin=Depends(get_current_admin)):
    """Distinct values available to power the filter dropdowns."""
    db = await _get_db()
    return {
        "categories": await db.audit_logs.distinct("category"),
        "actions": await db.audit_logs.distinct("action"),
        "severities": ["info", "warning", "critical"],
    }


# ============================================================
# EXECUTIVE OVERVIEW - read-only for owner/admin
# ============================================================
async def _get_current_owner(user=Depends(get_current_user)):
    """Access gate: owner or admin only. Owner is READ-ONLY globally."""
    if user.get("role") not in ("admin", "owner"):
        raise HTTPException(status_code=403, detail="Executive access only")
    return user


def _relative_months(n: int):
    """Return a timedelta-like approximation of N months (30d each)."""
    return timedelta(days=30 * n)


@router.get("/admin/owner/overview")
async def owner_overview(owner=Depends(_get_current_owner)):
    """Aggregate MRR/ARPU/churn/uptime/SLA into a single dashboard payload."""
    db = await _get_db()
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # ---- MRR: sum of active-service monthly prices ----
    services = await db.services.find({"status": "active"}).to_list(5000)
    mrr = sum(float(s.get("price_monthly") or 0) for s in services)
    active_services_count = len(services)
    total_clients = await db.users.count_documents({"role": "client"})
    clients_with_active = len({str(s["user_id"]) for s in services if s.get("user_id")})
    arpu = round(mrr / clients_with_active, 2) if clients_with_active else 0.0

    # ---- Churn: services terminated in last 30 days ----
    thirty_ago = (now - timedelta(days=30)).isoformat()
    churned = await db.services.count_documents({
        "status": "terminated", "terminated_at": {"$gte": thirty_ago},
    })
    churn_pct = round((churned / (churned + active_services_count)) * 100, 2) \
                if (churned + active_services_count) else 0.0

    # ---- Revenue: MTD + last 12 months trend ----
    paid_month = await db.invoices.find({"status": "paid",
                                         "paid_at": {"$gte": month_start.isoformat()}}).to_list(5000)
    revenue_month = sum(float(d.get("total") or 0) for d in paid_month)
    trend = []
    for i in range(11, -1, -1):
        start = month_start - _relative_months(i)
        end = start + _relative_months(1)
        docs = await db.invoices.find({
            "status": "paid",
            "paid_at": {"$gte": start.isoformat(), "$lt": end.isoformat()},
        }).to_list(5000)
        trend.append({"period": start.strftime("%Y-%m"),
                      "revenue": sum(float(d.get("total") or 0) for d in docs),
                      "invoices": len(docs)})

    # ---- Outstanding & overdue ----
    overdue_docs = await db.invoices.find({"status": "overdue"}).to_list(2000)
    unpaid_count = await db.invoices.count_documents({"status": {"$in": ["unpaid", "overdue"]}})
    overdue_total = sum(float(d.get("total") or 0) for d in overdue_docs)

    # ---- NOC uptime: fleet-wide 24h / 7d ----
    since_24h = (now - timedelta(hours=24)).isoformat()
    total_samples = await db.noc_probes.count_documents({"at": {"$gte": since_24h}})
    up_samples = await db.noc_probes.count_documents({"at": {"$gte": since_24h}, "ok": True})
    uptime_24h_pct = round((up_samples / total_samples) * 100, 2) if total_samples else None
    since_7d = (now - timedelta(days=7)).isoformat()
    samples_7d = await db.noc_probes.count_documents({"at": {"$gte": since_7d}})
    up_7d = await db.noc_probes.count_documents({"at": {"$gte": since_7d}, "ok": True})
    uptime_7d_pct = round((up_7d / samples_7d) * 100, 2) if samples_7d else None
    devices_down = await db.noc_device_state.count_documents({"status": "down"})
    devices_total = await db.mikrotik_devices.count_documents({})
    # SLA: rough outage minutes = down samples * 5-min cadence
    down_samples_30d = await db.noc_probes.count_documents({
        "at": {"$gte": (now - timedelta(days=30)).isoformat()}, "ok": False,
    })
    outage_minutes_30d = down_samples_30d * 5

    # ---- Ticket load ----
    open_tickets = await db.tickets.count_documents({"status": {"$in": ["open", "awaiting_staff"]}})
    critical_tickets = await db.tickets.count_documents({"status": {"$nin": ["resolved", "closed"]},
                                                          "priority": "critical"})

    # ---- Top clients by lifetime revenue ----
    paid_all = await db.invoices.find({"status": "paid"}).to_list(20000)
    per_client: dict = {}
    for d in paid_all:
        uid = str(d.get("user_id"))
        if not uid or uid == "None":
            continue
        per_client[uid] = per_client.get(uid, 0.0) + float(d.get("total") or 0)
    top_pairs = sorted(per_client.items(), key=lambda x: x[1], reverse=True)[:5]
    top_uids = []
    for uid, _rev in top_pairs:
        try: top_uids.append(ObjectId(uid))
        except Exception: pass
    top_users = {str(u["_id"]): u for u in await db.users.find(
        {"_id": {"$in": top_uids}}).to_list(20)} if top_uids else {}
    top_clients = [{
        "user_id": uid,
        "name": (top_users.get(uid) or {}).get("name", ""),
        "email": (top_users.get(uid) or {}).get("email", ""),
        "lifetime_revenue": rev,
    } for uid, rev in top_pairs]

    return {
        "generated_at": _now(),
        "mrr": mrr,
        "arr": mrr * 12,
        "arpu": arpu,
        "churn_pct_30d": churn_pct,
        "clients_total": total_clients,
        "clients_with_active_service": clients_with_active,
        "active_services": active_services_count,
        "revenue_month_to_date": revenue_month,
        "revenue_trend_12m": trend,
        "unpaid_invoices": unpaid_count,
        "overdue_total": overdue_total,
        "noc": {
            "uptime_24h_pct": uptime_24h_pct,
            "uptime_7d_pct": uptime_7d_pct,
            "devices_total": devices_total,
            "devices_down": devices_down,
            "outage_minutes_30d": outage_minutes_30d,
            "samples_24h": total_samples,
        },
        "support": {
            "open_tickets": open_tickets,
            "critical_open": critical_tickets,
        },
        "top_clients": top_clients,
    }


# ---------- System Health (Diagnostics - System Health + verifikasi pasca-install) ----------
import time as _hc_time  # noqa: E402


import shutil as _hc_shutil  # noqa: E402


import ssl as _hc_ssl  # noqa: E402


import socket as _hc_socket  # noqa: E402


import subprocess as _hc_subprocess  # noqa: E402


import asyncio as _hc_asyncio  # noqa: E402


_PROCESS_STARTED_AT = datetime.now(timezone.utc)


def _hc_row(key, name, status, detail, **metrics):
    return {"key": key, "name": name, "status": status, "detail": detail, "metrics": metrics}


def _hc_ssl_days_left(host: str, port: int = 443, timeout: float = 6.0):
    ctx = _hc_ssl.create_default_context()
    with _hc_socket.create_connection((host, port), timeout=timeout) as sock:
        with ctx.wrap_socket(sock, server_hostname=host) as ssock:
            cert = ssock.getpeercert()
    expires = datetime.strptime(cert["notAfter"], "%b %d %H:%M:%S %Y %Z").replace(tzinfo=timezone.utc)
    return (expires - datetime.now(timezone.utc)).days, expires.date().isoformat()


def _hc_systemd_state(unit: str) -> str:
    try:
        r = _hc_subprocess.run(["systemctl", "is-active", unit],
                               capture_output=True, text=True, timeout=5)
        state = (r.stdout or "").strip()
        if not state or "not been booted" in (r.stderr or ""):
            return "unavailable"
        return state
    except Exception:
        return "unavailable"


@router.get("/admin/system/health")
async def admin_system_health(admin=Depends(require_roles("admin", "support"))):
    """Kesehatan live host + aplikasi: DB, disk, memori, CPU, SSL, scheduler,
    service systemd. Dipakai di Admin - Diagnostics - System Health dan untuk
    verifikasi setelah scripts/install.sh selesai."""
    db = await _get_db()
    checks = []

    try:
        t0 = _hc_time.perf_counter()
        await db.command("ping")
        latency_ms = round((_hc_time.perf_counter() - t0) * 1000, 1)
        data_mb = None
        try:
            stats = await db.command("dbStats")
            data_mb = round((stats.get("dataSize") or 0) / (1024 * 1024), 1)
        except Exception:  # noqa: BLE001
            pass
        detail = f"Ping {latency_ms} ms"
        if data_mb is not None:
            detail += f" - data {data_mb} MB"
        checks.append(_hc_row("database", "MongoDB", "ok" if latency_ms < 250 else "warn",
                              detail, latency_ms=latency_ms, data_mb=data_mb))
    except Exception as e:  # noqa: BLE001
        checks.append(_hc_row("database", "MongoDB", "fail", f"Ping gagal: {e}"))

    try:
        du = _hc_shutil.disk_usage("/")
        pct = round(du.used / du.total * 100, 1)
        free_gb = round(du.free / 1024 ** 3, 1)
        total_gb = round(du.total / 1024 ** 3, 1)
        status = "ok" if pct < 80 else ("warn" if pct < 90 else "fail")
        checks.append(_hc_row("disk", "Disk", status,
                              f"{pct}% terpakai - {free_gb} GB bebas dari {total_gb} GB",
                              used_percent=pct, free_gb=free_gb, total_gb=total_gb))
    except Exception as e:  # noqa: BLE001
        checks.append(_hc_row("disk", "Disk", "fail", str(e)))

    try:
        mem = {}
        with open("/proc/meminfo") as f:
            for line in f:
                k, _, v = line.partition(":")
                if v.strip():
                    mem[k.strip()] = int(v.strip().split()[0])
        total_mb = round(mem["MemTotal"] / 1024)
        avail_mb = round(mem.get("MemAvailable", 0) / 1024)
        pct = round((total_mb - avail_mb) / total_mb * 100, 1) if total_mb else 0
        status = "ok" if pct < 85 else ("warn" if pct < 95 else "fail")
        checks.append(_hc_row("memory", "Memori", status,
                              f"{pct}% terpakai - {avail_mb} MB tersedia dari {total_mb} MB",
                              used_percent=pct, available_mb=avail_mb, total_mb=total_mb))
    except Exception as e:  # noqa: BLE001
        checks.append(_hc_row("memory", "Memori", "warn", f"Tidak terbaca: {e}"))

    try:
        load1, load5, load15 = os.getloadavg()
        cores = os.cpu_count() or 1
        ratio = load5 / cores
        status = "ok" if ratio < 0.8 else ("warn" if ratio < 1.5 else "fail")
        checks.append(_hc_row("cpu", "CPU Load", status,
                              f"load {load1:.2f} / {load5:.2f} / {load15:.2f} ({cores} core)",
                              load1=round(load1, 2), load5=round(load5, 2),
                              load15=round(load15, 2), cores=cores))
    except Exception as e:  # noqa: BLE001
        checks.append(_hc_row("cpu", "CPU Load", "warn", f"Tidak terbaca: {e}"))

    origin = (os.environ.get("REACT_APP_BACKEND_URL") or "").strip().rstrip("/")
    if origin.startswith("https://"):
        host = origin[8:].split("/")[0].split(":")[0]
        try:
            days, exp = await _hc_asyncio.wait_for(
                _hc_asyncio.to_thread(_hc_ssl_days_left, host), timeout=10)
            status = "ok" if days > 21 else ("warn" if days > 7 else "fail")
            checks.append(_hc_row("ssl", "SSL / HTTPS", status,
                                  f"{host} - sisa {days} hari (exp {exp})",
                                  days_left=days, host=host, expires=exp))
        except Exception as e:  # noqa: BLE001
            checks.append(_hc_row("ssl", "SSL / HTTPS", "warn",
                                  f"{host} - cek gagal: {type(e).__name__}", host=host))
    else:
        checks.append(_hc_row("ssl", "SSL / HTTPS", "off",
                              "Portal berjalan tanpa HTTPS (HTTP only)"))

    try:
        from portal import emails as _hc_emails
        sched = getattr(_hc_emails, "_scheduler", None)
        if sched is not None and getattr(sched, "running", False):
            jobs = len(sched.get_jobs())
            checks.append(_hc_row("scheduler", "Scheduler", "ok",
                                  f"Aktif - {jobs} job terjadwal (reminder, laporan, backup)",
                                  jobs=jobs))
        else:
            checks.append(_hc_row("scheduler", "Scheduler", "warn",
                                  "Tidak berjalan di proses ini"))
    except Exception as e:  # noqa: BLE001
        checks.append(_hc_row("scheduler", "Scheduler", "warn", str(e)))

    up = datetime.now(timezone.utc) - _PROCESS_STARTED_AT
    days_u, rem = divmod(int(up.total_seconds()), 86400)
    hrs, rem = divmod(rem, 3600)
    mins = rem // 60
    if days_u:
        uptxt = f"{days_u} hari {hrs} jam"
    elif hrs:
        uptxt = f"{hrs} jam {mins} menit"
    else:
        uptxt = f"{mins} menit"
    checks.append(_hc_row("uptime", "Backend Uptime", "ok",
                          f"Berjalan {uptxt} (sejak {_PROCESS_STARTED_AT.strftime('%Y-%m-%d %H:%M UTC')})",
                          seconds=int(up.total_seconds())))

    for unit, label in (("mongod", "Service mongod"), ("nginx", "Service nginx")):
        state = await _hc_asyncio.to_thread(_hc_systemd_state, unit)
        if state == "unavailable":
            continue
        checks.append(_hc_row(f"svc_{unit}", label,
                              "ok" if state == "active" else "fail",
                              f"systemd: {state}", state=state))

    overall = ("fail" if any(c["status"] == "fail" for c in checks)
               else "warn" if any(c["status"] == "warn" for c in checks) else "ok")
    return {"generated_at": datetime.now(timezone.utc).isoformat(),
            "overall": overall, "checks": checks}

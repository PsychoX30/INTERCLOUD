"""Transaction ledger for invoice payment lifecycle and reconciliation."""
from datetime import datetime, timezone
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException

from .. import models as m
from ..auth import require_roles
from .shared import _get_db, _iso, _oid, _sales_scope_filter, _pagination_params, _pagination_response
from .users import _paginate

router = APIRouter()


def _parse_dt(value: Optional[str], field: str):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid {field} date")


def _serialize_transaction(d: dict) -> dict:
    verified_at = d.get("verified_at")
    return {
        "id": str(d["_id"]),
        "invoice_id": str(d["invoice_id"]) if d.get("invoice_id") else None,
        "invoice_number": d.get("invoice_number"),
        "user_id": str(d["user_id"]) if d.get("user_id") else None,
        "user_name": d.get("user_name"),
        "customer_name": d.get("customer_name", ""),
        "amount": float(d.get("amount") or 0),
        "method": d.get("method"),
        "status": d.get("status", "pending"),
        "paid_at": _iso(d.get("paid_at")) if d.get("paid_at") else None,
        "verified": bool(verified_at),
        "verified_at": _iso(verified_at) if verified_at else None,
        "verified_by": str(d.get("verified_by")) if d.get("verified_by") else None,
        "invoice_date": _iso(d.get("invoice_date")) if d.get("invoice_date") else None,
        "due_date": _iso(d.get("due_date")) if d.get("due_date") else None,
        "reference": d.get("reference"),
        "notes": d.get("notes", ""),
        "source": d.get("source", "auto"),
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }


@router.get("/admin/transactions")
async def list_transactions(
    staff=Depends(require_roles("admin", "finance", "sales")),
    user_id: Optional[str] = None,
    status: Optional[str] = None,
    method: Optional[str] = None,
    search: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    page: Optional[int] = None,
    limit: int = 25,
    skip: int = 0,
    sort: str = "created_at",
    order: str = "desc",
    paginate: bool = False,
):
    db = await _get_db()
    query = _sales_scope_filter(staff, key="user_id")
    if user_id:
        uid = _oid(user_id)
        if staff.get("role") == "sales" and uid not in [ObjectId(x) for x in staff.get("assigned_client_ids") or []]:
            return _pagination_response([], 0, 0, limit, True) if paginate else []
        query["user_id"] = uid
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        query["status"] = {"$in": statuses} if len(statuses) > 1 else statuses[0]
    if method:
        methods = [m.strip() for m in method.split(",") if m.strip()]
        query["method"] = {"$in": methods} if len(methods) > 1 else methods[0]
    if search:
        query["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"user_name": {"$regex": search, "$options": "i"}},
            {"reference": {"$regex": search, "$options": "i"}},
        ]
    start_dt, end_dt = _parse_dt(start, "start"), _parse_dt(end, "end")
    if start_dt or end_dt:
        query["created_at"] = {}
        if start_dt:
            query["created_at"]["$gte"] = start_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        if end_dt:
            query["created_at"]["$lte"] = end_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    sort_field = sort if sort in {"created_at", "amount", "status", "method"} else "created_at"
    direction = 1 if order.lower() == "asc" else -1
    if paginate:
        # When paginate=True, map page (1-based) to skip if skip=0 and page provided
        if page is not None and skip == 0:
            skip = max(0, (int(page) - 1) * limit)
        skip_n, limit_n = _pagination_params(skip, limit)
        total = await db.transactions.count_documents(query)
        cursor = db.transactions.find(query).sort(sort_field, direction).skip(skip_n)
        if limit_n is not None:
            cursor = cursor.limit(limit_n)
        docs = await cursor.to_list(limit_n or 5000)
        items = [_serialize_transaction(d) for d in docs]
        return _pagination_response(items, total, skip_n, limit_n, True)
    # Legacy bare-array behavior. When page is provided, slice the page-sized window
    # with limit; otherwise return the full list. _paginate returns the original list
    # when page is None (compatible with old consumers like DDoSPanel).
    docs = await db.transactions.find(query).sort(sort_field, direction).to_list(5000)
    items = [_serialize_transaction(d) for d in docs]
    if page is not None:
        page_n = max(1, int(page))
        limit_n = max(1, min(int(limit), 200))
        offset = (page_n - 1) * limit_n
        return items[offset:offset + limit_n]
    return items


@router.get("/admin/transactions/summary")
async def transaction_summary(
    staff=Depends(require_roles("admin", "finance", "sales")),
    status: Optional[str] = None,
    method: Optional[str] = None,
    search: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
):
    """Reconciliation totals for transactions visible to the current staff role."""
    db = await _get_db()
    query = _sales_scope_filter(staff, key="user_id")
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        query["status"] = {"$in": statuses} if len(statuses) > 1 else statuses[0]
    if method:
        methods = [m.strip() for m in method.split(",") if m.strip()]
        query["method"] = {"$in": methods} if len(methods) > 1 else methods[0]
    if search:
        query["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"user_name": {"$regex": search, "$options": "i"}},
            {"reference": {"$regex": search, "$options": "i"}},
        ]
    start_dt, end_dt = _parse_dt(start, "start"), _parse_dt(end, "end")
    if start_dt or end_dt:
        query["created_at"] = {}
        if start_dt:
            query["created_at"]["$gte"] = start_dt.isoformat()
        if end_dt:
            query["created_at"]["$lte"] = end_dt.isoformat()

    rows = await db.transactions.aggregate([
        {"$match": query},
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
            "amount": {"$sum": "$amount"},
        }},
    ]).to_list(100)
    by_status = {
        row["_id"]: {"count": int(row["count"]), "amount": float(row["amount"] or 0)}
        for row in rows
    }
    paid = by_status.get("paid", {"count": 0, "amount": 0.0})
    outstanding = sum(
        float(row["amount"] or 0)
        for row in rows if row["_id"] in {"pending", "unpaid"}
    )
    return {
        "paid_count": paid["count"],
        "paid_amount": paid["amount"],
        "outstanding_amount": outstanding,
        "by_status": by_status,
    }


@router.get("/admin/transactions/export/xlsx")
async def export_transactions(
    staff=Depends(require_roles("admin", "finance", "sales")),
    user_id: Optional[str] = None,
    status: Optional[str] = None,
    method: Optional[str] = None,
    search: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
):
    """Export filtered transaction ledger as an Excel workbook with formulas."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    db = await _get_db()
    query = _sales_scope_filter(staff, key="user_id")
    if user_id:
        uid = _oid(user_id)
        if staff.get("role") == "sales" and uid not in [ObjectId(x) for x in staff.get("assigned_client_ids") or []]:
            query["user_id"] = None
        else:
            query["user_id"] = uid
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        query["status"] = {"$in": statuses} if len(statuses) > 1 else statuses[0]
    if method:
        methods = [m.strip() for m in method.split(",") if m.strip()]
        query["method"] = {"$in": methods} if len(methods) > 1 else methods[0]
    if search:
        query["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"user_name": {"$regex": search, "$options": "i"}},
            {"reference": {"$regex": search, "$options": "i"}},
        ]
    start_dt, end_dt = _parse_dt(start, "start"), _parse_dt(end, "end")
    if start_dt or end_dt:
        query["created_at"] = {}
        if start_dt:
            query["created_at"]["$gte"] = start_dt.isoformat()
        if end_dt:
            query["created_at"]["$lte"] = end_dt.isoformat()

    docs = await db.transactions.find(query).sort("created_at", -1).to_list(5000)
    rows = [_serialize_transaction(d) for d in docs]

    RP = '"Rp" #,##0'
    DATE = "yyyy-mm-dd hh:mm"
    thin = Side(style="thin", color="FFD8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF0A2350")
    tot_font = Font(bold=True)
    tot_fill = PatternFill("solid", fgColor="FFFEF3C7")

    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Ledger"

    headers = [
        "No", "Invoice #", "Customer", "User", "Amount", "Method",
        "Status", "Reference", "Paid At", "Verified", "Invoice Date",
        "Due Date", "Source", "Notes",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"

    def _excel_dt(iso_str):
        if not iso_str:
            return None
        try:
            dt = datetime.fromisoformat(str(iso_str).replace("Z", "+00:00"))
            return dt.replace(tzinfo=None)
        except Exception:
            return None

    for idx, r in enumerate(rows, start=1):
        ws.append([
            idx,
            r.get("invoice_number", ""),
            r.get("customer_name", ""),
            r.get("user_name", ""),
            float(r.get("amount") or 0),
            (r.get("method") or "").replace("_", " ").title(),
            (r.get("status") or "").upper(),
            r.get("reference", ""),
            _excel_dt(r.get("paid_at")),
            "Yes" if r.get("verified") else "No",
            _excel_dt(r.get("invoice_date")),
            _excel_dt(r.get("due_date")),
            r.get("source", ""),
            r.get("notes", ""),
        ])

    # Style body + formulas
    n = len(rows)
    total_row = n + 2
    if n:
        ws.append(["", "TOTAL", "", "", f"=SUM(E2:E{total_row - 1})", "", "", "", "", "", "", "", "", ""])
    else:
        ws.append(["", "TOTAL", "", "", 0, "", "", "", "", "", "", "", "", ""])

    for r in range(2, total_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c == 1:
                cell.alignment = Alignment(horizontal="center")
            elif c == 5:
                cell.number_format = RP
            elif c == 9:
                cell.number_format = DATE
            elif c == 11 or c == 12:
                cell.number_format = DATE
            if r == total_row:
                cell.font = tot_font
                cell.fill = tot_fill

    # Status summary table on the right (columns P-T)
    summary_headers = ["Status", "Count", "Amount"]
    ws.cell(row=1, column=16, value="Summary by Status")
    ws.cell(row=1, column=16).font = Font(bold=True, size=12, color="FF0A2350")
    ws.merge_cells(start_row=1, start_column=16, end_row=1, end_column=18)
    for c, h in enumerate(summary_headers, start=16):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    status_order = ["paid", "pending", "unpaid", "failed", "refunded", "cancelled"]
    status_counts = {}
    status_amounts = {}
    for r in rows:
        s = r.get("status", "unknown")
        status_counts[s] = status_counts.get(s, 0) + 1
        status_amounts[s] = status_amounts.get(s, 0) + float(r.get("amount") or 0)
    sr = 3
    for s in status_order:
        if s not in status_counts:
            continue
        ws.cell(row=sr, column=16, value=s.upper())
        ws.cell(row=sr, column=17, value=status_counts[s])
        ws.cell(row=sr, column=18, value=status_amounts[s])
        ws.cell(row=sr, column=18).number_format = RP
        for c in range(16, 19):
            ws.cell(row=sr, column=c).border = border
        sr += 1
    ws.cell(row=sr, column=16, value="TOTAL")
    ws.cell(row=sr, column=16).font = tot_font
    ws.cell(row=sr, column=17, value=f"=SUM(Q3:Q{sr - 1})" if sr > 3 else 0)
    ws.cell(row=sr, column=18, value=f"=SUM(R3:R{sr - 1})" if sr > 3 else 0)
    ws.cell(row=sr, column=18).number_format = RP
    for c in range(16, 19):
        ws.cell(row=sr, column=c).border = border
        ws.cell(row=sr, column=c).fill = tot_fill

    # Autofit
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Transaction_Ledger_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/admin/transactions/{tid}")
async def get_transaction(tid: str, staff=Depends(require_roles("admin", "finance", "sales"))):
    db = await _get_db()
    doc = await db.transactions.find_one({"_id": _oid(tid), **_sales_scope_filter(staff, key="user_id")})
    if not doc:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return _serialize_transaction(doc)


@router.post("/admin/transactions", response_model=m.TransactionOut)
async def create_transaction(payload: m.TransactionIn, staff=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    invoice_id = _oid(payload.invoice_id) if payload.invoice_id else None
    user_id = _oid(payload.user_id) if payload.user_id else None
    invoice = await db.invoices.find_one({"_id": invoice_id}) if invoice_id else None
    if payload.invoice_id and not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice and not user_id:
        user_id = invoice.get("user_id")
    user = await db.users.find_one({"_id": user_id}) if user_id else None
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "invoice_id": invoice_id,
        "invoice_number": (invoice or {}).get("number"),
        "user_id": user_id,
        "user_name": (user or {}).get("name"),
        "customer_name": payload.customer_name or (user or {}).get("name", ""),
        "amount": float(payload.amount or (invoice or {}).get("total") or 0),
        "method": payload.method or "manual",
        "status": payload.status,
        "paid_at": payload.paid_at or (now if payload.status == "paid" else None),
        "verified_at": payload.verified_at,
        "invoice_date": payload.invoice_date or (invoice or {}).get("created_at"),
        "due_date": payload.due_date or (invoice or {}).get("due_date"),
        "reference": payload.reference,
        "notes": payload.notes,
        "source": "manual",
        "created_at": now,
        "updated_at": now,
    }
    result = await db.transactions.insert_one(doc)
    doc["_id"] = result.inserted_id
    return _serialize_transaction(doc)


@router.post("/admin/transactions/{tid}/verify", response_model=m.TransactionOut)
async def verify_transaction(tid: str, payload: m.TransactionVerifyIn, staff=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    now = datetime.now(timezone.utc).isoformat()
    result = await db.transactions.update_one(
        {"_id": _oid(tid)},
        {"$set": {"verified_at": now, "verified_by": staff.get("id"), "updated_at": now},
         "$setOnInsert": {},
         **({"$push": {"verification_notes": payload.notes}} if payload.notes else {})},
    )
    if not result.matched_count:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return _serialize_transaction(await db.transactions.find_one({"_id": _oid(tid)}))

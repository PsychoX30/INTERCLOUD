"""Business ops: CRM, projects, content planner, follow-ups, documents, media library, content calendar.

Split from the former monolithic routes.py - behavior preserved 1:1.
"""
import os
import asyncio
import logging
import secrets
import re
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Request, WebSocket
from bson import ObjectId
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from .. import models as m
from ..auth import (
    verify_password, hash_password, create_access_token,
    get_current_user, get_current_admin, get_current_staff, get_current_content,
    require_roles, sales_can_access,
    STAFF_ROLES, FINANCE_ROLES, BILLING_ROLES, CATALOG_ROLES,
    OPS_ROLES, USER_MGMT_ROLES, TICKET_ROLES, CONTENT_ROLES,
)
from ..audit import log_audit, serialize as _serialize_audit
from ..secretbox import (dec_value as _sb_dec, enc_value as _sb_enc,
                         decrypt_config as _sb_dec_config)
from .. import integrations_v2 as iv2
from .shared import (_get_db, _iso, _now, _oid, _sales_scope_filter, _sales_visible_crm_ids,
                     _pagination_params, _pagination_response)  # noqa: E402
from .tickets import _deny_creative  # noqa: E402

router = APIRouter()


# ============================================================
# BUSINESS - CRM, Projects, Content Planner, Follow-ups, Documents
# ============================================================

# ---------- CRM (customers/prospects) ----------
def _serialize_crm(d):
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "email": d.get("email", ""),
        "phone": d.get("phone", ""),
        "company": d.get("company", ""),
        "position": d.get("position", ""),
        "industry": d.get("industry", ""),
        "status": d.get("status", "prospect"),
        "notes": d.get("notes", ""),
        "user_id": str(d["user_id"]) if d.get("user_id") else None,
        "assigned_to": str(d["assigned_to"]) if d.get("assigned_to") else None,
        "assigned_at": _iso(d.get("assigned_at", "")),
        "assignment_expires_at": _iso(d.get("assignment_expires_at", "")),
        "source": d.get("source", ""),
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }


def _staff_oid(staff: dict):
    """Extract staff ObjectId from either _id (test fixture) or id (real auth).

    Real JWT auth (auth.py line 95) pops _id and creates id string.
    Test fixtures use _id directly. This helper handles both.
    """
    if "_id" in staff:
        return ObjectId(str(staff["_id"]))
    if "id" in staff:
        return ObjectId(str(staff["id"]))
    return None


def _crm_sales_scope_filter(staff: dict) -> dict:
    """Mongo filter restricting CRM visibility for sales.

    Sales can see:
    - clients assigned to them (user_id in assigned_client_ids)
    - the shared lead pool: status prospect/partnership (unclaimed)
    - prospects they currently own (status='assigned', assigned_to=staff id)

    Returns {} for non-sales (no restriction). This is deliberately separate
    from the strict _sales_scope_filter used by billing/orders/tickets.
    """
    if staff.get("role") != "sales":
        return {}
    clauses = []
    assigned = [ObjectId(x) for x in (staff.get("assigned_client_ids") or [])]
    if assigned:
        clauses.append({"user_id": {"$in": assigned}})
    clauses.append({"status": {"$in": ["prospect", "partnership"]}})
    staff_id = _staff_oid(staff)
    if staff_id:
        clauses.append({"assigned_to": staff_id})
    if not clauses:
        return {"_id": None}  # matches nothing
    return {"$or": clauses}



# Order statuses that count as "in-progress" (needs attention) vs "won" vs "closed"
ORDER_TERMINAL_LOST = {"rejected", "cancelled"}


ORDER_IN_PROGRESS = {"pending", "pending_payment", "awaiting_verification",
                     "awaiting_quote", "payment_verified", "assigned", "provisioning"}


ORDER_WON = {"active"}


async def _crm_enrichment_by_uid(db, user_ids: list) -> dict:
    """Return {user_id_str: {latest_order, active_orders_count, lifetime_value, in_progress_count}}
    for the given user IDs, in one round-trip per collection."""
    if not user_ids:
        return {}
    result = {}
    # ---- Orders (grouped in-memory: small dataset per tenant) ----
    orders_cur = db.orders.find(
        {"user_id": {"$in": user_ids}},
        {"user_id": 1, "status": 1, "created_at": 1, "product_name": 1,
         "invoice_id": 1, "config": 1},
    ).sort("created_at", -1)
    async for o in orders_cur:
        key = str(o["user_id"])
        bucket = result.setdefault(key, {
            "latest_order": None,
            "active_orders_count": 0,
            "in_progress_count": 0,
            "won_orders_count": 0,
            "lifetime_value": 0.0,
        })
        if bucket["latest_order"] is None:
            bucket["latest_order"] = {
                "id": str(o["_id"]),
                "status": o.get("status", "pending"),
                "product_name": o.get("product_name", ""),
                "created_at": _iso(o.get("created_at", "")),
                "invoice_id": str(o["invoice_id"]) if o.get("invoice_id") else None,
            }
        st = o.get("status", "pending")
        if st not in ORDER_TERMINAL_LOST:
            bucket["active_orders_count"] += 1
        if st in ORDER_IN_PROGRESS:
            bucket["in_progress_count"] += 1
        if st in ORDER_WON:
            bucket["won_orders_count"] += 1
    # ---- Paid invoices → lifetime value ----
    inv_cur = db.invoices.find(
        {"user_id": {"$in": user_ids}, "status": "paid"},
        {"user_id": 1, "total": 1, "number": 1},
    )
    async for inv in inv_cur:
        key = str(inv["user_id"])
        bucket = result.setdefault(key, {
            "latest_order": None,
            "active_orders_count": 0,
            "in_progress_count": 0,
            "won_orders_count": 0,
            "lifetime_value": 0.0,
        })
        try:
            bucket["lifetime_value"] += float(inv.get("total") or 0)
        except Exception:
            pass
    return result


@router.get("/admin/crm")
async def crm_list(staff=Depends(get_current_staff),
                   skip: int = 0, limit: int = 50, sort: str = "updated_at",
                   order: str = "desc", q: Optional[str] = None,
                   status: Optional[str] = None,
                   paginate: Optional[bool] = None):
    """Server-side pagination + q-search + status filter for CRM.
    Default response stays bare array."""
    _deny_creative(staff)
    db = await _get_db()
    query = _crm_sales_scope_filter(staff)
    if status and status != "all":
        query["status"] = status
    if q:
        query["$or"] = [
            {field: {"$regex": q.strip(), "$options": "i"}}
            for field in ("name", "email", "phone", "company")
        ]
    sort_field = sort if sort in {
        "name", "email", "company", "status", "created_at", "updated_at"
    } else "updated_at"
    direction = 1 if order.lower() == "asc" else -1
    skip_n, limit_n = _pagination_params(skip, limit)
    cursor = db.crm_customers.find(query).sort(sort_field, direction)
    total = 0
    if bool(paginate):
        total = await db.crm_customers.count_documents(query)
        cursor = cursor.skip(skip_n).limit(limit_n if limit_n is not None else 500)
        docs = await cursor.to_list(None)
    else:
        docs = await cursor.to_list(2000)
    # Collect user_ids for enrichment
    uids = [d.get("user_id") for d in docs if d.get("user_id")]
    enrich = await _crm_enrichment_by_uid(db, uids)
    out = []
    for d in docs:
        row = _serialize_crm(d)
        e = enrich.get(str(d.get("user_id"))) if d.get("user_id") else None
        row["latest_order"] = (e or {}).get("latest_order")
        row["active_orders_count"] = (e or {}).get("active_orders_count", 0)
        row["in_progress_count"] = (e or {}).get("in_progress_count", 0)
        row["won_orders_count"] = (e or {}).get("won_orders_count", 0)
        row["lifetime_value"] = (e or {}).get("lifetime_value", 0.0)
        # Warm-lead heuristic: any prospect / lead with an in-progress order,
        # OR an existing customer with a fresh in-progress order (upsell signal)
        row["is_warm"] = row["in_progress_count"] > 0
        out.append(row)
    if bool(paginate):
        return _pagination_response(out, total, skip_n, limit_n, True)
    return out


@router.post("/admin/crm")
async def crm_create(payload: dict, staff=Depends(get_current_staff)):
    """Create a CRM row.

    Sales may create:
    - prospects (no user_id, shared pool)
    - CRM rows linked to one of their assigned clients
    """
    db = await _get_db()
    _deny_creative(staff)
    user_id = payload.get("user_id")
    if staff.get("role") == "sales" and user_id:
        assigned = {str(client_id) for client_id in (staff.get("assigned_client_ids") or [])}
        if str(user_id) not in assigned:
            raise HTTPException(status_code=403, detail="CRM client is not assigned to you")
        client = await db.users.find_one({"_id": _oid(str(user_id)), "role": "client"})
        if not client:
            raise HTTPException(status_code=403, detail="CRM target must be an assigned client")

    doc = {
        "name": payload.get("name", ""),
        "email": (payload.get("email") or "").lower(),
        "phone": payload.get("phone", ""),
        "company": payload.get("company", ""),
        "position": payload.get("position", ""),
        "industry": payload.get("industry", ""),
        "status": payload.get("status", "prospect"),
        "notes": payload.get("notes", ""),
        "created_at": _now(),
        "updated_at": _now(),
    }
    if user_id:
        doc["user_id"] = _oid(str(user_id))
    r = await db.crm_customers.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_crm(doc)


async def _assert_sales_can_touch_crm(db, staff: dict, cid: str) -> dict:
    """Load a CRM row and 403 if `staff` is a sales user who cannot access it.

    Sales can touch:
    - prospects/partnerships in the shared pool
    - CRM rows linked to their assigned clients
    - prospects currently assigned to them (status='assigned', assigned_to=their id)
    """
    _deny_creative(staff)
    d = await db.crm_customers.find_one({"_id": _oid(cid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    if staff.get("role") == "sales":
        status = d.get("status", "prospect")
        assigned = {str(x) for x in (staff.get("assigned_client_ids") or [])}
        staff_id = _staff_oid(staff)
        assigned_to = d.get("assigned_to")
        is_shared = status in ("prospect", "partnership")
        is_own_client = (d.get("user_id") and str(d["user_id"]) in assigned)
        is_own_assignment = (
            assigned_to is not None
            and staff_id is not None
            and str(assigned_to) == str(staff_id)
        )
        if not (is_shared or is_own_client or is_own_assignment):
            raise HTTPException(status_code=403, detail="Not your client")
    return d


@router.put("/admin/crm/{cid}")
async def crm_update(cid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    await _assert_sales_can_touch_crm(db, staff, cid)
    payload = {k: v for k, v in payload.items() if k in {
        "name", "email", "phone", "company", "position", "industry", "status", "notes"
    }}
    payload["updated_at"] = _now()
    if "email" in payload and payload["email"]:
        payload["email"] = payload["email"].lower()
    await db.crm_customers.update_one({"_id": _oid(cid)}, {"$set": payload})
    d = await db.crm_customers.find_one({"_id": _oid(cid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    return _serialize_crm(d)


@router.delete("/admin/crm/{cid}")
async def crm_delete(cid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    await _assert_sales_can_touch_crm(db, staff, cid)
    r = await db.crm_customers.delete_one({"_id": _oid(cid)})
    return {"deleted": r.deleted_count}


# ---------- CRM import/export XLSX ----------
from fastapi import UploadFile as _UploadFile, File as _File  # noqa: E402

_CRM_XLSX_HEADERS = ["Nama", "Nomor Telp", "E-Mail", "Perusahaan", "Jabatan",
                     "Segmen Industri", "Status"]
_CRM_STATUS_EXPORT = {"prospect": "PROSPECT", "partnership": "POSSIBLE PARTNERSHIP",
                      "existing": "EXISTING CLIENT", "ex_client": "EX CLIENT"}


def _crm_status_import(v) -> str:
    s = re.sub(r"[^a-z]+", " ", str(v or "").lower()).strip()
    if not s:
        return ""
    if "partner" in s:
        return "partnership"
    if s.startswith("ex ") or s.startswith("ex") and "exist" not in s:
        return "ex_client"
    if "exist" in s or s in ("client", "customer", "active"):
        return "existing"
    return "prospect"


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


@router.get("/admin/crm/search")
async def crm_search(staff=Depends(get_current_staff),
                    q: Optional[str] = None, limit: int = 20):
    """Lightweight CRM search for autocomplete widgets.

    Returns rows visible to `staff` (CRM shared-pool scope). Search matches
    name, email, phone, and company case-insensitively. Always returns a list.
    """
    _deny_creative(staff)
    db = await _get_db()
    query = _crm_sales_scope_filter(staff)
    if q and q.strip():
        query["$or"] = [
            {field: {"$regex": q.strip(), "$options": "i"}}
            for field in ("name", "email", "phone", "company")
        ]
    limit_n = max(1, min(int(limit or 20), 100))
    docs = await db.crm_customers.find(query).sort("updated_at", -1).to_list(limit_n)
    out = []
    for d in docs:
        s = _serialize_crm(d)
        out.append({
            "id": s["id"],
            "name": s["name"],
            "email": s["email"],
            "phone": s["phone"],
            "company": s["company"],
            "status": s["status"],
            "assigned_to": s["assigned_to"],
        })
    return out


async def _crm_export_queryset(db, staff: dict) -> list:
    """Rows visible to `staff` for CRM export, using the CRM shared-pool scope."""
    q = _crm_sales_scope_filter(staff)
    return await db.crm_customers.find(q).sort("name", 1).to_list(20000)


@router.get("/admin/crm/export.xlsx")
async def crm_export_xlsx(staff=Depends(get_current_staff)):
    """Export Customer DB ke .xlsx dengan format template Database Marketing
    (Nama, Nomor Telp, E-Mail, Perusahaan, Jabatan, Segmen Industri, Status)."""
    _deny_creative(staff)
    db = await _get_db()
    docs = await _crm_export_queryset(db, staff)

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Database Marketing"
    header_fill = PatternFill("solid", fgColor="1E7145")
    status_style = {"prospect": ("9DC3E6", "000000"), "partnership": ("FFD966", "000000"),
                    "existing": ("A9D08E", "000000"), "ex_client": ("FF0000", "FFFFFF")}
    for col, h in enumerate(_CRM_XLSX_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for i, w in enumerate([24, 18, 32, 36, 40, 26, 24], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for r, d in enumerate(docs, start=2):
        ws.cell(row=r, column=1, value=d.get("name", ""))
        ws.cell(row=r, column=2, value=d.get("phone", ""))
        ws.cell(row=r, column=3, value=d.get("email", ""))
        ws.cell(row=r, column=4, value=d.get("company", ""))
        ws.cell(row=r, column=5, value=d.get("position", ""))
        ws.cell(row=r, column=6, value=d.get("industry", ""))
        st = d.get("status", "prospect")
        c = ws.cell(row=r, column=7, value=_CRM_STATUS_EXPORT.get(st, str(st).upper()))
        fill, fg = status_style.get(st, (None, None))
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
            c.font = Font(bold=True, color=fg)
            c.alignment = Alignment(horizontal="center")
    buf = io.BytesIO()
    wb.save(buf)
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customer-database.xlsx"'})


@router.post("/admin/crm/import")
async def crm_import_xlsx(file: _UploadFile = _File(...), staff=Depends(get_current_staff)):
    """Import kontak dari .xlsx (format Database Marketing). Header dicocokkan
    fleksibel (Nama/Name, Nomor Telp/Phone, E-Mail, Perusahaan/Company, Jabatan,
    Segmen Industri, Status); tanpa header -> urutan kolom A-G. Upsert by email
    (fallback: nama+telepon)."""
    _deny_creative(staff)
    if staff.get("role") == "sales":
        # XLSX rows have no portal client ID, so they cannot be proven to fall
        # within a sales user's assigned-client scope.
        raise HTTPException(
            status_code=403,
            detail="Sales cannot import unscoped CRM contacts",
        )
    db = await _get_db()
    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File terlalu besar (maks 10 MB)")
    import io
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File bukan .xlsx yang valid")
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row or []))
        if len(all_rows) > 20000:
            raise HTTPException(status_code=400, detail="Maksimal 20.000 baris per import")

    def _col_key(h):
        h = str(h or "").strip().lower()
        if not h:
            return None
        if "nama" in h or h == "name":
            return "name"
        if "telp" in h or "phone" in h or "telepon" in h or h in ("hp", "wa"):
            return "phone"
        if "mail" in h:
            return "email"
        if "perusahaan" in h or "company" in h:
            return "company"
        if "jabatan" in h or "position" in h or "title" in h:
            return "position"
        if "industri" in h or "industry" in h or "segmen" in h:
            return "industry"
        if "status" in h:
            return "status"
        return None

    # Cari baris header di 10 baris pertama (template punya baris judul dulu)
    col_map, start_idx = None, 0
    for i, row in enumerate(all_rows[:10]):
        keys = {}
        for ci, cell in enumerate(row):
            k = _col_key(cell)
            if k and k not in keys:
                keys[k] = ci
        if "name" in keys and ("email" in keys or "phone" in keys):
            col_map, start_idx = keys, i + 1
            break
    if col_map is None:
        # Tanpa header: pakai posisi kolom A-G sesuai template
        col_map = {"name": 0, "phone": 1, "email": 2, "company": 3,
                   "position": 4, "industry": 5, "status": 6}

    created = updated = skipped = 0
    errors = []
    for rn, row in enumerate(all_rows[start_idx:], start=start_idx + 1):
        try:
            vals = {k: _cell_str(row[ci]) if ci < len(row) else ""
                    for k, ci in col_map.items()}
            name = vals.get("name", "")
            email = vals.get("email", "").lower()
            if not name and not email:
                skipped += 1
                continue
            status = _crm_status_import(vals.get("status"))
            fields = {k: vals.get(k, "") for k in
                      ("name", "phone", "company", "position", "industry")}
            fields["email"] = email
            query = {"email": email} if email else {"name": name, "phone": vals.get("phone", "")}
            existing = await db.crm_customers.find_one(query)
            if existing:
                upd = {k: v for k, v in fields.items() if v}
                if status:
                    upd["status"] = status
                upd["updated_at"] = _now()
                await db.crm_customers.update_one({"_id": existing["_id"]}, {"$set": upd})
                updated += 1
            else:
                await db.crm_customers.insert_one({
                    **fields, "status": status or "prospect", "notes": "",
                    "source": "xlsx_import",
                    "created_at": _now(), "updated_at": _now()})
                created += 1
        except Exception as e:
            errors.append({"row": rn, "error": str(e)[:120]})
            if len(errors) >= 20:
                break
    await log_audit(db, actor=staff, action="crm.import_xlsx", category="crm",
                    target_type="crm", target_id="bulk", target_label=file.filename or "xlsx",
                    metadata={"created": created, "updated": updated, "skipped": skipped,
                              "errors": len(errors)})
    return {"ok": True, "created": created, "updated": updated,
            "skipped": skipped, "errors": errors,
            "total_rows": max(0, len(all_rows) - start_idx)}


# ---------- Projects ----------
def _serialize_project(d):
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "customer_id": str(d.get("customer_id", "")) if d.get("customer_id") else None,
        "customer_name": d.get("customer_name", ""),
        "owner": d.get("owner", ""),
        "status": d.get("status", "planning"),
        "priority": d.get("priority", "medium"),
        "progress": d.get("progress", 0),
        "start_date": d.get("start_date", ""),
        "target_date": d.get("target_date", ""),
        "description": d.get("description", ""),
        "tasks": d.get("tasks", []),
        "created_at": _iso(d.get("created_at", "")),
        "updated_at": _iso(d.get("updated_at", "")),
    }


@router.get("/admin/projects")
async def projects_list(staff=Depends(get_current_staff),
                        skip: int = 0, limit: int = 50, sort: str = "updated_at",
                        order: str = "desc", status: Optional[str] = None,
                        paginate: Optional[bool] = None):
    """Server-side pagination + status filter for projects. Default stays bare array."""
    db = await _get_db()
    query = {"status": status} if status and status != "all" else {}
    sort_field = sort if sort in {
        "name", "customer_name", "owner", "status", "priority", "progress",
        "start_date", "target_date", "created_at", "updated_at"
    } else "updated_at"
    direction = 1 if order.lower() == "asc" else -1
    skip_n, limit_n = _pagination_params(skip, limit)
    cursor = db.projects.find(query).sort(sort_field, direction)
    total = 0
    if bool(paginate):
        total = await db.projects.count_documents(query)
        cursor = cursor.skip(skip_n).limit(limit_n if limit_n is not None else 500)
        docs = await cursor.to_list(None)
    else:
        docs = await cursor.to_list(1000)
    items = [_serialize_project(d) for d in docs]
    return _pagination_response(items, total, skip_n, limit_n, True) if bool(paginate) else items


@router.post("/admin/projects")
async def projects_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "name": payload.get("name", ""),
        "customer_id": _oid(payload["customer_id"]) if payload.get("customer_id") else None,
        "customer_name": payload.get("customer_name", ""),
        "owner": payload.get("owner", ""),
        "status": payload.get("status", "planning"),
        "priority": payload.get("priority", "medium"),
        "progress": int(payload.get("progress", 0)),
        "start_date": payload.get("start_date", ""),
        "target_date": payload.get("target_date", ""),
        "description": payload.get("description", ""),
        "tasks": payload.get("tasks", []),
        "created_at": _now(),
        "updated_at": _now(),
    }
    r = await db.projects.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_project(doc)


@router.put("/admin/projects/{pid}")
async def projects_update(pid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    upd = {k: v for k, v in payload.items() if k in {
        "name", "customer_name", "owner", "status", "priority", "progress",
        "start_date", "target_date", "description", "tasks"
    }}
    if "customer_id" in payload:
        upd["customer_id"] = _oid(payload["customer_id"]) if payload["customer_id"] else None
    upd["updated_at"] = _now()
    await db.projects.update_one({"_id": _oid(pid)}, {"$set": upd})
    d = await db.projects.find_one({"_id": _oid(pid)})
    return _serialize_project(d)


@router.delete("/admin/projects/{pid}")
async def projects_delete(pid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.projects.delete_one({"_id": _oid(pid)})
    return {"deleted": r.deleted_count}


# ---------- Content Planner ----------
def _serialize_content(d):
    return {
        "id": str(d["_id"]),
        "title": d.get("title", ""),
        "channel": d.get("channel", "blog"),
        "type": d.get("type", "post"),
        "status": d.get("status", "idea"),
        "owner": d.get("owner", ""),
        "publish_date": d.get("publish_date", ""),
        "hook": d.get("hook", ""),
        "url": d.get("url", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/content")
async def content_list(staff=Depends(get_current_staff),
                       skip: int = 0, limit: int = 50, sort: str = "publish_date",
                       order: str = "asc", paginate: Optional[bool] = None):
    """Server-side pagination for content planner. Default stays bare array."""
    db = await _get_db()
    sort_field = sort if sort in {
        "title", "channel", "type", "status", "owner", "publish_date", "created_at"
    } else "publish_date"
    direction = 1 if order.lower() == "asc" else -1
    skip_n, limit_n = _pagination_params(skip, limit)
    cursor = db.content_plan.find({}).sort(sort_field, direction)
    total = 0
    if bool(paginate):
        total = await db.content_plan.count_documents({})
        cursor = cursor.skip(skip_n).limit(limit_n if limit_n is not None else 500)
        docs = await cursor.to_list(None)
    else:
        docs = await cursor.to_list(1000)
    items = [_serialize_content(d) for d in docs]
    return _pagination_response(items, total, skip_n, limit_n, True) if bool(paginate) else items


@router.post("/admin/content")
async def content_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    doc = {
        "title": payload.get("title", ""),
        "channel": payload.get("channel", "blog"),
        "type": payload.get("type", "post"),
        "status": payload.get("status", "idea"),
        "owner": payload.get("owner", ""),
        "publish_date": payload.get("publish_date", ""),
        "hook": payload.get("hook", ""),
        "url": payload.get("url", ""),
        "created_at": _now(),
    }
    r = await db.content_plan.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_content(doc)


@router.put("/admin/content/{cid}")
async def content_update(cid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    upd = {k: v for k, v in payload.items() if k in {
        "title", "channel", "type", "status", "owner", "publish_date", "hook", "url"
    }}
    await db.content_plan.update_one({"_id": _oid(cid)}, {"$set": upd})
    d = await db.content_plan.find_one({"_id": _oid(cid)})
    return _serialize_content(d)


@router.delete("/admin/content/{cid}")
async def content_delete(cid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    r = await db.content_plan.delete_one({"_id": _oid(cid)})
    return {"deleted": r.deleted_count}


# ---------- Follow-ups ----------
def _serialize_followup(d):
    return {
        "id": str(d["_id"]),
        "customer_id": str(d.get("customer_id", "")) if d.get("customer_id") else None,
        "customer_name": d.get("customer_name", ""),
        "task": d.get("task", ""),
        "channel": d.get("channel", "whatsapp"),
        "due_date": d.get("due_date", ""),
        "done": bool(d.get("done", False)),
        "owner": d.get("owner", ""),
        "notes": _normalize_note_threads(d.get("notes")),
        "role_tags": d.get("role_tags", []) or [],
        "assignee_role": d.get("assignee_role", ""),
        "approvals": [_serialize_approval(a) for a in (d.get("approvals") or [])],
        "deal_action": d.get("deal_action"),
        "deal_registration_link": d.get("deal_registration_link", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


async def _sales_followup_filter(db, staff: dict) -> dict | None:
    """Return a Mongo filter that restricts follow-ups to CRM rows the sales
    staff can access. Returns {} for non-sales. Returns None if the caller is
    a sales user with zero visible CRM rows (endpoint should short-circuit)."""
    if staff.get("role") != "sales":
        return {}
    ids = await _sales_visible_crm_ids(db, staff)
    if not ids:
        return None
    return {"customer_id": {"$in": ids}}


async def _assert_sales_can_touch_followup(db, staff: dict, fid: str) -> dict:
    d = await db.followups.find_one({"_id": _oid(fid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    if staff.get("role") == "sales":
        visible = await _sales_visible_crm_ids(db, staff) or []
        cust_id = d.get("customer_id")
        if not (cust_id and any(str(cust_id) == str(x) for x in visible)):
            raise HTTPException(status_code=403, detail="Not your follow-up")
    return d


@router.get("/admin/followups")
async def followups_list(staff=Depends(get_current_staff),
                         skip: int = 0, limit: int = 50, sort: str = "due_date",
                         order: str = "asc", done: Optional[bool] = None,
                         paginate: Optional[bool] = None):
    """Server-side pagination for follow-ups with done filter.
    Sales-empty short-circuit and _sales_followup_filter scoping preserved."""
    _deny_creative(staff)
    db = await _get_db()
    q = await _sales_followup_filter(db, staff)
    if q is None:
        return []
    if done is not None:
        q["done"] = done
    sort_field = sort if sort in {
        "customer_name", "task", "channel", "due_date", "done", "owner", "created_at"
    } else "due_date"
    direction = 1 if order.lower() == "asc" else -1
    skip_n, limit_n = _pagination_params(skip, limit)
    cursor = db.followups.find(q).sort(sort_field, direction)
    total = 0
    if bool(paginate):
        total = await db.followups.count_documents(q)
        cursor = cursor.skip(skip_n).limit(limit_n if limit_n is not None else 500)
        docs = await cursor.to_list(None)
    else:
        docs = await cursor.to_list(1000)
    items = [_serialize_followup(d) for d in docs]
    return _pagination_response(items, total, skip_n, limit_n, True) if bool(paginate) else items


_PROSPECT_ASSIGNMENT_DAYS = 5

# Roles that can be tagged on / own a follow-up task.
_FOLLOWUP_ROLES = {"sales", "noc", "support", "finance", "admin"}


def _clean_role_tags(value) -> list:
    """Validate a role_tags payload into a de-duplicated ordered list."""
    if not isinstance(value, list):
        raise HTTPException(status_code=400, detail="role_tags harus berupa list")
    out = []
    for raw in value:
        role = str(raw or "").strip().lower()
        if role not in _FOLLOWUP_ROLES:
            raise HTTPException(status_code=400, detail=f"Invalid role tag: {raw}")
        if role not in out:
            out.append(role)
    return out


def _serialize_approval(a: dict) -> dict:
    """Serialize an approval sub-document for API response."""
    return {
        "id": str(a.get("id") or a.get("_id") or ""),
        "requested_by": a.get("requested_by", ""),
        "requested_by_id": str(a["requested_by_id"]) if a.get("requested_by_id") else None,
        "requested_role": a.get("requested_role", ""),
        "target_role": a.get("target_role", ""),
        "target_user_id": str(a["target_user_id"]) if a.get("target_user_id") else None,
        "message": a.get("message", ""),
        "status": a.get("status", "pending"),
        "responded_by": a.get("responded_by", ""),
        "responded_at": _iso(a.get("responded_at", "")),
        "response_note": a.get("response_note", ""),
        "created_at": _iso(a.get("created_at", "")),
    }


# ---------- Close-deal registration token ----------
import jwt as _jwt  # noqa: E402
from portal.auth import _secret as _jwt_secret, JWT_ALGORITHM as _jwt_alg  # noqa: E402

_CRM_TOKEN_TTL_DAYS = 7


def _make_crm_registration_token(crm_id: str) -> str:
    """Sign a short-lived JWT carrying the CRM row id for close-deal registration."""
    payload = {
        "crm_id": str(crm_id),
        "exp": datetime.now(timezone.utc) + timedelta(days=_CRM_TOKEN_TTL_DAYS),
        "type": "crm_register",
    }
    return _jwt.encode(payload, _jwt_secret(), algorithm=_jwt_alg)


def _verify_crm_registration_token(token: str) -> str:
    """Decode and validate a close-deal registration token.

    Returns the crm_id. Raises HTTPException(400) on invalid/expired tokens.
    """
    try:
        data = _jwt.decode(token, _jwt_secret(), algorithms=[_jwt_alg])
    except Exception:
        raise HTTPException(status_code=400, detail="Token registrasi tidak valid")
    if data.get("type") != "crm_register" or not data.get("crm_id"):
        raise HTTPException(status_code=400, detail="Token registrasi tidak valid")
    return str(data["crm_id"])


async def expire_overdue_assignments(db, now_iso: str | None = None) -> int:
    """Release expired prospect assignments back to the shared pool.

    Only rows with status='assigned' AND assignment_expires_at < now are
    touched; existing clients / shared leads are never modified. Intended to
    run as a scheduled job; returns the number of released rows.
    """
    now_value = now_iso or datetime.now(timezone.utc).isoformat()
    result = await db.crm_customers.update_many(
        {"status": "assigned", "assignment_expires_at": {"$lt": now_value}},
        {"$set": {
            "status": "prospect",
            "assigned_to": None,
            "assigned_at": None,
            "assignment_expires_at": None,
            "updated_at": now_value,
        }},
    )
    return getattr(result, "modified_count", 0)


async def sync_crm_after_registration(db, *, crm_token: str, new_user_id,
                                       reg_email: str, reg_name: str = "") -> None:
    """After a user registers with a close-deal crm_token, sync the CRM row.

    If the registered email/name matches the CRM row, the row is upgraded to
    'existing' and linked to the new user_id. If there's a mismatch (different
    person using the link), the original CRM row is left untouched — the new
    user is simply a new customer, not a silent merge.
    """
    crm_id_str = _verify_crm_registration_token(crm_token)
    crm = await db.crm_customers.find_one({"_id": _oid(crm_id_str)})
    if not crm:
        return
    crm_email = (crm.get("email") or "").lower().strip()
    crm_name = (crm.get("name") or "").strip()
    reg_email_norm = (reg_email or "").lower().strip()
    reg_name_norm = (reg_name or "").strip()
    email_match = crm_email and crm_email == reg_email_norm
    name_match = (not crm_name) or (not reg_name_norm) or crm_name == reg_name_norm
    if email_match and name_match:
        now = datetime.now(timezone.utc)
        await db.crm_customers.update_one(
            {"_id": crm["_id"]},
            {"$set": {
                "status": "existing",
                "user_id": new_user_id,
                "assigned_to": None,
                "assigned_at": None,
                "assignment_expires_at": None,
                "updated_at": now.isoformat(),
            }},
        )


async def crm_register_prefill(staff, token: str) -> dict:
    """Public endpoint: validate a crm_token and return CRM data for prefill.

    Returns {name, email, phone, company}. Raises HTTPException(400) if token
    is invalid/expired or CRM row not found.
    """
    crm_id = _verify_crm_registration_token(token)
    db = await _get_db()
    crm = await db.crm_customers.find_one({"_id": _oid(crm_id)})
    if not crm:
        raise HTTPException(status_code=400, detail="Token registrasi tidak valid")
    return {
        "name": crm.get("name", ""),
        "email": crm.get("email", ""),
        "phone": crm.get("phone", ""),
        "company": crm.get("company", ""),
    }


def _normalize_note_threads(raw) -> dict:
    """Normalize stored notes into an append-only thread per role.

    Two shapes exist on disk:
      legacy: {"sales": "free text"}          -> one synthetic legacy entry
      thread: {"sales": [{author, text, at}]} -> passed through

    Threads are append-only by design: no endpoint rewrites or deletes an
    entry, so one division can never edit another division's feedback.
    """
    out = {role: [] for role in sorted(_FOLLOWUP_ROLES)}
    if not isinstance(raw, dict):
        return out
    for raw_role, value in raw.items():
        role = str(raw_role or "").strip().lower()
        if role not in _FOLLOWUP_ROLES:
            continue
        if isinstance(value, str):
            text = value.strip()
            if text:
                out[role].append({
                    "author": "",
                    "author_role": role,
                    "text": text,
                    "at": "",
                    "legacy": True,
                })
            continue
        if isinstance(value, list):
            for entry in value:
                if not isinstance(entry, dict):
                    continue
                text = str(entry.get("text") or "").strip()
                if not text:
                    continue
                out[role].append({
                    "author": str(entry.get("author") or ""),
                    "author_role": str(entry.get("author_role") or role),
                    "text": text,
                    "at": _iso(entry.get("at", "")),
                    "legacy": bool(entry.get("legacy", False)),
                })
    return out


def _note_author_role(staff: dict) -> str:
    """Role bucket a staff member is allowed to post notes into.

    A user posts as their own role only. Nobody -- admin included -- may
    write into another division's thread, which is what the old editable
    per-role textarea allowed.
    """
    role = str(staff.get("role") or "").strip().lower()
    if role not in _FOLLOWUP_ROLES:
        raise HTTPException(
            status_code=403,
            detail=f"Role '{role}' tidak dapat menulis catatan follow-up",
        )
    return role


@router.post("/admin/followups")
async def followups_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    _deny_creative(staff)
    cust_id = _oid(payload["customer_id"]) if payload.get("customer_id") else None
    if staff.get("role") == "sales":
        visible = await _sales_visible_crm_ids(db, staff) or []
        if not (cust_id and any(str(cust_id) == str(x) for x in visible)):
            raise HTTPException(
                status_code=403,
                detail="Follow-up harus untuk pelanggan yang dapat Anda akses",
            )
    # Atomic claim: prospect must be prospect/partnership (unclaimed) or
    # already assigned to this staff. Prevents double-claim by another sales.
    if cust_id:
        crm = await db.crm_customers.find_one({"_id": cust_id})
        if crm and crm.get("status") == "assigned":
            assigned_to = crm.get("assigned_to")
            staff_id = _staff_oid(staff)
            if staff.get("role") == "sales" and (
                not assigned_to
                or not staff_id
                or str(assigned_to) != str(staff_id)
            ):
                raise HTTPException(
                    status_code=403,
                    detail="Prospect ini sedang di-assign ke sales lain",
                )
        # Claim the prospect: only "prospect" status (not partnership) is claimable.
        # The status guard lives in the filter so two concurrent sales cannot
        # both win the claim -- the loser matches zero documents.
        if crm and crm.get("status") == "prospect":
            now_dt = datetime.now(timezone.utc)
            expires = now_dt + timedelta(days=_PROSPECT_ASSIGNMENT_DAYS)
            claimed = await db.crm_customers.update_one(
                {"_id": cust_id, "status": "prospect"},
                {"$set": {
                    "status": "assigned",
                    "assigned_to": _staff_oid(staff),
                    "assigned_at": now_dt.isoformat(),
                    "assignment_expires_at": expires.isoformat(),
                    "updated_at": now_dt.isoformat(),
                }},
            )
            if not getattr(claimed, "modified_count", 0):
                raise HTTPException(
                    status_code=403,
                    detail="Prospect ini baru saja di-assign ke sales lain",
                )
        elif crm and crm.get("status") == "assigned" and _staff_oid(staff):
            # Already assigned to this sales — renew the expiry timer
            now_dt = datetime.now(timezone.utc)
            expires = now_dt + timedelta(days=_PROSPECT_ASSIGNMENT_DAYS)
            await db.crm_customers.update_one(
                {"_id": cust_id},
                {"$set": {
                    "assigned_at": now_dt.isoformat(),
                    "assignment_expires_at": expires.isoformat(),
                    "updated_at": now_dt.isoformat(),
                }},
            )
    doc = {
        "customer_id": cust_id,
        "customer_name": payload.get("customer_name", ""),
        "task": payload.get("task", ""),
        "channel": payload.get("channel", "whatsapp"),
        "due_date": payload.get("due_date", ""),
        "done": False,
        "owner": payload.get("owner", staff.get("name", "")),
        "created_at": _now(),
    }
    r = await db.followups.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_followup(doc)


@router.put("/admin/followups/{fid}")
async def followups_update(fid: str, payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    _deny_creative(staff)
    d = await _assert_sales_can_touch_followup(db, staff, fid)
    upd = {}
    MAPPED = {"task", "channel", "due_date", "done", "owner", "customer_name"}
    for k, v in payload.items():
        if k in MAPPED:
            upd[k] = v
        elif k == "notes":
            # Notes are append-only threads now. Rewriting them through the
            # generic update endpoint is what let one division overwrite
            # another division's feedback, so it is refused outright.
            raise HTTPException(
                status_code=400,
                detail="Catatan tidak bisa diubah lewat update. "
                       "Gunakan POST /admin/followups/{id}/notes untuk menambah catatan.",
            )
        elif k == "role_tags":
            upd["role_tags"] = _clean_role_tags(v)
        elif k == "assignee_role":
            role = str(v or "").strip().lower()
            if role not in _FOLLOWUP_ROLES:
                raise HTTPException(status_code=400, detail=f"Invalid assignee role: {v}")
            upd["assignee_role"] = role
    if not upd:
        raise HTTPException(status_code=400, detail="Tidak ada field yang valid untuk diupdate")
    await db.followups.update_one({"_id": _oid(fid)}, {"$set": upd})
    # If this follow-up has a linked CRM prospect, renew or release it.
    cust_id = d.get("customer_id")
    if cust_id:
        crm = await db.crm_customers.find_one({"_id": cust_id})
        if crm and crm.get("status") == "assigned":
            if payload.get("done") is True:
                # Release prospect back to shared pool
                now_dt = datetime.now(timezone.utc)
                await db.crm_customers.update_one(
                    {"_id": cust_id},
                    {"$set": {
                        "status": "prospect",
                        "assigned_to": None,
                        "assigned_at": None,
                        "assignment_expires_at": None,
                        "updated_at": now_dt.isoformat(),
                    }},
                )
            else:
                # Any other update renews the 5-day timer
                now_dt = datetime.now(timezone.utc)
                expires = now_dt + timedelta(days=_PROSPECT_ASSIGNMENT_DAYS)
                await db.crm_customers.update_one(
                    {"_id": cust_id},
                    {"$set": {
                        "assigned_at": now_dt.isoformat(),
                        "assignment_expires_at": expires.isoformat(),
                        "updated_at": now_dt.isoformat(),
                    }},
                )
    updated = await db.followups.find_one({"_id": _oid(fid)})
    return _serialize_followup(updated)


@router.get("/register/crm-prefill")
async def crm_register_prefill_route(token: str):
    """Public: validate a close-deal crm_token and return prefill data."""
    return await crm_register_prefill(staff=None, token=token)


@router.post("/admin/followups/{fid}/close-deal")
async def followup_close_deal(fid: str, staff=Depends(get_current_staff)):
    """Mark a follow-up as close-deal and generate a 7-day registration link."""
    db = await _get_db()
    _deny_creative(staff)
    fu = await _assert_sales_can_touch_followup(db, staff, fid)
    cust_id = fu.get("customer_id")
    if not cust_id:
        raise HTTPException(status_code=400, detail="Follow-up tidak memiliki customer CRM")
    crm = await db.crm_customers.find_one({"_id": cust_id})
    if not crm:
        raise HTTPException(status_code=404, detail="Customer CRM tidak ditemukan")
    token = _make_crm_registration_token(str(cust_id))
    frontend_base = (os.environ.get("PORTAL_FRONTEND_URL") or "").rstrip("/")
    path = f"/register?crm_token={quote(token)}"
    link = f"{frontend_base}{path}" if frontend_base else path
    now = _now()
    update = {
        "deal_action": "close_deal",
        "deal_registration_link": link,
        "deal_closed_at": now,
        "deal_closed_by": staff.get("name") or staff.get("email") or "staff",
        "updated_at": now,
    }
    await db.followups.update_one({"_id": _oid(fid)}, {"$set": update})
    return {**update, "crm_id": str(cust_id)}


@router.post("/admin/followups/{fid}/approval")
async def followup_approval_request(fid: str, payload: dict,
                                    staff=Depends(get_current_staff)):
    """Request an approval from another role/user. Pending request is appended."""
    db = await _get_db()
    _deny_creative(staff)
    await _assert_sales_can_touch_followup(db, staff, fid)
    target_role = str(payload.get("target_role") or "").strip().lower()
    target_user_id = payload.get("target_user_id")
    if target_role not in _FOLLOWUP_ROLES:
        raise HTTPException(status_code=400, detail=f"Invalid target role: {target_role}")
    approval = {
        "id": ObjectId(),
        "requested_by": staff.get("name") or staff.get("email") or "staff",
        "requested_by_id": _staff_oid(staff),
        "requested_role": staff.get("role", ""),
        "target_role": target_role,
        "target_user_id": _oid(target_user_id) if target_user_id else None,
        "message": payload.get("message", ""),
        "status": "pending",
        "responded_by": "",
        "responded_at": "",
        "response_note": "",
        "created_at": _now(),
    }
    await db.followups.update_one(
        {"_id": _oid(fid)},
        {"$push": {"approvals": approval}, "$set": {"updated_at": _now()}},
    )
    return _serialize_approval(approval)


@router.put("/admin/followups/{fid}/approval/{aid}")
async def followup_approval_respond(fid: str, aid: str, payload: dict,
                                    staff=Depends(get_current_staff)):
    """Accept or reject a pending approval.

    The caller must be the target role (or admin). Accepting/rejecting a
    non-pending approval is a no-op that surfaces the existing state.
    """
    db = await _get_db()
    _deny_creative(staff)
    await _assert_sales_can_touch_followup(db, staff, fid)
    fu = await db.followups.find_one({"_id": _oid(fid)})
    if not fu:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    approvals = fu.get("approvals") or []
    target = None
    for a in approvals:
        if str(a.get("id")) == aid:
            target = a
            break
    if not target:
        raise HTTPException(status_code=404, detail="Approval not found")
    if target.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Approval sudah direspon")
    caller_role = staff.get("role", "")
    if caller_role not in ("admin",) and caller_role != target.get("target_role"):
        # Allow targeted user if specified
        target_user = target.get("target_user_id")
        if not target_user or str(target_user) != str(_staff_oid(staff) or ""):
            raise HTTPException(
                status_code=403,
                detail="Anda tidak berhak merespon approval ini",
            )
    response = str(payload.get("response") or "").lower() if isinstance(payload, dict) else ""
    if response not in ("accepted", "rejected"):
        raise HTTPException(status_code=400, detail="response harus 'accepted' atau 'rejected'")
    note = payload.get("note", "") if isinstance(payload, dict) else ""
    now = _now()
    # Update the approval in-place in the array, then persist the whole array.
    target["status"] = response
    target["responded_by"] = staff.get("name") or staff.get("email") or "staff"
    target["responded_at"] = now
    target["response_note"] = note
    await db.followups.update_one(
        {"_id": _oid(fid)},
        {"$set": {"approvals": approvals, "updated_at": now}},
    )
    return _serialize_approval(target)


@router.delete("/admin/followups/{fid}")
async def followups_delete(fid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    await _assert_sales_can_touch_followup(db, staff, fid)
    r = await db.followups.delete_one({"_id": _oid(fid)})
    return {"deleted": r.deleted_count}


@router.post("/admin/followups/{fid}/notes")
async def followup_notes_add(fid: str, payload: dict, staff=Depends(get_current_staff)):
    """Append one note to the caller's own role thread.

    This is intentionally the only way to write a follow-up note: threads
    are append-only, and the author role is always derived from the
    authenticated caller, never from the request body. That is what
    guarantees admin (or any other role) cannot edit or inject content
    into another division's thread.
    """
    db = await _get_db()
    _deny_creative(staff)
    d = await _assert_sales_can_touch_followup(db, staff, fid)
    text = str((payload or {}).get("text") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Catatan tidak boleh kosong")
    author_role = _note_author_role(staff)
    entry = {
        "author": staff.get("name") or staff.get("email") or "staff",
        "author_role": author_role,
        "text": text,
        "at": _now(),
        "legacy": False,
    }
    threads = _normalize_note_threads(d.get("notes"))
    threads[author_role].append(entry)
    await db.followups.update_one(
        {"_id": _oid(fid)},
        {"$set": {"notes": threads, "updated_at": _now()}},
    )
    return {"notes": threads}


# ---------- Documents (metadata + file upload lokal) ----------
from pathlib import Path as _DocPath  # noqa: E402


DOCS_DIR = _DocPath(__file__).resolve().parent.parent.parent / "uploads" / "documents"


_DOC_ALLOWED_TYPES = {
    "application/pdf", "image/png", "image/jpeg", "image/webp",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/zip", "application/x-zip-compressed", "text/plain", "text/csv",
}


_DOC_MAX_BYTES = 15 * 1024 * 1024  # 15 MB


def _serialize_doc(d):
    return {
        "id": str(d["_id"]),
        "title": d.get("title", ""),
        "category": d.get("category", "contract"),
        "customer_name": d.get("customer_name", ""),
        "url": d.get("url", ""),
        "notes": d.get("notes", ""),
        "filename": d.get("filename", ""),
        "size_bytes": d.get("size_bytes", 0),
        "has_file": bool(d.get("stored_name")),
        "created_at": _iso(d.get("created_at", "")),
    }


def _require_internal_document_access(staff: dict) -> None:
    """Restrict unscoped internal documents to internal operations roles.

    Documents do not yet carry client ownership, therefore sales cannot be
    scoped safely. Creative is content-scoped and must not access business
    document operations.
    """
    _deny_creative(staff)
    if staff.get("role") == "sales":
        raise HTTPException(status_code=403, detail="Sales cannot access business documents")


@router.get("/admin/documents")
async def docs_list(staff=Depends(get_current_staff),
                    skip: int = 0, limit: int = 50, sort: str = "created_at",
                    order: str = "desc", q: Optional[str] = None,
                    paginate: Optional[bool] = None):
    """Server-side pagination + q-search for documents. Default stays bare array."""
    _require_internal_document_access(staff)
    db = await _get_db()
    query: dict = {}
    if q:
        query["$or"] = [
            {field: {"$regex": q.strip(), "$options": "i"}}
            for field in ("title", "category", "customer_name", "notes", "filename")
        ]
    sort_field = sort if sort in {
        "title", "category", "customer_name", "filename", "size_bytes", "created_at"
    } else "created_at"
    direction = 1 if order.lower() == "asc" else -1
    skip_n, limit_n = _pagination_params(skip, limit)
    cursor = db.documents.find(query).sort(sort_field, direction)
    total = 0
    if bool(paginate):
        total = await db.documents.count_documents(query)
        cursor = cursor.skip(skip_n).limit(limit_n if limit_n is not None else 500)
        docs = await cursor.to_list(None)
    else:
        docs = await cursor.to_list(1000)
    items = [_serialize_doc(d) for d in docs]
    return _pagination_response(items, total, skip_n, limit_n, True) if bool(paginate) else items


@router.post("/admin/documents")
async def docs_create(payload: dict, staff=Depends(get_current_staff)):
    _require_internal_document_access(staff)
    db = await _get_db()
    doc = {
        "title": payload.get("title", ""),
        "category": payload.get("category", "contract"),
        "customer_name": payload.get("customer_name", ""),
        "url": payload.get("url", ""),
        "notes": payload.get("notes", ""),
        "created_at": _now(),
    }
    r = await db.documents.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_doc(doc)


@router.delete("/admin/documents/{did}")
async def docs_delete(did: str, staff=Depends(get_current_staff)):
    _require_internal_document_access(staff)
    db = await _get_db()
    d = await db.documents.find_one({"_id": _oid(did)})
    if d and d.get("stored_name"):
        try:
            (DOCS_DIR / d["stored_name"]).unlink(missing_ok=True)
        except Exception:
            pass
    r = await db.documents.delete_one({"_id": _oid(did)})
    return {"deleted": r.deleted_count}


@router.get("/documents/file/{did}")
async def docs_file(did: str, staff=Depends(get_current_staff)):
    """Serve dokumen bisnis yang di-upload (URL ber-ObjectId, seperti media)."""
    _require_internal_document_access(staff)
    db = await _get_db()
    d = await db.documents.find_one({"_id": _oid(did)})
    if not d or not d.get("stored_name"):
        raise HTTPException(status_code=404, detail="Document not found")
    fp = DOCS_DIR / d["stored_name"]
    if not fp.exists():
        raise HTTPException(status_code=404, detail="File missing on disk")
    return FileResponse(fp, media_type=d.get("content_type") or "application/octet-stream",
                        filename=d.get("filename") or d["stored_name"])


# ============================================================
# Backup / Restore
# ============================================================
# ---------- UTM link persistence ----------

@router.get("/admin/utm-links")
async def utm_links_list(staff=Depends(get_current_staff)):
    db = await _get_db()
    docs = await db.utm_links.find({}).sort("created_at", -1).to_list(200)
    return [{"id": str(d["_id"]), "url": d.get("url", ""), "base": d.get("base", ""),
             "params": d.get("params", {}), "label": d.get("label", ""),
             "created_by": d.get("created_by", ""), "created_at": _iso(d.get("created_at", ""))}
            for d in docs]


@router.post("/admin/utm-links")
async def utm_links_create(payload: dict, staff=Depends(get_current_staff)):
    db = await _get_db()
    url = str(payload.get("url", "")).strip()
    if not url.startswith("http"):
        raise HTTPException(status_code=400, detail="URL tidak valid")
    doc = {"url": url, "base": payload.get("base", ""), "params": payload.get("params", {}),
           "label": payload.get("label", ""), "created_by": staff.get("email", ""),
           "created_at": _now()}
    res = await db.utm_links.insert_one(doc)
    return {"ok": True, "id": str(res.inserted_id)}


@router.delete("/admin/utm-links/{lid}")
async def utm_links_delete(lid: str, staff=Depends(get_current_staff)):
    db = await _get_db()
    await db.utm_links.delete_one({"_id": _oid(lid)})
    return {"ok": True}


# ---------- Content calendar: hari libur nasional ----------

ID_HOLIDAYS_2026 = [
    {"date": "2026-01-01", "name": "Tahun Baru Masehi"},
    {"date": "2026-01-16", "name": "Isra Mikraj Nabi Muhammad SAW"},
    {"date": "2026-02-17", "name": "Tahun Baru Imlek 2577"},
    {"date": "2026-03-19", "name": "Hari Suci Nyepi"},
    {"date": "2026-03-20", "name": "Idul Fitri 1447 H (perkiraan)"},
    {"date": "2026-03-21", "name": "Idul Fitri 1447 H (hari kedua)"},
    {"date": "2026-04-03", "name": "Wafat Isa Almasih"},
    {"date": "2026-05-01", "name": "Hari Buruh Internasional"},
    {"date": "2026-05-14", "name": "Kenaikan Isa Almasih"},
    {"date": "2026-05-27", "name": "Idul Adha 1447 H (perkiraan)"},
    {"date": "2026-05-31", "name": "Hari Raya Waisak"},
    {"date": "2026-06-01", "name": "Hari Lahir Pancasila"},
    {"date": "2026-06-16", "name": "Tahun Baru Islam 1448 H"},
    {"date": "2026-08-17", "name": "Hari Kemerdekaan RI"},
    {"date": "2026-08-25", "name": "Maulid Nabi Muhammad SAW"},
    {"date": "2026-12-25", "name": "Hari Raya Natal"},
]


@router.get("/admin/content-calendar/holidays")
async def content_calendar_holidays(year: int = 2026, staff=Depends(get_current_staff)):
    return {"year": year, "holidays": [h for h in ID_HOLIDAYS_2026 if h["date"].startswith(str(year))]}


# ============================================================
# MEDIA LIBRARY - shared assets for the Digital Creative team
# ============================================================
from fastapi import UploadFile, File, Form  # noqa: E402


from fastapi.responses import FileResponse  # noqa: E402


import uuid as _uuid  # noqa: E402


from pathlib import Path as _Path  # noqa: E402


MEDIA_DIR = _Path(__file__).resolve().parent.parent.parent / "uploads" / "media"


_MEDIA_ALLOWED_TYPES = {"image/png", "image/jpeg", "image/webp", "image/gif", "image/svg+xml"}


_MEDIA_MAX_BYTES = 8 * 1024 * 1024  # 8 MB


async def _media_usage(db, media_id: str) -> list:
    """Where is this asset referenced? Scans articles (cover/OG/body) and
    branding/landing settings. Computed live so it never goes stale."""
    needle = f"/media/file/{media_id}"
    used = []
    cur = db.articles.find({"$or": [
        {"cover_image_url": {"$regex": needle}},
        {"og_image_url": {"$regex": needle}},
        {"body_html": {"$regex": needle}},
    ]}, {"title": 1, "slug": 1})
    async for a in cur:
        used.append({"type": "article", "id": str(a["_id"]),
                     "label": a.get("title") or a.get("slug") or "article"})
    async for s in db.settings.find({"key": {"$in": ["branding", "landing_content"]}}):
        if needle in str(s.get("value", "")):
            used.append({"type": "settings", "id": s.get("key"),
                         "label": f"Settings: {s.get('key')}"})
    return used


def _serialize_media(d: dict, used_in=None) -> dict:
    return {
        "id": str(d["_id"]),
        "filename": d.get("filename", ""),
        "url": d.get("url", ""),
        "content_type": d.get("content_type", ""),
        "size_bytes": int(d.get("size_bytes") or 0),
        "alt_text": d.get("alt_text", ""),
        "tags": d.get("tags", []),
        "uploaded_by": d.get("uploaded_by", ""),
        "created_at": d.get("created_at", ""),
        "used_in": used_in if used_in is not None else d.get("used_in", []),
    }


@router.get("/admin/media")
async def media_list(staff=Depends(get_current_staff),
                     tag: Optional[str] = None, q: Optional[str] = None):
    db = await _get_db()
    query: dict = {}
    if tag:
        query["tags"] = tag.strip().lower()
    if q:
        query["$or"] = [
            {"filename": {"$regex": q.strip(), "$options": "i"}},
            {"alt_text": {"$regex": q.strip(), "$options": "i"}},
        ]
    docs = await db.media_assets.find(query).sort("created_at", -1).to_list(500)
    out = []
    for d in docs:
        used = await _media_usage(db, str(d["_id"]))
        if used != d.get("used_in"):
            await db.media_assets.update_one({"_id": d["_id"]}, {"$set": {"used_in": used}})
        out.append(_serialize_media(d, used))
    return out


@router.post("/admin/media")
async def media_upload(file: UploadFile = File(...),
                       alt_text: str = Form(""),
                       tags: str = Form(""),
                       staff=Depends(get_current_content)):
    db = await _get_db()
    if file.content_type not in _MEDIA_ALLOWED_TYPES:
        raise HTTPException(status_code=400,
                            detail=f"Unsupported type {file.content_type}. Allowed: PNG, JPEG, WebP, GIF, SVG.")
    raw = await file.read()
    if len(raw) > _MEDIA_MAX_BYTES:
        raise HTTPException(status_code=400, detail="File exceeds 8 MB limit")
    MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    ext = _Path(file.filename or "upload.bin").suffix.lower() or ".bin"
    mid = ObjectId()
    stored_name = f"{mid}{ext}"
    (MEDIA_DIR / stored_name).write_bytes(raw)
    doc = {
        "_id": mid,
        "filename": file.filename or stored_name,
        "stored_name": stored_name,
        "url": f"/api/portal/media/file/{mid}",
        "content_type": file.content_type,
        "size_bytes": len(raw),
        "alt_text": (alt_text or "").strip(),
        "tags": sorted({t.strip().lower() for t in (tags or "").split(",") if t.strip()}),
        "uploaded_by": staff["email"],
        "used_in": [],
        "created_at": _now(),
    }
    await db.media_assets.insert_one(doc)
    return _serialize_media(doc)


@router.post("/admin/documents/upload")
async def docs_upload(file: UploadFile = File(...), title: str = Form(""),
                      category: str = Form("contract"), customer_name: str = Form(""),
                      notes: str = Form(""), staff=Depends(get_current_staff)):
    """UAT-003: upload dokumen lokal (drag & drop) selain link URL."""
    _require_internal_document_access(staff)
    db = await _get_db()
    ctype = file.content_type or "application/octet-stream"
    if ctype not in _DOC_ALLOWED_TYPES:
        raise HTTPException(status_code=400,
                            detail=f"Tipe file {ctype} tidak didukung. Gunakan PDF, Word, Excel, gambar, ZIP, atau teks.")
    raw = await file.read()
    if len(raw) > _DOC_MAX_BYTES:
        raise HTTPException(status_code=400, detail="Ukuran file melebihi 15 MB")
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    did = ObjectId()
    ext = _DocPath(file.filename or "dokumen.bin").suffix.lower() or ".bin"
    stored_name = f"{did}{ext}"
    (DOCS_DIR / stored_name).write_bytes(raw)
    doc = {
        "_id": did,
        "title": (title or "").strip() or (file.filename or "Dokumen"),
        "category": category or "contract",
        "customer_name": customer_name or "",
        "url": f"/api/portal/documents/file/{did}",
        "notes": notes or "",
        "filename": file.filename or stored_name,
        "stored_name": stored_name,
        "content_type": ctype,
        "size_bytes": len(raw),
        "uploaded_by": staff["email"],
        "created_at": _now(),
    }
    await db.documents.insert_one(doc)
    return _serialize_doc(doc)


@router.put("/admin/media/{mid}")
async def media_update(mid: str, payload: dict, staff=Depends(get_current_content)):
    db = await _get_db()
    d = await db.media_assets.find_one({"_id": _oid(mid)})
    if not d:
        raise HTTPException(status_code=404, detail="Media not found")
    upd: dict = {}
    if "alt_text" in payload:
        upd["alt_text"] = (payload.get("alt_text") or "").strip()
    if "tags" in payload:
        raw_tags = payload.get("tags") or []
        if isinstance(raw_tags, str):
            raw_tags = raw_tags.split(",")
        upd["tags"] = sorted({str(t).strip().lower() for t in raw_tags if str(t).strip()})
    if upd:
        await db.media_assets.update_one({"_id": d["_id"]}, {"$set": upd})
    d = await db.media_assets.find_one({"_id": d["_id"]})
    return _serialize_media(d)


@router.delete("/admin/media/{mid}")
async def media_delete(mid: str, staff=Depends(get_current_content)):
    db = await _get_db()
    d = await db.media_assets.find_one({"_id": _oid(mid)})
    if not d:
        raise HTTPException(status_code=404, detail="Media not found")
    used = await _media_usage(db, mid)
    if used:
        raise HTTPException(status_code=409, detail={
            "message": "Asset is still in use - detach it first.",
            "used_in": used,
        })
    try:
        (MEDIA_DIR / d.get("stored_name", "")).unlink(missing_ok=True)
    except Exception:
        pass
    await db.media_assets.delete_one({"_id": d["_id"]})
    return {"deleted": 1}


@router.get("/media/file/{mid}", include_in_schema=False)
async def media_file(mid: str):
    """Public file serve - media is referenced from public articles."""
    db = await _get_db()
    d = await db.media_assets.find_one({"_id": _oid(mid)})
    if not d:
        raise HTTPException(status_code=404, detail="Media not found")
    fp = MEDIA_DIR / d.get("stored_name", "")
    if not fp.exists():
        raise HTTPException(status_code=404, detail="File missing on disk")
    return FileResponse(fp, media_type=d.get("content_type") or "application/octet-stream",
                        headers={"Cache-Control": "public, max-age=86400"})


# ============================================================
# CONTENT CALENDAR - plan articles / campaigns / social posts
# ============================================================
def _serialize_calendar(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "source": "calendar",
        "title": d.get("title", ""),
        "type": d.get("type", "article"),
        "scheduled_at": d.get("scheduled_at", ""),
        "status": d.get("status", "draft"),
        "linked_article_id": d.get("linked_article_id"),
        "owner_id": d.get("owner_id"),
        "notes": d.get("notes", ""),
        "created_at": d.get("created_at", ""),
    }


# Mapping Content Planner (content_plan) -> entri Content Calendar agar apa
# yang di-schedule di planner otomatis muncul di kalender.
_PLANNER_TYPE_MAP = {"blog": "article", "email_campaign": "campaign",
                     "instagram": "social_post", "linkedin": "social_post",
                     "youtube": "social_post", "tiktok": "social_post"}
_PLANNER_STATUS_MAP = {"idea": "draft", "draft": "draft",
                       "scheduled": "scheduled", "published": "published"}


def _planner_as_calendar(d: dict) -> dict:
    return {
        "id": f"plan-{d['_id']}",
        "planner_id": str(d["_id"]),
        "source": "planner",
        "title": d.get("title", ""),
        "type": _PLANNER_TYPE_MAP.get(d.get("channel", "blog"), "article"),
        "channel": d.get("channel", "blog"),
        "scheduled_at": f"{d.get('publish_date', '')}T09:00:00",
        "status": _PLANNER_STATUS_MAP.get(d.get("status", "idea"), "draft"),
        "owner": d.get("owner", ""),
        "notes": d.get("hook", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


_CAL_TYPES = {"article", "campaign", "social_post"}


_CAL_STATUSES = {"draft", "scheduled", "published"}


@router.get("/admin/content-calendar")
async def calendar_list(staff=Depends(get_current_staff),
                        date_from: Optional[str] = None,
                        date_to: Optional[str] = None):
    db = await _get_db()
    query: dict = {}
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to + "T23:59:59"
        query["scheduled_at"] = rng
    docs = await db.content_calendar.find(query).sort("scheduled_at", 1).to_list(1000)
    out = [_serialize_calendar(d) for d in docs]
    # Gabungkan item Content Planner yang punya publish_date (yyyy-mm-dd)
    prng: dict = {"$ne": ""}
    if date_from:
        prng["$gte"] = date_from[:10]
    if date_to:
        prng["$lte"] = date_to[:10]
    pdocs = await db.content_plan.find({"publish_date": prng}).sort("publish_date", 1).to_list(1000)
    out.extend(_planner_as_calendar(d) for d in pdocs)
    out.sort(key=lambda x: x.get("scheduled_at") or "")
    return out


@router.post("/admin/content-calendar")
async def calendar_create(payload: dict, staff=Depends(get_current_content)):
    db = await _get_db()
    title = (payload.get("title") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title is required")
    ctype = payload.get("type") or "article"
    if ctype not in _CAL_TYPES:
        raise HTTPException(status_code=400, detail=f"type must be one of {sorted(_CAL_TYPES)}")
    status = payload.get("status") or "draft"
    if status not in _CAL_STATUSES:
        raise HTTPException(status_code=400, detail=f"status must be one of {sorted(_CAL_STATUSES)}")
    doc = {
        "title": title,
        "type": ctype,
        "scheduled_at": payload.get("scheduled_at") or _now(),
        "status": status,
        "linked_article_id": payload.get("linked_article_id"),
        "owner_id": staff["id"],
        "notes": payload.get("notes") or "",
        "created_at": _now(),
    }
    r = await db.content_calendar.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_calendar(doc)


@router.put("/admin/content-calendar/{cid}")
async def calendar_update(cid: str, payload: dict, staff=Depends(get_current_content)):
    db = await _get_db()
    d = await db.content_calendar.find_one({"_id": _oid(cid)})
    if not d:
        raise HTTPException(status_code=404, detail="Calendar entry not found")
    upd: dict = {}
    for k in ("title", "scheduled_at", "linked_article_id", "notes"):
        if k in payload:
            upd[k] = payload[k]
    if "type" in payload:
        if payload["type"] not in _CAL_TYPES:
            raise HTTPException(status_code=400, detail=f"type must be one of {sorted(_CAL_TYPES)}")
        upd["type"] = payload["type"]
    if "status" in payload:
        if payload["status"] not in _CAL_STATUSES:
            raise HTTPException(status_code=400, detail=f"status must be one of {sorted(_CAL_STATUSES)}")
        upd["status"] = payload["status"]
    if upd:
        await db.content_calendar.update_one({"_id": d["_id"]}, {"$set": upd})
    d = await db.content_calendar.find_one({"_id": d["_id"]})
    return _serialize_calendar(d)


@router.delete("/admin/content-calendar/{cid}")
async def calendar_delete(cid: str, staff=Depends(get_current_content)):
    db = await _get_db()
    r = await db.content_calendar.delete_one({"_id": _oid(cid)})
    if not r.deleted_count:
        raise HTTPException(status_code=404, detail="Calendar entry not found")
    return {"deleted": 1}


async def _sync_article_calendar(db, article: dict, staff) -> None:
    """When an article is published, upsert its calendar entry to published.
    Fire-and-forget: never blocks the article save."""
    try:
        if not article or article.get("status") != "published":
            return
        aid = str(article["_id"])
        await db.content_calendar.update_one(
            {"linked_article_id": aid},
            {"$set": {"title": article.get("title", ""),
                      "type": "article",
                      "status": "published",
                      "scheduled_at": article.get("published_at") or _now(),
                      "linked_article_id": aid},
             "$setOnInsert": {"owner_id": staff.get("id"), "notes": "",
                               "created_at": _now()}},
            upsert=True,
        )
    except Exception:
        logging.getLogger("portal.calendar").exception("calendar sync failed")

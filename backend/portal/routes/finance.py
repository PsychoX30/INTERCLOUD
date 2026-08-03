"""Finance: assets, expenses, ledgers, salary/fee slips, reports, cashflow, credit notes.

Split from the former monolithic routes.py - behavior preserved 1:1.
"""
import os
import asyncio
import logging
import secrets
import re
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Request, WebSocket
from bson import ObjectId
from datetime import datetime, timezone, timedelta
from typing import List, Optional


from ..auth import (
    verify_password, hash_password, create_access_token,
    get_current_user, get_current_admin, get_current_staff, get_current_content,
    require_roles, sales_can_access,
    STAFF_ROLES, FINANCE_ROLES, BILLING_ROLES, CATALOG_ROLES,
    OPS_ROLES, USER_MGMT_ROLES, TICKET_ROLES, CONTENT_ROLES,
)
from ..audit import log_audit, serialize as _serialize_audit
from ..secretbox import (dec_value as _sb_dec, enc_value as _sb_enc,
                         decrypt_config as _sb_dec_config)
from .. import integrations_v2 as iv2
from .billing import _apply_pending_upgrade  # noqa: E402
from .provision import _provision_order_from_invoice  # noqa: E402
from .documents import _idr, _long_date, _render_pdf_bytes  # noqa: E402
from .shared import _get_db, _iso, _next_number, _now, _oid, _sum_applied_credit  # noqa: E402
from fastapi.responses import HTMLResponse  # noqa: E402
from fastapi.responses import Response  # noqa: E402
from portal.branding import get_branding as _get_branding_dict  # noqa: E402

router = APIRouter()


# Finance
@router.get("/admin/finance/summary")
async def admin_finance_summary(admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    # Aggregate by month for last 12 months
    paid = await db.invoices.find({"status": "paid"}).to_list(5000)
    by_month = {}
    for inv in paid:
        p = inv.get("paid_at") or inv.get("created_at", "")
        if not p:
            continue
        key = p[:7]  # YYYY-MM
        by_month[key] = by_month.get(key, 0) + inv.get("total", 0)
    series = sorted(
        [{"month": k, "revenue": v} for k, v in by_month.items()],
        key=lambda x: x["month"],
    )
    unpaid = await db.invoices.find({"status": {"$in": ["unpaid", "overdue"]}}).to_list(2000)
    outstanding = sum(d.get("total", 0) for d in unpaid)
    total_revenue = sum(d.get("total", 0) for d in paid)
    return {
        "total_revenue": total_revenue,
        "outstanding": outstanding,
        "paid_invoices": len(paid),
        "monthly_series": series,
    }


# ============================================================
# ASSETS (native asset tracking + STRAIGHT-LINE depreciation)
# Formula (Metode Garis Lurus):
#   Penyusutan per Tahun = (Harga Perolehan − Nilai Sisa) / Umur Ekonomis
#   Penyusutan per Bulan = Penyusutan per Tahun / 12
#   Akumulasi Penyusutan = Penyusutan per Bulan × bulan_terpakai
#   Nilai Buku            = max(Harga Perolehan − Akumulasi Penyusutan, Nilai Sisa)
# ============================================================
def _asset_life_years(a: dict) -> int:
    """Effective useful-life in years. Prefer explicit field, fall back to
    legacy fields (`useful_life_months`, `depreciation_percent`) so we stay
    backwards-compatible with data seeded before the straight-line rewrite."""
    life_y = int(a.get("useful_life_years", 0) or 0)
    if life_y > 0:
        return life_y
    life_m = int(a.get("useful_life_months", 0) or 0)
    if life_m > 0:
        return max(1, round(life_m / 12))
    dep_pct = float(a.get("depreciation_percent", 0) or 0)
    if dep_pct > 0:
        return max(1, round(100.0 / dep_pct))
    return 0


def _asset_depreciation(a: dict) -> dict:
    """Compute straight-line depreciation snapshot for an asset document."""
    value = float(a.get("value", 0) or 0)
    salvage = float(a.get("salvage_value", 0) or 0)
    life_y = _asset_life_years(a)
    purchase = a.get("purchase_date", "") or ""

    if life_y <= 0 or not purchase:
        return {
            "life_years": life_y,
            "depreciable_base": max(value - salvage, 0.0),
            "annual_depreciation": 0.0,
            "monthly_depreciation": 0.0,
            "months_elapsed": 0,
            "total_months": life_y * 12,
            "accumulated_depreciation": 0.0,
            "book_value": round(value, 2),
            "is_fully_depreciated": False,
        }

    base = max(value - salvage, 0.0)
    annual = base / life_y
    monthly = annual / 12.0
    total_months = life_y * 12

    try:
        p = datetime.fromisoformat(purchase[:10])
    except Exception:
        return {
            "life_years": life_y, "depreciable_base": base,
            "annual_depreciation": round(annual, 2), "monthly_depreciation": round(monthly, 2),
            "months_elapsed": 0, "total_months": total_months,
            "accumulated_depreciation": 0.0, "book_value": round(value, 2),
            "is_fully_depreciated": False,
        }

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    months = (now.year - p.year) * 12 + (now.month - p.month)
    if now.day >= p.day:
        months += 1
    months = max(0, min(months, total_months))

    accumulated = round(monthly * months, 2)
    book = max(round(value - accumulated, 2), salvage)
    return {
        "life_years": life_y,
        "depreciable_base": round(base, 2),
        "annual_depreciation": round(annual, 2),
        "monthly_depreciation": round(monthly, 2),
        "months_elapsed": months,
        "total_months": total_months,
        "accumulated_depreciation": accumulated,
        "book_value": book,
        "is_fully_depreciated": months >= total_months,
    }


def _asset_book_value(a: dict) -> float:
    """Current book value using the straight-line method (floored at salvage)."""
    return _asset_depreciation(a)["book_value"]


def _asset_schedule(a: dict) -> list:
    """Yearly schedule from purchase year through end of useful life."""
    value = float(a.get("value", 0) or 0)
    salvage = float(a.get("salvage_value", 0) or 0)
    life_y = _asset_life_years(a)
    purchase = a.get("purchase_date", "") or ""
    if life_y <= 0 or not purchase:
        return []
    base = max(value - salvage, 0.0)
    annual = base / life_y
    try:
        start_year = datetime.fromisoformat(purchase[:10]).year
    except Exception:
        return []
    rows, accumulated = [], 0.0
    for i in range(life_y):
        accumulated = min(accumulated + annual, base)
        book = max(value - accumulated, salvage)
        rows.append({
            "period": i + 1,
            "year": start_year + i,
            "depreciation": round(annual, 2),
            "accumulated_depreciation": round(accumulated, 2),
            "book_value": round(book, 2),
        })
    return rows


def _serialize_asset(d):
    dep = _asset_depreciation(d)
    return {
        "id": str(d["_id"]),
        "name": d.get("name", ""),
        "category": d.get("category", "server"),
        "serial_number": d.get("serial_number", ""),
        "location": d.get("location", ""),
        "vendor": d.get("vendor", ""),
        "value": float(d.get("value", 0)),
        "salvage_value": float(d.get("salvage_value", 0) or 0),
        "useful_life_years": dep["life_years"],
        # legacy fields kept for backward compat with UI/tests still referencing them
        "depreciation_percent": float(d.get("depreciation_percent", 0) or 0),
        "useful_life_months": int(d.get("useful_life_months", 0) or 0),
        "purchase_date": d.get("purchase_date", ""),
        "annual_depreciation": dep["annual_depreciation"],
        "monthly_depreciation": dep["monthly_depreciation"],
        "accumulated_depreciation": dep["accumulated_depreciation"],
        "book_value": dep["book_value"],
        # kept for compat with old frontend field name
        "depreciated_amount": dep["accumulated_depreciation"],
        "months_elapsed": dep["months_elapsed"],
        "total_months": dep["total_months"],
        "is_fully_depreciated": dep["is_fully_depreciated"],
        "status": d.get("status", "active"),
        "disposed_at": d.get("disposed_at", ""),
        "notes": d.get("notes", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/assets")
async def assets_list(admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    docs = await db.assets.find({}).sort("created_at", -1).to_list(2000)
    return [_serialize_asset(d) for d in docs]


@router.get("/admin/assets/{aid}")
async def assets_get(aid: str, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    d = await db.assets.find_one({"_id": _oid(aid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    payload = _serialize_asset(d)
    payload["schedule"] = _asset_schedule(d)
    return payload


def _coerce_asset_payload(payload: dict) -> dict:
    """Normalize incoming asset payload. Falls back to legacy fields when
    salvage/useful_life_years are omitted."""
    life_y = payload.get("useful_life_years")
    if life_y in (None, "", 0, "0"):
        # derive from legacy fields if provided in same payload
        dep_pct = float(payload.get("depreciation_percent", 0) or 0)
        life_m = int(payload.get("useful_life_months", 0) or 0)
        if life_m > 0:
            life_y = max(1, round(life_m / 12))
        elif dep_pct > 0:
            life_y = max(1, round(100.0 / dep_pct))
        else:
            life_y = 0
    return {
        "salvage_value": float(payload.get("salvage_value", 0) or 0),
        "useful_life_years": int(life_y or 0),
    }


@router.post("/admin/assets")
async def assets_create(payload: dict, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    coerced = _coerce_asset_payload(payload)
    doc = {
        "name": payload.get("name", ""),
        "category": payload.get("category", "server"),
        "serial_number": payload.get("serial_number", ""),
        "location": payload.get("location", ""),
        "vendor": payload.get("vendor", ""),
        "value": float(payload.get("value", 0) or 0),
        "salvage_value": coerced["salvage_value"],
        "useful_life_years": coerced["useful_life_years"],
        # legacy fields retained if the client still sends them
        "depreciation_percent": float(payload.get("depreciation_percent", 0) or 0),
        "useful_life_months": int(payload.get("useful_life_months", 0) or 0),
        "purchase_date": payload.get("purchase_date", ""),
        "status": payload.get("status") if payload.get("status") in ("active", "disposed") else "active",
        "notes": payload.get("notes", ""),
        "created_at": _now(),
    }
    if doc["status"] == "disposed":
        doc["disposed_at"] = payload.get("disposed_at") or datetime.now(timezone.utc).date().isoformat()
    r = await db.assets.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_asset(doc)


@router.put("/admin/assets/{aid}")
async def assets_update(aid: str, payload: dict, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    allowed = {
        "name", "category", "serial_number", "location", "vendor", "value",
        "salvage_value", "useful_life_years",
        "depreciation_percent", "useful_life_months",
        "purchase_date", "notes", "status", "disposed_at",
    }
    upd = {k: v for k, v in payload.items() if k in allowed}
    if "status" in upd:
        if upd["status"] not in ("active", "disposed"):
            upd.pop("status")
        elif upd["status"] == "disposed" and not upd.get("disposed_at"):
            upd["disposed_at"] = datetime.now(timezone.utc).date().isoformat()
        elif upd["status"] == "active":
            upd["disposed_at"] = ""
    coerced = _coerce_asset_payload({**upd})
    upd["salvage_value"] = coerced["salvage_value"]
    if coerced["useful_life_years"] > 0:
        upd["useful_life_years"] = coerced["useful_life_years"]
    for k in ("value", "depreciation_percent"):
        if k in upd:
            upd[k] = float(upd[k] or 0)
    if "useful_life_months" in upd:
        upd["useful_life_months"] = int(upd["useful_life_months"] or 0)
    await db.assets.update_one({"_id": _oid(aid)}, {"$set": upd})
    d = await db.assets.find_one({"_id": _oid(aid)})
    if not d:
        raise HTTPException(status_code=404, detail="Not found")
    return _serialize_asset(d)


@router.delete("/admin/assets/{aid}")
async def assets_delete(aid: str, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    r = await db.assets.delete_one({"_id": _oid(aid)})
    return {"deleted": r.deleted_count}


# ============================================================
# EXPENSES (manual bookkeeping)
# ============================================================
def _serialize_expense(d):
    return {
        "id": str(d["_id"]),
        "date": d.get("date", ""),
        "category": d.get("category", "other"),
        "vendor": d.get("vendor", ""),
        "amount": float(d.get("amount", 0)),
        "description": d.get("description", ""),
        "created_at": _iso(d.get("created_at", "")),
    }


@router.get("/admin/expenses")
async def expenses_list(admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    docs = await db.expenses.find({}).sort("date", -1).to_list(5000)
    return [_serialize_expense(d) for d in docs]


@router.post("/admin/expenses")
async def expenses_create(payload: dict, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    doc = {
        "date": payload.get("date", datetime.now(timezone.utc).date().isoformat()),
        "category": payload.get("category", "other"),
        "vendor": payload.get("vendor", ""),
        "amount": float(payload.get("amount", 0) or 0),
        "description": payload.get("description", ""),
        "created_at": _now(),
    }
    r = await db.expenses.insert_one(doc)
    doc["_id"] = r.inserted_id
    return _serialize_expense(doc)


@router.delete("/admin/expenses/{eid}")
async def expenses_delete(eid: str, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    r = await db.expenses.delete_one({"_id": _oid(eid)})
    return {"deleted": r.deleted_count}


# Extended finance report (revenue + expenses + assets)
@router.get("/admin/assets/report/depreciation")
async def assets_depreciation_report(months: int = 12, admin=Depends(require_roles("admin", "finance"))):
    """Laporan total beban depresiasi aset aktif per periode (bulanan, N bulan terakhir)."""
    db = await _get_db()
    months = max(1, min(months, 60))
    assets = await db.assets.find({}).to_list(2000)
    now = datetime.now(timezone.utc)
    periods = []
    for i in range(months - 1, -1, -1):
        y = now.year + (now.month - 1 - i) // 12
        mo = (now.month - 1 - i) % 12 + 1
        periods.append(f"{y:04d}-{mo:02d}")
    rows = []
    for pkey in periods:
        cur_idx = int(pkey[:4]) * 12 + int(pkey[5:7]) - 1
        total, count = 0.0, 0
        for a in assets:
            dep = _asset_depreciation(a)
            monthly = dep["monthly_depreciation"]
            purchase = (a.get("purchase_date") or "")[:10]
            if monthly <= 0 or not purchase or pkey < purchase[:7]:
                continue
            try:
                pd_ = datetime.strptime(purchase, "%Y-%m-%d")
            except Exception:
                continue
            end_idx = pd_.year * 12 + (pd_.month - 1) + dep["total_months"]
            if cur_idx >= end_idx:
                continue
            disposed = (a.get("disposed_at") or "")[:7]
            if a.get("status") == "disposed" and disposed and pkey > disposed:
                continue
            total += monthly
            count += 1
        rows.append({"period": pkey, "depreciation": round(total, 2), "active_assets": count})
    return {"months": months, "rows": rows,
            "total_depreciation": round(sum(r["depreciation"] for r in rows), 2)}


@router.get("/admin/finance/report")
async def admin_finance_report(admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    paid = await db.invoices.find({"status": "paid"}).to_list(5000)
    total_revenue = sum(d.get("total", 0) for d in paid)
    expenses = await db.expenses.find({}).to_list(5000)
    total_expenses = sum(d.get("amount", 0) for d in expenses)
    net_profit = total_revenue - total_expenses

    assets = await db.assets.find({}).to_list(2000)
    total_assets_value = sum(float(a.get("value", 0)) for a in assets)
    net_assets_value = sum(_asset_book_value(a) for a in assets)
    total_depreciation = round(total_assets_value - net_assets_value, 2)

    # Revenue & expenses by month (last 12 months)
    by_month_rev, by_month_exp = {}, {}
    for inv in paid:
        p = inv.get("paid_at") or inv.get("created_at", "")
        if not p:
            continue
        key = p[:7]
        by_month_rev[key] = by_month_rev.get(key, 0) + inv.get("total", 0)
    for e in expenses:
        key = (e.get("date") or "")[:7]
        if not key:
            continue
        by_month_exp[key] = by_month_exp.get(key, 0) + e.get("amount", 0)

    all_keys = sorted(set(by_month_rev.keys()) | set(by_month_exp.keys()))
    monthly = [
        {"month": k, "revenue": by_month_rev.get(k, 0), "expenses": by_month_exp.get(k, 0),
         "profit": by_month_rev.get(k, 0) - by_month_exp.get(k, 0)}
        for k in all_keys
    ]

    return {
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "total_assets_value": total_assets_value,
        "net_assets_value": net_assets_value,
        "total_depreciation": total_depreciation,
        "asset_count": len(assets),
        "monthly": monthly,
    }


@router.post("/admin/credit-notes/preview")
async def credit_notes_preview(payload: dict, admin=Depends(require_roles("admin", "finance"))):
    """Render a DRAFT credit note as an inline PDF without saving anything."""
    db = await _get_db()
    try:
        inv_oid = ObjectId(payload.get("invoice_id") or "")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid invoice_id")
    invoice = await db.invoices.find_one({"_id": inv_oid})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be > 0")
    user = await db.users.find_one({"_id": invoice["user_id"]}) or {}
    cn = {
        "number": "CN-PREVIEW",
        "invoice_id": inv_oid,
        "invoice_number": invoice.get("number", ""),
        "user_id": invoice.get("user_id"),
        "amount": amount,
        "reason": (payload.get("reason") or "").strip() or "-",
        "notes": payload.get("notes") or "",
        "status": "draft",
        "created_at": _now(),
    }
    branding = await _get_branding_dict(db)
    html = _credit_note_html(cn=cn, invoice=invoice, billed_to=user, for_pdf=True,
                             logo_url=branding["logo_dark"])
    return Response(content=_render_pdf_bytes(html), media_type="application/pdf",
                    headers={"Content-Disposition": 'inline; filename="CreditNote-PREVIEW.pdf"'})


# ============================================================
# FINANCE V2 - Kas Kecil / Salaries / Sales Fees / Excel reports
# ============================================================
from fastapi.responses import StreamingResponse  # noqa: E402


import io as _io  # noqa: E402


def _generic_ledger_serialize(d: dict) -> dict:
    return {
        "id": str(d["_id"]),
        "date": d.get("date", ""),
        "amount": float(d.get("amount") or 0),
        "category": d.get("category", ""),
        "notes": d.get("notes", ""),
        "vendor": d.get("vendor", ""),
        "employee": d.get("employee", ""),
        "sales_person": d.get("sales_person", ""),
        "invoice_number": d.get("invoice_number", ""),
        "period_yyyy_mm": d.get("period_yyyy_mm") or (d.get("date", "")[:7]),
        "created_at": _iso(d.get("created_at", "")),
    }


def _month_locked(period_yyyy_mm: str) -> bool:
    """A month is 'locked' once its data is frozen for reporting.

    Rule: a month M/Y is locked as soon as we're in month M+1/Y or later.
    Additionally, once the calendar year Y completes, ALL months of Y stay
    locked until January 5th of Y+1 (audit window). Only strictly future
    months are always mutable.
    """
    today = datetime.now(timezone.utc).date()
    try:
        y, m = int(period_yyyy_mm[:4]), int(period_yyyy_mm[5:7])
    except Exception:
        return False
    # Future months → not locked
    if (y, m) > (today.year, today.month):
        return False
    # Prior calendar year: locked until Jan 5 of Y+1
    if y < today.year:
        if today.year == y + 1 and today.month == 1 and today.day <= 5:
            return False   # 5-day amendment window
        return True
    # Same year, prior month: locked
    if y == today.year and m < today.month:
        return True
    return False


def _mk_ledger_router(*, collection: str, label: str, extra_fields: list):
    """Factory that creates a set of endpoints for a simple ledger table.

    Each ledger has: date (YYYY-MM-DD), amount, notes, plus `extra_fields`.
    We register 3 endpoints per ledger: list / create / delete.
    """

    async def _list(admin=Depends(get_current_admin)):
        db = await _get_db()
        docs = await db[collection].find({}).sort("date", -1).to_list(5000)
        return [_generic_ledger_serialize(d) for d in docs]

    async def _create(payload: dict, admin=Depends(get_current_admin)):
        db = await _get_db()
        date_str = payload.get("date") or datetime.now(timezone.utc).date().isoformat()
        period = date_str[:7]
        if _month_locked(period):
            raise HTTPException(status_code=403, detail=f"Cannot add {label} for locked month {period}. Contact finance to unlock.")
        doc = {"date": date_str, "amount": float(payload.get("amount", 0) or 0),
               "notes": payload.get("notes", ""), "period_yyyy_mm": period,
               "created_at": _now()}
        for k in extra_fields:
            doc[k] = payload.get(k, "")
        r = await db[collection].insert_one(doc)
        doc["_id"] = r.inserted_id
        return _generic_ledger_serialize(doc)

    async def _delete(item_id: str, admin=Depends(get_current_admin)):
        db = await _get_db()
        d = await db[collection].find_one({"_id": _oid(item_id)})
        if not d:
            raise HTTPException(status_code=404, detail="Not found")
        if _month_locked(d.get("period_yyyy_mm") or d.get("date", "")[:7]):
            raise HTTPException(status_code=403, detail=f"Cannot delete {label} from a locked month")
        r = await db[collection].delete_one({"_id": _oid(item_id)})
        return {"deleted": r.deleted_count}

    return _list, _create, _delete


# --- kas kecil (petty cash) ---
_kk_list, _kk_create, _kk_delete = _mk_ledger_router(
    collection="kas_kecil", label="petty cash", extra_fields=["category", "vendor"],
)


router.get("/admin/kas-kecil")(_kk_list)


router.post("/admin/kas-kecil")(_kk_create)


router.delete("/admin/kas-kecil/{item_id}")(_kk_delete)


# --- salaries ---
_sal_list, _sal_create, _sal_delete = _mk_ledger_router(
    collection="salaries", label="salary", extra_fields=["employee", "category"],
)


router.get("/admin/salaries")(_sal_list)


router.post("/admin/salaries")(_sal_create)


router.delete("/admin/salaries/{item_id}")(_sal_delete)


# --- sales fees ---
_sf_list, _sf_create, _sf_delete = _mk_ledger_router(
    collection="sales_fees", label="sales fee", extra_fields=["sales_person", "invoice_number"],
)


router.get("/admin/sales-fees")(_sf_list)


router.post("/admin/sales-fees")(_sf_create)


router.delete("/admin/sales-fees/{item_id}")(_sf_delete)


@router.get("/documents/salary-slip/{sid}")
async def render_salary_slip(sid: str, format: str = "pdf", admin=Depends(get_current_admin)):
    """UAT-034: slip gaji PDF per entri salary (WeasyPrint)."""
    db = await _get_db()
    d = await db.salaries.find_one({"_id": _oid(sid)})
    if not d:
        raise HTTPException(status_code=404, detail="Salary entry not found")
    period = d.get("period_yyyy_mm") or (d.get("date") or "")[:7]
    amount = float(d.get("amount") or 0)
    amount_str = "Rp " + f"{amount:,.0f}".replace(",", ".")
    employee = d.get("employee") or "-"
    category = d.get("category") or "Gaji pokok"
    issued = datetime.now(timezone.utc).date().isoformat()
    html = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 24mm 18mm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; color: #0f172a; font-size: 13px; }}
      .head {{ display: flex; justify-content: space-between; border-bottom: 3px solid #0a2350; padding-bottom: 14px; }}
      .co {{ font-size: 19px; font-weight: 800; color: #0a2350; }}
      .co small {{ display:block; font-size: 10px; color:#64748b; font-weight: 400; margin-top: 3px; }}
      h1 {{ font-size: 15px; letter-spacing: 2px; color: #0a2350; margin: 26px 0 4px; text-transform: uppercase; }}
      .meta {{ color: #64748b; font-size: 11px; margin-bottom: 20px; }}
      table {{ width: 100%; border-collapse: collapse; }}
      td, th {{ padding: 10px 12px; border: 1px solid #e2e8f0; text-align: left; }}
      th {{ background: #f8fafc; font-size: 10px; text-transform: uppercase; letter-spacing: 1px; color: #64748b; width: 34%; }}
      .amt {{ font-size: 17px; font-weight: 800; color: #0a2350; }}
      .foot {{ margin-top: 44px; display: flex; justify-content: space-between; font-size: 11px; color: #64748b; }}
      .sig {{ text-align: center; }}
      .sig .line {{ margin-top: 56px; border-top: 1px solid #94a3b8; padding-top: 5px; width: 190px; }}
      .conf {{ margin-top: 26px; font-size: 9.5px; color: #94a3b8; }}
    </style></head><body>
      <div class="head">
        <div class="co">INTERCLOUD DIGITAL INOVASI<small>Jakarta, Indonesia · support@intercloud-digital.com · +62 878-1239-7187</small></div>
        <div style="text-align:right"><div style="font-size:11px;color:#64748b">SLIP GAJI</div>
          <div style="font-weight:800;color:#0a2350">{period}</div></div>
      </div>
      <h1>Slip Gaji Karyawan</h1>
      <div class="meta">Diterbitkan {issued} · No. ref {str(d["_id"])[-8:].upper()}</div>
      <table>
        <tr><th>Nama karyawan</th><td style="font-weight:700">{employee}</td></tr>
        <tr><th>Periode</th><td>{period}</td></tr>
        <tr><th>Kategori</th><td>{category}</td></tr>
        <tr><th>Tanggal pembayaran</th><td>{(d.get("date") or "")[:10]}</td></tr>
        <tr><th>Jumlah diterima (net)</th><td class="amt">{amount_str}</td></tr>
        <tr><th>Catatan</th><td>{d.get("notes") or "-"}</td></tr>
      </table>
      <div class="foot">
        <div class="sig">Diserahkan oleh,<div class="line">Finance - Intercloud</div></div>
        <div class="sig">Diterima oleh,<div class="line">{employee}</div></div>
      </div>
      <div class="conf">Dokumen ini bersifat rahasia dan dihasilkan otomatis oleh Intercloud Portal.</div>
    </body></html>"""
    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        emp_slug = re.sub(r"[^A-Za-z0-9_-]+", "-", employee).strip("-") or "karyawan"
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="SlipGaji-{emp_slug}-{period}.pdf"'})
    return HTMLResponse(content=html)


@router.get("/documents/sales-fee-slip/{sid}")
async def render_sales_fee_slip(sid: str, format: str = "pdf", admin=Depends(get_current_admin)):
    """Slip fee sales PDF per entri sales_fees (WeasyPrint)."""
    db = await _get_db()
    d = await db.sales_fees.find_one({"_id": _oid(sid)})
    if not d:
        raise HTTPException(status_code=404, detail="Sales fee entry not found")
    period = d.get("period_yyyy_mm") or (d.get("date") or "")[:7]
    amount = float(d.get("amount") or 0)
    amount_str = "Rp " + f"{amount:,.0f}".replace(",", ".")
    person = d.get("sales_person") or "-"
    invoice_no = d.get("invoice_number") or "-"
    issued = datetime.now(timezone.utc).date().isoformat()
    html = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 24mm 18mm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; color: #0f172a; font-size: 13px; }}
      .head {{ display: flex; justify-content: space-between; border-bottom: 3px solid #0a2350; padding-bottom: 14px; }}
      .co {{ font-size: 19px; font-weight: 800; color: #0a2350; }}
      .co small {{ display:block; font-size: 10px; color:#64748b; font-weight: 400; margin-top: 3px; }}
      h1 {{ font-size: 15px; letter-spacing: 2px; color: #0a2350; margin: 26px 0 4px; text-transform: uppercase; }}
      .meta {{ color: #64748b; font-size: 11px; margin-bottom: 20px; }}
      table {{ width: 100%; border-collapse: collapse; }}
      td, th {{ padding: 10px 12px; border: 1px solid #e2e8f0; text-align: left; }}
      th {{ background: #f8fafc; font-size: 10px; text-transform: uppercase; letter-spacing: 1px; color: #64748b; width: 34%; }}
      .amt {{ font-size: 17px; font-weight: 800; color: #0a2350; }}
      .foot {{ margin-top: 44px; display: flex; justify-content: space-between; font-size: 11px; color: #64748b; }}
      .sig {{ text-align: center; }}
      .sig .line {{ margin-top: 56px; border-top: 1px solid #94a3b8; padding-top: 5px; width: 190px; }}
      .conf {{ margin-top: 26px; font-size: 9.5px; color: #94a3b8; }}
    </style></head><body>
      <div class="head">
        <div class="co">INTERCLOUD DIGITAL INOVASI<small>Jakarta, Indonesia · support@intercloud-digital.com · +62 878-1239-7187</small></div>
        <div style="text-align:right"><div style="font-size:11px;color:#64748b">SLIP FEE SALES</div>
          <div style="font-weight:800;color:#0a2350">{period}</div></div>
      </div>
      <h1>Slip Fee Penjualan</h1>
      <div class="meta">Diterbitkan {issued} · No. ref {str(d["_id"])[-8:].upper()}</div>
      <table>
        <tr><th>Nama sales</th><td style="font-weight:700">{person}</td></tr>
        <tr><th>Periode</th><td>{period}</td></tr>
        <tr><th>Invoice terkait</th><td>{invoice_no}</td></tr>
        <tr><th>Tanggal pembayaran</th><td>{(d.get("date") or "")[:10]}</td></tr>
        <tr><th>Jumlah fee (net)</th><td class="amt">{amount_str}</td></tr>
        <tr><th>Catatan</th><td>{d.get("notes") or "-"}</td></tr>
      </table>
      <div class="foot">
        <div class="sig">Diserahkan oleh,<div class="line">Finance - Intercloud</div></div>
        <div class="sig">Diterima oleh,<div class="line">{person}</div></div>
      </div>
      <div class="conf">Dokumen ini bersifat rahasia dan dihasilkan otomatis oleh Intercloud Portal.</div>
    </body></html>"""
    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        slug = re.sub(r"[^A-Za-z0-9_-]+", "-", person).strip("-") or "sales"
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="SlipFeeSales-{slug}-{period}.pdf"'})
    return HTMLResponse(content=html)


# ---------------- Finance detailed report ----------------
@router.get("/admin/finance/detailed")
async def finance_detailed(admin=Depends(require_roles("admin", "finance"))):
    """Returns paid-invoice detail + all four expense ledgers + assets + depreciation.

    The frontend Finance page uses this to render tabbed detailed tables.
    """
    db = await _get_db()
    paid = await db.invoices.find({"status": "paid"}).sort("paid_at", -1).to_list(5000)
    revenue_rows = [{
        "id": str(inv["_id"]),
        "number": inv.get("number", ""),
        "paid_at": inv.get("paid_at", "")[:10],
        "customer": inv.get("customer_name") or "",
        "total": float(inv.get("total") or 0),
        "period_yyyy_mm": (inv.get("paid_at") or inv.get("created_at", ""))[:7],
    } for inv in paid]

    async def _fetch(coll):
        docs = await db[coll].find({}).sort("date", -1).to_list(5000)
        return [_generic_ledger_serialize(d) for d in docs]

    expenses_rows = []
    async for d in db.expenses.find({}).sort("date", -1):
        expenses_rows.append({
            "id": str(d["_id"]),
            "date": d.get("date", ""),
            "category": d.get("category", ""),
            "vendor": d.get("vendor", ""),
            "amount": float(d.get("amount", 0)),
            "description": d.get("description", ""),
            "period_yyyy_mm": (d.get("date") or "")[:7],
        })
    kk_rows = await _fetch("kas_kecil")
    sal_rows = await _fetch("salaries")
    sf_rows = await _fetch("sales_fees")

    assets_rows = []
    async for a in db.assets.find({}):
        dep = _asset_depreciation(a)
        assets_rows.append({
            "id": str(a["_id"]),
            "name": a.get("name", ""),
            "category": a.get("category", ""),
            "purchase_date": a.get("purchase_date", ""),
            "value": float(a.get("value", 0)),
            "salvage_value": float(a.get("salvage_value", 0) or 0),
            "useful_life_years": dep["life_years"],
            "useful_life_months": int(a.get("useful_life_months", 0) or 0),
            "annual_depreciation": dep["annual_depreciation"],
            "monthly_depreciation": dep["monthly_depreciation"],
            "book_value": dep["book_value"],
            "accumulated_depreciation": dep["accumulated_depreciation"],
        })

    total_revenue = sum(r["total"] for r in revenue_rows)
    total_expenses = sum(e["amount"] for e in expenses_rows)
    total_kas_kecil = sum(r["amount"] for r in kk_rows)
    total_salaries = sum(r["amount"] for r in sal_rows)
    total_sales_fees = sum(r["amount"] for r in sf_rows)
    total_all_expenses = total_expenses + total_kas_kecil + total_salaries + total_sales_fees
    total_depreciation = sum(a["accumulated_depreciation"] for a in assets_rows)
    return {
        "revenue_rows": revenue_rows,
        "expenses_rows": expenses_rows,
        "kas_kecil_rows": kk_rows,
        "salaries_rows": sal_rows,
        "sales_fees_rows": sf_rows,
        "assets_rows": assets_rows,
        "totals": {
            "revenue": total_revenue,
            "expenses_recurring": total_expenses,
            "kas_kecil": total_kas_kecil,
            "salaries": total_salaries,
            "sales_fees": total_sales_fees,
            "expenses_all": total_all_expenses,
            "depreciation_accumulated": total_depreciation,
            "net_profit": total_revenue - total_all_expenses - total_depreciation,
        },
    }


# ---------------- Excel report generation ----------------
def _idr_fmt(v):
    return f"Rp {float(v or 0):,.0f}".replace(",", ".")


def _write_xlsx(sheets: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF0A2350")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="FFFEF3C7")
    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, cell in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=cell)
                if isinstance(cell, (int, float)) or (isinstance(cell, str) and cell.startswith("=")):
                    c.number_format = "#,##0"
                if r_idx == 1:
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal="center")
                elif r_idx == len(rows) and str(row[0]).lower().startswith(("total", "net ")):
                    c.font = total_font
                    c.fill = total_fill
        # auto-width
        for col in ws.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, 40)
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _gather_period_data(db, *, year: int, month: Optional[int] = None) -> dict:
    def in_period(dt: str) -> bool:
        if not dt or len(dt) < 7:
            return False
        y = int(dt[:4])
        if y != year:
            return False
        if month is not None:
            return int(dt[5:7]) == month
        return True

    paid = await db.invoices.find({"status": "paid"}).to_list(10000)
    revenue = [i for i in paid if in_period(i.get("paid_at") or i.get("created_at", ""))]
    expenses = [e for e in await db.expenses.find({}).to_list(10000)
                if in_period(e.get("date", ""))]
    kk = [e for e in await db.kas_kecil.find({}).to_list(10000) if in_period(e.get("date", ""))]
    sal = [e for e in await db.salaries.find({}).to_list(10000) if in_period(e.get("date", ""))]
    sf = [e for e in await db.sales_fees.find({}).to_list(10000) if in_period(e.get("date", ""))]
    assets = await db.assets.find({}).to_list(10000)
    return {"revenue": revenue, "expenses": expenses, "kk": kk, "sal": sal, "sf": sf, "assets": assets}


@router.get("/admin/finance/report/monthly/{period}")
async def finance_monthly_xlsx(period: str, admin=Depends(require_roles("admin", "finance"))):
    """`period` is YYYY-MM. Returns an .xlsx with 6 sheets:
    Summary / Revenue / Expenses / Kas Kecil / Salaries / Sales Fees.
    Also freezes the month into `finalized_reports` so it becomes read-only.
    """
    try:
        y, m = int(period[:4]), int(period[5:7])
    except Exception:
        raise HTTPException(status_code=400, detail="Bad period, use YYYY-MM")
    db = await _get_db()
    d = await _gather_period_data(db, year=y, month=m)
    rev_total = sum(float(i.get("total") or 0) for i in d["revenue"])
    exp_total = sum(float(e.get("amount") or 0) for e in d["expenses"])
    kk_total = sum(float(e.get("amount") or 0) for e in d["kk"])
    sal_total = sum(float(e.get("amount") or 0) for e in d["sal"])
    sf_total = sum(float(e.get("amount") or 0) for e in d["sf"])
    all_exp = exp_total + kk_total + sal_total + sf_total
    net_profit = rev_total - all_exp
    summary = [
        ["Line", "Amount (IDR)"],
        ["Revenue (paid invoices)", float(rev_total)],
        ["Expenses (recurring)", float(exp_total)],
        ["Kas Kecil (petty cash)", float(kk_total)],
        ["Salaries", float(sal_total)],
        ["Sales Fees", float(sf_total)],
        ["Total expenses", "=SUM(B3:B6)"],
        ["Net profit (before depreciation)", "=B2-B7"],
    ]
    rev_rows = [["Paid at", "Invoice #", "Customer", "Amount"]] + [
        [i.get("paid_at", "")[:10], i.get("number", ""), i.get("customer_name") or "",
         float(i.get("total") or 0)] for i in d["revenue"]
    ] + [["TOTAL", "", "", f"=SUM(D2:D{len(d['revenue']) + 1})"]]
    exp_rows = [["Date", "Category", "Vendor", "Description", "Amount"]] + [
        [e.get("date", ""), e.get("category", ""), e.get("vendor", ""),
         e.get("description", ""), float(e.get("amount") or 0)] for e in d["expenses"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['expenses']) + 1})"]]
    kk_rows = [["Date", "Category", "Vendor", "Notes", "Amount"]] + [
        [e.get("date", ""), e.get("category", ""), e.get("vendor", ""),
         e.get("notes", ""), float(e.get("amount") or 0)] for e in d["kk"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['kk']) + 1})"]]
    sal_rows = [["Date", "Employee", "Category", "Notes", "Amount"]] + [
        [e.get("date", ""), e.get("employee", ""), e.get("category", ""),
         e.get("notes", ""), float(e.get("amount") or 0)] for e in d["sal"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['sal']) + 1})"]]
    sf_rows = [["Date", "Sales person", "Invoice #", "Notes", "Amount"]] + [
        [e.get("date", ""), e.get("sales_person", ""), e.get("invoice_number", ""),
         e.get("notes", ""), float(e.get("amount") or 0)] for e in d["sf"]
    ] + [["TOTAL", "", "", "", f"=SUM(E2:E{len(d['sf']) + 1})"]]

    xlsx = _write_xlsx([
        (f"Summary {period}", summary),
        ("Revenue", rev_rows),
        ("Expenses", exp_rows),
        ("Kas Kecil", kk_rows),
        ("Salaries", sal_rows),
        ("Sales Fees", sf_rows),
    ])

    # Freeze: save into finalized_reports for the month (idempotent)
    await db.finalized_reports.update_one(
        {"period": period, "kind": "monthly"},
        {"$set": {"period": period, "kind": "monthly", "totals": {
            "revenue": rev_total, "expenses_all": all_exp, "net_profit": net_profit,
        }, "generated_at": _now()}},
        upsert=True,
    )
    filename = f"Intercloud_Finance_{period}.xlsx"
    return StreamingResponse(
        _io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/admin/finance/report/annual/{year}")
async def finance_annual_xlsx(year: int, admin=Depends(require_roles("admin", "finance"))):
    """One workbook with per-month AND cumulative Jan-Dec P&L + assets."""
    db = await _get_db()
    d = await _gather_period_data(db, year=year)

    # Per-month buckets
    months = [f"{year}-{m:02d}" for m in range(1, 13)]
    buckets = {mm: {"rev": 0.0, "exp": 0.0, "kk": 0.0, "sal": 0.0, "sf": 0.0} for mm in months}
    for i in d["revenue"]:
        k = (i.get("paid_at") or i.get("created_at", ""))[:7]
        if k in buckets: buckets[k]["rev"] += float(i.get("total") or 0)
    for e in d["expenses"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["exp"] += float(e.get("amount") or 0)
    for e in d["kk"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["kk"] += float(e.get("amount") or 0)
    for e in d["sal"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["sal"] += float(e.get("amount") or 0)
    for e in d["sf"]:
        k = e.get("date", "")[:7]
        if k in buckets: buckets[k]["sf"] += float(e.get("amount") or 0)

    # Monthly + cumulative sheet
    monthly_rows = [["Month", "Revenue", "Recurring exp.", "Kas Kecil", "Salaries",
                     "Sales Fees", "Total expenses", "Net profit",
                     "Cumulative revenue", "Cumulative net"]]
    cum_rev = cum_net = 0.0
    for idx, mm in enumerate(months):
        b = buckets[mm]
        exps = b["exp"] + b["kk"] + b["sal"] + b["sf"]
        net = b["rev"] - exps
        cum_rev += b["rev"]; cum_net += net
        rn = idx + 2
        monthly_rows.append([mm, float(b["rev"]), float(b["exp"]), float(b["kk"]),
                             float(b["sal"]), float(b["sf"]),
                             f"=SUM(C{rn}:F{rn})", f"=B{rn}-G{rn}",
                             float(cum_rev), float(cum_net)])
    total_rev = sum(b["rev"] for b in buckets.values())
    total_exp_all = sum(b["exp"] + b["kk"] + b["sal"] + b["sf"] for b in buckets.values())
    monthly_rows.append(["TOTAL", "=SUM(B2:B13)", "", "", "", "",
                         "=SUM(G2:G13)", "=B14-G14",
                         float(total_rev), float(total_rev - total_exp_all)])

    # Assets sheet (straight-line depreciation)
    asset_rows = [["Asset", "Category", "Purchased", "Cost", "Salvage",
                   "Useful life (yr)", "Annual depreciation",
                   "Book value", "Accumulated depreciation"]]
    total_cost = total_book = 0.0
    for a in d["assets"]:
        dep = _asset_depreciation(a)
        cost = float(a.get("value", 0)); book = dep["book_value"]
        total_cost += cost; total_book += book
        asset_rows.append([a.get("name", ""), a.get("category", ""),
                           a.get("purchase_date", ""), _idr_fmt(cost),
                           _idr_fmt(float(a.get("salvage_value", 0) or 0)),
                           dep["life_years"],
                           _idr_fmt(dep["annual_depreciation"]),
                           _idr_fmt(book),
                           _idr_fmt(dep["accumulated_depreciation"])])
    asset_rows.append(["TOTAL", "", "", _idr_fmt(total_cost), "", "", "",
                       _idr_fmt(total_book), _idr_fmt(total_cost - total_book)])

    # Details
    rev_rows = [["Paid at", "Invoice #", "Customer", "Amount"]] + [
        [i.get("paid_at", "")[:10], i.get("number", ""), i.get("customer_name") or "",
         _idr_fmt(i.get("total"))] for i in d["revenue"]] + [["TOTAL", "", "", _idr_fmt(total_rev)]]

    xlsx = _write_xlsx([
        (f"P&L {year}", monthly_rows),
        (f"Assets {year}", asset_rows),
        (f"Revenue {year}", rev_rows),
    ])

    await db.finalized_reports.update_one(
        {"period": str(year), "kind": "annual"},
        {"$set": {"period": str(year), "kind": "annual",
                  "totals": {"revenue": total_rev, "expenses_all": total_exp_all,
                             "net_profit": total_rev - total_exp_all},
                  "generated_at": _now()}},
        upsert=True,
    )
    return StreamingResponse(
        _io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Intercloud_Finance_Annual_{year}.xlsx"'},
    )


@router.get("/admin/finance/cashflow-forecast")
async def finance_cashflow_forecast(admin=Depends(require_roles("admin", "finance"))):
    """Proyeksi arus kas 30/60/90 hari: inflow dari invoice unpaid/overdue + renewal
    layanan aktif; outflow dari run-rate 3 bulan terakhir keempat buku beban."""
    db = await _get_db()
    return await _compute_cashflow_forecast(db)


async def _compute_cashflow_forecast(db) -> dict:
    today = datetime.now(timezone.utc).date()
    horizon = today + timedelta(days=91)

    events = []  # (date, amount, kind)
    async for inv in db.invoices.find({"status": {"$in": ["unpaid", "overdue"]}}):
        try:
            due = datetime.strptime((inv.get("due_date") or "")[:10], "%Y-%m-%d").date()
        except Exception:
            continue
        due = max(due, today)
        if due < horizon:
            events.append((due, float(inv.get("total") or 0), "invoice"))
    async for svc in db.services.find({"status": "active"}):
        try:
            due = datetime.strptime((svc.get("next_renewal") or "")[:10], "%Y-%m-%d").date()
        except Exception:
            continue
        if today <= due < horizon:
            events.append((due, float(svc.get("price_monthly") or 0), "renewal"))

    three_months_ago = (today - timedelta(days=90)).isoformat()
    monthly_expense = 0.0
    for coll in ("expenses", "kas_kecil", "salaries", "sales_fees"):
        total = 0.0
        async for e in db[coll].find({"date": {"$gte": three_months_ago}}):
            total += float(e.get("amount") or 0)
        monthly_expense += total / 3.0
    daily_expense = monthly_expense / 30.0

    weekly = []
    cumulative = 0.0
    for w in range(13):
        ws = today + timedelta(days=w * 7)
        we = ws + timedelta(days=7)
        inflow = sum(a for (dt, a, _k) in events if ws <= dt < we)
        outflow = daily_expense * 7
        net = inflow - outflow
        cumulative += net
        weekly.append({"week_start": ws.isoformat(), "inflow": round(inflow, 2),
                       "outflow": round(outflow, 2), "net": round(net, 2),
                       "cumulative": round(cumulative, 2)})

    def _bucket(days: int) -> dict:
        end = today + timedelta(days=days)
        inflow = sum(a for (dt, a, _k) in events if dt < end)
        outflow = daily_expense * days
        return {"inflow": round(inflow, 2), "outflow": round(outflow, 2),
                "net": round(inflow - outflow, 2)}

    return {
        "as_of": today.isoformat(),
        "monthly_expense_run_rate": round(monthly_expense, 2),
        "buckets": {"d30": _bucket(30), "d60": _bucket(60), "d90": _bucket(90)},
        "weekly": weekly,
        "sources": {
            "unpaid_invoices": sum(1 for e in events if e[2] == "invoice"),
            "upcoming_renewals": sum(1 for e in events if e[2] == "renewal"),
        },
    }


def _rp(v) -> str:
    return "Rp " + f"{float(v or 0):,.0f}".replace(",", ".")


@router.get("/admin/finance/cashflow-forecast/export")
async def finance_cashflow_export(format: str = "pdf", token: str = "",
                                  admin=Depends(require_roles("admin", "finance"))):
    """Unduh proyeksi arus kas 30/60/90 hari sebagai PDF atau Excel (xlsx)."""
    db = await _get_db()
    f = await _compute_cashflow_forecast(db)
    b = f["buckets"]
    stamp = f["as_of"]
    fname = f"Proyeksi-Arus-Kas-{stamp}"

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyeksi Arus Kas"
        navy = PatternFill("solid", fgColor="0A2350")
        white_bold = Font(color="FFFFFF", bold=True)
        ws["A1"] = "Intercloud Digital - Proyeksi Arus Kas"
        ws["A1"].font = Font(bold=True, size=14, color="0A2350")
        ws["A2"] = f"Per {stamp} · run-rate beban {_rp(f['monthly_expense_run_rate'])}/bulan"
        ws["A2"].font = Font(italic=True, size=9, color="64748B")
        # Buckets summary
        ws["A4"] = "Ringkasan"
        ws["A4"].font = Font(bold=True, color="0A2350")
        hdr = ["Horizon", "Perkiraan Masuk", "Perkiraan Keluar", "Net"]
        for i, h in enumerate(hdr):
            c = ws.cell(row=5, column=1 + i, value=h)
            c.fill = navy
            c.font = white_bold
            c.alignment = Alignment(horizontal="center")
        for r, (lbl, key) in enumerate([("30 hari", "d30"), ("60 hari", "d60"), ("90 hari", "d90")], start=6):
            ws.cell(row=r, column=1, value=lbl)
            ws.cell(row=r, column=2, value=b[key]["inflow"])
            ws.cell(row=r, column=3, value=b[key]["outflow"])
            ws.cell(row=r, column=4, value=b[key]["net"])
        # Weekly detail
        ws["A10"] = "Rincian Mingguan (90 hari)"
        ws["A10"].font = Font(bold=True, color="0A2350")
        whdr = ["Minggu (mulai)", "Masuk", "Keluar", "Net", "Kumulatif"]
        for i, h in enumerate(whdr):
            c = ws.cell(row=11, column=1 + i, value=h)
            c.fill = navy
            c.font = white_bold
        for r, wk in enumerate(f["weekly"], start=12):
            ws.cell(row=r, column=1, value=wk["week_start"])
            ws.cell(row=r, column=2, value=wk["inflow"])
            ws.cell(row=r, column=3, value=wk["outflow"])
            ws.cell(row=r, column=4, value=wk["net"])
            ws.cell(row=r, column=5, value=wk["cumulative"])
        for col, w in zip("ABCDE", (18, 18, 18, 18, 18)):
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=6, max_row=11 + len(f["weekly"]), min_col=2, max_col=5):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        import io
        buf = io.BytesIO()
        wb.save(buf)
        return Response(content=buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})

    # default: PDF
    def _card(lbl, bk):
        net_color = "#059669" if bk["net"] >= 0 else "#dc2626"
        return (f"<td style='padding:14px;border:1px solid #e2e8f0;border-radius:10px;width:33%'>"
                f"<div style='font-size:10px;text-transform:uppercase;letter-spacing:1px;color:#64748b'>Proyeksi {lbl}</div>"
                f"<div style='font-size:20px;font-weight:800;color:{net_color};margin:4px 0'>{_rp(bk['net'])}</div>"
                f"<div style='font-size:11px;color:#334155'>Masuk {_rp(bk['inflow'])}<br/>Keluar {_rp(bk['outflow'])}</div></td>")
    rows = "".join(
        f"<tr><td style='padding:6px 8px;border-bottom:1px solid #eef2f7'>{w['week_start']}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;color:#059669'>{_rp(w['inflow'])}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;color:#dc2626'>{_rp(w['outflow'])}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;font-weight:700'>{_rp(w['net'])}</td>"
        f"<td style='padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:right;font-weight:700;color:#0a2350'>{_rp(w['cumulative'])}</td></tr>"
        for w in f["weekly"])
    html = f"""<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 20mm 16mm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; color:#0f172a; font-size:12px; }}
      h1 {{ color:#0a2350; font-size:18px; margin:0; }}
      .sub {{ color:#64748b; font-size:11px; margin:2px 0 18px; }}
      table {{ width:100%; border-collapse:collapse; }}
      th {{ background:#0a2350; color:#fff; padding:7px 8px; font-size:10px; text-transform:uppercase; letter-spacing:1px; text-align:right; }}
      th:first-child {{ text-align:left; }}
    </style></head><body>
      <h1>Proyeksi Arus Kas - Intercloud Digital</h1>
      <div class="sub">Per {stamp} · run-rate beban {_rp(f['monthly_expense_run_rate'])}/bulan ·
        {f['sources']['unpaid_invoices']} invoice belum lunas · {f['sources']['upcoming_renewals']} perpanjangan akan datang</div>
      <table style="margin-bottom:18px"><tr>{_card('30 hari', b['d30'])}{_card('60 hari', b['d60'])}{_card('90 hari', b['d90'])}</tr></table>
      <div style="font-weight:800;color:#0a2350;margin-bottom:6px">Rincian Mingguan (90 hari ke depan)</div>
      <table><thead><tr><th>Minggu (mulai)</th><th>Masuk</th><th>Keluar</th><th>Net</th><th>Kumulatif</th></tr></thead>
      <tbody>{rows}</tbody></table>
      <div style="margin-top:22px;font-size:9px;color:#94a3b8">Estimasi otomatis dari data invoice, jadwal perpanjangan layanan, dan rata-rata beban 3 bulan terakhir. Bukan angka final akuntansi.</div>
    </body></html>"""
    return Response(content=_render_pdf_bytes(html), media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'})


@router.get("/admin/finance/reports")
async def finance_finalized_reports(admin=Depends(require_roles("admin", "finance"))):
    """List all previously-generated monthly/annual reports (audit trail)."""
    db = await _get_db()
    docs = await db.finalized_reports.find({}).sort("period", -1).to_list(500)
    return [{
        "id": str(r["_id"]), "period": r["period"], "kind": r["kind"],
        "totals": r.get("totals", {}), "generated_at": _iso(r.get("generated_at", "")),
        "locked": _month_locked(r["period"]) if r["kind"] == "monthly" else True,
    } for r in docs]


# ============================================================
# CREDIT NOTES - refunds & adjustments applied to invoices
# ============================================================
def _credit_note_serialize(d: dict, invoice: dict | None = None, user: dict | None = None) -> dict:
    return {
        "id": str(d["_id"]),
        "number": d.get("number", ""),
        "invoice_id": str(d["invoice_id"]) if d.get("invoice_id") else None,
        "invoice_number": (invoice or {}).get("number", d.get("invoice_number", "")),
        "user_id": str(d["user_id"]) if d.get("user_id") else None,
        "user_name": (user or {}).get("name", d.get("user_name", "")),
        "user_email": (user or {}).get("email", d.get("user_email", "")),
        "amount": float(d.get("amount") or 0),
        "reason": d.get("reason") or "",
        "notes": d.get("notes") or "",
        "status": d.get("status") or "draft",   # draft / applied / cancelled
        "applied_at": d.get("applied_at"),
        "applied_by": str(d["applied_by"]) if d.get("applied_by") else None,
        "created_at": d.get("created_at", ""),
    }


async def _settle_invoice_from_credit(db, invoice: dict, request, admin) -> int:
    """If the total applied credit meets/exceeds the invoice total, flip the
    invoice to paid and reactivate suspended services - same effect as a
    Duitku webhook, but source=credit_note. Returns number of services
    reactivated. Idempotent: filters status != paid."""
    total_credit = await _sum_applied_credit(db, invoice["_id"])
    inv_total = float(invoice.get("total") or 0)
    if total_credit + 0.001 < inv_total:
        return 0
    r = await db.invoices.update_one(
        {"_id": invoice["_id"], "status": {"$ne": "paid"}},
        {"$set": {"status": "paid", "paid_at": _now(),
                  "payment_method": "credit_note",
                  "payment_ref": f"CN>={inv_total:.0f}"}},
    )
    if not r.modified_count:
        return 0
    # Reactivate services suspended for THIS invoice, mirroring the webhook logic
    inv_no = invoice.get("number", "")
    reactivated = 0
    import re as _re
    ors = [{"user_id": invoice["user_id"],
            "suspended_reason": {"$regex": f"invoice {_re.escape(inv_no)} overdue",
                                  "$options": "i"}}]
    if invoice.get("service_id"):
        try:
            ors.append({"_id": _oid(invoice["service_id"])})
        except Exception:
            pass
    res = await db.services.update_many(
        {"status": "suspended", "$or": ors},
        {"$set": {"status": "active", "reactivated_at": _now(),
                  "reactivated_reason": f"invoice {inv_no} settled via credit note"},
         "$unset": {"suspended_at": "", "suspended_reason": ""}},
    )
    reactivated = res.modified_count
    try:
        await _apply_pending_upgrade(db, invoice)
    except Exception:
        pass
    # Invoice order yang dilunasi credit note juga harus memicu auto-provisioning
    # (sama seperti webhook Duitku / admin mark-paid) agar service muncul di klien.
    try:
        await _provision_order_from_invoice(db, invoice)
    except Exception:
        logging.getLogger("portal.finance").exception(
            "auto-provision setelah pelunasan credit note gagal")
    await log_audit(db, actor=admin, action="invoice.settled_by_credit", category="billing",
                    target_type="invoice", target_id=str(invoice["_id"]),
                    target_label=inv_no,
                    metadata={"total_credit": total_credit,
                              "invoice_total": inv_total,
                              "reactivated_services": reactivated},
                    severity="warning", request=request)
    return reactivated


@router.get("/admin/credit-notes")
async def credit_notes_list(admin=Depends(require_roles("admin", "finance")),
                            invoice_id: Optional[str] = None,
                            user_id: Optional[str] = None,
                            status: Optional[str] = None):
    db = await _get_db()
    query: dict = {}
    if invoice_id:
        try: query["invoice_id"] = ObjectId(invoice_id)
        except Exception: query["invoice_id"] = None
    if user_id:
        try: query["user_id"] = ObjectId(user_id)
        except Exception: query["user_id"] = None
    if status:
        query["status"] = status
    docs = await db.credit_notes.find(query).sort("created_at", -1).to_list(500)
    # bulk-fetch related invoices + users
    inv_ids = [d["invoice_id"] for d in docs if d.get("invoice_id")]
    user_ids = [d["user_id"] for d in docs if d.get("user_id")]
    invs = {i["_id"]: i for i in await db.invoices.find({"_id": {"$in": inv_ids}}).to_list(500)}
    users = {u["_id"]: u for u in await db.users.find({"_id": {"$in": user_ids}}).to_list(500)}
    return [_credit_note_serialize(d, invs.get(d.get("invoice_id")), users.get(d.get("user_id")))
            for d in docs]


@router.post("/admin/credit-notes")
async def credit_notes_create(payload: dict, request: Request, admin=Depends(require_roles("admin", "finance"))):
    """Issue a credit note against an invoice.

    Body: `{invoice_id, amount, reason, notes?, auto_apply?: bool}`.
    If `auto_apply` is true, immediately applies (which may settle the invoice)."""
    db = await _get_db()
    inv_id_raw = payload.get("invoice_id") or ""
    try:
        inv_oid = ObjectId(inv_id_raw)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid invoice_id")
    invoice = await db.invoices.find_one({"_id": inv_oid})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be > 0")
    reason = (payload.get("reason") or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Reason is required")
    inv_total = float(invoice.get("total") or 0)
    already_applied = await _sum_applied_credit(db, inv_oid)
    if amount + already_applied > inv_total + 0.001:
        raise HTTPException(status_code=400,
                            detail=f"Credit ({amount:.0f}) + already applied ({already_applied:.0f}) exceeds invoice total ({inv_total:.0f}).")
    user = await db.users.find_one({"_id": invoice["user_id"]}) or {}
    number = await _next_number(db, "credit_notes", "CN")
    doc = {
        "number": number,
        "invoice_id": inv_oid,
        "invoice_number": invoice.get("number", ""),
        "user_id": invoice["user_id"],
        "user_name": user.get("name", ""),
        "user_email": user.get("email", ""),
        "amount": amount,
        "reason": reason,
        "notes": payload.get("notes") or "",
        "status": "draft",
        "created_at": _now(),
        "created_by": ObjectId(admin["id"]),
    }
    r = await db.credit_notes.insert_one(doc)
    doc["_id"] = r.inserted_id
    await log_audit(db, actor=admin, action="credit_note.create", category="billing",
                    target_type="credit_note", target_id=str(r.inserted_id),
                    target_label=number,
                    after={"invoice_number": invoice.get("number"), "amount": amount,
                           "reason": reason},
                    severity="warning", request=request)
    # Auto-apply if requested (single POST for the "quick refund" workflow)
    if bool(payload.get("auto_apply")):
        await _apply_credit_note_inner(db, doc, admin, request)
        doc = await db.credit_notes.find_one({"_id": r.inserted_id})
    return _credit_note_serialize(doc, invoice, user)


async def _apply_credit_note_inner(db, cn: dict, admin, request):
    """Mark the credit note applied, then settle the invoice if fully covered."""
    if cn.get("status") == "applied":
        return {"already_applied": True}
    await db.credit_notes.update_one(
        {"_id": cn["_id"], "status": {"$ne": "applied"}},
        {"$set": {"status": "applied", "applied_at": _now(),
                  "applied_by": ObjectId(admin["id"])}},
    )
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]})
    reactivated = 0
    if invoice:
        reactivated = await _settle_invoice_from_credit(db, invoice, request, admin)
    await log_audit(db, actor=admin, action="credit_note.apply", category="billing",
                    target_type="credit_note", target_id=str(cn["_id"]),
                    target_label=cn.get("number", ""),
                    metadata={"invoice_number": (invoice or {}).get("number"),
                              "amount": cn.get("amount"),
                              "reactivated_services": reactivated},
                    severity="warning", request=request)
    return {"applied": True, "reactivated_services": reactivated}


@router.post("/admin/credit-notes/{cid}/apply")
async def credit_notes_apply(cid: str, request: Request, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    return await _apply_credit_note_inner(db, cn, admin, request)


@router.post("/admin/credit-notes/{cid}/cancel")
async def credit_notes_cancel(cid: str, request: Request, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    if cn.get("status") == "applied":
        raise HTTPException(status_code=400,
                            detail="Cannot cancel an already-applied credit note (invoice may be paid).")
    await db.credit_notes.update_one({"_id": cn["_id"]},
                                     {"$set": {"status": "cancelled",
                                               "cancelled_at": _now()}})
    await log_audit(db, actor=admin, action="credit_note.cancel", category="billing",
                    target_type="credit_note", target_id=cid,
                    target_label=cn.get("number", ""),
                    severity="info", request=request)
    return {"ok": True}


@router.put("/admin/credit-notes/{cid}")
async def credit_notes_update(cid: str, payload: dict, request: Request,
                              admin=Depends(require_roles("admin", "finance"))):
    """Edit amount/reason/notes credit note. CN yang sudah applied pada invoice
    LUNAS tidak bisa diedit (ubah status invoice dulu)."""
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]}) if cn.get("invoice_id") else None
    if cn.get("status") == "applied" and (invoice or {}).get("status") == "paid":
        raise HTTPException(status_code=400,
                            detail="Credit note sudah diterapkan dan invoice sudah LUNAS - "
                                   "ubah status invoice ke unpaid dulu bila perlu koreksi.")
    upd, before = {}, {}
    if "amount" in payload:
        amount = float(payload.get("amount") or 0)
        if amount <= 0:
            raise HTTPException(status_code=400, detail="Amount must be > 0")
        if invoice:
            inv_total = float(invoice.get("total") or 0)
            other_applied = await _sum_applied_credit(db, cn["invoice_id"]) - (
                float(cn.get("amount") or 0) if cn.get("status") == "applied" else 0)
            if amount + other_applied > inv_total + 0.001:
                raise HTTPException(status_code=400,
                                    detail=f"Credit ({amount:.0f}) + credit lain ({other_applied:.0f}) "
                                           f"melebihi total invoice ({inv_total:.0f}).")
        before["amount"], upd["amount"] = float(cn.get("amount") or 0), amount
    if "reason" in payload:
        reason = (payload.get("reason") or "").strip()
        if not reason:
            raise HTTPException(status_code=400, detail="Reason is required")
        before["reason"], upd["reason"] = cn.get("reason"), reason
    if "notes" in payload:
        before["notes"], upd["notes"] = cn.get("notes"), payload.get("notes") or ""
    if not upd:
        raise HTTPException(status_code=400, detail="Nothing to update")
    upd["updated_at"] = _now()
    await db.credit_notes.update_one({"_id": cn["_id"]}, {"$set": upd})
    await log_audit(db, actor=admin, action="credit_note.update", category="billing",
                    target_type="credit_note", target_id=cid,
                    target_label=cn.get("number", ""), before=before, after=upd,
                    severity="warning", request=request)
    # Bila CN applied dan setelah edit menutup total invoice -> settle otomatis
    if cn.get("status") == "applied" and invoice and invoice.get("status") != "paid":
        try:
            await _settle_invoice_from_credit(db, invoice, request, admin)
        except Exception:
            pass
    doc = await db.credit_notes.find_one({"_id": cn["_id"]})
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]}) if cn.get("invoice_id") else None
    user = await db.users.find_one({"_id": cn["user_id"]}) if cn.get("user_id") else None
    return _credit_note_serialize(doc, invoice, user)


@router.delete("/admin/credit-notes/{cid}")
async def credit_notes_delete(cid: str, request: Request,
                              admin=Depends(require_roles("admin", "finance"))):
    """Hapus credit note. CN applied pada invoice LUNAS tidak bisa dihapus."""
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]}) if cn.get("invoice_id") else None
    if cn.get("status") == "applied" and (invoice or {}).get("status") == "paid":
        raise HTTPException(status_code=400,
                            detail="Credit note sudah diterapkan dan invoice sudah LUNAS - "
                                   "ubah status invoice ke unpaid dulu bila ingin menghapus.")
    await db.credit_notes.delete_one({"_id": cn["_id"]})
    await log_audit(db, actor=admin, action="credit_note.delete", category="billing",
                    target_type="credit_note", target_id=cid,
                    target_label=cn.get("number", ""),
                    before={"invoice_number": cn.get("invoice_number"),
                            "amount": cn.get("amount"), "status": cn.get("status")},
                    severity="warning", request=request)
    return {"ok": True}


@router.get("/admin/credit-notes/{cid}")
async def credit_notes_detail(cid: str, admin=Depends(require_roles("admin", "finance"))):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]}) if cn.get("invoice_id") else None
    user = await db.users.find_one({"_id": cn["user_id"]}) if cn.get("user_id") else None
    return _credit_note_serialize(cn, invoice, user)


# ---- Client-visible credit notes (read-only, own only) ----
@router.get("/client/credit-notes")
async def client_credit_notes(user=Depends(get_current_user)):
    db = await _get_db()
    docs = await db.credit_notes.find({"user_id": ObjectId(user["id"])}).sort("created_at", -1).to_list(200)
    inv_ids = [d["invoice_id"] for d in docs if d.get("invoice_id")]
    invs = {i["_id"]: i for i in await db.invoices.find({"_id": {"$in": inv_ids}}).to_list(500)}
    return [_credit_note_serialize(d, invs.get(d.get("invoice_id"))) for d in docs]


# ---- PDF render (HTML or PDF, mirrors invoice endpoint) ----
@router.get("/documents/credit-note/{cid}")
async def render_credit_note_pdf(cid: str, format: str = "html", user=Depends(get_current_user)):
    db = await _get_db()
    cn = await db.credit_notes.find_one({"_id": _oid(cid)})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    # Access: owner or staff
    if user["role"] == "client" and str(cn["user_id"]) != str(user["id"]):
        raise HTTPException(status_code=403, detail="Not your credit note")
    invoice = await db.invoices.find_one({"_id": cn["invoice_id"]}) if cn.get("invoice_id") else None
    u = await db.users.find_one({"_id": cn["user_id"]}) or {}
    branding = await _get_branding_dict(db)
    html = _credit_note_html(
        cn=cn,
        invoice=invoice,
        billed_to=u,
        for_pdf=(format == "pdf"),
        logo_url=branding["logo_dark"],
    )
    if format == "pdf":
        pdf_bytes = _render_pdf_bytes(html)
        filename = f"CreditNote-{cn.get('number','credit-note')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    token = user.get("_token", "")
    html = html.replace("{TOKEN_PLACEHOLDER}", token)
    return HTMLResponse(content=html)


def _credit_note_html(*, cn: dict, invoice: dict | None, billed_to: dict,
                       for_pdf: bool, logo_url: str) -> str:
    number = cn.get("number", "")
    amount = float(cn.get("amount") or 0)
    status = (cn.get("status") or "draft").lower()
    status_colors = {"draft": "#94a3b8", "applied": "#059669", "cancelled": "#dc2626"}
    status_color = status_colors.get(status, "#64748b")
    inv_no = (invoice or {}).get("number", cn.get("invoice_number", ""))
    inv_total = float((invoice or {}).get("total") or 0)
    generated_on = _long_date(datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    actions_bar = "" if for_pdf else (
        f'<div class="actions">'
        f'<button onclick="window.print()">Print</button>'
        f'<a class="dl" href="?format=pdf&token={{TOKEN_PLACEHOLDER}}">Download PDF</a>'
        f'</div>'
    )
    bill_to = _billing_block(billed_to)
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Credit Note #{number}</title>
<style>
  @page {{ size: A4; margin: 14mm 14mm 16mm 14mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
         color:#334155; margin:0; padding:0; background:#f1f5f9; font-size:12px; line-height:1.5; }}
  .paper {{ background:#fff; padding:34px 40px 30px; max-width:800px; margin:20px auto;
           position:relative; box-shadow:0 6px 30px rgba(2,6,23,.08); }}
  .header {{ display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:22px;
             padding-bottom:18px; border-bottom:1px solid #e2e8f0; }}
  .logo {{ height:64px; width:auto; }}
  h1 {{ margin:0; font-size:26px; letter-spacing:.02em; color:#0a2540; }}
  .sub {{ font-size:11px; color:#64748b; margin-top:4px; }}
  .status {{ display:inline-block; margin-top:8px; padding:6px 14px; border-radius:999px;
             color:#fff; font-weight:700; letter-spacing:.08em; font-size:10px;
             text-transform:uppercase; background:{status_color}; }}
  .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:24px; margin-bottom:22px; }}
  .card {{ background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:14px 16px; }}
  .card h3 {{ margin:0 0 8px 0; font-size:11px; color:#475569; letter-spacing:.14em;
              text-transform:uppercase; font-weight:700; }}
  .amount-box {{ background:#0a2540; color:#fff; padding:18px 22px; border-radius:12px;
                 display:flex; justify-content:space-between; align-items:center; margin-bottom:22px; }}
  .amount-box .lbl {{ font-size:11px; letter-spacing:.14em; text-transform:uppercase;
                       color:#cbd5e1; margin-bottom:4px; }}
  .amount-box .val {{ font-size:26px; font-weight:800; color:#f5b120; }}
  .notes {{ background:#fef3c7; border:1px solid #fbbf24; padding:12px 14px; border-radius:8px;
            margin-top:16px; color:#78350f; font-size:12px; line-height:1.6; }}
  .footer {{ margin-top:28px; padding-top:14px; border-top:1px dashed #cbd5e1;
             font-size:10.5px; color:#64748b; line-height:1.6; }}
  .actions {{ text-align:center; margin:16px 0; }}
  .actions button, .actions .dl {{ display:inline-block; margin:0 6px; padding:8px 18px;
       background:#0a2540; color:#fff; border:0; border-radius:6px; font-size:12px;
       text-decoration:none; cursor:pointer; }}
  .actions .dl {{ background:#f5b120; color:#0a2540; }}
  @media print {{ body{{background:#fff}} .paper{{box-shadow:none;margin:0;max-width:100%}} .actions{{display:none}} }}
</style></head><body>
{actions_bar}
<div class="paper">
  <div class="header">
    <div>
      <img src="{logo_url}" alt="Intercloud" class="logo" onerror="this.style.display='none'"/>
      <div class="sub">PT Intercloud Digital Inovasi<br>Cyber 1 Building, Kuningan · Jakarta 12950</div>
    </div>
    <div style="text-align:right">
      <h1>Credit Note</h1>
      <div class="sub"><b>#{number}</b></div>
      <div class="status">{status}</div>
      <div class="sub" style="margin-top:6px">Issued {_long_date((cn.get('created_at') or '')[:10])}</div>
    </div>
  </div>
  <div class="grid">
    <div class="card">
      <h3>Issued To</h3>
      {bill_to}
    </div>
    <div class="card">
      <h3>Reference Invoice</h3>
      <div style="font-weight:700; color:#0a2540; font-size:14px">#{inv_no or '-'}</div>
      <div class="sub">Invoice total: {_idr(inv_total)}</div>
      {'<div class="sub">Applied on: ' + _long_date((cn.get('applied_at') or '')[:10]) + '</div>' if cn.get('applied_at') else ''}
    </div>
  </div>
  <div class="amount-box">
    <div>
      <div class="lbl">Credit Amount</div>
      <div class="sub" style="color:#cbd5e1; margin-top:2px">{cn.get('reason', '')}</div>
    </div>
    <div class="val">{_idr(amount)}</div>
  </div>
  {'<div class="notes"><b>Notes.</b> ' + (cn.get('notes') or '').replace(chr(10), '<br>') + '</div>' if cn.get('notes') else ''}
  <div class="footer">
    <b>Reason</b><br>{cn.get('reason', '')}
    <br><br>
    Generated on {generated_on}. Credit notes reduce the outstanding amount of the referenced invoice.
    If the total applied credit meets or exceeds the invoice total, the invoice is automatically
    marked as paid and any suspended services are reactivated.
  </div>
</div>
</body></html>"""


def _billing_block(u: dict) -> str:
    """Small helper - same shape used by invoice PDF sidebar."""
    lines = []
    if u.get("attention"): lines.append(f"<b>{u['attention']}</b>")
    elif u.get("name"):    lines.append(f"<b>{u['name']}</b>")
    if u.get("company"):   lines.append(u["company"])
    if u.get("address_line1"): lines.append(u["address_line1"])
    if u.get("address_line2"): lines.append(u["address_line2"])
    city_line = " ".join(x for x in [u.get("city"), u.get("province"), u.get("postal_code")] if x)
    if city_line: lines.append(city_line)
    if u.get("country"): lines.append(u["country"])
    if u.get("email"):   lines.append(f"<span style='color:#64748b'>{u['email']}</span>")
    if u.get("phone"):   lines.append(f"<span style='color:#64748b'>{u['phone']}</span>")
    return "<br>".join(lines) if lines else "<i>No billing address on file</i>"
